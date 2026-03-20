"""
数据导出服务

提供资质矩阵、工时报表、排班计划的 Excel 导出能力。
使用 openpyxl 生成 .xlsx 文件。
"""
import io
import logging
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.lab_personnel.models import (
    LabStaffProfile, MethodQualification, MethodQualLevel,
)
from apps.lab_personnel.models_worktime import WorkTimeSummary
from apps.lab_personnel.models_scheduling import ShiftSlot

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADER_FILL = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)

LEVEL_FILLS = {
    MethodQualLevel.LEARNING: PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid'),
    MethodQualLevel.PROBATION: PatternFill(start_color='BFDBFE', end_color='BFDBFE', fill_type='solid'),
    MethodQualLevel.INDEPENDENT: PatternFill(start_color='BBF7D0', end_color='BBF7D0', fill_type='solid'),
    MethodQualLevel.MENTOR: PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid'),
}


def export_qualification_matrix() -> io.BytesIO:
    """导出资质矩阵为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '资质矩阵'

    profiles = LabStaffProfile.objects.filter(
        is_active=True, staff__is_deleted=False,
    ).select_related('staff').order_by('staff__name')

    from apps.resource.models_detection_method import DetectionMethodTemplate
    methods = list(DetectionMethodTemplate.objects.filter(
        status='active', is_deleted=False,
    ).order_by('code'))

    ws.cell(row=1, column=1, value='人员/方法').font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.column_dimensions['A'].width = 15

    for j, method in enumerate(methods, 2):
        cell = ws.cell(row=1, column=j, value=f'{method.code}\n{method.name}')
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = 14

    for i, profile in enumerate(profiles, 2):
        name_cell = ws.cell(row=i, column=1, value=profile.staff.name)
        name_cell.font = Font(bold=True)
        name_cell.border = THIN_BORDER

        for j, method in enumerate(methods, 2):
            qual = MethodQualification.objects.filter(
                staff=profile.staff, method=method,
            ).first()

            cell = ws.cell(row=i, column=j)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            if qual:
                cell.value = qual.get_level_display()
                fill = LEVEL_FILLS.get(qual.level)
                if fill:
                    cell.fill = fill
            else:
                cell.value = '-'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_worktime(week_start_date: date = None) -> io.BytesIO:
    """导出工时报表为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '工时报表'

    if not week_start_date:
        today = date.today()
        week_start_date = today - timedelta(days=today.weekday())

    headers = ['人员', '周起始日', '总工时', '工单工时', '培训工时', '其他工时', '可用工时', '利用率(%)']
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = 14

    summaries = WorkTimeSummary.objects.filter(
        week_start_date=week_start_date,
    ).select_related('staff').order_by('staff__name')

    for i, s in enumerate(summaries, 2):
        values = [
            s.staff.name,
            s.week_start_date.isoformat(),
            float(s.total_hours),
            float(s.workorder_hours),
            float(s.training_hours),
            float(s.other_hours),
            float(s.available_hours),
            float(s.utilization_rate),
        ]
        for j, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            if j >= 3:
                cell.number_format = '0.0'

        if float(s.utilization_rate) > 90:
            ws.cell(row=i, column=8).fill = PatternFill(
                start_color='FCA5A5', end_color='FCA5A5', fill_type='solid',
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_schedule(week_start_date: date = None) -> io.BytesIO:
    """导出排班计划为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '排班计划'

    if not week_start_date:
        today = date.today()
        week_start_date = today - timedelta(days=today.weekday())

    week_end_date = week_start_date + timedelta(days=6)

    headers = ['人员', '日期', '开始', '结束', '计划工时', '项目', '任务描述', '确认状态']
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25

    slots = ShiftSlot.objects.filter(
        shift_date__gte=week_start_date,
        shift_date__lte=week_end_date,
    ).select_related('staff').order_by('shift_date', 'start_time')

    for i, slot in enumerate(slots, 2):
        values = [
            slot.staff.name,
            slot.shift_date.isoformat(),
            slot.start_time.strftime('%H:%M'),
            slot.end_time.strftime('%H:%M'),
            float(slot.planned_hours),
            slot.project_name or '-',
            slot.tasks_description or '-',
            slot.get_confirm_status_display(),
        ]
        for j, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
