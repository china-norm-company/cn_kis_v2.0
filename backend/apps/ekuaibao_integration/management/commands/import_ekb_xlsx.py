"""
import_ekb_xlsx — 将 NAS 易快报管理后台导出的 Excel 历史数据导入 V2

NAS 文件路径: /private/tmp/nas_cn_kis/易快报历史数据/
文件特征:
  - Sheet: 报销单
  - 第1行: 分组标题（"单据信息"）
  - 第2行: 列名（125 列）
  - 第3行起: 数据行（B 开头的报销单号，2018-2026）

与 V1 API 数据的关系:
  - API 数据（V1）覆盖 2018-2026，包含 flowId，审批链完整
  - Excel 数据覆盖 2018-2026，无 flowId，但有银行账号、收款方等支付信息
  - 去重规则: 通过单号（ekuaibao_no / 单号列）匹配
  - 本命令优先用 Excel 数据填充 API 数据缺失的支付字段，不重复创建记录

用法:
  python manage.py import_ekb_xlsx [--file PATH] [--dry-run] [--enrich-only]

参数:
  --file: Excel 文件路径，默认查找 NAS 挂载目录
  --dry-run: 只统计，不写入
  --enrich-only: 只补充已有 ExpenseRequest 的支付字段，不新建记录
"""
import hashlib
import json
import logging
import os
from datetime import date as _date_type
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone

logger = logging.getLogger('cn_kis.ekuaibao.xlsx_import')

NAS_BASE = Path('/private/tmp/nas_cn_kis/易快报历史数据')
DEFAULT_SEARCH_PATHS = [
    NAS_BASE,
    Path('/mnt/nas/易快报历史数据'),
    Path('/opt/cn-kis-v2/backend/data/ekuaibao_backup'),
]

# Excel 列名 → 系统字段映射
COLUMN_MAP = {
    '单号':           'bill_no',
    '标题':           'title',
    '提交人名称':     'submitter_name',
    '提交人部门名称': 'dept_name',
    '提交日期':       'submit_date',
    '审批状态':       'approval_status_raw',
    '单据模板名称':   'template_name',
    '报销金额':       'amount',
    '报销日期':       'expense_date',
    '支付金额':       'payment_amount',
    '支付方式':       'payment_method',
    '支付日期':       'payment_date',
    '企业账户名称':   'company_account',
    '付款人名称':     'payer_name',
    '收款帐号':       'payee_account',
    '户名':           'payee_name',
    '开户行':         'payee_bank',
    '开户网点':       'payee_bank_branch',
    '开户行所在省':   'payee_province',
    '开户行所在市':   'payee_city',
    '账户类型':       'payee_account_type',
    '费用承担部门名称': 'cost_department',
    '项目名称':       'project_name',
    '项目编码':       'project_code',
    '客户名称':       'client_name',
    '关联申请名称':   'linked_req_name',
    '发票张数/张':    'invoice_count',
    '描述':           'description',
    '消费事由':       'consumption_reason',
    '凭证号':         'voucher_no',
    '会计期间':       'account_period',
    '单据审核通过日期': 'approval_date',
    '核销金额':       'write_off_amount',
    '税率/%':         'tax_rate',
    '税额':           'tax_amount',
    '不计税金额':     'pre_tax_amount',
    '发票类型名称':   'invoice_type',
    '业务板块名称':   'sector_name',
    '板块名称':       'sector_name2',
    '版块名称':       'sector_name3',
    '项目执行台账名称': 'ledger_name',
    '项目执行台账新名称': 'ledger_new_name',
    '客户名称部门':   'client_name_dept',
    '项目标的':       'project_target',
    '利润率/％':      'profit_rate',
    '样本数量/人':    'sample_count',
    '客户经理名称':   'account_manager',
    '项目经理':       'project_manager',
    '委托来源名称':   'commission_source',
    '结算方式名称':   'settlement_method',
    '申请单号':       'requisition_no',
    '当前审批人':     'current_approver',
}

APPROVAL_STATUS_MAP = {
    '已完成': 'reimbursed',
    '审批中': 'submitted',
    '已驳回': 'rejected',
    '待支付': 'approved',
    '待审批': 'submitted',
    '草稿':   'draft',
}


def _find_excel_file() -> Path | None:
    for base in DEFAULT_SEARCH_PATHS:
        if not base.exists():
            continue
        for f in base.glob('易快报全量历史费用*.xlsx'):
            return f
    return None


def _parse_date(val) -> _date_type | None:
    if not val:
        return None
    if isinstance(val, _date_type):
        return val
    try:
        from datetime import datetime as _dt
        if hasattr(val, 'date'):
            return val.date()
        s = str(val).strip()[:10]
        return _dt.strptime(s, '%Y-%m-%d').date()
    except Exception:
        return None


def _safe_decimal(val) -> Decimal | None:
    if val is None or str(val).strip() in ('', 'None'):
        return None
    try:
        return Decimal(str(val)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


def _safe_int(val) -> int:
    try:
        return int(float(str(val or 0)))
    except (ValueError, TypeError):
        return 0


class Command(BaseCommand):
    help = '将 NAS 易快报 Excel 历史数据导入/补充到 V2 系统'

    def add_arguments(self, parser):
        parser.add_argument('--file', default=None,
                            help='Excel 文件路径（默认自动查找 NAS 挂载）')
        parser.add_argument('--dry-run', action='store_true',
                            help='只统计，不写入')
        parser.add_argument('--enrich-only', action='store_true',
                            help='只补充已有记录的支付字段，不新建 EkbRawRecord 和 ExpenseRequest')
        parser.add_argument('--batch-size', type=int, default=200,
                            help='每批处理条数（默认 200）')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('❌ 缺少 openpyxl，请运行: pip install openpyxl')
            return

        from apps.ekuaibao_integration.models import EkbImportBatch, EkbRawRecord, EkbBatchStatus
        from apps.finance.models_expense import ExpenseRequest

        dry_run = options['dry_run']
        enrich_only = options['enrich_only']
        batch_size = options['batch_size']

        # 1. 定位 Excel 文件
        file_path = Path(options['file']) if options['file'] else _find_excel_file()
        if not file_path or not file_path.exists():
            self.stderr.write(f'❌ 找不到 Excel 文件，请用 --file 指定路径')
            return
        self.stdout.write(f'📂 Excel 文件: {file_path}')

        # 2. 加载 Excel
        self.stdout.write('加载 Excel（可能需要 30 秒）...')
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.stdout.write(f'工作表: {ws.title}, 共 {ws.max_row:,} 行 × {ws.max_column} 列')

        # 解析列名（第2行）
        headers = [cell.value for cell in ws[2]]
        col_idx = {}
        for i, h in enumerate(headers):
            if h and h in COLUMN_MAP:
                col_idx[COLUMN_MAP[h]] = i

        self.stdout.write(f'识别到 {len(col_idx)} 个有效列映射')

        # 3. 创建或找到 xlsx 导入批次
        batch_no = f'xlsx_import_{timezone.now().strftime("%Y%m%d_%H%M%S")}'

        if not dry_run and not enrich_only:
            batch = EkbImportBatch.objects.create(
                batch_no=batch_no,
                phase='xlsx_import',
                status=EkbBatchStatus.COLLECTING,
                modules=['expense_xlsx'],
                operator='auto',
                notes=f'NAS Excel 历史数据导入: {file_path.name}',
            )
            self.stdout.write(f'批次号: {batch_no}')
        else:
            batch = None

        # 4. 逐行处理
        stats = {'new_raw': 0, 'enriched': 0, 'skipped': 0, 'no_match': 0, 'failed': 0}
        total_data_rows = ws.max_row - 2  # 减去 2 行表头

        def get_cell(row, field_name):
            idx = col_idx.get(field_name)
            if idx is None:
                return None
            return row[idx]

        rows_iter = ws.iter_rows(min_row=3, values_only=True)
        buffer = []

        for row_num, row in enumerate(rows_iter, start=3):
            bill_no = get_cell(row, 'bill_no')
            if not bill_no:
                continue
            bill_no = str(bill_no).strip()
            if not bill_no.startswith('B'):
                stats['skipped'] += 1
                continue

            # 构造结构化行数据
            row_data = {
                'bill_no':            bill_no,
                'title':              get_cell(row, 'title') or '',
                'submitter_name':     get_cell(row, 'submitter_name') or '',
                'dept_name':          get_cell(row, 'dept_name') or '',
                'submit_date':        _parse_date(get_cell(row, 'submit_date')),
                'approval_status':    APPROVAL_STATUS_MAP.get(
                                          str(get_cell(row, 'approval_status_raw') or ''), 'submitted'),
                'template_name':      get_cell(row, 'template_name') or '',
                'amount':             _safe_decimal(get_cell(row, 'amount')),
                'payment_amount':     _safe_decimal(get_cell(row, 'payment_amount')),
                'payment_date':       _parse_date(get_cell(row, 'payment_date')),
                'payment_method':     get_cell(row, 'payment_method') or '',
                'payee_account':      str(get_cell(row, 'payee_account') or '').strip(),
                'payee_name':         get_cell(row, 'payee_name') or '',
                'payee_bank':         get_cell(row, 'payee_bank') or '',
                'payee_bank_branch':  get_cell(row, 'payee_bank_branch') or '',
                'payee_province':     get_cell(row, 'payee_province') or '',
                'payee_city':         get_cell(row, 'payee_city') or '',
                'payee_account_type': get_cell(row, 'payee_account_type') or '',
                'cost_department':    get_cell(row, 'cost_department') or '',
                'project_name':       get_cell(row, 'project_name') or '',
                'project_code':       get_cell(row, 'project_code') or '',
                'client_name':        get_cell(row, 'client_name') or '',
                'linked_req_name':    get_cell(row, 'linked_req_name') or '',
                'invoice_count':      _safe_int(get_cell(row, 'invoice_count')),
                'description':        get_cell(row, 'description') or '',
                'voucher_no':         get_cell(row, 'voucher_no') or '',
                'account_period':     get_cell(row, 'account_period') or '',
                'sector_name':        (get_cell(row, 'sector_name') or
                                       get_cell(row, 'sector_name2') or
                                       get_cell(row, 'sector_name3') or ''),
                'ledger_name':        (get_cell(row, 'ledger_new_name') or
                                       get_cell(row, 'ledger_name') or ''),
                'project_target':     _safe_decimal(get_cell(row, 'project_target')),
            }

            # 完整的125列原始数据（JSON）
            excel_raw = {headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row) if headers[i]}

            buffer.append((bill_no, row_data, excel_raw))

            # 每 batch_size 条处理一次
            if len(buffer) >= batch_size:
                _process_batch(self, buffer, batch, dry_run, enrich_only, stats)
                buffer.clear()
                pct = (row_num - 2) / total_data_rows * 100
                self.stdout.write(
                    f'  进度: {row_num-2:,}/{total_data_rows:,} ({pct:.1f}%) '
                    f'新增:{stats["new_raw"]:,} 补充:{stats["enriched"]:,} '
                    f'跳过:{stats["skipped"]:,}',
                    ending='\r'
                )
                self.stdout.flush()

        # 处理剩余
        if buffer:
            _process_batch(self, buffer, batch, dry_run, enrich_only, stats)

        # 更新批次状态
        if batch and not dry_run:
            batch.status = EkbBatchStatus.COLLECTED
            batch.total_records = stats['new_raw']
            batch.collected_at = timezone.now()
            batch.save()

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Excel 导入完成:\n'
            f'  新建 EkbRawRecord: {stats["new_raw"]:,}\n'
            f'  补充支付字段:       {stats["enriched"]:,}\n'
            f'  无匹配跳过:         {stats["no_match"]:,}\n'
            f'  跳过(非B单):        {stats["skipped"]:,}\n'
            f'  失败:               {stats["failed"]:,}'
        ))

        if not enrich_only and not dry_run:
            self.stdout.write(
                '\n提示: 运行以下命令将 Excel 数据注入到 ExpenseRequest:\n'
                f'  python manage.py export_ekuaibao_full --inject-only'
            )


def _process_batch(cmd, buffer, batch, dry_run, enrich_only, stats):
    """处理一批行数据"""
    from apps.ekuaibao_integration.models import EkbRawRecord
    from apps.finance.models_expense import ExpenseRequest

    bill_nos = [b[0] for b in buffer]

    # 查找已有的 ExpenseRequest（按单号）
    existing_expenses = {
        e.ekuaibao_no: e
        for e in ExpenseRequest.objects.filter(ekuaibao_no__in=bill_nos)
    }

    # 查找已有的 EkbRawRecord（按 ekb_id = bill_no）
    existing_raws = {
        r.ekb_id
        for r in EkbRawRecord.objects.filter(
            ekb_id__in=bill_nos, module='expense_xlsx'
        )
    }

    for bill_no, row_data, excel_raw in buffer:
        try:
            # ── 补充已有 ExpenseRequest 的支付字段 ────────────────────────
            expense = existing_expenses.get(bill_no)
            if expense:
                _enrich_expense(expense, row_data, excel_raw, dry_run)
                stats['enriched'] += 1
            else:
                stats['no_match'] += 1

            # ── 新建 EkbRawRecord（保留完整 Excel 原始行）─────────────────
            if not enrich_only and bill_no not in existing_raws and batch:
                if not dry_run:
                    raw_json = {
                        '_source': 'xlsx_export',
                        '_bill_no': bill_no,
                        **row_data,
                        '_excel_raw': excel_raw,
                    }
                    # JSON 序列化时处理 date/Decimal
                    def json_default(obj):
                        if hasattr(obj, 'isoformat'):
                            return obj.isoformat()
                        if hasattr(obj, '__str__'):
                            return str(obj)
                        raise TypeError

                    raw_str = json.dumps(raw_json, default=json_default, ensure_ascii=False)
                    checksum = hashlib.sha256(raw_str.encode()).hexdigest()

                    EkbRawRecord.objects.create(
                        batch=batch,
                        module='expense_xlsx',
                        ekb_id=bill_no,
                        raw_data=json.loads(raw_str),
                        scraped_at=timezone.now(),
                        checksum=checksum,
                        injection_status='pending',
                    )
                    existing_raws.add(bill_no)
                stats['new_raw'] += 1

        except Exception as e:
            logger.warning('处理 %s 失败: %s', bill_no, e)
            stats['failed'] += 1


def _enrich_expense(expense, row_data: dict, excel_raw: dict, dry_run: bool):
    """用 Excel 数据补充 ExpenseRequest 的支付/收款字段"""
    fields_updated = []

    def _set_if_empty(field, value):
        current = getattr(expense, field, None)
        if value and (current is None or str(current).strip() in ('', 'None', '0', '0.00')):
            setattr(expense, field, value)
            fields_updated.append(field)

    _set_if_empty('submit_date',       row_data.get('submit_date'))
    _set_if_empty('payment_amount',    row_data.get('payment_amount'))
    _set_if_empty('payment_date',      row_data.get('payment_date'))
    _set_if_empty('payment_method',    row_data.get('payment_method'))
    _set_if_empty('payee_name',        row_data.get('payee_name'))
    _set_if_empty('payee_account',     row_data.get('payee_account'))
    _set_if_empty('payee_bank',        row_data.get('payee_bank'))
    _set_if_empty('payee_bank_branch', row_data.get('payee_bank_branch'))
    _set_if_empty('payee_province',    row_data.get('payee_province'))
    _set_if_empty('payee_city',        row_data.get('payee_city'))
    _set_if_empty('payee_account_type',row_data.get('payee_account_type'))
    _set_if_empty('invoice_count',     row_data.get('invoice_count') or None)
    _set_if_empty('voucher_no',        row_data.get('voucher_no'))
    _set_if_empty('account_period',    row_data.get('account_period'))
    _set_if_empty('cost_department',   row_data.get('cost_department'))
    _set_if_empty('client_name',       row_data.get('client_name'))

    # 如果没有记录 excel_raw_data，就补充进去
    if not expense.excel_raw_data and excel_raw:
        expense.excel_raw_data = excel_raw
        fields_updated.append('excel_raw_data')

    if fields_updated and not dry_run:
        expense.save(update_fields=fields_updated)
