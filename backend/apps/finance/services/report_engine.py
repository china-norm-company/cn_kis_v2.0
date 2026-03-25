"""
报表引擎：结构化数据采集 + Excel/PDF 导出

支持 6 种报表模板：
- project_profit: 项目损益报表
- monthly_operation: 月度经营报表
- quarterly_operation: 季度经营报表
- annual_operation: 年度经营报表
- client_statement: 客户对账报表
- ar_aging: 应收账龄报表
"""
import io
import logging
from decimal import Decimal
from datetime import date

from dateutil.relativedelta import relativedelta
from django.db.models import Sum, DecimalField
from django.db.models.functions import TruncMonth, Coalesce

from apps.finance.models import (
    Contract, ContractStatus,
    Invoice, InvoiceStatus, PaymentRecord, PaymentRecordStatus,
    PaymentPlan, PaymentPlanStatus,
    CostRecord, CostRecordStatus, CostType,
    ProjectBudget, BudgetItem,
)

logger = logging.getLogger(__name__)
ZERO = Decimal('0')


def collect_project_profit_report(protocol_id: int, period_start: date = None, period_end: date = None) -> dict:
    """项目损益报表数据"""
    today = date.today()
    period_start = period_start or today.replace(month=1, day=1)
    period_end = period_end or today

    contracts = Contract.objects.filter(
        protocol_id=protocol_id, is_deleted=False,
    )
    contract_amount = contracts.aggregate(
        total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField())
    )['total']
    project_name = contracts.first().project if contracts.exists() else ''

    invoiced = Invoice.objects.filter(
        contract__protocol_id=protocol_id, is_deleted=False,
    ).exclude(status=InvoiceStatus.DRAFT).aggregate(
        total=Coalesce(Sum('total'), ZERO, output_field=DecimalField())
    )['total']

    received = PaymentRecord.objects.filter(
        protocol_id=protocol_id, status=PaymentRecordStatus.CONFIRMED,
    ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']

    deferred = max(received - invoiced, ZERO)

    costs = CostRecord.objects.filter(
        protocol_id=protocol_id, status=CostRecordStatus.CONFIRMED,
    )

    cost_breakdown = {}
    for ct in CostType:
        amount = costs.filter(cost_type=ct.value).aggregate(
            total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField())
        )['total']
        cost_breakdown[ct.value] = float(amount)

    total_cost = sum(cost_breakdown.values())
    gross_profit = float(contract_amount) - total_cost
    gross_margin = (gross_profit / float(contract_amount) * 100) if contract_amount > 0 else 0

    budget = ProjectBudget.objects.filter(protocol_id=protocol_id).order_by('-version').first()
    budget_comparison = []
    if budget:
        for bi in BudgetItem.objects.filter(budget=budget).select_related('category'):
            budget_comparison.append({
                'category': bi.category.name,
                'budget': float(bi.budget_amount),
                'actual': float(bi.actual_amount),
                'variance': float(bi.actual_amount - bi.budget_amount),
            })

    monthly_trend = (
        costs
        .annotate(month=TruncMonth('cost_date'))
        .values('month')
        .annotate(cost=Sum('amount'))
        .order_by('month')
    )
    inv_trend = (
        Invoice.objects.filter(
            contract__protocol_id=protocol_id, is_deleted=False,
        ).exclude(status=InvoiceStatus.DRAFT)
        .annotate(month=TruncMonth('invoice_date'))
        .values('month')
        .annotate(revenue=Sum('total'))
        .order_by('month')
    )
    cost_map = {r['month'].isoformat(): float(r['cost']) for r in monthly_trend}
    rev_map = {r['month'].isoformat(): float(r['revenue']) for r in inv_trend}
    all_months = sorted(set(list(cost_map.keys()) + list(rev_map.keys())))
    trend = [{
        'month': m,
        'revenue': rev_map.get(m, 0),
        'cost': cost_map.get(m, 0),
        'margin': rev_map.get(m, 0) - cost_map.get(m, 0),
    } for m in all_months]

    return {
        'report_type': 'project_profit',
        'project_name': project_name,
        'protocol_id': protocol_id,
        'period': {'start': str(period_start), 'end': str(period_end)},
        'revenue': {
            'contract_amount': float(contract_amount),
            'invoiced': float(invoiced),
            'received': float(received),
            'deferred': float(deferred),
        },
        'cost_breakdown': cost_breakdown,
        'total_cost': total_cost,
        'profit': {
            'gross_profit': round(gross_profit, 2),
            'gross_margin': round(gross_margin, 2),
        },
        'budget_comparison': budget_comparison,
        'trend': trend,
    }


def collect_monthly_operation_report(year: int, month: int) -> dict:
    """月度经营报表数据"""
    period_start = date(year, month, 1)
    period_end = period_start + relativedelta(months=1) - relativedelta(days=1)

    prev_start = period_start - relativedelta(months=1)
    prev_end = period_start - relativedelta(days=1)
    yoy_start = period_start - relativedelta(years=1)
    yoy_end = period_end - relativedelta(years=1)

    ytd_start = date(year, 1, 1)

    def _period_metrics(start, end):
        invoiced = Invoice.objects.filter(
            is_deleted=False, invoice_date__gte=start, invoice_date__lte=end,
        ).exclude(status=InvoiceStatus.DRAFT).aggregate(
            total=Coalesce(Sum('total'), ZERO, output_field=DecimalField())
        )['total']

        received = PaymentRecord.objects.filter(
            status=PaymentRecordStatus.CONFIRMED,
            payment_date__gte=start, payment_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']

        cost = CostRecord.objects.filter(
            status=CostRecordStatus.CONFIRMED,
            cost_date__gte=start, cost_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']

        signed = Contract.objects.filter(
            is_deleted=False, signed_date__gte=start, signed_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']

        return {
            'invoiced': float(invoiced),
            'received': float(received),
            'cost': float(cost),
            'profit': float(invoiced - cost),
            'margin': round(float((invoiced - cost) / invoiced * 100), 2) if invoiced > 0 else 0,
            'signed': float(signed),
        }

    current = _period_metrics(period_start, period_end)
    previous = _period_metrics(prev_start, prev_end)
    yoy = _period_metrics(yoy_start, yoy_end)
    ytd = _period_metrics(ytd_start, period_end)

    return {
        'report_type': 'monthly_operation',
        'year': year,
        'month': month,
        'period': {'start': str(period_start), 'end': str(period_end)},
        'current': current,
        'previous': previous,
        'yoy': yoy,
        'ytd': ytd,
        'active_contracts': Contract.objects.filter(
            is_deleted=False, status=ContractStatus.ACTIVE,
        ).count(),
        'overdue_amount': float(PaymentPlan.objects.filter(
            status=PaymentPlanStatus.OVERDUE,
        ).aggregate(total=Coalesce(Sum('remaining_amount'), ZERO, output_field=DecimalField()))['total']),
    }


def collect_client_statement(client_id: int, period_start: date = None, period_end: date = None) -> dict:
    """客户对账报表"""
    today = date.today()
    period_start = period_start or today.replace(month=1, day=1)
    period_end = period_end or today

    contracts = Contract.objects.filter(
        client_id=client_id, is_deleted=False,
    ).order_by('-create_time')

    client_name = contracts.first().client if contracts.exists() else ''

    contract_list = []
    for c in contracts:
        invoiced = Invoice.objects.filter(
            contract=c, is_deleted=False,
        ).exclude(status=InvoiceStatus.DRAFT).aggregate(
            total=Coalesce(Sum('total'), ZERO, output_field=DecimalField())
        )['total']

        received = PaymentRecord.objects.filter(
            protocol_id=c.protocol_id, status=PaymentRecordStatus.CONFIRMED,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']

        contract_list.append({
            'contract_no': c.code,
            'project': c.project,
            'amount': float(c.amount),
            'status': c.status,
            'invoiced': float(invoiced),
            'received': float(received),
            'balance': float(invoiced - received),
        })

    total_contract = sum(c['amount'] for c in contract_list)
    total_invoiced = sum(c['invoiced'] for c in contract_list)
    total_received = sum(c['received'] for c in contract_list)

    return {
        'report_type': 'client_statement',
        'client_id': client_id,
        'client_name': client_name,
        'period': {'start': str(period_start), 'end': str(period_end)},
        'contracts': contract_list,
        'summary': {
            'total_contract': total_contract,
            'total_invoiced': total_invoiced,
            'total_received': total_received,
            'balance': total_invoiced - total_received,
        },
    }


def collect_quarterly_operation_report(year: int, quarter: int) -> dict:
    """季度经营报表数据"""
    month_start = (quarter - 1) * 3 + 1
    period_start = date(year, month_start, 1)
    period_end = period_start + relativedelta(months=3) - relativedelta(days=1)

    prev_q_start = period_start - relativedelta(months=3)
    prev_q_end = period_start - relativedelta(days=1)
    yoy_start = period_start - relativedelta(years=1)
    yoy_end = period_end - relativedelta(years=1)

    def _period_metrics(start, end):
        invoiced = Invoice.objects.filter(
            is_deleted=False, invoice_date__gte=start, invoice_date__lte=end,
        ).exclude(status=InvoiceStatus.DRAFT).aggregate(
            total=Coalesce(Sum('total'), ZERO, output_field=DecimalField())
        )['total']
        received = PaymentRecord.objects.filter(
            status=PaymentRecordStatus.CONFIRMED,
            payment_date__gte=start, payment_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        cost = CostRecord.objects.filter(
            status=CostRecordStatus.CONFIRMED,
            cost_date__gte=start, cost_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        signed = Contract.objects.filter(
            is_deleted=False, signed_date__gte=start, signed_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        return {
            'invoiced': float(invoiced), 'received': float(received),
            'cost': float(cost), 'profit': float(invoiced - cost),
            'margin': round(float((invoiced - cost) / invoiced * 100), 2) if invoiced > 0 else 0,
            'signed': float(signed),
        }

    # Monthly breakdown within the quarter
    monthly = []
    for i in range(3):
        m_start = period_start + relativedelta(months=i)
        m_end = m_start + relativedelta(months=1) - relativedelta(days=1)
        m_data = _period_metrics(m_start, m_end)
        m_data['month'] = m_start.strftime('%Y-%m')
        monthly.append(m_data)

    return {
        'report_type': 'quarterly_operation',
        'year': year, 'quarter': quarter,
        'period': {'start': str(period_start), 'end': str(period_end)},
        'current': _period_metrics(period_start, period_end),
        'previous_quarter': _period_metrics(prev_q_start, prev_q_end),
        'yoy': _period_metrics(yoy_start, yoy_end),
        'monthly_breakdown': monthly,
        'active_contracts': Contract.objects.filter(
            is_deleted=False, status=ContractStatus.ACTIVE,
        ).count(),
        'overdue_amount': float(PaymentPlan.objects.filter(
            status=PaymentPlanStatus.OVERDUE,
        ).aggregate(total=Coalesce(Sum('remaining_amount'), ZERO, output_field=DecimalField()))['total']),
    }


def collect_annual_operation_report(year: int) -> dict:
    """年度经营报表数据"""
    period_start = date(year, 1, 1)
    period_end = date(year, 12, 31)
    prev_start = date(year - 1, 1, 1)
    prev_end = date(year - 1, 12, 31)

    def _period_metrics(start, end):
        invoiced = Invoice.objects.filter(
            is_deleted=False, invoice_date__gte=start, invoice_date__lte=end,
        ).exclude(status=InvoiceStatus.DRAFT).aggregate(
            total=Coalesce(Sum('total'), ZERO, output_field=DecimalField())
        )['total']
        received = PaymentRecord.objects.filter(
            status=PaymentRecordStatus.CONFIRMED,
            payment_date__gte=start, payment_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        cost = CostRecord.objects.filter(
            status=CostRecordStatus.CONFIRMED,
            cost_date__gte=start, cost_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        signed = Contract.objects.filter(
            is_deleted=False, signed_date__gte=start, signed_date__lte=end,
        ).aggregate(total=Coalesce(Sum('amount'), ZERO, output_field=DecimalField()))['total']
        return {
            'invoiced': float(invoiced), 'received': float(received),
            'cost': float(cost), 'profit': float(invoiced - cost),
            'margin': round(float((invoiced - cost) / invoiced * 100), 2) if invoiced > 0 else 0,
            'signed': float(signed),
        }

    # Quarterly breakdown
    quarterly = []
    for q in range(1, 5):
        q_start = date(year, (q - 1) * 3 + 1, 1)
        q_end = q_start + relativedelta(months=3) - relativedelta(days=1)
        q_data = _period_metrics(q_start, q_end)
        q_data['quarter'] = f'Q{q}'
        quarterly.append(q_data)

    return {
        'report_type': 'annual_operation',
        'year': year,
        'period': {'start': str(period_start), 'end': str(period_end)},
        'current': _period_metrics(period_start, period_end),
        'previous_year': _period_metrics(prev_start, prev_end),
        'quarterly_breakdown': quarterly,
        'contracts_signed': Contract.objects.filter(
            is_deleted=False, signed_date__gte=period_start, signed_date__lte=period_end,
        ).count(),
    }


def collect_ar_aging_report(as_of: date = None) -> dict:
    """应收账龄报表"""
    as_of = as_of or date.today()

    plans = PaymentPlan.objects.filter(
        status__in=[PaymentPlanStatus.PENDING, PaymentPlanStatus.PARTIAL, PaymentPlanStatus.OVERDUE],
    ).select_related()

    buckets = {'current': ZERO, '1_30': ZERO, '31_60': ZERO, '61_90': ZERO, 'over_90': ZERO}
    client_detail = {}

    for plan in plans:
        days = (as_of - plan.planned_date).days
        remaining = plan.remaining_amount

        if days <= 0:
            bucket = 'current'
        elif days <= 30:
            bucket = '1_30'
        elif days <= 60:
            bucket = '31_60'
        elif days <= 90:
            bucket = '61_90'
        else:
            bucket = 'over_90'

        buckets[bucket] += remaining

        client = plan.client_name or 'Unknown'
        if client not in client_detail:
            client_detail[client] = {
                'client': client,
                'current': ZERO, '1_30': ZERO, '31_60': ZERO, '61_90': ZERO, 'over_90': ZERO,
                'total': ZERO,
            }
        client_detail[client][bucket] += remaining
        client_detail[client]['total'] += remaining

    total = sum(float(v) for v in buckets.values())

    # Bad debt estimate using probability-weighted approach
    bad_debt_rates = {'current': 0.01, '1_30': 0.05, '31_60': 0.15, '61_90': 0.30, 'over_90': 0.50}
    bad_debt_estimate = sum(float(buckets[b]) * bad_debt_rates[b] for b in buckets)

    clients = sorted(client_detail.values(), key=lambda x: float(x['total']), reverse=True)
    for c in clients:
        for k in ['current', '1_30', '31_60', '61_90', 'over_90', 'total']:
            c[k] = float(c[k])

    return {
        'report_type': 'ar_aging',
        'as_of': str(as_of),
        'summary': {k: float(v) for k, v in buckets.items()},
        'total': total,
        'bad_debt_estimate': round(bad_debt_estimate, 2),
        'clients': clients,
    }


def export_report_excel(report_data: dict) -> bytes:
    """导出报表为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.error('openpyxl 未安装, 无法导出 Excel')
        raise ImportError('需要安装 openpyxl: pip install openpyxl')

    wb = openpyxl.Workbook()
    ws = wb.active

    report_type = report_data.get('report_type', 'unknown')
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_text = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def _write_header(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def _write_row(ws, row, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0.00'

    if report_type == 'project_profit':
        ws.title = '项目损益'
        ws.cell(row=1, column=1, value=f'项目损益报表 - {report_data.get("project_name", "")}').font = title_font
        ws.cell(row=2, column=1, value=f'期间: {report_data["period"]["start"]} ~ {report_data["period"]["end"]}')

        row = 4
        ws.cell(row=row, column=1, value='收入信息').font = header_font
        row += 1
        revenue = report_data['revenue']
        for label, key in [('合同金额', 'contract_amount'), ('已开票', 'invoiced'),
                           ('已回款', 'received'), ('递延收入', 'deferred')]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=revenue[key]).number_format = '#,##0.00'
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='成本明细').font = header_font
        row += 1
        type_labels = {'labor': '人工', 'material': '材料', 'equipment': '设备',
                       'outsource': '外包', 'travel': '差旅', 'other': '其他'}
        for ct, amount in report_data['cost_breakdown'].items():
            ws.cell(row=row, column=1, value=type_labels.get(ct, ct))
            ws.cell(row=row, column=2, value=amount).number_format = '#,##0.00'
            row += 1
        ws.cell(row=row, column=1, value='成本合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_data['total_cost']).number_format = '#,##0.00'
        row += 2

        ws.cell(row=row, column=1, value='盈利').font = header_font
        row += 1
        ws.cell(row=row, column=1, value='毛利')
        ws.cell(row=row, column=2, value=report_data['profit']['gross_profit']).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value='毛利率(%)')
        ws.cell(row=row, column=2, value=report_data['profit']['gross_margin']).number_format = '0.00'

        if report_data.get('budget_comparison'):
            row += 2
            ws.cell(row=row, column=1, value='预算对比').font = header_font
            row += 1
            _write_header(ws, row, ['科目', '预算金额', '实际金额', '偏差'])
            row += 1
            for item in report_data['budget_comparison']:
                _write_row(ws, row, [item['category'], item['budget'], item['actual'], item['variance']])
                row += 1

    elif report_type == 'monthly_operation':
        ws.title = '月度经营'
        ws.cell(row=1, column=1, value=f'{report_data["year"]}年{report_data["month"]}月 经营报表').font = title_font

        row = 3
        _write_header(ws, row, ['指标', '当月', '上月', '环比', '去年同期', '同比', '本年累计'])
        row += 1

        cur = report_data['current']
        prev = report_data['previous']
        yoy = report_data['yoy']
        ytd = report_data['ytd']

        for label, key in [('开票收入', 'invoiced'), ('回款金额', 'received'),
                           ('成本', 'cost'), ('利润', 'profit'), ('签约金额', 'signed')]:
            mom = round((cur[key] - prev[key]) / prev[key] * 100, 2) if prev[key] != 0 else None
            yoy_chg = round((cur[key] - yoy[key]) / yoy[key] * 100, 2) if yoy[key] != 0 else None
            _write_row(ws, row, [
                label, cur[key], prev[key],
                f'{mom}%' if mom is not None else '-',
                yoy[key],
                f'{yoy_chg}%' if yoy_chg is not None else '-',
                ytd[key],
            ])
            row += 1

    elif report_type == 'client_statement':
        ws.title = '客户对账'
        ws.cell(row=1, column=1, value=f'客户对账单 - {report_data.get("client_name", "")}').font = title_font
        ws.cell(row=2, column=1, value=f'期间: {report_data["period"]["start"]} ~ {report_data["period"]["end"]}')

        row = 4
        _write_header(ws, row, ['合同编号', '项目名称', '合同金额', '已开票', '已回款', '余额', '状态'])
        row += 1
        for c in report_data['contracts']:
            _write_row(ws, row, [
                c['contract_no'], c['project'], c['amount'],
                c['invoiced'], c['received'], c['balance'], c['status'],
            ])
            row += 1

        row += 1
        s = report_data['summary']
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=3, value=s['total_contract']).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=s['total_invoiced']).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=s['total_received']).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=s['balance']).number_format = '#,##0.00'

    elif report_type == 'quarterly_operation':
        ws.title = '季度经营'
        ws.cell(row=1, column=1, value=f'{report_data["year"]}年Q{report_data["quarter"]} 季度经营报表').font = title_font

        row = 3
        _write_header(ws, row, ['指标', '本季度', '上季度', '环比', '去年同期', '同比'])
        row += 1
        cur = report_data['current']
        prev = report_data['previous_quarter']
        yoy = report_data['yoy']
        for label, key in [('开票收入', 'invoiced'), ('回款金额', 'received'),
                           ('成本', 'cost'), ('利润', 'profit'), ('签约金额', 'signed')]:
            qoq = round((cur[key] - prev[key]) / prev[key] * 100, 2) if prev[key] != 0 else None
            yoy_chg = round((cur[key] - yoy[key]) / yoy[key] * 100, 2) if yoy[key] != 0 else None
            _write_row(ws, row, [label, cur[key], prev[key],
                f'{qoq}%' if qoq is not None else '-',
                yoy[key], f'{yoy_chg}%' if yoy_chg is not None else '-'])
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='月度明细').font = header_font
        row += 1
        _write_header(ws, row, ['月份', '开票收入', '回款', '成本', '利润', '毛利率(%)'])
        row += 1
        for m in report_data.get('monthly_breakdown', []):
            _write_row(ws, row, [m['month'], m['invoiced'], m['received'],
                                 m['cost'], m['profit'], m['margin']])
            row += 1

    elif report_type == 'annual_operation':
        ws.title = '年度经营'
        ws.cell(row=1, column=1, value=f'{report_data["year"]}年度经营报表').font = title_font

        row = 3
        _write_header(ws, row, ['指标', '本年', '上年', '同比'])
        row += 1
        cur = report_data['current']
        prev = report_data['previous_year']
        for label, key in [('开票收入', 'invoiced'), ('回款金额', 'received'),
                           ('成本', 'cost'), ('利润', 'profit'), ('签约金额', 'signed')]:
            yoy_chg = round((cur[key] - prev[key]) / prev[key] * 100, 2) if prev[key] != 0 else None
            _write_row(ws, row, [label, cur[key], prev[key],
                f'{yoy_chg}%' if yoy_chg is not None else '-'])
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='季度明细').font = header_font
        row += 1
        _write_header(ws, row, ['季度', '开票收入', '回款', '成本', '利润', '毛利率(%)'])
        row += 1
        for q in report_data.get('quarterly_breakdown', []):
            _write_row(ws, row, [q['quarter'], q['invoiced'], q['received'],
                                 q['cost'], q['profit'], q['margin']])
            row += 1

    elif report_type == 'ar_aging':
        ws.title = '应收账龄'
        ws.cell(row=1, column=1, value=f'应收账龄分析报表 截至 {report_data.get("as_of", "")}').font = title_font

        row = 3
        ws.cell(row=row, column=1, value='账龄汇总').font = header_font
        row += 1
        _write_header(ws, row, ['账龄段', '金额'])
        row += 1
        labels = {'current': '未到期', '1_30': '1-30天', '31_60': '31-60天',
                  '61_90': '61-90天', 'over_90': '90天以上'}
        for key, label in labels.items():
            _write_row(ws, row, [label, report_data['summary'].get(key, 0)])
            row += 1
        ws.cell(row=row, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_data.get('total', 0)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value='坏账准备估算')
        ws.cell(row=row, column=2, value=report_data.get('bad_debt_estimate', 0)).number_format = '#,##0.00'

        row += 2
        ws.cell(row=row, column=1, value='客户明细').font = header_font
        row += 1
        _write_header(ws, row, ['客户', '未到期', '1-30天', '31-60天', '61-90天', '90天以上', '合计'])
        row += 1
        for c in report_data.get('clients', []):
            _write_row(ws, row, [c['client'], c['current'], c['1_30'],
                                 c['31_60'], c['61_90'], c['over_90'], c['total']])
            row += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_report_pdf(report_data: dict) -> bytes:
    """导出报表为 PDF（基于 reportlab）"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics  # noqa: F401
        from reportlab.pdfbase.ttfonts import TTFont  # noqa: F401
    except ImportError:
        logger.error('reportlab 未安装, 无法导出 PDF')
        raise ImportError('需要安装 reportlab: pip install reportlab')

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title_CN', parent=styles['Title'], fontSize=16)
    heading_style = ParagraphStyle('Heading_CN', parent=styles['Heading2'], fontSize=12)
    normal_style = styles['Normal']

    report_type = report_data.get('report_type', '')

    if report_type == 'project_profit':
        elements.append(Paragraph(f"Project Profit Report - {report_data.get('project_name', '')}", title_style))
        elements.append(Spacer(1, 10 * mm))

        rev = report_data['revenue']
        data = [
            ['Item', 'Amount'],
            ['Contract Amount', f"{rev['contract_amount']:,.2f}"],
            ['Invoiced', f"{rev['invoiced']:,.2f}"],
            ['Received', f"{rev['received']:,.2f}"],
            ['Deferred', f"{rev['deferred']:,.2f}"],
        ]
        t = Table(data, colWidths=[120, 120])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 8 * mm))

        elements.append(Paragraph('Cost Breakdown', heading_style))
        cost_data = [['Type', 'Amount']]
        for ct, amt in report_data['cost_breakdown'].items():
            cost_data.append([ct, f'{amt:,.2f}'])
        cost_data.append(['Total', f"{report_data['total_cost']:,.2f}"])
        t2 = Table(cost_data, colWidths=[120, 120])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 8 * mm))

        p = report_data['profit']
        elements.append(Paragraph(
            f"Gross Profit: {p['gross_profit']:,.2f} | Margin: {p['gross_margin']:.2f}%",
            heading_style,
        ))

    elif report_type == 'monthly_operation':
        elements.append(Paragraph(
            f"{report_data['year']}-{report_data['month']:02d} Monthly Report", title_style,
        ))
        elements.append(Spacer(1, 10 * mm))

        cur = report_data['current']
        prev = report_data['previous']
        data = [['Metric', 'Current', 'Previous', 'MoM', 'YTD']]
        ytd = report_data['ytd']
        for label, key in [('Invoiced', 'invoiced'), ('Received', 'received'),
                           ('Cost', 'cost'), ('Profit', 'profit'), ('Signed', 'signed')]:
            mom = f"{(cur[key] - prev[key]) / prev[key] * 100:.1f}%" if prev[key] != 0 else '-'
            data.append([label, f'{cur[key]:,.2f}', f'{prev[key]:,.2f}', mom, f'{ytd[key]:,.2f}'])

        t = Table(data, colWidths=[80, 90, 90, 60, 90])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(t)

    elif report_type == 'quarterly_operation':
        elements.append(Paragraph(
            f"{report_data['year']} Q{report_data['quarter']} Quarterly Report", title_style,
        ))
        elements.append(Spacer(1, 10 * mm))
        cur = report_data['current']
        prev = report_data.get('previous_quarter', {})
        data = [['Metric', 'Current Q', 'Previous Q', 'QoQ']]
        for label, key in [('Invoiced', 'invoiced'), ('Received', 'received'),
                           ('Cost', 'cost'), ('Profit', 'profit'), ('Signed', 'signed')]:
            prev_val = prev.get(key, 0)
            qoq = f"{(cur[key] - prev_val) / prev_val * 100:.1f}%" if prev_val != 0 else '-'
            data.append([label, f'{cur[key]:,.2f}', f'{prev_val:,.2f}', qoq])
        t = Table(data, colWidths=[80, 90, 90, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(t)

    elif report_type == 'annual_operation':
        elements.append(Paragraph(
            f"{report_data['year']} Annual Report", title_style,
        ))
        elements.append(Spacer(1, 10 * mm))
        cur = report_data['current']
        prev = report_data.get('previous_year', {})
        data = [['Metric', 'Current Year', 'Previous Year', 'YoY']]
        for label, key in [('Invoiced', 'invoiced'), ('Received', 'received'),
                           ('Cost', 'cost'), ('Profit', 'profit'), ('Signed', 'signed')]:
            prev_val = prev.get(key, 0)
            yoy = f"{(cur[key] - prev_val) / prev_val * 100:.1f}%" if prev_val != 0 else '-'
            data.append([label, f'{cur[key]:,.2f}', f'{prev_val:,.2f}', yoy])
        t = Table(data, colWidths=[80, 100, 100, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(t)

    elif report_type == 'ar_aging':
        elements.append(Paragraph(f"AR Aging Report as of {report_data.get('as_of', '')}", title_style))
        elements.append(Spacer(1, 10 * mm))
        data = [['Aging Bucket', 'Amount']]
        labels = {'current': 'Current', '1_30': '1-30 days', '31_60': '31-60 days',
                  '61_90': '61-90 days', 'over_90': 'Over 90 days'}
        for key, label in labels.items():
            data.append([label, f"{report_data['summary'].get(key, 0):,.2f}"])
        data.append(['Total', f"{report_data.get('total', 0):,.2f}"])
        data.append(['Bad Debt Estimate', f"{report_data.get('bad_debt_estimate', 0):,.2f}"])
        t = Table(data, colWidths=[120, 120])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(t)

    else:
        elements.append(Paragraph('Financial Report', title_style))
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph(str(report_data), normal_style))

    doc.build(elements)
    return output.getvalue()
