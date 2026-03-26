"""
人事能力服务

封装资质管理、胜任力模型、能力评估、培训跟踪的业务逻辑。

飞书集成：
- 创建/更新培训记录时同步到飞书 GCP 培训日历（对应 FEISHU_NATIVE_SETUP.md 5.2）
"""
import os
import logging
import io
import uuid
from typing import Optional
from datetime import date, datetime, timedelta
from django.db.models import Q
from django.utils import timezone
import openpyxl
from apps.identity.filters import filter_queryset_by_scope
from .models import (
    Staff,
    CompetencyModel,
    Assessment,
    Training,
    StaffArchive,
    StaffContract,
    HrStaffCertificate,
    StaffChangeLog,
    StaffExitRecord,
    RecruitmentDemand,
    RecruitmentCandidate,
    PerformanceCycle,
    PerformanceRecord,
    PayrollRecord,
    IncentiveRecord,
    CultureActivity,
    EngagementPulse,
    HrCollaborationSnapshot,
)

logger = logging.getLogger(__name__)

FEISHU_CALENDAR_TRAINING_ID = os.getenv('FEISHU_CALENDAR_TRAINING_ID', '')
HR_ADMIN_ROLES = {'admin', 'superadmin', 'general_manager', 'hr'}
HR_DEPARTMENT_ROLES = {'hr_manager'}


def _paginate_queryset(qs, page: int = 1, page_size: int = 20) -> dict:
    total = qs.count()
    offset = (page - 1) * page_size
    items = list(qs[offset:offset + page_size])
    return {'items': items, 'total': total, 'page': page, 'page_size': page_size}


def _get_role_names(account) -> set:
    if not account or not getattr(account, 'id', None):
        return set()
    try:
        from apps.identity.authz import get_authz_service
        return set(get_authz_service().get_account_role_names(account.id) or set())
    except Exception as e:
        logger.warning('获取 HR 数据域角色失败: account_id=%s error=%s', getattr(account, 'id', None), e)
        return set()


def _get_linked_staff(account) -> Optional[Staff]:
    if not account or not getattr(account, 'id', None):
        return None
    return (
        Staff.objects.filter(is_deleted=False)
        .filter(Q(account_fk_id=account.id) | Q(account_id=account.id))
        .order_by('id')
        .first()
    )


def _get_hr_scope_mode(account) -> str:
    """
    HR 模块数据域：
    - admin/superadmin/general_manager: global
    - hr_manager: department
    - 其余角色: personal
    """
    roles = _get_role_names(account)
    if roles & HR_ADMIN_ROLES:
        return 'global'
    if roles & HR_DEPARTMENT_ROLES:
        return 'department'
    return 'personal'


def _get_department_scope(account) -> str:
    staff = _get_linked_staff(account)
    return (staff.department or '').strip() if staff else ''


def _apply_hr_scope(queryset, account, *, personal_field: Optional[str] = None, department_field: Optional[str] = None):
    if not account:
        return queryset

    scope_mode = _get_hr_scope_mode(account)
    if scope_mode == 'global':
        return queryset

    if scope_mode == 'department':
        department = _get_department_scope(account)
        if not department or not department_field:
            return queryset.none()
        return queryset.filter(**{f'{department_field}__icontains': department})

    if personal_field:
        return filter_queryset_by_scope(
            queryset,
            account,
            scope_override='personal',
            field_mapping={'personal': personal_field},
        )
    return queryset.none()


def _ensure_department_write_allowed(account, department: str) -> None:
    if not account:
        return
    if _get_hr_scope_mode(account) != 'department':
        return
    allowed_department = _get_department_scope(account)
    if not allowed_department:
        raise PermissionError('当前账号未绑定部门，不能执行 HR 写操作')
    if (department or '').strip() != allowed_department:
        raise PermissionError(f'仅允许操作本部门数据: {allowed_department}')


# ============================================================================
# 资质管理
# ============================================================================
def list_staff(
    account=None,
    department: str = None,
    gcp_status: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = Staff.objects.filter(is_deleted=False).select_related('archive')
    qs = _apply_hr_scope(qs, account, personal_field='account_fk_id', department_field='department')
    if department:
        qs = qs.filter(department__icontains=department)
    if gcp_status:
        qs = qs.filter(gcp_status=gcp_status)
    return _paginate_queryset(qs, page, page_size)


def get_staff(staff_id: int, account=None) -> Optional[Staff]:
    qs = Staff.objects.filter(id=staff_id, is_deleted=False).select_related('archive')
    qs = _apply_hr_scope(qs, account, personal_field='account_fk_id', department_field='department')
    return qs.first()


def create_staff(name: str, position: str, department: str,
                 employee_no: str = '', email: str = '', phone: str = '',
                 gcp_cert: str = '', gcp_expiry: date = None,
                 gcp_status: str = 'none', other_certs: str = '',
                 training_status: str = '未开始', account=None,
                 feishu_open_id: str = '') -> Staff:
    _ensure_department_write_allowed(account, department)
    resolved_open_id = (feishu_open_id or '').strip()
    if not resolved_open_id:
        # Staff.feishu_open_id 为唯一字段；Excel/手工新增无飞书ID时生成稳定占位，避免 '' 触发唯一约束
        resolved_open_id = f'manual_{uuid.uuid4().hex[:24]}'
    staff = Staff.objects.create(
        name=name, position=position, department=department,
        employee_no=employee_no, email=email, phone=phone,
        gcp_cert=gcp_cert, gcp_expiry=gcp_expiry,
        gcp_status=gcp_status, other_certs=other_certs,
        training_status=training_status,
        feishu_open_id=resolved_open_id,
    )
    StaffArchive.objects.get_or_create(
        staff=staff,
        defaults={
            'department': staff.department or '',
            'employment_status': 'active',
            'employment_type': 'full_time',
            'sync_source': 'manual',
        },
    )
    return staff


def update_staff(staff_id: int, account=None, **kwargs) -> Optional[Staff]:
    s = get_staff(staff_id, account=account)
    if not s:
        return None
    target_department = kwargs.get('department')
    if target_department is not None:
        _ensure_department_write_allowed(account, target_department)
    for k, v in kwargs.items():
        if v is not None and hasattr(s, k):
            setattr(s, k, v)
    s.save()
    return s


def delete_staff(staff_id: int, account=None) -> bool:
    s = get_staff(staff_id, account=account)
    if not s:
        return False
    s.is_deleted = True
    s.save(update_fields=['is_deleted', 'update_time'])
    return True


def get_staff_stats(account=None) -> dict:
    from django.db.models import Count
    qs = Staff.objects.filter(is_deleted=False)
    qs = _apply_hr_scope(qs, account, personal_field='account_fk_id', department_field='department')
    by_gcp = qs.values('gcp_status').annotate(count=Count('id'))
    return {
        'by_gcp_status': {item['gcp_status']: item['count'] for item in by_gcp},
        'total': qs.count(),
    }


def sync_staff_from_feishu_contacts() -> dict:
    """通过飞书通讯录同步员工主数据（Staff + StaffArchive）。"""
    from apps.hr.services.sync_service import FeishuContactSyncService
    return FeishuContactSyncService.sync_all()


def import_staff_rows(items: list, account=None) -> dict:
    """
    批量导入员工基础信息（按工号/姓名+部门匹配 upsert）。
    仅处理主数据：Staff + StaffArchive。
    """
    created = 0
    updated = 0
    skipped = 0

    for row in items or []:
        name = str((row or {}).get('name', '')).strip()
        department = str((row or {}).get('department', '')).strip()
        position = str((row or {}).get('position', '')).strip()
        if not name or not department:
            skipped += 1
            continue

        _ensure_department_write_allowed(account, department)
        employee_no = str((row or {}).get('employee_no', '')).strip()
        email = str((row or {}).get('email', '')).strip()
        phone = str((row or {}).get('phone', '')).strip()

        staff = None
        if employee_no:
            staff = Staff.objects.filter(employee_no=employee_no, is_deleted=False).first()
        if not staff:
            staff = Staff.objects.filter(name=name, department=department, is_deleted=False).first()

        if not staff:
            staff = create_staff(
                name=name,
                position=position or '待完善',
                department=department,
                employee_no=employee_no,
                email=email,
                phone=phone,
                account=account,
            )
            created += 1
            continue

        updated_fields = {}
        if position:
            updated_fields['position'] = position
        if employee_no:
            updated_fields['employee_no'] = employee_no
        if email:
            updated_fields['email'] = email
        if phone:
            updated_fields['phone'] = phone
        updated_fields['department'] = department

        update_staff(staff.id, account=account, **updated_fields)
        StaffArchive.objects.get_or_create(
            staff=staff,
            defaults={
                'department': department,
                'employment_status': 'active',
                'employment_type': 'full_time',
                'sync_source': 'excel_import',
            },
        )
        updated += 1

    return {'created': created, 'updated': updated, 'skipped': skipped}


def import_staff_excel(content: bytes, filename: str = '', account=None) -> dict:
    """从 Excel（.xlsx）导入员工主数据，默认读取“人员信息表”或首个工作表。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f'无法读取 Excel 文件: {e}')

    ws = wb['人员信息表'] if '人员信息表' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True))

    header_row = None
    header_idx = None
    required = {'姓名', '岗位'}
    for idx, row in enumerate(rows, start=1):
        vals = [str(v).strip() if v is not None else '' for v in row]
        if required.issubset(set(vals)):
            header_row = vals
            header_idx = idx
            break
    if not header_row or not header_idx:
        raise ValueError('未找到表头，请确保含“姓名/岗位”列')

    def col(name: str) -> int:
        try:
            return header_row.index(name)
        except ValueError:
            return -1

    idx_name = col('姓名')
    idx_emp = col('工号')
    idx_dept = col('组别')
    idx_center = col('中心')
    idx_position = col('岗位')
    idx_phone = col('手机')
    idx_email = col('邮箱')

    items = []
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in row]
        name = vals[idx_name] if idx_name >= 0 and idx_name < len(vals) else ''
        position = vals[idx_position] if idx_position >= 0 and idx_position < len(vals) else ''
        department = ''
        if idx_dept >= 0 and idx_dept < len(vals):
            department = vals[idx_dept]
        if not department and idx_center >= 0 and idx_center < len(vals):
            department = vals[idx_center]
        if not name:
            continue
        items.append({
            'name': name,
            'employee_no': vals[idx_emp] if idx_emp >= 0 and idx_emp < len(vals) else '',
            'department': department or '未分组',
            'position': position or '待完善',
            'phone': vals[idx_phone] if idx_phone >= 0 and idx_phone < len(vals) else '',
            'email': vals[idx_email] if idx_email >= 0 and idx_email < len(vals) else '',
        })

    if not items:
        raise ValueError('未识别到可导入的数据行')
    result = import_staff_rows(items, account=account)
    result['sheet'] = ws.title
    result['filename'] = filename
    return result


# ============================================================================
# 胜任力模型
# ============================================================================
def list_competency_models() -> list:
    return list(CompetencyModel.objects.all().order_by('sort_order'))


def get_competency_model(model_id: int) -> Optional[CompetencyModel]:
    return CompetencyModel.objects.filter(id=model_id).first()


def create_competency_model(name: str, description: str = '', icon: str = '',
                            levels: list = None, sort_order: int = 0) -> CompetencyModel:
    return CompetencyModel.objects.create(
        name=name, description=description, icon=icon,
        levels=levels or [], sort_order=sort_order,
    )


def update_competency_model(model_id: int, **kwargs) -> Optional[CompetencyModel]:
    m = get_competency_model(model_id)
    if not m:
        return None
    for k, v in kwargs.items():
        if v is not None and hasattr(m, k):
            setattr(m, k, v)
    m.save()
    return m


def delete_competency_model(model_id: int) -> bool:
    m = get_competency_model(model_id)
    if not m:
        return False
    m.delete()
    return True


# ============================================================================
# 能力评估
# ============================================================================
def list_assessments(
    account=None,
    staff_id: int = None,
    period: str = None,
    status: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = Assessment.objects.filter(is_deleted=False).select_related('staff')
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    if period:
        qs = qs.filter(period=period)
    if status:
        qs = qs.filter(status=status)
    return _paginate_queryset(qs, page, page_size)


def get_assessment(assessment_id: int, account=None) -> Optional[Assessment]:
    qs = Assessment.objects.filter(id=assessment_id, is_deleted=False).select_related('staff')
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    return qs.first()


def create_assessment(staff_id: int, period: str, assessor: str,
                      scores: dict = None, overall: str = '',
                      assessor_id: int = None, account=None) -> Assessment:
    staff = get_staff(staff_id, account=account)
    if not staff:
        raise PermissionError('无权为该员工创建评估')
    return Assessment.objects.create(
        staff_id=staff_id, period=period, assessor=assessor,
        scores=scores or {}, overall=overall, assessor_id=assessor_id,
    )


def update_assessment(assessment_id: int, account=None, **kwargs) -> Optional[Assessment]:
    a = get_assessment(assessment_id, account=account)
    if not a:
        return None
    for k, v in kwargs.items():
        if v is not None and hasattr(a, k):
            setattr(a, k, v)
    a.save()
    return a


def delete_assessment(assessment_id: int, account=None) -> bool:
    a = get_assessment(assessment_id, account=account)
    if not a:
        return False
    a.is_deleted = True
    a.save(update_fields=['is_deleted', 'update_time'])
    return True


# ============================================================================
# 培训跟踪
# ============================================================================
def list_trainings(
    account=None,
    trainee_id: int = None,
    status: str = None,
    category: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = Training.objects.filter(is_deleted=False).select_related('trainee')
    qs = _apply_hr_scope(qs, account, personal_field='trainee__account_fk_id', department_field='trainee__department')
    if trainee_id:
        qs = qs.filter(trainee_id=trainee_id)
    if status:
        qs = qs.filter(status=status)
    if category:
        qs = qs.filter(category__icontains=category)
    return _paginate_queryset(qs, page, page_size)


def get_training(training_id: int, account=None) -> Optional[Training]:
    qs = Training.objects.filter(id=training_id, is_deleted=False).select_related('trainee')
    qs = _apply_hr_scope(qs, account, personal_field='trainee__account_fk_id', department_field='trainee__department')
    return qs.first()


def _sync_training_to_calendar(training: Training) -> None:
    """
    同步培训记录到飞书 GCP 培训日历

    对应 FEISHU_NATIVE_SETUP.md 5.2：GCP 培训计划日历
    """
    if not FEISHU_CALENDAR_TRAINING_ID:
        return

    try:
        from libs.feishu_client import feishu_client
        import time as time_module

        # 将 date 转为 Unix 时间戳
        start_ts = int(datetime.combine(training.start_date, datetime.min.time()).timestamp())
        end_date = training.end_date or training.start_date
        end_ts = int(datetime.combine(end_date, datetime.min.time()).timestamp()) + 86400

        summary = f"[培训] {training.course_name}"
        description = (
            f"课程: {training.course_name}\n"
            f"类别: {training.category}\n"
            f"讲师: {training.trainer}\n"
            f"学时: {training.hours}h"
        )

        if training.feishu_calendar_id:
            feishu_client.update_calendar_event(
                calendar_id=FEISHU_CALENDAR_TRAINING_ID,
                event_id=training.feishu_calendar_id,
                summary=summary,
                start_time=start_ts,
                end_time=end_ts,
                description=description,
            )
            logger.info(f"培训#{training.id} 日历事件已更新")
        else:
            data = feishu_client.create_calendar_event(
                calendar_id=FEISHU_CALENDAR_TRAINING_ID,
                summary=summary,
                start_time=start_ts,
                end_time=end_ts,
                description=description,
            )
            event_id = data.get('event', {}).get('event_id', '')
            if event_id:
                training.feishu_calendar_id = event_id
                training.save(update_fields=['feishu_calendar_id'])
                logger.info(f"培训#{training.id} 日历事件已创建: {event_id}")
    except Exception as e:
        logger.error(f"培训#{training.id} 日历同步失败: {e}")


def create_training(course_name: str, category: str, trainee_id: int,
                    trainer: str, start_date: date, hours: int,
                    end_date: date = None, score: str = '', account=None) -> Training:
    """创建培训记录并同步到飞书 GCP 培训日历"""
    trainee = get_staff(trainee_id, account=account)
    if not trainee:
        raise PermissionError('无权为该员工创建培训')
    training = Training.objects.create(
        course_name=course_name, category=category, trainee_id=trainee_id,
        trainer=trainer, start_date=start_date, hours=hours,
        end_date=end_date, score=score,
    )
    _sync_training_to_calendar(training)
    return training


def update_training(training_id: int, account=None, **kwargs) -> Optional[Training]:
    """更新培训记录并同步飞书日历"""
    t = get_training(training_id, account=account)
    if not t:
        return None
    for k, v in kwargs.items():
        if v is not None and hasattr(t, k):
            setattr(t, k, v)
    t.save()
    _sync_training_to_calendar(t)
    return t


def delete_training(training_id: int, account=None) -> bool:
    t = get_training(training_id, account=account)
    if not t:
        return False
    t.is_deleted = True
    t.save(update_fields=['is_deleted', 'update_time'])
    return True


def get_training_stats(account=None) -> dict:
    from django.db.models import Count, Sum
    qs = Training.objects.filter(is_deleted=False)
    qs = _apply_hr_scope(qs, account, personal_field='trainee__account_fk_id', department_field='trainee__department')
    by_status = qs.values('status').annotate(count=Count('id'))
    total_hours = qs.filter(status='completed').aggregate(total=Sum('hours'))['total'] or 0
    return {
        'by_status': {item['status']: item['count'] for item in by_status},
        'total': qs.count(),
        'total_completed_hours': total_hours,
    }


# ============================================================================
# P1：人事档案中心
# ============================================================================
def list_archives(
    account=None,
    keyword: str = None,
    employment_status: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = StaffArchive.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='department')
    if keyword:
        qs = qs.filter(staff__name__icontains=keyword)
    if employment_status:
        qs = qs.filter(employment_status=employment_status)
    return _paginate_queryset(qs.order_by('-update_time'), page, page_size)


def get_archive(staff_id: int, account=None):
    staff = get_staff(staff_id, account=account)
    if not staff:
        return None
    archive = StaffArchive.objects.select_related('staff').filter(staff_id=staff_id).first()
    if archive:
        return archive
    return StaffArchive.objects.create(
        staff=staff,
        department=staff.department or '',
        employment_status='active',
        employment_type='full_time',
        sync_source='bootstrap',
    )


def upsert_archive(staff_id: int, account=None, **kwargs):
    archive = get_archive(staff_id, account=account)
    if not archive:
        return None
    target_department = kwargs.get('department')
    if target_department is not None:
        _ensure_department_write_allowed(account, target_department)
    for k, v in kwargs.items():
        if v is not None and hasattr(archive, k):
            setattr(archive, k, v)
    archive.save()
    return archive


def list_contracts(staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = StaffContract.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('-create_time'), page, page_size)


def create_contract(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建合同')
    return StaffContract.objects.create(**kwargs)


def list_certificates(staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = HrStaffCertificate.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('expiry_date', '-create_time'), page, page_size)


def create_certificate(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建证照')
    return HrStaffCertificate.objects.create(**kwargs)


def list_change_logs(staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = StaffChangeLog.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('-change_date', '-id'), page, page_size)


def create_change_log(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建异动记录')
    return StaffChangeLog.objects.create(**kwargs)


def list_exit_records(staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = StaffExitRecord.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('-exit_date', '-id'), page, page_size)


def create_exit_record(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建离职记录')
    return StaffExitRecord.objects.create(**kwargs)


# ============================================================================
# P2：招聘、绩效、薪酬激励、文化
# ============================================================================
def list_recruitment_demands(status: str = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = RecruitmentDemand.objects.all()
    qs = _apply_hr_scope(qs, account, department_field='department')
    if status:
        qs = qs.filter(status=status)
    return _paginate_queryset(qs.order_by('-create_time'), page, page_size)


def create_recruitment_demand(account=None, **kwargs):
    department = kwargs.get('department', '')
    if department:
        _ensure_department_write_allowed(account, department)
    return RecruitmentDemand.objects.create(**kwargs)


def list_candidates(demand_id: int = None, stage: str = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = RecruitmentCandidate.objects.select_related('demand').all()
    qs = _apply_hr_scope(qs, account, department_field='demand__department')
    if demand_id:
        qs = qs.filter(demand_id=demand_id)
    if stage:
        qs = qs.filter(stage=stage)
    return _paginate_queryset(qs.order_by('-create_time'), page, page_size)


def create_candidate(**kwargs):
    return RecruitmentCandidate.objects.create(**kwargs)


def list_performance_records(cycle_id: int = None, staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = PerformanceRecord.objects.select_related('staff', 'cycle').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if cycle_id:
        qs = qs.filter(cycle_id=cycle_id)
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('-create_time'), page, page_size)


def create_performance_cycle(**kwargs):
    return PerformanceCycle.objects.create(**kwargs)


def list_performance_cycles(page: int = 1, page_size: int = 20) -> dict:
    return _paginate_queryset(PerformanceCycle.objects.all().order_by('-period_start'), page, page_size)


def create_performance_record(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建绩效记录')
    return PerformanceRecord.objects.create(**kwargs)


def list_payroll_records(staff_id: int = None, pay_month: str = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = PayrollRecord.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    if pay_month:
        qs = qs.filter(pay_month=pay_month)
    return _paginate_queryset(qs.order_by('-pay_month', '-create_time'), page, page_size)


def create_payroll_record(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建薪资记录')
    return PayrollRecord.objects.create(**kwargs)


def list_incentives(staff_id: int = None, page: int = 1, page_size: int = 20, account=None) -> dict:
    qs = IncentiveRecord.objects.select_related('staff').all()
    qs = _apply_hr_scope(qs, account, personal_field='staff__account_fk_id', department_field='staff__department')
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    return _paginate_queryset(qs.order_by('-grant_date', '-id'), page, page_size)


def create_incentive_record(account=None, **kwargs):
    if not get_staff(kwargs.get('staff_id'), account=account):
        raise PermissionError('无权为该员工创建激励记录')
    return IncentiveRecord.objects.create(**kwargs)


def list_culture_activities(status: str = None, page: int = 1, page_size: int = 20) -> dict:
    qs = CultureActivity.objects.all()
    if status:
        qs = qs.filter(status=status)
    return _paginate_queryset(qs.order_by('-planned_date', '-id'), page, page_size)


def create_culture_activity(**kwargs):
    return CultureActivity.objects.create(**kwargs)


def list_engagement_pulse(page: int = 1, page_size: int = 20) -> dict:
    return _paginate_queryset(EngagementPulse.objects.all().order_by('-survey_month'), page, page_size)


def create_engagement_pulse(**kwargs):
    return EngagementPulse.objects.create(**kwargs)


# ============================================================================
# P3：跨台协同治理
# ============================================================================
def list_collaboration_snapshots(
    source_workstation: str = None,
    data_type: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = HrCollaborationSnapshot.objects.all()
    if source_workstation:
        qs = qs.filter(source_workstation=source_workstation)
    if data_type:
        qs = qs.filter(data_type=data_type)
    return _paginate_queryset(qs.order_by('-create_time'), page, page_size)


def create_collaboration_snapshot(**kwargs):
    return HrCollaborationSnapshot.objects.create(**kwargs)


def get_ops_overview(month: str = None, department: str = None, account=None) -> dict:
    """HRD 经营驾驶舱汇总"""
    from django.db.models import Count, Sum

    if not department and account:
        scope_mode = _get_hr_scope_mode(account)
        if scope_mode != 'global':
            department = _get_department_scope(account) or '__no_access_department__'

    archive_qs = StaffArchive.objects.all()
    if department:
        archive_qs = archive_qs.filter(department__icontains=department)

    active_archive_count = archive_qs.filter(employment_status='active').count()
    exited_archive_count = archive_qs.filter(employment_status='exited').count()

    recruitment_qs = RecruitmentDemand.objects.all()
    if department:
        recruitment_qs = recruitment_qs.filter(department__icontains=department)
    recruitment_open_count = recruitment_qs.exclude(status__in=['closed', 'cancelled']).count()

    candidate_qs = RecruitmentCandidate.objects.all()
    if department:
        candidate_qs = candidate_qs.filter(demand__department__icontains=department)
    candidate_pipeline_count = candidate_qs.exclude(stage__in=['rejected', 'joined']).count()

    performance_cycle_active_count = PerformanceCycle.objects.exclude(status__in=['closed']).count()
    performance_qs = PerformanceRecord.objects.select_related('staff').all()
    if department:
        performance_qs = performance_qs.filter(staff__department__icontains=department)
    performance_record_count = performance_qs.count()

    payroll_qs = PayrollRecord.objects.select_related('staff').all()
    if month:
        payroll_qs = payroll_qs.filter(pay_month=month)
    if department:
        payroll_qs = payroll_qs.filter(staff__department__icontains=department)
    payroll_month_count = payroll_qs.values('pay_month').distinct().count()
    payroll_total_net = payroll_qs.aggregate(total=Sum('net_salary'))['total'] or 0

    incentive_qs = IncentiveRecord.objects.select_related('staff').all()
    if month:
        incentive_qs = incentive_qs.filter(grant_date__year=int(month[:4]), grant_date__month=int(month[5:7]))
    if department:
        incentive_qs = incentive_qs.filter(staff__department__icontains=department)
    incentive_total = incentive_qs.aggregate(total=Sum('amount'))['total'] or 0

    culture_activity_count = CultureActivity.objects.count()
    pulse_qs = EngagementPulse.objects.all()
    if month:
        pulse_qs = pulse_qs.filter(survey_month=month)
    latest_pulse = pulse_qs.order_by('-survey_month').first()
    collaboration_open_count = HrCollaborationSnapshot.objects.exclude(sync_status='done').count()

    exit_qs = StaffExitRecord.objects.select_related('staff').all()
    if month:
        exit_qs = exit_qs.filter(exit_date__year=int(month[:4]), exit_date__month=int(month[5:7]))
    if department:
        exit_qs = exit_qs.filter(staff__department__icontains=department)
    recent_exits = list(exit_qs.order_by('-exit_date')[:5])

    recent_recruitment = list(
        recruitment_qs.order_by('-create_time')[:5]
    )

    # 联合预警规则：
    # 1) 个体层：连续两期绩效下降且最新分<70
    # 2) 组织层：敬业度脉冲低于70或被标记 high/medium
    org_pulse_risk = False
    if latest_pulse:
        org_pulse_risk = latest_pulse.score < 70 or latest_pulse.risk_level in ['high', 'medium']

    risk_candidates = []
    staff_ids = list(
        performance_qs.values_list('staff_id', flat=True).distinct()
    )
    for sid in staff_ids:
        records = list(
            performance_qs.filter(staff_id=sid).order_by('-create_time')[:2]
        )
        if len(records) < 2:
            continue
        latest, prev = records[0], records[1]
        perf_declining = latest.score < prev.score and latest.score < 70
        if perf_declining:
            reasons = ['绩效连续下降']
            if org_pulse_risk:
                reasons.append('组织敬业度偏低')
            risk_candidates.append({
                'staff_id': latest.staff_id,
                'staff_name': latest.staff.name if latest.staff else '',
                'latest_score': float(latest.score),
                'previous_score': float(prev.score),
                'severity': 'high' if org_pulse_risk else 'medium',
                'reasons': reasons,
            })

    # 风险动作闭环指标（基于协同快照）
    actions_qs = HrCollaborationSnapshot.objects.filter(data_type='risk_followup').order_by('-create_time')
    if month:
        actions_qs = actions_qs.filter(period=month)
    if department:
        dept_staff_ids = set(
            Staff.objects.filter(department__icontains=department, is_deleted=False).values_list('id', flat=True)
        )
        actions_qs = [a for a in actions_qs if (a.payload or {}).get('staff_id') in dept_staff_ids]
    else:
        actions_qs = list(actions_qs)

    action_total = len(actions_qs)
    done_count = 0
    overdue_count = 0
    recent_actions = []
    now_date = timezone.localdate()
    for action in actions_qs[:20]:
        status = action.sync_status or 'pending'
        if status == 'done':
            done_count += 1
        payload = action.payload or {}
        due_text = payload.get('due_date')
        if due_text:
            try:
                due = date.fromisoformat(due_text)
            except ValueError:
                due = action.create_time.date() + timedelta(days=7)
        else:
            due = action.create_time.date() + timedelta(days=7)
        is_overdue = status in ['pending', 'in_progress'] and due < now_date
        if is_overdue:
            overdue_count += 1
        recent_actions.append({
            'id': action.id,
            'staff_id': payload.get('staff_id') or 0,
            'staff_name': payload.get('staff_name') or '',
            'action_type': payload.get('action_type') or '',
            'owner': payload.get('owner') or '',
            'due_date': due.isoformat(),
            'sync_status': status,
            'create_time': action.create_time.isoformat(),
            'is_overdue': is_overdue,
        })

    completion_rate = round((done_count / action_total) * 100, 1) if action_total else 0
    overdue_rate = round((overdue_count / action_total) * 100, 1) if action_total else 0

    return {
        'workforce': {
            'active': active_archive_count,
            'exited': exited_archive_count,
            'net_change': active_archive_count - exited_archive_count,
        },
        'recruitment': {
            'open_demands': recruitment_open_count,
            'pipeline_candidates': candidate_pipeline_count,
            'recent_demands': [{
                'id': d.id, 'title': d.title, 'department': d.department, 'status': d.status,
            } for d in recent_recruitment],
        },
        'performance': {
            'active_cycles': performance_cycle_active_count,
            'records': performance_record_count,
        },
        'compensation': {
            'payroll_months': payroll_month_count,
            'total_net_salary': float(payroll_total_net),
            'total_incentive': float(incentive_total),
        },
        'culture': {
            'activity_count': culture_activity_count,
            'latest_pulse': {
                'survey_month': latest_pulse.survey_month if latest_pulse else '',
                'score': float(latest_pulse.score) if latest_pulse else 0,
                'risk_level': latest_pulse.risk_level if latest_pulse else '',
            },
        },
        'collaboration': {
            'open_snapshots': collaboration_open_count,
        },
        'risks': {
            'recent_exits': [{
                'staff_name': r.staff.name if r.staff else '',
                'exit_date': r.exit_date.isoformat(),
                'exit_type': r.exit_type,
                'reason': r.reason,
            } for r in recent_exits],
            'risk_candidates': risk_candidates[:10],
            'org_pulse_risk': org_pulse_risk,
            'action_metrics': {
                'total': action_total,
                'done': done_count,
                'overdue': overdue_count,
                'completion_rate': completion_rate,
                'overdue_rate': overdue_rate,
            },
            'recent_actions': recent_actions[:8],
        },
    }


def create_risk_followup_action(
    staff_id: int,
    action_type: str,
    operator: str = '',
    owner: str = '',
    due_date: date = None,
    note: str = '',
) -> dict:
    """为联合预警对象创建跟进行动。"""
    staff = Staff.objects.filter(id=staff_id, is_deleted=False).first()
    if not staff:
        return {}

    action_type = action_type or 'interview'
    today = date.today()
    due = due_date or (today + timedelta(days=7))
    action_owner = owner or operator or 'HRBP'
    result = {
        'staff_id': staff.id,
        'staff_name': staff.name,
        'action_type': action_type,
        'owner': action_owner,
        'due_date': due.isoformat(),
    }
    if action_type == 'training':
        training = Training.objects.create(
            course_name='风险干预培训',
            category='风险干预',
            trainee=staff,
            trainer=action_owner,
            start_date=today,
            hours=2,
            status='scheduled',
            score='',
        )
        result['training_id'] = training.id
    else:
        change_log = StaffChangeLog.objects.create(
            staff=staff,
            change_type='风险面谈',
            change_date=today,
            before_data={},
            after_data={'action': 'risk_interview', 'note': note, 'owner': action_owner, 'due_date': due.isoformat()},
            operated_by=operator or 'HRBP',
            reason=note or '联合预警触发面谈',
        )
        result['change_log_id'] = change_log.id

    # 记录跨台协同快照，便于后续治理追踪
    HrCollaborationSnapshot.objects.create(
        source_workstation='hr',
        data_type='risk_followup',
        period=today.strftime('%Y-%m'),
        payload=result,
        sync_status='pending',
    )
    return result


def list_risk_followup_actions(
    sync_status: str = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    qs = HrCollaborationSnapshot.objects.filter(data_type='risk_followup').order_by('-create_time')
    if sync_status:
        qs = qs.filter(sync_status=sync_status)
    return _paginate_queryset(qs, page, page_size)


def update_risk_followup_action(action_id: int, sync_status: str, note: str = '') -> Optional[HrCollaborationSnapshot]:
    action = HrCollaborationSnapshot.objects.filter(id=action_id, data_type='risk_followup').first()
    if not action:
        return None
    action.sync_status = sync_status
    payload = dict(action.payload or {})
    if note:
        payload['note'] = note
    if sync_status in ['done', 'cancelled']:
        payload['closed_at'] = timezone.now().isoformat()
    action.payload = payload
    action.save(update_fields=['sync_status', 'payload'])
    return action


def update_risk_followup_action_meta(action_id: int, owner: str = '', due_date: date = None) -> Optional[HrCollaborationSnapshot]:
    action = HrCollaborationSnapshot.objects.filter(id=action_id, data_type='risk_followup').first()
    if not action:
        return None
    payload = dict(action.payload or {})
    if owner:
        payload['owner'] = owner
    if due_date:
        payload['due_date'] = due_date.isoformat()
    action.payload = payload
    action.save(update_fields=['payload'])
    return action
