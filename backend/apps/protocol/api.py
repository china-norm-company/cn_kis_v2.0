"""
协议管理 API

端点：
- GET  /protocol/list           协议列表
- PUT  /protocol/{id}           更新协议基本信息（项目名称/编号，编号唯一）
- GET  /protocol/{id}           协议详情
- DELETE /protocol/{id}         软删除协议（打标，可恢复）
- POST /protocol/create         创建协议
- POST /protocol/upload         上传协议文件
- POST /protocol/upload-create  上传文件直接创建协议（知情管理用）
- POST /protocol/reorder-consent 调整知情管理展示顺序
- POST /protocol/{id}/parse     触发 AI 解析
- GET  /protocol/{id}/logs      解析日志
"""
import io
import logging
import mimetypes
import os
import re
import threading
import time
from ninja import Router, Schema, Query, File, Form, Body
from ninja.files import UploadedFile
from typing import Optional, List, Any
from datetime import date, datetime
from django.conf import settings as django_settings
from django.db import IntegrityError
from django.http import FileResponse
from django.utils import timezone
from . import services
from .consent_signing_names import normalize_consent_signing_staff_storage, split_consent_signing_staff_names
from .models import Protocol, ConsentGlobalConfig, ConsentConfigMode, WitnessDualSignAuthToken, WitnessStaff
from apps.identity.decorators import _get_account_from_request, require_permission, require_any_permission
from apps.identity.filters import get_visible_object

router = Router()
logger = logging.getLogger(__name__)


def _parse_consent_date_query(value: Optional[str]) -> Optional[date]:
    """签署记录日期筛选：YYYY-MM-DD。"""
    if not value or not str(value).strip():
        return None
    try:
        return datetime.strptime(str(value).strip()[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


# ============================================================================
# Schema
# ============================================================================
class ProtocolOut(Schema):
    id: int
    title: str
    code: Optional[str] = None
    file_path: Optional[str] = None
    status: str
    parsed_data: Optional[dict] = None
    efficacy_type: Optional[str] = None
    sample_size: Optional[int] = None
    create_time: datetime
    update_time: datetime


class ScreeningDayIn(Schema):
    """现场筛选计划：日期 + 目标筛选人数（分母按 人数×ICF 文档数）；可选测试筛选（须早于最早正式筛选日）"""
    date: str
    target_count: int = 1
    is_test_screening: bool = False
    signing_staff_name: Optional[str] = None


class ProtocolCreateIn(Schema):
    title: str
    code: Optional[str] = None
    efficacy_type: Optional[str] = None
    sample_size: Optional[int] = None
    screening_schedule: Optional[List[ScreeningDayIn]] = None
    # 治理台账号 ID（全局角色 qa）；每项目至多一人
    consent_config_account_id: Optional[int] = None
    consent_signing_staff_name: Optional[str] = None
    # 质量台项目监察/测试：与列表展示字段对齐，写入 parsed_data / team_members
    group_label: Optional[str] = None
    backup_sample_label: Optional[str] = None
    visits_summary: Optional[str] = None
    execution_start: Optional[str] = None
    execution_end: Optional[str] = None
    principal_investigator: Optional[str] = None
    # True：质量台本地测试项目，parsed_data.quality_origin=manual_test，不进入「项目管理」维周列表
    quality_manual_test: Optional[bool] = False


class ProtocolBasicUpdateIn(Schema):
    """更新项目名称/编号（至少填一项）；项目编号非空时全局唯一。"""
    title: Optional[str] = None
    code: Optional[str] = None
    consent_config_account_id: Optional[int] = None  # 0 表示清空；须为 qa（QA质量管理）全局角色


class ProtocolQueryParams(Schema):
    status: Optional[str] = None
    title: Optional[str] = None
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 20


class StaffReturnIn(Schema):
    """执行台退回重签：可选填写退回原因（供小程序展示，最长 500 字）。"""
    reason: Optional[str] = None


class ProtocolUploadIn(Schema):
    protocol_id: int
    file_path: str


class ICFVersionOut(Schema):
    id: int
    protocol_id: int
    version: str
    file_path: Optional[str] = None
    content: Optional[str] = None
    is_active: bool
    required_reading_duration_seconds: int = 0
    create_time: datetime
    update_time: datetime


class ICFVersionCreateIn(Schema):
    version: str
    content: Optional[str] = ''
    file_path: Optional[str] = ''
    is_active: bool = True
    required_reading_duration_seconds: int = 0
    node_title: Optional[str] = ''


class ICFVersionUpdateIn(Schema):
    version: Optional[str] = None
    content: Optional[str] = None
    is_active: Optional[bool] = None
    required_reading_duration_seconds: Optional[int] = None
    node_title: Optional[str] = None


class ICFReorderIn(Schema):
    id_order: List[int]


class ConsentRecordOut(Schema):
    id: int
    subject_id: int
    subject_no: Optional[str] = None
    subject_name: Optional[str] = None
    icf_version_id: int
    icf_version: Optional[str] = None
    is_signed: bool
    signed_at: Optional[datetime] = None
    investigator_signed_at: Optional[datetime] = None
    receipt_no: Optional[str] = None
    receipt_pdf_path: Optional[str] = None
    create_time: datetime


class ConsentStatsOut(Schema):
    total: int
    signed_count: int
    pending_count: int


def _protocol_to_dict(p) -> dict:
    return {
        'id': p.id,
        'title': p.title,
        'code': p.code,
        'file_path': p.file_path,
        'status': p.status,
        'parsed_data': p.parsed_data,
        'efficacy_type': p.efficacy_type,
        'sample_size': p.sample_size,
        'team_members': getattr(p, 'team_members', None) or [],
        'consent_config_account_id': getattr(p, 'consent_config_account_id', None),
        'create_time': p.create_time.isoformat(),
        'update_time': p.update_time.isoformat(),
    }


# ============================================================================
# 端点
# ============================================================================
@router.get('/list', summary='协议列表')
@require_permission('protocol.protocol.read')
def list_protocols(request, params: ProtocolQueryParams = Query(...)):
    """分页查询协议列表（数据权限过滤）"""
    account = _get_account_from_request(request)
    result = services.list_protocols(
        status=params.status,
        title=params.title,
        keyword=(params.keyword or '').strip() or None,
        page=params.page,
        page_size=params.page_size,
        account=account,
    )
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': [_protocol_to_dict(item) for item in result['items']],
            'total': result['total'],
            'page': result['page'],
            'page_size': result['page_size'],
        },
    }


@router.post('/create', summary='创建协议', response={200: dict, 400: dict, 500: dict})
@require_any_permission(['protocol.protocol.create', 'quality.deviation.create'])
def create_protocol(request, data: ProtocolCreateIn):
    """创建新协议（质量台项目监察与维周执行台同源；具备其一即可）"""
    account = _get_account_from_request(request)
    created_by_id = getattr(account, 'id', None) if account else None
    sched = None
    if getattr(data, 'screening_schedule', None):
        from apps.subject.services.consent_service import (
            _normalize_screening_schedule_for_stats as _norm_ss,
            validate_screening_schedule_test_rules,
        )

        sched = _norm_ss(
            [
                x.model_dump() if hasattr(x, 'model_dump') else x.dict()
                for x in data.screening_schedule
            ]
        )
        verr = validate_screening_schedule_test_rules(sched)
        if verr:
            return 400, {'code': 400, 'msg': verr, 'data': None}
        if any((x.get('signing_staff_name') or '').strip() for x in (sched or [])):
            return 400, {
                'code': 400,
                'msg': '请先在「知情配置」中添加双签工作人员并保存后，再指定各现场日知情签署人员',
                'data': None,
            }
    try:
        protocol = services.create_protocol(
            title=data.title,
            code=data.code or '',
            efficacy_type=data.efficacy_type or '',
            sample_size=data.sample_size,
            screening_schedule=sched,
            consent_config_account_id=getattr(data, 'consent_config_account_id', None),
            consent_signing_staff_name=getattr(data, 'consent_signing_staff_name', None),
            group_label=getattr(data, 'group_label', None),
            backup_sample_label=getattr(data, 'backup_sample_label', None),
            visits_summary=getattr(data, 'visits_summary', None),
            execution_start=getattr(data, 'execution_start', None),
            execution_end=getattr(data, 'execution_end', None),
            principal_investigator=getattr(data, 'principal_investigator', None),
            created_by_id=created_by_id,
            quality_manual_test=bool(getattr(data, 'quality_manual_test', False)),
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e), 'data': None}
    except IntegrityError as e:
        logger.warning('创建协议唯一约束冲突: %s', e)
        return 400, {'code': 400, 'msg': '项目编号可能已被占用，请更换后重试', 'data': None}
    except Exception as e:
        logger.exception('创建协议失败: %s', e)
        detail = str(e)[:500] if getattr(django_settings, 'DEBUG', False) else '创建协议失败，请查看服务端日志或稍后重试'
        return 500, {'code': 500, 'msg': detail, 'data': None}
    return {
        'code': 200,
        'msg': 'OK',
        'data': {'id': protocol.id, 'title': protocol.title, 'status': protocol.status},
    }


@router.post('/upload', summary='上传协议文件')
@require_permission('protocol.protocol.update')
def upload_protocol(request, data: ProtocolUploadIn):
    """上传协议文件"""
    protocol = services.upload_protocol_file(data.protocol_id, data.file_path)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    return {
        'code': 200,
        'msg': 'OK',
        'data': {'id': protocol.id, 'file_path': protocol.file_path, 'status': protocol.status},
    }


@router.post('/upload-create', summary='上传文件创建协议（知情管理）')
@require_any_permission(['protocol.protocol.create', 'protocol.protocol.update', 'protocol.protocol.read'])
def upload_create_protocol(request, file: UploadedFile = File(...), title: Optional[str] = Form(None)):
    """上传协议文件直接创建协议，支持多份上传（多次调用）"""
    account = _get_account_from_request(request)
    protocol = services.create_protocol_from_upload(file, title=title, created_by_id=getattr(account, 'id', None))
    return {
        'code': 200,
        'msg': 'OK',
        'data': _protocol_to_dict(protocol),
    }


def _parse_screening_schedule_from_import_row(row: dict):
    """
    从导入行解析 screening_schedule（可选）。
    支持：1）列「现场筛选计划」文本，格式 2026-03-18(10)|2026-03-20(12) 或 日期:人数；
    2）成对列 现场筛选日1 + 目标人数1 … 现场筛选日8 + 目标人数8（或 预约人数n）。
    """
    from apps.subject.services.consent_service import _normalize_screening_schedule_for_stats

    date_re = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    parts = []
    raw = (row.get('现场筛选计划') or row.get('screening_schedule') or row.get('现场筛选日期计划') or '').strip()
    if raw:
        for seg in re.split(r'\s*\|\s*', raw):
            seg = seg.strip()
            if not seg:
                continue
            m = re.match(r'^(\d{4}-\d{2}-\d{2})\s*[\(（](\d+)[\)）]\s*$', seg)
            if m and date_re.match(m.group(1)):
                parts.append({'date': m.group(1), 'target_count': max(1, int(m.group(2)))})
                continue
            m2 = re.match(r'^(\d{4}-\d{2}-\d{2})\s*[:：]\s*(\d+)\s*$', seg)
            if m2 and date_re.match(m2.group(1)):
                parts.append({'date': m2.group(1), 'target_count': max(1, int(m2.group(2)))})
    if not parts:
        for i in range(1, 9):
            ds = (row.get(f'现场筛选日{i}') or row.get(f'筛选日{i}') or '').strip()[:10]
            tc_raw = (
                row.get(f'目标人数{i}')
                or row.get(f'目标筛选人数{i}')
                or row.get(f'预约人数{i}')
                or ''
            ).strip()
            if not ds or not date_re.match(ds):
                continue
            try:
                tc = int(tc_raw) if tc_raw else 1
            except ValueError:
                tc = 1
            parts.append({'date': ds, 'target_count': max(1, tc)})
    if not parts:
        return None
    return _normalize_screening_schedule_for_stats(parts)


def _append_import_protocol_row(rows: List[dict], row: dict) -> None:
    """由一行键值对追加到 rows；项目名称、项目编号为导入必填（空行跳过）。"""
    title = (row.get('项目名称') or row.get('title') or row.get('项目名') or '').strip()
    code = (row.get('项目编号') or row.get('code') or row.get('编号') or '').strip() or None
    if not title and not code:
        return
    entry = {'title': title, 'code': code}
    sched = _parse_screening_schedule_from_import_row(row)
    if sched:
        entry['screening_schedule'] = sched
    rows.append(entry)


def _parse_import_file(uploaded_file) -> List[dict]:
    """解析 CSV/Excel 导入文件，返回 [{'title', 'code', 'screening_schedule'?}, ...]"""
    name = getattr(uploaded_file, 'name', '') or ''
    ext = (name.rsplit('.', 1)[-1] or '').lower()
    rows = []
    if ext in ('csv', 'txt'):
        import csv
        content = b''.join(uploaded_file.chunks()).decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        for r in reader:
            if not r:
                continue
            row = {str(k).strip(): (v.strip() if isinstance(v, str) else (v or '')) for k, v in r.items() if k}
            _append_import_protocol_row(rows, row)
    elif ext == 'xlsx':
        try:
            import openpyxl
        except ImportError:
            raise ValueError('Excel 导入依赖 openpyxl，请安装')
        content = b''.join(uploaded_file.chunks())
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for col, cell in enumerate(ws[1], 1):
            h = str(cell.value or '').strip()
            headers.append((col, h))
        for row in ws.iter_rows(min_row=2):
            cells = list(row)
            row_dict = {}
            for col_num, h in headers:
                if not h:
                    continue
                idx = col_num - 1
                val = cells[idx].value if idx < len(cells) else None
                row_dict[h] = '' if val is None else str(val).strip()
            _append_import_protocol_row(rows, row_dict)
        wb.close()
    else:
        raise ValueError('仅支持 CSV 或 Excel (.xlsx) 文件')
    return rows


@router.post('/batch-import', summary='批量导入项目（新建项目）')
@require_any_permission(['protocol.protocol.create', 'protocol.protocol.update', 'protocol.protocol.read'])
def batch_import_protocols(request, file: UploadedFile = File(...)):
    """
    批量导入项目。必填列：项目名称、项目编号。
    可选：现场筛选计划（如 2026-03-18(10)|2026-03-20(12)），或 现场筛选日1+目标人数1 … 成对列。
    支持 CSV、Excel (.xlsx)。
    """
    account = _get_account_from_request(request)
    try:
        rows = _parse_import_file(file)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e), 'data': None}
    if not rows:
        return 400, {'code': 400, 'msg': '文件中无有效数据行', 'data': None}
    result = services.batch_create_protocols(rows, created_by_id=getattr(account, 'id', None))
    return {
        'code': 200,
        'msg': 'OK',
        'data': result,
    }


class ReorderConsentIn(Schema):
    id_order: List[int]


class DualSignStaffIn(Schema):
    staff_id: Optional[str] = ''
    name: str
    id_card_no: Optional[str] = ''
    email: Optional[str] = ''
    phone: Optional[str] = ''
    identity_verified: bool = False


def _clamp_1_or_2(v, default: int = 1) -> int:
    try:
        n = int(v)
        return 2 if n >= 2 else 1
    except (TypeError, ValueError):
        return default


class ConsentSettingsIn(Schema):
    require_face_verify: bool = False
    require_dual_sign: bool = False
    require_comprehension_quiz: bool = False
    enable_min_reading_duration: bool = True
    min_reading_duration_seconds: int = 30
    dual_sign_staffs: List[DualSignStaffIn] = []
    collect_id_card: bool = False
    collect_screening_number: bool = False
    collect_initials: bool = False
    collect_subject_name: bool = False
    collect_other_information: bool = False
    enable_checkbox_recognition: bool = False
    enable_staff_signature: bool = False
    staff_signature_times: int = 1
    enable_subject_signature: bool = False
    subject_signature_times: int = 1
    enable_guardian_signature: bool = False
    guardian_parent_count: int = 1
    guardian_signature_times: int = 1
    enable_auto_sign_date: bool = False
    planned_screening_dates: List[str] = []
    screening_schedule: Optional[List[ScreeningDayIn]] = None
    consent_signing_staff_name: Optional[str] = None


class MiniSignRulesIn(Schema):
    """单签署节点小程序签署规则（不含现场计划）"""

    require_face_verify: bool = False
    require_dual_sign: bool = False
    require_comprehension_quiz: bool = False
    enable_min_reading_duration: bool = True
    min_reading_duration_seconds: int = 30
    dual_sign_staffs: List[DualSignStaffIn] = []
    collect_id_card: bool = False
    collect_screening_number: bool = False
    collect_initials: bool = False
    collect_subject_name: bool = False
    collect_other_information: bool = False
    # 补充说明类采集项的自定义标签（非空则视为需采集「其他」类信息）；最多 20 条
    supplemental_collect_labels: Optional[List[str]] = None
    enable_checkbox_recognition: bool = False
    enable_staff_signature: bool = False
    staff_signature_times: int = 1
    enable_subject_signature: bool = False
    subject_signature_times: int = 1
    enable_guardian_signature: bool = False
    guardian_parent_count: int = 1
    guardian_signature_times: int = 1
    enable_auto_sign_date: bool = False


class ConsentLaunchIn(Schema):
    """发布/取消发布"""
    launched: bool


class ConsentWitnessSignIn(Schema):
    staff_id: Optional[str] = ''
    staff_name: str
    staff_phone: Optional[str] = ''
    staff_email: Optional[str] = ''


def _merge_consent_screening_schedule(raw: dict) -> list:
    """合并 screening_schedule 与旧版 planned_screening_dates（目标量默认 1）。"""
    from apps.subject.services.consent_service import _normalize_screening_schedule_for_stats
    if not isinstance(raw, dict):
        return []
    rs = raw.get('screening_schedule')
    if rs:
        return _normalize_screening_schedule_for_stats(rs)
    return [
        {'date': d, 'target_count': 1}
        for d in _normalize_planned_screening_dates(raw.get('planned_screening_dates') or [])
    ]


def _screening_schedule_input_to_plain_dicts(sched_raw: Optional[List]) -> List[dict]:
    """将请求体中的 screening_schedule 转为纯 dict 列表；跳过 None，避免 dict(None) 导致 500。"""
    if not sched_raw:
        return []
    out: List[dict] = []
    for x in sched_raw:
        if x is None:
            continue
        if isinstance(x, dict):
            out.append(dict(x))
        elif hasattr(x, 'model_dump'):
            out.append(x.model_dump())
        elif hasattr(x, 'dict'):
            out.append(x.dict())
    return out


def _screening_schedule_summary(sched: list) -> str:
    if not sched:
        return ''
    parts = []
    for x in sched:
        if isinstance(x, dict) and x.get('date'):
            suf = '测' if x.get('is_test_screening') else ''
            parts.append(f"{str(x['date'])[:10]}({int(x.get('target_count') or 1)}){suf}")
    return ' | '.join(parts)


def _display_protocol_title_for_export(title: Optional[str], code: Optional[str]) -> str:
    """与知情管理列表「项目名称」列一致：去掉末尾「（编号）」后缀。"""
    raw = (title or '').strip()
    c = (code or '').strip()
    if not raw:
        return '-'
    if c:
        for suf in (f'（{c}）', f'({c})'):
            if raw.endswith(suf):
                stripped = raw[: -len(suf)].strip()
                return stripped or raw
    return raw


def _first_configured_screening_date_for_export(
    screening_schedule: list,
    earliest_screening_date: Optional[str],
) -> str:
    """与前端一致：有计划时取「正式筛选」升序第一日（不含测试筛选）；否则取接口最早现场日。"""
    sched = screening_schedule or []
    if sched:
        formal = [
            x for x in sched
            if isinstance(x, dict) and x.get('date') and not x.get('is_test_screening')
        ]
        if formal:
            sorted_rows = sorted(formal, key=lambda a: str(a.get('date', '')).strip()[:10])
            return str(sorted_rows[0].get('date', '')).strip()[:10]
    if earliest_screening_date:
        return str(earliest_screening_date).strip()[:10]
    return ''


def _last_configured_screening_date_for_export(
    screening_schedule: list,
    latest_screening_date: Optional[str],
) -> str:
    """与前端「筛选结束日期」一致：正式筛选计划降序第一日；否则取接口最晚现场日。"""
    sched = screening_schedule or []
    if sched:
        formal = [
            x for x in sched
            if isinstance(x, dict) and x.get('date') and not x.get('is_test_screening')
        ]
        if formal:
            sorted_rows = sorted(formal, key=lambda a: str(a.get('date', '')).strip()[:10], reverse=True)
            return str(sorted_rows[0].get('date', '')).strip()[:10]
    if latest_screening_date:
        return str(latest_screening_date).strip()[:10]
    return ''


def _target_count_for_batch_date(settings: dict, date_str: str) -> Optional[int]:
    for x in settings.get('screening_schedule') or []:
        if not isinstance(x, dict):
            continue
        ds = str(x.get('date', '')).strip()[:10]
        if ds == date_str:
            try:
                return max(1, int(x.get('target_count') or 1))
            except (TypeError, ValueError):
                return 1
    return None


def _batch_progress_components(b: dict) -> tuple:
    """与前端 batchProgressPair 一致。"""
    den = b.get('progress_total')
    if den is None:
        den = max(1, int(b.get('total') or 0), int(b.get('expected_consent_rows') or 0))
    num = b.get('progress_signed')
    if num is None:
        num = int(b.get('signed_count') or 0)
    pend = b.get('pending_progress')
    if pend is None:
        pend = int(b.get('pending_count') or 0)
    return int(num), int(den), int(pend)


def _is_test_screening_batch_for_export(b: dict, settings: dict) -> bool:
    """与前端 isTestScreeningForBatch 一致：批次 is_test_screening 或与 screening_schedule 日期对齐。"""
    if b.get('is_test_screening'):
        return True
    ds = str(b.get('screening_date', '')).strip().replace('/', '-')[:10]
    if not ds:
        return False
    for x in settings.get('screening_schedule') or []:
        if not isinstance(x, dict):
            continue
        row_ds = str(x.get('date', '')).strip().replace('/', '-')[:10]
        if row_ds == ds:
            return bool(x.get('is_test_screening'))
    return False


def _format_staff_verification_for_export(settings: dict, v_snap: dict) -> str:
    """与列表「工作人员核验」列一致：状态 + 启用双签且有人员时附「已核验数/总人数」（不再单独导出核验进度列）。"""
    st = (v_snap.get('staff_verification_status') or '').strip() or '—'
    if not bool(settings.get('require_dual_sign')):
        return st
    t = int(v_snap.get('dual_sign_staff_total') or 0)
    if t <= 0:
        return st
    v = int(v_snap.get('verified_staff_count') or 0)
    return f'{st} {v}/{t}'


def _batch_attendance_label_for_export(settings: dict, b: dict, today_iso: str) -> str:
    """与前端 batchAttendanceLabel 一致（到场/待签到说明）。"""
    d = str(b.get('screening_date', ''))[:10]
    cohort = int(b.get('cohort_subject_count') or 0)
    target = _target_count_for_batch_date(settings, d)
    if d <= today_iso:
        return f'{cohort}人已签到'
    remain = max(0, int(target or 0) - cohort) if target is not None else 0
    return f'{remain}人待签到'


def _format_consent_signing_progress_export(
    config_status: str,
    settings: dict,
    batches: list,
) -> str:
    """与列表「签署进度」列展示等价：合计区分正式/测试筛选；各现场日标注测试筛选（导出 Excel 单格换行）。"""
    today_iso = timezone.localdate().strftime('%Y-%m-%d')
    blist = list(batches or [])
    if not blist:
        if config_status == '待配置':
            return (
                '暂无签署进度（待配置）。请先上传签署节点并完善规则；登记现场计划或产生到场数据后，将展示进度。'
            )
        return '暂无安排。可在「配置」登记现场筛选计划（日期+目标量），或录入粗筛/接待到场日。'

    sorted_batches = sorted(blist, key=lambda x: str(x.get('screening_date', ''))[:10])
    formal_n = formal_d = formal_p = 0
    test_n = test_d = test_p = 0
    for b in sorted_batches:
        n, d, p = _batch_progress_components(b)
        if _is_test_screening_batch_for_export(b, settings):
            test_n += n
            test_d += d
            test_p += p
        else:
            formal_n += n
            formal_d += d
            formal_p += p

    has_test = any(_is_test_screening_batch_for_export(b, settings) for b in sorted_batches)
    has_formal = any(not _is_test_screening_batch_for_export(b, settings) for b in sorted_batches)

    lines: List[str] = []
    if has_test and has_formal:
        # 与前端列表首行同一行：正式/测试进度与待签署合计，不换行
        lines.append(
            f'正式 {formal_n}/{formal_d} · 测试 {test_n}/{test_d}；待签署 {formal_p} · {test_p}（正式·测试）'
        )
    elif has_test and not has_formal:
        lines.append(f'知情签署总计（测试筛选） {test_n}/{test_d}；待签署 {test_p}')
    else:
        num_t = formal_n + test_n
        den_t = formal_d + test_d
        pend_t = formal_p + test_p
        lines.append(f'知情签署总计 {num_t}/{den_t}；待签署 {pend_t}')

    for b in sorted_batches:
        n, d, p = _batch_progress_components(b)
        ds = str(b.get('screening_date', ''))[:10]
        is_test = _is_test_screening_batch_for_export(b, settings)
        date_label = f'{ds}（测试筛选）' if is_test else ds
        att = _batch_attendance_label_for_export(settings, b, today_iso)
        lines.append(f'{date_label} 当日小计 {n}/{d}；待签署 {p}；{att}')
        if (
            not b.get('is_planned_placeholder')
            and int(b.get('total') or 0) == 0
            and int(b.get('expected_consent_rows') or 0) > 0
        ):
            lines.append(f'  待生成 {int(b.get("expected_consent_rows") or 0)} 份文档')
    return '\n'.join(lines)


def _normalize_planned_screening_dates(items) -> List[str]:
    """最多 4 条 YYYY-MM-DD，去重、按日期排序。"""
    date_re = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    out: List[str] = []
    for x in items or []:
        s = str(x).strip()[:10]
        if not date_re.match(s):
            continue
        try:
            datetime.strptime(s, '%Y-%m-%d')
        except ValueError:
            continue
        out.append(s)
    uniq = sorted(set(out))
    return uniq[:4]


def _validate_screening_signing_staff_names(norm_sched: list, dual_staffs_normalized: List[dict]) -> Optional[str]:
    """现场日 signing_staff_name 须为双签工作人员档案中的姓名（与 list_witness_staff 一致），并兼容请求内双签名单。"""
    from apps.protocol.services.witness_staff_service import witness_staff_allowed_name_set

    allowed = witness_staff_allowed_name_set()
    for s in dual_staffs_normalized or []:
        n = (s.get('name') or '').strip()
        if n:
            allowed.add(n)
    for row in norm_sched or []:
        if not isinstance(row, dict):
            continue
        sn = (row.get('signing_staff_name') or '').strip()
        if not sn:
            continue
        if sn not in allowed:
            return '请去之前配置中选择工作人员'
    return None


def _normalize_dual_sign_staffs(items) -> List[dict]:
    normalized = []
    for raw in items or []:
        name = str(getattr(raw, 'name', '') if hasattr(raw, 'name') else (raw.get('name') if isinstance(raw, dict) else '')).strip()
        if not name:
            continue
        staff = {
            'staff_id': str(getattr(raw, 'staff_id', '') if hasattr(raw, 'staff_id') else (raw.get('staff_id') if isinstance(raw, dict) else '')).strip(),
            'name': name,
            'id_card_no': str(getattr(raw, 'id_card_no', '') if hasattr(raw, 'id_card_no') else (raw.get('id_card_no') if isinstance(raw, dict) else '')).strip(),
            'email': str(getattr(raw, 'email', '') if hasattr(raw, 'email') else (raw.get('email') if isinstance(raw, dict) else '')).strip(),
            'phone': str(getattr(raw, 'phone', '') if hasattr(raw, 'phone') else (raw.get('phone') if isinstance(raw, dict) else '')).strip(),
            'identity_verified': bool(getattr(raw, 'identity_verified', False) if hasattr(raw, 'identity_verified') else (raw.get('identity_verified') if isinstance(raw, dict) else False)),
        }
        normalized.append(staff)
    return normalized


def _default_consent_settings() -> dict:
    return {
        'require_face_verify': False,
        'require_dual_sign': False,
        'require_comprehension_quiz': False,
        'enable_min_reading_duration': True,
        'min_reading_duration_seconds': 30,
        'dual_sign_staffs': [],
        'consent_launched': False,
        'consent_locked_at': None,
        'collect_id_card': False,
        'collect_screening_number': False,
        'collect_initials': False,
        'collect_subject_name': False,
        'collect_other_information': False,
        'enable_checkbox_recognition': False,
        'enable_staff_signature': False,
        'staff_signature_times': 1,
        'enable_subject_signature': False,
        'subject_signature_times': 1,
        'enable_guardian_signature': False,
        'guardian_parent_count': 1,
        'guardian_signature_times': 1,
        'enable_auto_sign_date': False,
        'planned_screening_dates': [],
        'screening_schedule': [],
        'consent_signing_staff_name': '',
        'consent_verify_test_staff_name': '',
        # 邮件「签名授权」已同意：列表为「已授权待测试/已测试待开始」，否则双签核验完成前仍为「待认证授权」
        'consent_verify_signature_authorized': False,
    }


def _get_consent_scope() -> dict:
    """获取知情配置模式与全局配置（单例）"""
    row = ConsentGlobalConfig.objects.first()
    if not row:
        return {
            'config_mode': ConsentConfigMode.PER_PROTOCOL,
            'global_settings': _default_consent_settings(),
        }
    raw = row.settings if isinstance(row.settings, dict) else {}
    merged = _default_consent_settings()
    merged['require_face_verify'] = bool(raw.get('require_face_verify', False))
    merged['require_dual_sign'] = bool(raw.get('require_dual_sign', False))
    merged['require_comprehension_quiz'] = bool(raw.get('require_comprehension_quiz', False))
    merged['enable_min_reading_duration'] = raw.get('enable_min_reading_duration') is not False
    if 'min_reading_duration_seconds' in raw:
        merged['min_reading_duration_seconds'] = max(0, int(raw.get('min_reading_duration_seconds') or 0))
    merged['dual_sign_staffs'] = _normalize_dual_sign_staffs(raw.get('dual_sign_staffs') or [])
    merged['consent_launched'] = False
    merged['consent_locked_at'] = None
    merged['collect_id_card'] = bool(raw.get('collect_id_card', False))
    merged['collect_screening_number'] = bool(raw.get('collect_screening_number', False))
    merged['collect_initials'] = bool(raw.get('collect_initials', False))
    merged['collect_subject_name'] = bool(raw.get('collect_subject_name', False))
    merged['collect_other_information'] = bool(raw.get('collect_other_information', False))
    merged['enable_checkbox_recognition'] = bool(raw.get('enable_checkbox_recognition', False))
    merged['enable_staff_signature'] = bool(raw.get('enable_staff_signature', False))
    merged['staff_signature_times'] = _clamp_1_or_2(raw.get('staff_signature_times'), 1)
    merged['enable_subject_signature'] = bool(raw.get('enable_subject_signature', False))
    merged['subject_signature_times'] = _clamp_1_or_2(raw.get('subject_signature_times'), 1)
    merged['enable_guardian_signature'] = bool(raw.get('enable_guardian_signature', False))
    merged['guardian_parent_count'] = _clamp_1_or_2(raw.get('guardian_parent_count'), 1)
    merged['guardian_signature_times'] = _clamp_1_or_2(raw.get('guardian_signature_times'), 1)
    merged['enable_auto_sign_date'] = bool(raw.get('enable_auto_sign_date', False))
    merged['planned_screening_dates'] = _normalize_planned_screening_dates(raw.get('planned_screening_dates'))
    return {
        'config_mode': row.config_mode,
        'global_settings': merged,
    }


def _save_consent_scope(config_mode: str = None, global_settings: dict = None) -> None:
    """保存知情配置模式与全局配置"""
    row = ConsentGlobalConfig.objects.first()
    if not row:
        row = ConsentGlobalConfig(config_mode=ConsentConfigMode.PER_PROTOCOL, settings={})
    if config_mode is not None:
        row.config_mode = config_mode
    if global_settings is not None:
        row.settings = {
            'require_face_verify': global_settings.get('require_face_verify', False),
            'require_dual_sign': global_settings.get('require_dual_sign', False),
            'require_comprehension_quiz': global_settings.get('require_comprehension_quiz', False),
            'enable_min_reading_duration': global_settings.get('enable_min_reading_duration') is not False,
            'min_reading_duration_seconds': max(0, int(global_settings.get('min_reading_duration_seconds', 30) or 0)),
            'dual_sign_staffs': _normalize_dual_sign_staffs(global_settings.get('dual_sign_staffs') or []),
            'collect_id_card': bool(global_settings.get('collect_id_card', False)),
            'collect_screening_number': bool(global_settings.get('collect_screening_number', False)),
            'collect_initials': bool(global_settings.get('collect_initials', False)),
            'collect_subject_name': bool(global_settings.get('collect_subject_name', False)),
            'collect_other_information': bool(global_settings.get('collect_other_information', False)),
            'enable_checkbox_recognition': bool(global_settings.get('enable_checkbox_recognition', False)),
            'enable_staff_signature': bool(global_settings.get('enable_staff_signature', False)),
            'staff_signature_times': _clamp_1_or_2(global_settings.get('staff_signature_times'), 1),
            'enable_subject_signature': bool(global_settings.get('enable_subject_signature', False)),
            'subject_signature_times': _clamp_1_or_2(global_settings.get('subject_signature_times'), 1),
            'enable_guardian_signature': bool(global_settings.get('enable_guardian_signature', False)),
            'guardian_parent_count': _clamp_1_or_2(global_settings.get('guardian_parent_count'), 1),
            'guardian_signature_times': _clamp_1_or_2(global_settings.get('guardian_signature_times'), 1),
            'enable_auto_sign_date': bool(global_settings.get('enable_auto_sign_date', False)),
            'planned_screening_dates': _normalize_planned_screening_dates(global_settings.get('planned_screening_dates')),
        }
    row.save()


def _merge_witness_staff_verification(staffs: List[dict]) -> List[dict]:
    """双签名单与 t_witness_staff 同步核验状态与联系方式。"""
    out: List[dict] = []
    for s in staffs or []:
        sid = str(s.get('staff_id') or '').strip()
        if sid.isdigit():
            ws = WitnessStaff.objects.filter(id=int(sid), is_deleted=False).first()
            if ws:
                ns = dict(s)
                ns['identity_verified'] = bool(ws.identity_verified)
                ns['name'] = ws.name
                ns['email'] = ws.email or ns.get('email', '')
                ns['phone'] = ws.phone or ns.get('phone', '')
                ns['id_card_no'] = ws.id_card_no or ns.get('id_card_no', '')
                out.append(ns)
                continue
        out.append(dict(s))
    return out


def _get_consent_settings(protocol: Protocol) -> dict:
    """获取协议知情配置（仅协议自身 parsed_data，不考虑全局模式）"""
    parsed_data = protocol.parsed_data if isinstance(protocol.parsed_data, dict) else {}
    raw = parsed_data.get('consent_settings') if isinstance(parsed_data, dict) else {}
    raw = raw if isinstance(raw, dict) else {}
    merged = _default_consent_settings()
    merged['require_face_verify'] = bool(raw.get('require_face_verify', False))
    merged['require_dual_sign'] = bool(raw.get('require_dual_sign', False))
    merged['require_comprehension_quiz'] = bool(raw.get('require_comprehension_quiz', False))
    merged['enable_min_reading_duration'] = raw.get('enable_min_reading_duration') is not False
    if 'min_reading_duration_seconds' in raw:
        merged['min_reading_duration_seconds'] = max(0, int(raw.get('min_reading_duration_seconds') or 0))
    merged['dual_sign_staffs'] = _merge_witness_staff_verification(
        _normalize_dual_sign_staffs(raw.get('dual_sign_staffs') or [])
    )
    merged['consent_launched'] = bool(raw.get('consent_launched', False))
    merged['consent_locked_at'] = raw.get('consent_locked_at')
    merged['collect_id_card'] = bool(raw.get('collect_id_card', False))
    merged['collect_screening_number'] = bool(raw.get('collect_screening_number', False))
    merged['collect_initials'] = bool(raw.get('collect_initials', False))
    merged['collect_subject_name'] = bool(raw.get('collect_subject_name', False))
    merged['collect_other_information'] = bool(raw.get('collect_other_information', False))
    merged['enable_checkbox_recognition'] = bool(raw.get('enable_checkbox_recognition', False))
    merged['enable_staff_signature'] = bool(raw.get('enable_staff_signature', False))
    merged['staff_signature_times'] = _clamp_1_or_2(raw.get('staff_signature_times'), 1)
    merged['enable_subject_signature'] = bool(raw.get('enable_subject_signature', False))
    merged['subject_signature_times'] = _clamp_1_or_2(raw.get('subject_signature_times'), 1)
    merged['enable_guardian_signature'] = bool(raw.get('enable_guardian_signature', False))
    merged['guardian_parent_count'] = _clamp_1_or_2(raw.get('guardian_parent_count'), 1)
    merged['guardian_signature_times'] = _clamp_1_or_2(raw.get('guardian_signature_times'), 1)
    merged['enable_auto_sign_date'] = bool(raw.get('enable_auto_sign_date', False))
    merged['consent_signing_staff_name'] = normalize_consent_signing_staff_storage(str(raw.get('consent_signing_staff_name') or ''))
    merged['consent_verify_test_staff_name'] = str(raw.get('consent_verify_test_staff_name') or '').strip()
    merged['consent_verify_signature_authorized'] = bool(raw.get('consent_verify_signature_authorized', False))
    merged['screening_schedule'] = _merge_consent_screening_schedule(raw)
    merged['planned_screening_dates'] = [x['date'] for x in merged['screening_schedule']]
    return merged


def _get_effective_consent_settings(protocol: Protocol) -> dict:
    """获取协议生效的知情配置（一律按协议独立存储；不再使用全局模式）。"""
    return _get_consent_settings(protocol)


def _staff_verification_snapshot(settings: dict) -> dict:
    """
    工作人员认证核验（依赖 dual_sign_staffs 已合并 WitnessStaff 核验结果）：
    - 无需核验：未启用双签
    - 待核验：已启用双签但尚无人员，或有人员但 0 人已核验
    - 核验中：部分已核验
    - 核验完成：全部人员已核验（且至少 1 人）
    """
    req = bool(settings.get('require_dual_sign'))
    staffs = settings.get('dual_sign_staffs') or []
    total = len(staffs)
    verified = sum(1 for s in staffs if s.get('identity_verified'))
    if not req:
        return {
            'dual_sign_staff_total': total,
            'verified_staff_count': verified,
            'staff_verification_status': '无需核验',
        }
    if total == 0:
        status = '待核验'
    elif verified >= total:
        status = '核验完成'
    elif verified == 0:
        status = '待核验'
    else:
        status = '核验中'
    return {
        'dual_sign_staff_total': total,
        'verified_staff_count': verified,
        'staff_verification_status': status,
    }


def _config_ready_for_consent_overview(protocol: Protocol, icf_count: int) -> bool:
    """每个签署节点均已至少保存过一次小程序规则；启用双签的节点须至少登记一名见证人员（不要求全员完成核验）。"""
    from apps.subject.models import ICFVersion
    from apps.subject.services.consent_service import get_effective_mini_sign_rules

    if icf_count <= 0:
        return False
    for icf in ICFVersion.objects.filter(protocol_id=protocol.id):
        if not getattr(icf, 'mini_sign_rules_saved', False):
            return False
        rules = get_effective_mini_sign_rules(protocol, icf)
        if rules.get('require_dual_sign'):
            snap = _staff_verification_snapshot(rules)
            if snap['dual_sign_staff_total'] <= 0:
                return False
    return True


def _any_icf_effective_require_dual_sign(protocol: Protocol) -> bool:
    """任一生效节点（含节点级 mini_sign_rules 覆盖）启用双签时视为需要双签核验，与 _config_ready_for_consent_overview 一致。"""
    from apps.subject.models import ICFVersion
    from apps.subject.services.consent_service import get_effective_mini_sign_rules

    for icf in ICFVersion.objects.filter(protocol_id=protocol.id):
        rules = get_effective_mini_sign_rules(protocol, icf)
        if rules.get('require_dual_sign'):
            return True
    return False


def _effective_require_dual_sign_for_protocol(protocol: Protocol, settings: dict) -> bool:
    """协议级开关与节点级生效规则取并集，避免仅节点启用双签时列表仍按「未启用双签」误判为配置中。"""
    return bool(settings.get('require_dual_sign')) or _any_icf_effective_require_dual_sign(protocol)


def _compute_consent_config_status(
    settings: dict,
    icf_count: int,
    config_ready: bool,
    earliest_screening_date: Optional[str],
    latest_screening_date: Optional[str],
    v_snap: dict,
    *,
    has_test_signing: bool = False,
    effective_require_dual_sign: Optional[bool] = None,
) -> str:
    """
    知情管理列表「知情配置状态」：
    - 未发布：待配置 / 配置中 / 待认证授权 / 已授权待测试 / 已测试待开始（未发布且已产生测试签署时）
    - 已发布：已测试待开始 / 进行中 / 已结束（按筛选日期与今日比较；「已测试待开始」与未发布同名，表示尚未进入现场筛选窗口）
    """
    from django.utils import timezone

    today_s = timezone.localdate().strftime('%Y-%m-%d')
    launched = bool(settings.get('consent_launched'))
    earliest = (str(earliest_screening_date).strip()[:10] if earliest_screening_date else '') or ''
    latest = (str(latest_screening_date).strip()[:10] if latest_screening_date else '') or ''

    if launched:
        if not earliest and not latest:
            return '已测试待开始'
        if earliest and latest:
            if today_s > latest:
                return '已结束'
            if today_s < earliest:
                return '已测试待开始'
            return '进行中'
        if earliest and not latest:
            if today_s < earliest:
                return '已测试待开始'
            if today_s == earliest:
                return '进行中'
            return '已结束'
        if latest and not earliest:
            if today_s > latest:
                return '已结束'
            return '进行中'
        return '已测试待开始'

    if icf_count <= 0:
        return '待配置'

    req_dual = (
        bool(effective_require_dual_sign)
        if effective_require_dual_sign is not None
        else bool(settings.get('require_dual_sign'))
    )

    if config_ready and req_dual:
        st = (v_snap.get('staff_verification_status') or '').strip()
        if st in ('待核验', '核验中'):
            return '待认证授权'
        if st == '核验完成':
            if settings.get('consent_verify_signature_authorized'):
                if has_test_signing:
                    return '已测试待开始'
                return '已授权待测试'
            return '待认证授权'

    # 未启用双签时：列表「授权核验测试」邮件同意后不走上文分支，否则 consent_verify_signature_authorized 已写入仍长期为「配置中」
    if (
        icf_count > 0
        and config_ready
        and not launched
        and settings.get('consent_verify_signature_authorized')
        and not req_dual
    ):
        if has_test_signing:
            return '已测试待开始'
        return '已授权待测试'

    if not config_ready:
        return '配置中'

    return '配置中'


def get_consent_config_status_for_protocol(protocol: Protocol) -> str:
    """与 consent-overview 列表一致的「知情配置状态」（供核验测试扫码、落地页校验复用）。"""
    from apps.subject.services.consent_service import (
        get_effective_mini_sign_rules,
        get_icf_versions,
        get_screening_batch_consent_stats,
    )
    from apps.protocol.services import witness_staff_service as _ws_svc

    icf_versions = get_icf_versions(protocol.id)
    icf_count = len(icf_versions)
    settings = _get_effective_consent_settings(protocol)
    dual_eff = _effective_require_dual_sign_for_protocol(protocol, settings)
    v_snap = _staff_verification_snapshot({**settings, 'require_dual_sign': dual_eff})
    batch_stats = get_screening_batch_consent_stats(
        protocol.id,
        screening_schedule=settings.get('screening_schedule'),
        planned_screening_dates=settings.get('planned_screening_dates'),
        icf_n=icf_count,
    )
    config_ready = _config_ready_for_consent_overview(protocol, icf_count)
    from apps.protocol.services import witness_staff_service as _ws_svc

    has_test_signing = _ws_svc.protocol_has_test_signing(protocol.id)
    return _compute_consent_config_status(
        settings,
        icf_count,
        config_ready,
        batch_stats.get('earliest_screening_date'),
        batch_stats.get('latest_screening_date'),
        v_snap,
        has_test_signing=has_test_signing,
        effective_require_dual_sign=dual_eff,
    )


def _execution_base_for_consent_scan_qr(base: str) -> str:
    """二维码常配 Django :8001；执行台 H5 在 :3007（同主机替换端口）。"""
    import re

    b = base.rstrip('/')
    if re.search(r':8001$', b):
        return re.sub(r':8001$', ':3007', b)
    return b


def _build_consent_test_scan_url(request, protocol_id: int, token: str) -> str:
    """执行台「核验测试」二维码：直接打开 H5（/execution/#/consent-test-scan），不拉起小程序。"""
    import re
    from urllib.parse import quote

    from .consent_scan_url_utils import normalize_consent_test_scan_public_base

    q = quote(token, safe='')
    frag = f'/consent-test-scan?p={protocol_id}&t={q}'
    base = (getattr(django_settings, 'CONSENT_TEST_SCAN_PUBLIC_BASE', None) or '').strip()
    if base:
        base = _execution_base_for_consent_scan_qr(normalize_consent_test_scan_public_base(base).rstrip('/'))
        return f'{base}/execution/#{frag}'
    scheme = 'https' if request.is_secure() else 'http'
    host = request.get_host()
    if re.search(r':8001$', host):
        host = re.sub(r':8001$', ':3007', host)
    return f'{scheme}://{host}/execution/#{frag}'


def _build_consent_test_scan_icf_items(protocol: Protocol) -> tuple[list | None, str | None]:
    """与 witness 联调队列结构一致，自协议首节点起全部待签 ICF。"""
    from apps.subject.models import ICFVersion
    from apps.protocol.services import protocol_service as protocol_svc
    from apps.subject.services.consent_service import (
        effective_required_reading_seconds_for_icf,
        get_effective_mini_sign_rules,
    )

    qs = (
        ICFVersion.objects.filter(
            protocol_id=protocol.id,
            is_active=True,
            mini_sign_rules_saved=True,
        ).order_by('display_order', '-create_time')
    )
    icf_list = list(qs)
    if not icf_list:
        return None, '当前协议无已保存小程序签署规则的节点，请在执行台知情配置中维护各节点规则并至少保存一次后再试'

    items = []
    for icf in icf_list:
        sec = effective_required_reading_seconds_for_icf(icf, protocol)
        body_html = protocol_svc.resolve_icf_body_html_for_witness_dev(icf)
        rules = get_effective_mini_sign_rules(protocol, icf)
        enable_subj = bool(rules.get('enable_subject_signature', False))
        subj_times = _clamp_1_or_2(rules.get('subject_signature_times'), 1) if enable_subj else 0
        sup_labels = rules.get('supplemental_collect_labels')
        if not isinstance(sup_labels, list):
            sup_labels = []
        sup_labels = [str(x).strip() for x in sup_labels if str(x).strip()][:20]
        items.append(
            {
                'icf_version_id': icf.id,
                'node_title': (icf.node_title or '').strip() or f'签署节点 {icf.version}',
                'version': icf.version,
                'required_reading_duration_seconds': sec,
                'content': body_html,
                'enable_subject_signature': enable_subj,
                'subject_signature_times': subj_times,
                'enable_auto_sign_date': bool(rules.get('enable_auto_sign_date', False)),
                'enable_checkbox_recognition': bool(rules.get('enable_checkbox_recognition', False)),
                'supplemental_collect_labels': sup_labels,
                'collect_other_information': bool(rules.get('collect_other_information', False)),
            }
        )
    return items, None


def _is_consent_launched(protocol: Protocol) -> bool:
    """协议知情是否已发布（发布后禁止编辑节点）"""
    return bool(_get_consent_settings(protocol).get('consent_launched'))


def _save_consent_settings(protocol: Protocol, settings_data: dict) -> None:
    parsed_data = protocol.parsed_data if isinstance(protocol.parsed_data, dict) else {}
    parsed_data = dict(parsed_data)
    parsed_data['consent_settings'] = settings_data
    protocol.parsed_data = parsed_data
    protocol.save(update_fields=['parsed_data', 'update_time'])


def _propagate_dual_sign_after_consent_save(protocol: Protocol, settings_data: dict) -> None:
    """将双签开关与名单同步到同产品线/委托方下的其他协议（及本协议各已保存节点）。"""
    from .services.dual_sign_project_sync import propagate_dual_sign_across_project

    propagate_dual_sign_across_project(
        protocol,
        bool(settings_data.get('require_dual_sign', False)),
        settings_data.get('dual_sign_staffs'),
    )


@router.post('/reorder-consent', summary='调整知情管理协议展示顺序')
@require_any_permission(['protocol.protocol.update', 'protocol.protocol.create', 'protocol.protocol.read'])
def reorder_consent(request, data: ReorderConsentIn):
    """按 id_order 顺序更新协议的 consent_display_order"""
    account = _get_account_from_request(request)
    services.reorder_consent_protocols(account, data.id_order)
    return {'code': 200, 'msg': 'OK'}


# ============================================================================
# 知情管理（执行台）：协议概览、ICF 版本与签署记录
# ============================================================================
@router.get('/consent-config-assignees', summary='知情配置负责人候选（治理台 QA质量管理）')
@require_any_permission(['protocol.protocol.read', 'protocol.protocol.create', 'protocol.protocol.update', 'subject.subject.read'])
def consent_config_assignees(request):
    """具备全局角色 qa（QA质量管理）的治理台账号列表，供新建/编辑项目时单选。"""
    items = services.list_consent_config_assignee_accounts()
    return {'code': 200, 'msg': 'OK', 'data': {'items': items}}


@router.get('/consent-overview', summary='协议知情概览（知情管理入口）')
@require_any_permission(['protocol.protocol.read', 'protocol.protocol.create', 'protocol.protocol.update', 'subject.subject.read'])
def consent_overview(
    request,
    page: int = 1,
    page_size: int = 10,
    keyword: str = None,
    config_status: str = None,
    date_start: str = None,
    date_end: str = None,
    focus_protocol_id: int = None,
):
    """返回协议列表及每个协议的知情配置状态与签署统计。支持 keyword、config_status、date_start/date_end 筛选。

    focus_protocol_id：邮件/深链定位某协议所在分页时传入，服务端计算其所在页并返回对应 items 与 data.page（列表仍保持项目级视图，非进入单项目配置）。
    """
    # 临时策略：知情管理入口不做项目级 scope 过滤，避免工作台联调断点导致列表为空
    account = None
    kw = keyword.strip() if keyword and keyword.strip() else None
    ds = date_start.strip() if date_start and date_start.strip() else None
    de = date_end.strip() if date_end and date_end.strip() else None
    cfg = config_status.strip() if config_status and config_status.strip() else None

    # 性能优化：
    # - 常规列表（无 config_status / 无 focus_protocol_id）只计算当前页，避免 500 条全量聚合导致超时
    # - 仅在需要“全量计算后再筛选/定位”时走全量模式
    focus_only_mode = (focus_protocol_id is not None) and not bool(cfg)
    sql_filter_wait_config_mode = cfg == '待配置'
    full_scan_mode = (bool(cfg) and not sql_filter_wait_config_mode) or False
    req_page = max(1, int(page or 1))
    req_page_size = max(1, min(int(page_size or 10), 100))

    # 常规分页场景做短 TTL 缓存，减少首页反复刷新时的逐协议计算压力。
    if not full_scan_mode and not focus_only_mode:
        try:
            from django.core.cache import cache

            uid = None
            auth_payload = getattr(request, 'auth', None)
            if isinstance(auth_payload, dict):
                uid = auth_payload.get('user_id')
            try:
                gen = int(cache.get('protocol:consent_overview:cache_gen') or 0)
            except (TypeError, ValueError):
                gen = 0
            cache_key = (
                f'protocol:consent_overview:v3:g={gen}:uid={uid or "-"}:'
                f'p={req_page}:s={req_page_size}:kw={kw or ""}:ds={ds or ""}:de={de or ""}:cfg={cfg or ""}'
            )
            cached = cache.get(cache_key)
            if isinstance(cached, dict):
                return cached
        except Exception:
            cache = None
            cache_key = None
    else:
        cache = None
        cache_key = None

    if focus_only_mode:
        # 深链定位只为算出目标页码：先取有序 id 列表，再仅查询命中页，避免全量逐协议聚合。
        locate_res = services.list_protocols(
            page=1,
            page_size=10000,
            keyword=kw,
            date_start=ds,
            date_end=de,
            account=account,
        )
        ordered_ids = [int(getattr(p, 'id')) for p in locate_res.get('items', []) if getattr(p, 'id', None)]
        total_from_db = int(locate_res.get('total') or 0)
        effective_page = req_page
        try:
            fpid = int(focus_protocol_id)
        except (TypeError, ValueError):
            fpid = None
        if fpid and fpid in ordered_ids:
            effective_page = ordered_ids.index(fpid) // req_page_size + 1
        max_page = max(1, (total_from_db + req_page_size - 1) // req_page_size) if total_from_db else 1
        if effective_page < 1:
            effective_page = 1
        if effective_page > max_page:
            effective_page = max_page
        result = services.list_protocols(
            page=effective_page,
            page_size=req_page_size,
            keyword=kw,
            date_start=ds,
            date_end=de,
            account=account,
        )
    else:
        query_page = 1 if full_scan_mode else req_page
        query_page_size = 500 if full_scan_mode else req_page_size
        result = services.list_protocols(
            page=query_page,
            page_size=query_page_size,
            keyword=kw,
            date_start=ds,
            date_end=de,
            account=account,
            without_icf=sql_filter_wait_config_mode,
        )

    total_from_db = int(result.get('total') or 0)
    if not full_scan_mode and not focus_only_mode:
        max_page = max(1, (total_from_db + req_page_size - 1) // req_page_size) if total_from_db else 1
        effective_page = req_page if req_page <= max_page else max_page
        if effective_page != req_page:
            result = services.list_protocols(
                page=effective_page,
                page_size=req_page_size,
                keyword=kw,
                date_start=ds,
                date_end=de,
                account=account,
                without_icf=sql_filter_wait_config_mode,
            )
    from apps.subject.services.consent_service import (
        get_effective_mini_sign_rules,
        get_screening_batch_consent_stats,
    )
    from apps.subject.models import ICFVersion, SubjectConsent
    import hashlib
    import json

    try:
        from django.core.cache import cache as _cache_backend
    except Exception:
        _cache_backend = None

    protocol_ids = [int(getattr(p, 'id')) for p in result['items'] if getattr(p, 'id', None)]
    icf_map = {}
    if protocol_ids:
        icf_rows = list(
            ICFVersion.objects.filter(protocol_id__in=protocol_ids)
            .only('id', 'protocol_id', 'mini_sign_rules_saved', 'mini_sign_rules', 'create_time')
            .order_by('protocol_id', 'id')
        )
        for icf in icf_rows:
            icf_map.setdefault(int(icf.protocol_id), []).append(icf)
    test_signed_protocol_ids = set()
    if protocol_ids:
        # SQL 批量拉取后端所需最小字段，避免 N 次 protocol_has_test_signing()。
        test_rows = SubjectConsent.objects.filter(
            icf_version__protocol_id__in=protocol_ids,
            is_signed=True,
            is_deleted=False,
        ).values_list('icf_version__protocol_id', 'signature_data')
        for pid, sig in test_rows:
            sd = sig if isinstance(sig, dict) else {}
            if str(sd.get('signing_kind') or '').strip().lower() == 'test':
                test_signed_protocol_ids.add(int(pid))
    items = []
    for p in result['items']:
        icf_versions = icf_map.get(int(p.id), [])
        icf_count = len(icf_versions)
        settings = _get_effective_consent_settings(p)
        # 优先复用当前循环已取到的 ICF，避免同一协议重复拉取/重复计算。
        icf_model_rows = [x for x in icf_versions if hasattr(x, 'id')]
        if icf_count > 0 and len(icf_model_rows) == icf_count:
            any_icf_dual_sign = False
            config_ready = True
            for icf in icf_model_rows:
                if not getattr(icf, 'mini_sign_rules_saved', False):
                    config_ready = False
                    continue
                rules = get_effective_mini_sign_rules(p, icf)
                req_dual = bool(rules.get('require_dual_sign'))
                if req_dual:
                    any_icf_dual_sign = True
                    snap = _staff_verification_snapshot(rules)
                    if snap['dual_sign_staff_total'] <= 0:
                        config_ready = False
            dual_eff = bool(settings.get('require_dual_sign')) or any_icf_dual_sign
        else:
            # 兼容历史迁移未齐的降级路径
            dual_eff = _effective_require_dual_sign_for_protocol(p, settings)
            config_ready = _config_ready_for_consent_overview(p, icf_count)

        v_snap = _staff_verification_snapshot({**settings, 'require_dual_sign': dual_eff})
        sched = settings.get('screening_schedule')
        planned = settings.get('planned_screening_dates')
        stats_sig = hashlib.md5(
            json.dumps(
                {'schedule': sched or [], 'planned': planned or []},
                ensure_ascii=False,
                sort_keys=True,
                separators=(',', ':'),
            ).encode('utf-8')
        ).hexdigest()
        batch_cache_key = (
            f'protocol:consent_batch_stats:v3:pid={p.id}:icf={icf_count}:u='
            f'{int(p.update_time.timestamp()) if getattr(p, "update_time", None) else 0}:sig={stats_sig}'
        )
        batch_stats = _cache_backend.get(batch_cache_key) if _cache_backend is not None else None
        if not isinstance(batch_stats, dict):
            batch_stats = get_screening_batch_consent_stats(
                p.id,
                screening_schedule=sched,
                planned_screening_dates=planned,
                icf_n=icf_count,
            )
            if _cache_backend is not None:
                try:
                    _cache_backend.set(batch_cache_key, batch_stats, timeout=60)
                except Exception:
                    pass
        verified_staff_count = v_snap['verified_staff_count']
        has_test_signing = int(p.id) in test_signed_protocol_ids
        config_status_val = _compute_consent_config_status(
            settings,
            icf_count,
            config_ready,
            batch_stats.get('earliest_screening_date'),
            batch_stats.get('latest_screening_date'),
            v_snap,
            has_test_signing=has_test_signing,
            effective_require_dual_sign=dual_eff,
        )
        # 列表页当前主要展示 screening_batches 的签署进度，避免逐项目触发重型全量签署统计。
        signed_total = sum(int(b.get('signed_count') or 0) for b in batch_stats['batches'])
        pending_total = sum(int(b.get('pending_count') or 0) for b in batch_stats['batches'])
        total_rows = signed_total + pending_total
        from .consent_test_tokens import sign_consent_test_scan_token

        consent_test_token = sign_consent_test_scan_token(p.id)
        consent_test_scan_url = _build_consent_test_scan_url(request, p.id, consent_test_token)
        items.append({
            'id': p.id,
            'code': getattr(p, 'code', None) or '',
            'title': p.title,
            'create_time': p.create_time.isoformat() if p.create_time else None,
            'consent_last_update_at': p.update_time.isoformat() if getattr(p, 'update_time', None) else None,
            'consent_display_order': getattr(p, 'consent_display_order', 0),
            'icf_count': icf_count,
            'config_status': config_status_val,
            'consent_test_scan_url': consent_test_scan_url,
            'total': total_rows,
            'signed_count': signed_total,
            'pending_count': pending_total,
            'require_dual_sign': dual_eff,
            'verified_staff_count': verified_staff_count,
            'dual_sign_staff_total': v_snap['dual_sign_staff_total'],
            'staff_verification_status': v_snap['staff_verification_status'],
            'mini_app_ready': config_ready,
            'screening_batches': batch_stats['batches'],
            'screening_batch_count': batch_stats['batch_count'],
            'earliest_screening_date': batch_stats['earliest_screening_date'],
            'latest_screening_date': batch_stats['latest_screening_date'],
            'screening_batch_source': batch_stats.get('batch_source') or 'none',
            'planned_screening_dates': settings.get('planned_screening_dates') or [],
            'screening_schedule': settings.get('screening_schedule') or [],
            'screening_schedule_summary': _screening_schedule_summary(settings.get('screening_schedule') or []),
            'consent_launched': bool(settings.get('consent_launched')),
            'consent_config_account_id': getattr(p, 'consent_config_account_id', None),
            'consent_signing_staff_name': (settings.get('consent_signing_staff_name') or '').strip() or None,
            'consent_verify_test_staff_name': (settings.get('consent_verify_test_staff_name') or '').strip() or None,
        })
    acc_ids = {x['consent_config_account_id'] for x in items if x.get('consent_config_account_id')}
    name_map = {}
    if acc_ids:
        from apps.identity.models import Account

        for a in Account.objects.filter(id__in=acc_ids, is_deleted=False).only('id', 'display_name', 'username'):
            name_map[a.id] = ((a.display_name or '').strip() or a.username)
    for x in items:
        aid = x.get('consent_config_account_id')
        x['consent_config_display_name'] = name_map.get(aid) if aid else None
    if full_scan_mode and cfg:
        items = [x for x in items if x['config_status'] == cfg]

    if not full_scan_mode:
        resp = {
            'code': 200,
            'msg': 'OK',
            'data': {
                'items': items,
                'total': total_from_db,
                'page': effective_page,
                'page_size': req_page_size,
            },
        }
        if cache is not None and cache_key:
            try:
                cache.set(cache_key, resp, timeout=20)
            except Exception:
                pass
        return resp

    total = len(items)
    effective_page = req_page
    if focus_protocol_id is not None:
        try:
            fpid = int(focus_protocol_id)
        except (TypeError, ValueError):
            fpid = None
        if fpid and fpid > 0:
            for idx, it in enumerate(items):
                if it.get('id') == fpid:
                    effective_page = idx // req_page_size + 1
                    break
    max_page = max(1, (total + req_page_size - 1) // req_page_size) if total else 1
    if effective_page < 1:
        effective_page = 1
    if effective_page > max_page:
        effective_page = max_page
    offset = (effective_page - 1) * req_page_size
    items = items[offset:offset + req_page_size]
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': items,
            'total': total,
            'page': effective_page,
            'page_size': req_page_size,
        },
    }


@router.get('/consent-overview/export', summary='导出知情管理项目列表')
@require_any_permission(['protocol.protocol.read', 'protocol.protocol.create', 'protocol.protocol.update', 'subject.subject.read'])
def consent_overview_export(
    request,
    keyword: str = None,
    config_status: str = None,
    date_start: str = None,
    date_end: str = None,
):
    """导出知情管理项目列表为 Excel，支持与列表相同的筛选条件"""
    from django.http import HttpResponse
    from datetime import date
    account = None
    result = services.list_protocols(
        page=1,
        page_size=10000,
        keyword=keyword.strip() if keyword and keyword.strip() else None,
        date_start=date_start.strip() if date_start and date_start.strip() else None,
        date_end=date_end.strip() if date_end and date_end.strip() else None,
        account=account,
    )
    from apps.subject.services.consent_service import (
        get_icf_versions,
        get_screening_batch_consent_stats,
    )
    items = []
    for p in result['items']:
        icf_versions = get_icf_versions(p.id)
        icf_count = len(icf_versions)
        settings = _get_effective_consent_settings(p)
        dual_eff = _effective_require_dual_sign_for_protocol(p, settings)
        v_snap = _staff_verification_snapshot({**settings, 'require_dual_sign': dual_eff})
        batch_stats = get_screening_batch_consent_stats(
            p.id,
            screening_schedule=settings.get('screening_schedule'),
            planned_screening_dates=settings.get('planned_screening_dates'),
            icf_n=icf_count,
        )
        config_ready = _config_ready_for_consent_overview(p, icf_count)
        from apps.protocol.services import witness_staff_service as _ws_svc_exp

        config_status_val = _compute_consent_config_status(
            settings,
            icf_count,
            config_ready,
            batch_stats.get('earliest_screening_date'),
            batch_stats.get('latest_screening_date'),
            v_snap,
            has_test_signing=_ws_svc_exp.protocol_has_test_signing(p.id),
            effective_require_dual_sign=dual_eff,
        )
        if config_status and config_status.strip() and config_status_val != config_status.strip():
            continue
        items.append({
            'protocol': p,
            'settings': settings,
            'config_status': config_status_val,
            'staff_verification_status': v_snap['staff_verification_status'],
            'verified_staff_count': v_snap['verified_staff_count'],
            'dual_sign_staff_total': v_snap['dual_sign_staff_total'],
            'batches': batch_stats['batches'],
            'earliest_screening_date': batch_stats['earliest_screening_date'],
            'latest_screening_date': batch_stats['latest_screening_date'],
        })
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return 500, {'code': 500, 'msg': 'Excel 导出依赖 openpyxl 未安装', 'data': None}
    export_acc_ids = set()
    for row_data in items:
        pr = row_data['protocol']
        aid = getattr(pr, 'consent_config_account_id', None)
        if aid:
            export_acc_ids.add(aid)
    export_name_map = {}
    if export_acc_ids:
        from apps.identity.models import Account

        for a in Account.objects.filter(id__in=export_acc_ids, is_deleted=False).only('id', 'display_name', 'username'):
            export_name_map[a.id] = ((a.display_name or '').strip() or a.username)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '知情管理项目列表'
    # 与执行台知情管理列表「导出」列顺序一致（不含勾选列、操作列）
    headers = [
        '序号',
        '项目编号',
        '项目名称',
        '知情配置人员',
        '知情配置状态',
        '筛选开始日期',
        '筛选结束日期',
        '签署进度',
    ]
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = header_align
    for row_idx, row_data in enumerate(items, 2):
        p = row_data['protocol']
        settings = row_data['settings']
        code = getattr(p, 'code', None) or ''
        title_disp = _display_protocol_title_for_export(getattr(p, 'title', None), code)
        earliest = _first_configured_screening_date_for_export(
            settings.get('screening_schedule') or [],
            row_data.get('earliest_screening_date'),
        )
        latest = _last_configured_screening_date_for_export(
            settings.get('screening_schedule') or [],
            row_data.get('latest_screening_date'),
        )
        progress_text = _format_consent_signing_progress_export(
            row_data['config_status'],
            settings,
            row_data['batches'],
        )
        ca_id = getattr(p, 'consent_config_account_id', None)
        assignee_name = export_name_map.get(ca_id, '—') if ca_id else '—'
        serial = row_idx - 1
        ws.cell(row=row_idx, column=1, value=serial).alignment = Alignment(vertical='top')
        ws.cell(row=row_idx, column=2, value=code).alignment = Alignment(vertical='top')
        ws.cell(row=row_idx, column=3, value=title_disp).alignment = wrap_top
        ws.cell(row=row_idx, column=4, value=assignee_name).alignment = Alignment(vertical='top')
        ws.cell(row=row_idx, column=5, value=row_data['config_status']).alignment = Alignment(vertical='top')
        ws.cell(row=row_idx, column=6, value=earliest or '—').alignment = Alignment(vertical='top')
        ws.cell(row=row_idx, column=7, value=latest or '—').alignment = Alignment(vertical='top')
        prog_cell = ws.cell(row=row_idx, column=8, value=progress_text)
        prog_cell.alignment = wrap_top
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"知情管理项目列表_{date.today():%Y%m%d}.xlsx"
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _check_protocol_visible(request, protocol_id: int):
    """校验协议可见性，返回 (account, protocol_or_none)"""
    account = _get_account_from_request(request)
    # 临时策略：知情管理链路仅校验权限码，不做项目级 scope 过滤
    protocol = Protocol.objects.filter(id=protocol_id, is_deleted=False).first()
    return account, protocol


class ConsentConfigScopeIn(Schema):
    config_mode: Optional[str] = None
    global_settings: Optional[dict] = None


@router.get('/consent-config/scope', summary='获取知情配置模式与全局配置（固定返回按协议；global_settings 仅兼容旧客户端）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def get_consent_config_scope(request):
    scope = _get_consent_scope()
    # 产品已取消「全局配置」：接口始终声明为 per_protocol，避免旧数据误导
    scope = dict(scope)
    scope['config_mode'] = ConsentConfigMode.PER_PROTOCOL
    scope['consent_config_policy'] = 'per_protocol_only'
    return {'code': 200, 'msg': 'OK', 'data': scope}


@router.put('/consent-config/scope', summary='更新知情配置模式与全局配置（已废弃全局模式，固定为按协议）')
@require_permission('protocol.protocol.update')
def update_consent_config_scope(request, data: ConsentConfigScopeIn):
    if data.config_mode is not None and data.config_mode not in (ConsentConfigMode.GLOBAL, ConsentConfigMode.PER_PROTOCOL):
        return 400, {'code': 400, 'msg': 'config_mode 必须为 global 或 per_protocol', 'data': None}
    if data.config_mode is not None and data.config_mode == ConsentConfigMode.GLOBAL:
        return 400, {'code': 400, 'msg': '已统一为按协议独立配置，不再支持全局模式。请直接在各个协议下配置，或使用「引用其他协议配置」。', 'data': None}
    global_settings = None
    if data.global_settings is not None:
        global_settings = {
            'require_face_verify': bool(data.global_settings.get('require_face_verify', False)),
            'require_dual_sign': bool(data.global_settings.get('require_dual_sign', False)),
            'require_comprehension_quiz': bool(data.global_settings.get('require_comprehension_quiz', False)),
            'enable_min_reading_duration': data.global_settings.get('enable_min_reading_duration') is not False,
            'min_reading_duration_seconds': max(0, int(data.global_settings.get('min_reading_duration_seconds', 30) or 0)),
            'dual_sign_staffs': _normalize_dual_sign_staffs(data.global_settings.get('dual_sign_staffs') or []),
            'collect_id_card': bool(data.global_settings.get('collect_id_card', False)),
            'collect_screening_number': bool(data.global_settings.get('collect_screening_number', False)),
            'collect_initials': bool(data.global_settings.get('collect_initials', False)),
            'collect_subject_name': bool(data.global_settings.get('collect_subject_name', False)),
            'collect_other_information': bool(data.global_settings.get('collect_other_information', False)),
            'enable_checkbox_recognition': bool(data.global_settings.get('enable_checkbox_recognition', False)),
            'enable_staff_signature': bool(data.global_settings.get('enable_staff_signature', False)),
            'staff_signature_times': _clamp_1_or_2(data.global_settings.get('staff_signature_times'), 1),
            'enable_subject_signature': bool(data.global_settings.get('enable_subject_signature', False)),
            'subject_signature_times': _clamp_1_or_2(data.global_settings.get('subject_signature_times'), 1),
            'enable_guardian_signature': bool(data.global_settings.get('enable_guardian_signature', False)),
            'guardian_parent_count': _clamp_1_or_2(data.global_settings.get('guardian_parent_count'), 1),
            'guardian_signature_times': _clamp_1_or_2(data.global_settings.get('guardian_signature_times'), 1),
            'planned_screening_dates': _normalize_planned_screening_dates(data.global_settings.get('planned_screening_dates')),
        }
        if global_settings['require_dual_sign']:
            global_settings['dual_sign_staffs'] = _merge_witness_staff_verification(
                _normalize_dual_sign_staffs(global_settings.get('dual_sign_staffs') or [])
            )
    _save_consent_scope(config_mode=data.config_mode, global_settings=global_settings)
    scope = _get_consent_scope()
    return {'code': 200, 'msg': 'OK', 'data': scope}


@router.get(
    '/{protocol_id}/consent-settings',
    summary='获取知情配置（执行台配置中心）',
    response={200: dict, 404: dict},
)
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def get_consent_settings(request, protocol_id: int):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    settings_data = _get_effective_consent_settings(protocol)
    return {'code': 200, 'msg': 'OK', 'data': settings_data}


@router.put(
    '/{protocol_id}/consent-settings',
    summary='更新知情配置（执行台配置中心）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def update_consent_settings(request, protocol_id: int, data: ConsentSettingsIn):
    trace_id = f"consent_settings_update_{protocol_id}_{int(time.time() * 1000)}"
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再修改配置'}
    logger.info(
        '[%s] start update_consent_settings protocol_id=%s require_dual_sign=%s',
        trace_id,
        protocol_id,
        bool(getattr(data, 'require_dual_sign', False)),
    )
    from apps.subject.services.consent_service import (
        _normalize_screening_schedule_for_stats as _norm_ss,
        validate_screening_schedule_test_rules,
    )

    sched_raw = getattr(data, 'screening_schedule', None)
    if sched_raw is not None:
        norm_sched = _norm_ss(_screening_schedule_input_to_plain_dicts(sched_raw))
    else:
        norm_sched = [
            {'date': d, 'target_count': 1, 'is_test_screening': False}
            for d in _normalize_planned_screening_dates(getattr(data, 'planned_screening_dates', None))
        ]
    verr = validate_screening_schedule_test_rules(norm_sched)
    if verr:
        logger.warning('[%s] invalid screening_schedule protocol_id=%s err=%s', trace_id, protocol_id, verr)
        return 400, {'code': 400, 'msg': verr}
    norm_dual_for_row = _normalize_dual_sign_staffs(data.dual_sign_staffs)
    name_err = _validate_screening_signing_staff_names(norm_sched, norm_dual_for_row)
    if name_err:
        logger.warning('[%s] invalid signing_staff_name protocol_id=%s err=%s', trace_id, protocol_id, name_err)
        return 400, {'code': 400, 'msg': name_err}
    csn = normalize_consent_signing_staff_storage(getattr(data, 'consent_signing_staff_name', None) or '')
    if csn:
        from apps.protocol.services.witness_staff_service import witness_staff_allowed_name_set

        allowed_csn = witness_staff_allowed_name_set()
        for s in norm_dual_for_row:
            n = (s.get('name') or '').strip()
            if n:
                allowed_csn.add(n)
        for name in split_consent_signing_staff_names(csn):
            if name not in allowed_csn:
                logger.warning(
                    '[%s] consent_signing_staff_name out of whitelist protocol_id=%s name=%s',
                    trace_id,
                    protocol_id,
                    name,
                )
                return 400, {'code': 400, 'msg': '知情签署工作人员须从双签工作人员名单中选择'}
    settings_data = {
        # 人脸认证签署暂未开放，固定关闭（忽略客户端传入）
        'require_face_verify': False,
        'require_dual_sign': bool(data.require_dual_sign),
        'require_comprehension_quiz': bool(data.require_comprehension_quiz),
        'enable_min_reading_duration': getattr(data, 'enable_min_reading_duration', True) is not False,
        'min_reading_duration_seconds': max(0, int(data.min_reading_duration_seconds)),
        'dual_sign_staffs': _merge_witness_staff_verification(norm_dual_for_row),
        'collect_id_card': bool(getattr(data, 'collect_id_card', False)),
        'collect_screening_number': bool(getattr(data, 'collect_screening_number', False)),
        'collect_initials': bool(getattr(data, 'collect_initials', False)),
        'collect_subject_name': bool(getattr(data, 'collect_subject_name', False)),
        'collect_other_information': bool(getattr(data, 'collect_other_information', False)),
        'enable_checkbox_recognition': bool(getattr(data, 'enable_checkbox_recognition', False)),
        'enable_staff_signature': bool(getattr(data, 'enable_staff_signature', False)),
        'staff_signature_times': _clamp_1_or_2(getattr(data, 'staff_signature_times', None), 1),
        'enable_subject_signature': bool(getattr(data, 'enable_subject_signature', False)),
        'subject_signature_times': _clamp_1_or_2(getattr(data, 'subject_signature_times', None), 1),
        'enable_guardian_signature': bool(getattr(data, 'enable_guardian_signature', False)),
        'guardian_parent_count': _clamp_1_or_2(getattr(data, 'guardian_parent_count', None), 1),
        'guardian_signature_times': _clamp_1_or_2(getattr(data, 'guardian_signature_times', None), 1),
        'enable_auto_sign_date': bool(getattr(data, 'enable_auto_sign_date', False)),
        'screening_schedule': norm_sched,
        'planned_screening_dates': [x['date'] for x in norm_sched],
        'consent_signing_staff_name': csn,
    }
    existing = _get_consent_settings(protocol)
    settings_data['consent_launched'] = existing.get('consent_launched', False)
    settings_data['consent_locked_at'] = existing.get('consent_locked_at')
    settings_data['consent_verify_test_staff_name'] = (existing.get('consent_verify_test_staff_name') or '').strip()
    settings_data['consent_verify_signature_authorized'] = bool(existing.get('consent_verify_signature_authorized', False))
    _save_consent_settings(protocol, settings_data)
    logger.info(
        '[%s] saved consent_settings protocol_id=%s schedule_rows=%s',
        trace_id,
        protocol_id,
        len(settings_data.get('screening_schedule') or []),
    )
    try:
        _propagate_dual_sign_after_consent_save(protocol, settings_data)
    except Exception:
        logger.exception('[%s] 双签配置跨协议同步失败 protocol_id=%s', trace_id, protocol_id)
    protocol.refresh_from_db()
    logger.info('[%s] finish update_consent_settings protocol_id=%s', trace_id, protocol_id)
    return {'code': 200, 'msg': 'OK', 'data': _get_effective_consent_settings(protocol)}


@router.post('/{protocol_id}/consent-launch', summary='发布/取消发布知情（执行台）')
@require_permission('protocol.protocol.update')
def consent_launch(request, protocol_id: int, data: ConsentLaunchIn):
    """统一发布或取消发布。发布后禁止编辑节点与配置；取消发布后可恢复编辑。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from django.utils import timezone
    from apps.subject.models import ICFVersion
    from apps.subject.services.consent_service import (
        get_icf_versions,
        get_effective_mini_sign_rules,
        _normalize_screening_schedule_for_stats as _norm_launch_sched,
    )
    settings_data = dict(_get_consent_settings(protocol))
    if bool(data.launched):
        icf_list = get_icf_versions(protocol_id)
        if len(icf_list) <= 0:
            return 400, {'code': 400, 'msg': '发布前请至少创建一个签署节点'}
        for icf in ICFVersion.objects.filter(protocol_id=protocol_id):
            if not getattr(icf, 'mini_sign_rules_saved', False):
                return 400, {'code': 400, 'msg': '发布前请为每个签署节点保存小程序签署规则（至少保存一次）'}
            rules = get_effective_mini_sign_rules(protocol, icf)
            if rules.get('require_dual_sign'):
                snap = _staff_verification_snapshot(rules)
                if snap['dual_sign_staff_total'] <= 0:
                    return 400, {'code': 400, 'msg': '某签署节点启用了双签，请在该节点规则中添加至少一名见证工作人员'}
        sched_chk = _norm_launch_sched(settings_data.get('screening_schedule'))
        if any(x.get('is_test_screening') for x in sched_chk):
            return 400, {
                'code': 400,
                'msg': '发布前请删除或取消所有「测试筛选」计划行（测试完成后须移除后方可发布）',
            }
    settings_data['consent_launched'] = bool(data.launched)
    settings_data['consent_locked_at'] = timezone.now().isoformat() if data.launched else None
    _save_consent_settings(protocol, settings_data)
    if data.launched:
        ICFVersion.objects.filter(protocol_id=protocol_id).update(is_active=True)
    protocol.refresh_from_db()
    return {'code': 200, 'msg': 'OK', 'data': _get_effective_consent_settings(protocol)}


@router.get('/{protocol_id}/icf-versions', summary='ICF 版本列表（知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def list_icf_versions(request, protocol_id: int):
    """执行台知情管理：按协议列出 ICF 版本"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import get_icf_versions
    items = get_icf_versions(protocol_id)
    rows = []
    for i in items:
        is_dict = isinstance(i, dict)
        _get = lambda k, d=None: i.get(k, d) if is_dict else getattr(i, k, d)
        try:
            reading_seconds = int(_get('required_reading_duration_seconds', 0) or 0)
        except Exception:
            reading_seconds = 0
        ct, ut = _get('create_time'), _get('update_time')
        mr = _get('mini_sign_rules', None)
        if not isinstance(mr, dict):
            mr = {}
        rows.append({
            'id': _get('id'),
            'protocol_id': _get('protocol_id'),
            'version': _get('version'),
            'node_title': _get('node_title') or '',
            'display_order': int(_get('display_order', 0) or 0),
            'file_path': _get('file_path') or '',
            'content': _get('content') or '',
            'is_active': _get('is_active'),
            'required_reading_duration_seconds': reading_seconds,
            'create_time': ct.isoformat() if ct else '',
            'update_time': ut.isoformat() if ut else '',
            'mini_sign_rules_saved': bool(_get('mini_sign_rules_saved', False)),
            'mini_sign_rules': mr,
        })
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': rows,
        },
    }


@router.put('/{protocol_id}/icf-versions/{icf_id}/mini-sign-rules', summary='保存单签署节点小程序签署规则')
@require_permission('protocol.protocol.update')
def update_icf_mini_sign_rules(request, protocol_id: int, icf_id: int, data: MiniSignRulesIn):
    """每节点独立保存；保存后 mini_sign_rules_saved=true。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再修改配置'}
    from apps.subject.models import ICFVersion

    icf = ICFVersion.objects.filter(id=icf_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': '签署节点不存在'}
    raw_lbl = data.supplemental_collect_labels if getattr(data, 'supplemental_collect_labels', None) is not None else []
    clean_lbl = [str(x).strip() for x in (raw_lbl or []) if str(x).strip()][:20]
    mini = {
        'require_face_verify': False,
        'require_dual_sign': bool(data.require_dual_sign),
        'require_comprehension_quiz': bool(data.require_comprehension_quiz),
        'enable_min_reading_duration': getattr(data, 'enable_min_reading_duration', True) is not False,
        'min_reading_duration_seconds': max(0, int(data.min_reading_duration_seconds)),
        'dual_sign_staffs': _merge_witness_staff_verification(_normalize_dual_sign_staffs(data.dual_sign_staffs)),
        'collect_id_card': bool(data.collect_id_card),
        'collect_screening_number': bool(data.collect_screening_number),
        'collect_initials': bool(data.collect_initials),
        'collect_subject_name': bool(getattr(data, 'collect_subject_name', False)),
        'collect_other_information': bool(data.collect_other_information) or bool(clean_lbl),
        'supplemental_collect_labels': clean_lbl,
        'enable_checkbox_recognition': bool(getattr(data, 'enable_checkbox_recognition', False)),
        'enable_staff_signature': bool(getattr(data, 'enable_staff_signature', False)),
        'staff_signature_times': _clamp_1_or_2(getattr(data, 'staff_signature_times', None), 1),
        'enable_subject_signature': bool(getattr(data, 'enable_subject_signature', False)),
        'subject_signature_times': _clamp_1_or_2(getattr(data, 'subject_signature_times', None), 1),
        'enable_guardian_signature': bool(getattr(data, 'enable_guardian_signature', False)),
        'guardian_parent_count': _clamp_1_or_2(getattr(data, 'guardian_parent_count', None), 1),
        'guardian_signature_times': _clamp_1_or_2(getattr(data, 'guardian_signature_times', None), 1),
        'enable_auto_sign_date': bool(getattr(data, 'enable_auto_sign_date', False)),
    }
    icf.mini_sign_rules = mini
    icf.mini_sign_rules_saved = True
    icf.save(update_fields=['mini_sign_rules', 'mini_sign_rules_saved', 'update_time'])
    # 协议级双签字段与节点一致（完整协议配置与跨协议同步由随后 PUT consent-settings 统一写入，避免连续两次传播超时/竞态）
    try:
        existing_cs = _get_consent_settings(protocol)
        existing_cs['require_face_verify'] = False
        existing_cs['require_dual_sign'] = bool(mini['require_dual_sign'])
        existing_cs['dual_sign_staffs'] = mini['dual_sign_staffs']
        existing_cs['enable_auto_sign_date'] = bool(mini.get('enable_auto_sign_date', False))
        _save_consent_settings(protocol, existing_cs)
    except Exception:
        logger.exception('同步协议级双签字段失败 protocol_id=%s icf_id=%s', protocol_id, icf_id)
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'id': icf.id,
            'mini_sign_rules_saved': True,
            'mini_sign_rules': mini,
        },
    }


@router.get('/{protocol_id}/icf-versions/{icf_id}/file', summary='签署节点原始文件（预览/下载）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def get_icf_version_file(request, protocol_id: int, icf_id: int):
    """返回上传的 ICF 文件流，供执行台知情配置内嵌预览（PDF 等）。"""
    from apps.subject.models import ICFVersion

    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    icf = ICFVersion.objects.filter(id=icf_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': '签署节点不存在'}
    rel = (icf.file_path or '').strip()
    if not rel:
        return 404, {'code': 404, 'msg': '该节点未上传文件'}
    if '..' in rel or os.path.isabs(rel):
        return 400, {'code': 400, 'msg': '非法文件路径'}
    media_root = os.path.abspath(os.path.normpath(django_settings.MEDIA_ROOT))
    abs_path = os.path.abspath(os.path.normpath(os.path.join(media_root, rel)))
    if not abs_path.startswith(media_root + os.sep) and abs_path != media_root:
        return 400, {'code': 400, 'msg': '非法文件路径'}
    if not os.path.isfile(abs_path):
        return 404, {'code': 404, 'msg': '文件不存在'}
    content_type, _ = mimetypes.guess_type(abs_path)
    if not content_type:
        content_type = 'application/octet-stream'
    basename = os.path.basename(abs_path) or 'file'
    # 禁止手写含中文的 Content-Disposition，否则 WSGI 编码 header 时会抛错导致 500
    fh = open(abs_path, 'rb')
    return FileResponse(
        fh,
        content_type=content_type,
        as_attachment=False,
        filename=basename,
    )


def _get_icf_version_preview_pdf_impl(request, protocol_id: int, icf_id: int):
    """
    返回用于 iframe 内嵌预览的内容：
    - 原 PDF：直接返回；
    - Word：优先 *_preview.pdf（LibreOffice），否则 .docx 可用 *_preview.html（python-docx 抽取正文）。
    """
    from apps.subject.models import ICFVersion

    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    icf = ICFVersion.objects.filter(id=icf_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': '签署节点不存在'}
    rel = (icf.file_path or '').strip()
    if not rel:
        return 404, {'code': 404, 'msg': '该节点未上传文件'}
    if '..' in rel or os.path.isabs(rel):
        return 400, {'code': 400, 'msg': '非法文件路径'}
    media_root = os.path.abspath(os.path.normpath(django_settings.MEDIA_ROOT))
    ext = os.path.splitext(rel)[1].lower()
    if ext in ('.doc', '.docx'):
        if ext == '.docx':
            services.ensure_icf_preview_for_http_request(rel)
        else:
            services.ensure_icf_preview(rel)
        # .docx：优先 HTML；.doc：在 LO 将 .doc→__icf_autoconv.docx 后同样可有 HTML，其次 PDF
        if ext == '.docx':
            candidates = [
                services.icf_preview_html_relative_path(rel),
                services.icf_preview_pdf_relative_path(rel),
            ]
        else:
            autoconv = services.icf_autoconv_docx_relative_path(rel)
            candidates = []
            if autoconv:
                h_auto = services.icf_preview_html_relative_path(autoconv)
                if h_auto:
                    candidates.append(h_auto)
            candidates.append(services.icf_preview_pdf_relative_path(rel))
        serve_rel = None
        for candidate in candidates:
            if not candidate:
                continue
            abs_c = os.path.abspath(os.path.normpath(os.path.join(media_root, candidate)))
            if abs_c.startswith(media_root + os.sep) and os.path.isfile(abs_c):
                serve_rel = candidate
                break
    else:
        serve_rel = rel
    if not serve_rel:
        return 404, {'code': 404, 'msg': '无法生成预览'}
    abs_path = os.path.abspath(os.path.normpath(os.path.join(media_root, serve_rel)))
    if not abs_path.startswith(media_root + os.sep) and abs_path != media_root:
        return 400, {'code': 400, 'msg': '非法文件路径'}
    if not os.path.isfile(abs_path):
        if ext == '.docx':
            return 404, {'code': 404, 'msg': '无法生成预览，请确认文件为有效 .docx。'}
        if ext == '.doc':
            return 404, {
                'code': 404,
                'msg': (
                    '无法从 .doc 生成预览：macOS 可依赖自带 textutil；'
                    'Linux/Windows 请安装 LibreOffice 或设置环境变量 LIBREOFFICE_PATH 指向 soffice，'
                    '并确认文件可被正常打开。上传时系统会尽量自动转为 .docx。'
                ),
            }
        return 404, {'code': 404, 'msg': '文件不存在'}
    if abs_path.lower().endswith('.html'):
        content_type = 'text/html; charset=utf-8'
    elif abs_path.lower().endswith('.pdf'):
        content_type = 'application/pdf'
    else:
        content_type, _ = mimetypes.guess_type(abs_path)
        if not content_type:
            content_type = 'application/octet-stream'
    basename = os.path.basename(abs_path) or 'preview'
    fh = open(abs_path, 'rb')
    return FileResponse(
        fh,
        content_type=content_type,
        as_attachment=False,
        filename=basename,
    )


@router.get('/{protocol_id}/icf-versions/{icf_id}/preview', summary='签署节点内嵌预览（PDF 或自动生成的 HTML）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def get_icf_version_preview_pdf(request, protocol_id: int, icf_id: int):
    """内嵌预览；异常时返回 404 JSON，避免未捕获异常导致 5xx 与 axios 重试「服务暂时不可用」。"""
    try:
        return _get_icf_version_preview_pdf_impl(request, protocol_id, icf_id)
    except Exception:
        logger.exception(
            '签署节点预览失败 protocol_id=%s icf_id=%s',
            protocol_id,
            icf_id,
        )
        return 404, {
            'code': 404,
            'msg': (
                '无法生成预览，请下载原文件。'
                '本地开发可将 .doc 另存为 .docx，或检查 MEDIA 与 python-docx；'
                '亦可使用执行台开发模式下的浏览器端预览。'
            ),
            'data': {'use_client_preview_fallback': True},
        }


@router.post(
    '/{protocol_id}/icf-versions/upload',
    summary='上传文件创建签署节点（知情管理）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def upload_icf_version(request, protocol_id: int, file: UploadedFile = File(...), node_title: Optional[str] = Form(None)):
    """上传 ICF 文件创建签署节点。文件名自动解析为节点标题，可传 node_title 覆盖。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再添加节点'}
    from apps.subject.services.consent_service import create_icf_version as do_create

    file_path = services.save_icf_upload_file(file, protocol_id=protocol_id)
    if file_path.lower().endswith('.doc'):
        converted = services.try_convert_icf_doc_file_to_docx_inplace(file_path)
        if converted:
            file_path = converted

    parsed_title = services.parse_filename_as_node_title(getattr(file, 'name', '') or '')
    title = (node_title or '').strip() or parsed_title

    # 自动版本号：v1.0, v2.0, ...
    from apps.subject.models import ICFVersion
    count = ICFVersion.objects.filter(protocol_id=protocol_id).count()
    version = f'v{count + 1}.0'

    icf = do_create(
        protocol_id=protocol_id,
        version=version,
        file_path=file_path,
        content='',
        is_active=True,
        required_reading_duration_seconds=0,
        node_title=title,
    )

    def _ensure_icf_preview_background() -> None:
        try:
            services.ensure_icf_preview(file_path)
        except Exception:
            logger.exception('上传签署节点后后台生成预览失败 protocol_id=%s file_path=%s', protocol_id, file_path)

    threading.Thread(target=_ensure_icf_preview_background, daemon=True).start()
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'id': icf.id,
            'version': icf.version,
            'node_title': getattr(icf, 'node_title', '') or title,
            'file_path': file_path,
            'is_active': icf.is_active,
            'create_time': icf.create_time.isoformat(),
        },
    }


@router.post('/{protocol_id}/icf-versions', summary='创建 ICF 版本（知情管理，表单方式）')
@require_permission('protocol.protocol.update')
def create_icf_version(request, protocol_id: int, data: ICFVersionCreateIn):
    """执行台知情管理：创建 ICF 版本（无文件时用表单）"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import create_icf_version as do_create
    icf = do_create(
        protocol_id=protocol_id,
        version=data.version.strip(),
        file_path=data.file_path or '',
        content=data.content or '',
        is_active=data.is_active,
        required_reading_duration_seconds=getattr(data, 'required_reading_duration_seconds', 0) or 0,
        node_title=getattr(data, 'node_title', '') or '',
    )
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'id': icf.id,
            'version': icf.version,
            'is_active': icf.is_active,
            'create_time': icf.create_time.isoformat(),
        },
    }


@router.put(
    '/{protocol_id}/icf-versions/{icf_id}',
    summary='更新 ICF 版本（知情管理）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def update_icf_version(request, protocol_id: int, icf_id: int, data: ICFVersionUpdateIn):
    """执行台知情管理：更新 ICF 版本（版本号、内容、激活状态）"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再编辑节点'}
    from apps.subject.models import ICFVersion
    icf = ICFVersion.objects.filter(id=icf_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': 'ICF 版本不存在'}
    from apps.subject.services.consent_service import update_icf_version as do_update
    updated = do_update(
        icf_id=icf_id,
        version=data.version.strip() if data.version is not None else None,
        content=data.content if data.content is not None else None,
        is_active=data.is_active,
        required_reading_duration_seconds=getattr(data, 'required_reading_duration_seconds', None),
        node_title=getattr(data, 'node_title', None),
    )
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'id': updated.id,
            'version': updated.version,
            'content': (updated.content or '')[:200],
            'is_active': updated.is_active,
            'update_time': updated.update_time.isoformat(),
        },
    }


@router.delete(
    '/{protocol_id}/icf-versions/{icf_id}',
    summary='删除签署节点（知情管理）',
    # 须声明非 200，否则 Ninja 抛 ConfigError，全局处理器将正文替换为「操作失败」
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def delete_icf_version(request, protocol_id: int, icf_id: int):
    """删除 ICF 版本（签署节点）；已发布或已有签署/见证记录时不允许删除。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再删除节点'}
    from apps.subject.services.consent_service import delete_icf_version as do_delete
    ok, err = do_delete(protocol_id, icf_id)
    if not ok:
        return 400, {'code': 400, 'msg': err or '删除失败'}
    return {'code': 200, 'msg': 'OK', 'data': {'id': icf_id}}


@router.post(
    '/{protocol_id}/icf-versions/reorder',
    summary='调整 ICF 签署顺序（知情管理）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def reorder_icf_versions(request, protocol_id: int, data: ICFReorderIn):
    """按 id_order 更新 ICF 版本的签署顺序"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {'code': 400, 'msg': '已发布，请先取消发布后再调整顺序'}
    from apps.subject.services.consent_service import reorder_icf_versions as do_reorder
    do_reorder(protocol_id, data.id_order)
    return {'code': 200, 'msg': 'OK'}


@router.get('/{protocol_id}/consents/stats', summary='签署统计（知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def consent_stats(request, protocol_id: int):
    """执行台知情管理：签署统计与「签署记录」列表一致（按受试者合并行，非文档条数）。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import get_consents_stats
    stats = get_consents_stats(protocol_id)
    return {'code': 200, 'msg': 'OK', 'data': stats}


@router.get('/{protocol_id}/consents/export', summary='导出受试者基础信息 Excel（知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def export_consents(
    request,
    protocol_id: int,
    status: str = 'all',
    icf_version_id: int = None,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None, description='关键字：受试者编号/姓名/手机号（存于受试者表，列表不展示）/SC号/回执号等子串匹配'),
):
    """执行台知情管理：按当前筛选导出受试者基础信息（按受试者去重），字段 SC号、姓名、手机号、身份证号。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import get_subjects_basic_export_rows
    from django.http import HttpResponse
    df = _parse_consent_date_query(date_from)
    dt = _parse_consent_date_query(date_to)
    rows = get_subjects_basic_export_rows(
        protocol_id,
        status_filter=status or 'all',
        icf_version_id=icf_version_id,
        date_from=df,
        date_to=dt,
        search=search,
    )
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return 500, {'code': 500, 'msg': 'Excel 导出依赖 openpyxl 未安装', 'data': None}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '受试者基础信息'
    headers = ['SC号', '姓名', '手机号', '身份证号']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r.get('sc_number') or '')
        ws.cell(row=row_idx, column=2, value=r.get('subject_name') or '')
        ws.cell(row=row_idx, column=3, value=r.get('phone') or '')
        ws.cell(row=row_idx, column=4, value=r.get('id_card') or '')
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    code = getattr(protocol, 'code', None) or str(protocol_id)
    filename = f"{code}_受试者基础信息_{date.today():%Y%m%d}.xlsx"
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get('/{protocol_id}/consents/export-pdf', summary='批量导出签署回执 PDF（ZIP，知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def export_consents_pdf_zip(
    request,
    protocol_id: int,
    status: str = 'all',
    icf_version_id: int = None,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None, description='与列表 search 一致'),
):
    """将当前筛选下已有回执 PDF 的签署记录打包为 zip 下载。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import zip_consent_receipt_pdfs_for_protocol
    from django.http import HttpResponse
    df = _parse_consent_date_query(date_from)
    dt = _parse_consent_date_query(date_to)
    buf, err = zip_consent_receipt_pdfs_for_protocol(
        protocol_id,
        status_filter=status or 'all',
        icf_version_id=icf_version_id,
        date_from=df,
        date_to=dt,
        search=search,
        protocol_code=getattr(protocol, 'code', None) or str(protocol_id),
    )
    if err:
        return 400, {'code': 400, 'msg': err, 'data': None}
    code = getattr(protocol, 'code', None) or str(protocol_id)
    filename = f"{code}_知情签署文件_{date.today():%Y%m%d}.zip"
    response = HttpResponse(buf.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get('/{protocol_id}/consents', summary='签署记录列表（知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def list_consents(
    request,
    protocol_id: int,
    status: str = 'all',
    icf_version_id: int = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort: str = Query('signed_at'),
    order: str = Query('desc'),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None, description='关键字：受试者编号、姓名、手机号（存于受试者表，列表不展示）、SC号、回执号、节点标题等子串匹配'),
    group_by: Optional[str] = Query(
        None,
        description='subject：按受试者合并多知情节点为一行；签署结果按全部节点汇总（未签齐为「-」，任一为否则「否」，全为是则「是」）',
    ),
):
    """执行台知情管理：分页签署记录；status: all | signed | pending | result_no；sort 为字段名，order 为 asc|desc；默认按签署时间新到旧；可选 date_from/date_to（YYYY-MM-DD）、search、group_by=subject。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import (
        _dt_to_local_date,
        consent_list_api_rows_from_subject_groups,
        consent_list_display_fields,
        consent_staff_display_status,
        list_consents_grouped_by_subject_page,
        list_consents_page_for_protocol,
        safe_subject_consent_icf_version,
        signing_staff_name_for_screening_date,
        subject_name_display_for_consent,
        subject_no_display_for_consent,
    )
    from django.conf import settings
    df = _parse_consent_date_query(date_from)
    dt = _parse_consent_date_query(date_to)
    sort_field = (sort or 'signed_at').strip() or 'signed_at'
    sort_order = (order or 'desc').strip() or 'desc'
    settings_data = _get_consent_settings(protocol)
    sched = settings_data.get('screening_schedule') or []
    single_ref_date = df if (df and dt and df == dt) else None

    if (group_by or '').strip().lower() == 'subject':
        total, groups = list_consents_grouped_by_subject_page(
            protocol_id,
            status_filter=status or 'all',
            icf_version_id=icf_version_id,
            page=page,
            page_size=page_size,
            sort_field=sort_field,
            sort_order=sort_order,
            date_from=df,
            date_to=dt,
            search=search,
        )
        rows = consent_list_api_rows_from_subject_groups(
            groups,
            settings_data,
            sched,
            single_ref_date,
            settings.MEDIA_URL,
        )
        return {
            'code': 200,
            'msg': 'OK',
            'data': {
                'items': rows,
                'total': total,
                'page': page,
                'page_size': page_size,
                'group_by': 'subject',
            },
        }

    total, items = list_consents_page_for_protocol(
        protocol_id,
        status_filter=status or 'all',
        icf_version_id=icf_version_id,
        page=page,
        page_size=page_size,
        sort_field=sort_field,
        sort_order=sort_order,
        date_from=df,
        date_to=dt,
        search=search,
    )
    rows = []
    for c in items:
        receipt_pdf_path = (c.signature_data or {}).get('receipt_pdf_path') if c.signature_data else None
        receipt_pdf_url = f"{settings.MEDIA_URL}{receipt_pdf_path}" if receipt_pdf_path else None
        witness_meta = ((c.signature_data or {}).get('investigator_sign') if c.signature_data else None) or {}
        try:
            investigator_signed_at = c.investigator_signed_at.isoformat() if getattr(c, 'investigator_signed_at', None) else None
        except Exception:
            investigator_signed_at = None
        icf_row = safe_subject_consent_icf_version(c)
        node_title = (getattr(icf_row, 'node_title', None) or '') if icf_row else ''
        icf_version = icf_row.version if icf_row else ''
        extra = consent_list_display_fields(c, None)
        if single_ref_date:
            ref_day = single_ref_date
        else:
            ref_day = _dt_to_local_date(c.signed_at) or _dt_to_local_date(c.create_time)
        screening_signing_staff = signing_staff_name_for_screening_date(sched, ref_day) if ref_day else ''
        sig_for_row = c.signature_data or {}
        if sig_for_row.get('witness_dev_flow') and (sig_for_row.get('witness_staff_name') or '').strip():
            screening_signing_staff = str(sig_for_row.get('witness_staff_name')).strip()
        rows.append({
            'id': c.id,
            'subject_id': c.subject_id,
            'subject_no': subject_no_display_for_consent(c),
            'subject_name': subject_name_display_for_consent(c),
            'phone': extra.get('phone_masked') or '-',
            'id_card': extra.get('id_card_masked') or '-',
            'sc_number': extra.get('sc_number') or '-',
            'name_pinyin_initials': extra.get('name_pinyin_initials') or '-',
            'signing_result': extra.get('signing_result') or '-',
            'signing_type': extra.get('signing_type') or '正式',
            'icf_version_id': c.icf_version_id,
            'icf_version': icf_version,
            'node_title': node_title,
            'is_signed': c.is_signed,
            'signed_at': c.signed_at.isoformat() if c.signed_at else None,
            'investigator_signed_at': investigator_signed_at,
            'investigator_sign_staff_name': witness_meta.get('staff_name') or '',
            'screening_signing_staff': screening_signing_staff,
            'receipt_no': c.receipt_no or '',
            'receipt_pdf_path': receipt_pdf_path,
            'receipt_pdf_url': receipt_pdf_url,
            'create_time': c.create_time.isoformat(),
            'require_dual_sign': settings_data.get('require_dual_sign', False),
            'consent_status_label': consent_staff_display_status(c),
            'staff_audit_status': getattr(c, 'staff_audit_status', '') or '',
            'auth_verified_at': (
                c.subject.identity_verified_at.isoformat()
                if getattr(c.subject, 'identity_verified_at', None)
                else None
            ),
            'witness_dev_batch_id': (sig_for_row.get('witness_dev_batch_id') or '') or None,
        })
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': rows,
            'total': total,
            'page': page,
            'page_size': page_size,
        },
    }


@router.get('/{protocol_id}/consents/{consent_id}/preview', summary='执行台：签署内容预览（知情管理）')
@require_any_permission(['protocol.protocol.read', 'subject.subject.read'])
def consent_preview(request, protocol_id: int, consent_id: int):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import get_consent_preview_for_execution
    data = get_consent_preview_for_execution(protocol_id, consent_id)
    if not data:
        return 404, {'code': 404, 'msg': '签署记录不存在'}
    return {'code': 200, 'msg': 'OK', 'data': data}


@router.post('/{protocol_id}/consents/{consent_id}/staff-return', summary='执行台：退回重签（知情管理）')
@require_permission('protocol.protocol.update')
def staff_return_consent_api(
    request,
    protocol_id: int,
    consent_id: int,
    payload: Optional[StaffReturnIn] = Body(default=None),
):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import consent_staff_display_status, staff_return_consent_for_resign
    p = payload or StaffReturnIn()
    reason_raw = (p.reason or '').strip()
    reason = reason_raw[:500] if reason_raw else None
    try:
        c = staff_return_consent_for_resign(protocol_id, consent_id, reason=reason)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'consent_id': c.id,
            'is_signed': c.is_signed,
            'staff_audit_status': getattr(c, 'staff_audit_status', '') or '',
            'consent_status_label': consent_staff_display_status(c),
        },
    }


@router.post('/{protocol_id}/consents/{consent_id}/staff-approve', summary='执行台：通过审核（知情管理）')
@require_permission('protocol.protocol.update')
def staff_approve_consent_api(request, protocol_id: int, consent_id: int):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import consent_staff_display_status, staff_approve_consent
    try:
        c = staff_approve_consent(protocol_id, consent_id)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'consent_id': c.id,
            'is_signed': c.is_signed,
            'staff_audit_status': getattr(c, 'staff_audit_status', '') or '',
            'consent_status_label': consent_staff_display_status(c),
        },
    }


@router.delete('/{protocol_id}/consents/{consent_id}', summary='执行台：软删除签署记录（知情管理）')
@require_permission('protocol.protocol.update')
def soft_delete_consent_record(request, protocol_id: int, consent_id: int):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.services.consent_service import soft_delete_consent_for_execution
    try:
        soft_delete_consent_for_execution(protocol_id, consent_id)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': {'consent_id': consent_id}}


@router.post('/{protocol_id}/consents/{consent_id}/investigator-sign', summary='执行台：见证签署（双签）')
@require_permission('protocol.protocol.update')
def investigator_sign_consent(request, protocol_id: int, consent_id: int, data: ConsentWitnessSignIn):
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.subject.models import SubjectConsent
    consent = (
        SubjectConsent.objects
        .filter(id=consent_id, icf_version__protocol_id=protocol_id)
        .select_related('icf_version')
        .first()
    )
    if not consent:
        return 404, {'code': 404, 'msg': '签署记录不存在'}
    if not consent.is_signed:
        return 400, {'code': 400, 'msg': '受试者尚未签署，无法见证'}
    if consent.investigator_signed_at:
        return 409, {'code': 409, 'msg': '该记录已见证签署'}

    settings_data = _get_consent_settings(protocol)
    if settings_data.get('require_dual_sign'):
        verified_staffs = [s for s in settings_data.get('dual_sign_staffs', []) if s.get('identity_verified')]
        matched = None
        for s in verified_staffs:
            if data.staff_id and s.get('staff_id') == data.staff_id:
                matched = s
                break
            if not data.staff_id and s.get('name') == data.staff_name:
                matched = s
                break
        if not matched:
            return 400, {'code': 400, 'msg': '见证人员未在已核身双签名单中，请先在配置中心完成设置'}

    signature_data = dict(consent.signature_data or {})
    signature_data['investigator_sign'] = {
        'staff_id': (data.staff_id or '').strip(),
        'staff_name': data.staff_name.strip(),
        'staff_phone': (data.staff_phone or '').strip(),
        'staff_email': (data.staff_email or '').strip(),
        'signed_at': timezone.now().isoformat(),
    }
    consent.investigator_signed_at = timezone.now()
    consent.signature_data = signature_data
    consent.save(update_fields=['investigator_signed_at', 'signature_data', 'update_time'])
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'consent_id': consent.id,
            'investigator_signed_at': consent.investigator_signed_at.isoformat(),
            'investigator_sign_staff_name': signature_data.get('investigator_sign', {}).get('staff_name', ''),
            'status': 'witness_signed',
        },
    }


@router.put('/{protocol_id}', summary='更新协议基本信息（项目名称/编号）', response={200: dict, 400: dict, 404: dict})
@require_permission('protocol.protocol.update')
def update_protocol_basic(request, protocol_id: int, data: ProtocolBasicUpdateIn):
    """新建项目外的编辑入口：修改项目编号时需保证与现有项目不重复。"""
    # 与知情管理 consent-settings、consent-overview 一致：仅校验协议存在 + 权限码，不套用项目级 scope
    # （否则列表可见协议但 PUT 基本信息失败，现场计划保存会先调本接口）
    protocol = Protocol.objects.filter(id=protocol_id, is_deleted=False).first()
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    patch = data.model_dump(exclude_unset=True) if hasattr(data, 'model_dump') else data.dict(exclude_unset=True)
    if not patch:
        return 400, {'code': 400, 'msg': '无更新内容', 'data': None}
    kwargs = {}
    if 'title' in patch:
        t = (patch.get('title') or '').strip()
        if not t:
            return 400, {'code': 400, 'msg': '项目名称不能为空', 'data': None}
        kwargs['title'] = t
    if 'code' in patch:
        kwargs['code'] = patch.get('code')
    if 'consent_config_account_id' in patch:
        raw = patch.get('consent_config_account_id')
        if raw is None or raw == 0:
            kwargs['consent_config_account_id'] = None
        else:
            try:
                services.assert_consent_config_account_allowed(int(raw))
            except ValueError as e:
                return 400, {'code': 400, 'msg': str(e), 'data': None}
            kwargs['consent_config_account_id'] = int(raw)
    if not kwargs:
        return 400, {'code': 400, 'msg': '无有效更新字段', 'data': None}
    try:
        updated = services.update_protocol(protocol_id, **kwargs)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e), 'data': None}
    if not updated:
        return 404, {'code': 404, 'msg': '协议不存在'}
    return {'code': 200, 'msg': 'OK', 'data': _protocol_to_dict(updated)}


@router.get('/{protocol_id}', summary='协议详情')
@require_permission('protocol.protocol.read')
def get_protocol(request, protocol_id: int):
    """获取协议详细信息（与知情管理链路一致：存在性 + 权限码，不套用项目级 scope）。"""
    protocol = Protocol.objects.filter(id=protocol_id, is_deleted=False).first()
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    return {'code': 200, 'msg': 'OK', 'data': _protocol_to_dict(protocol)}


@router.delete('/{protocol_id}', summary='软删除协议')
@require_permission('protocol.protocol.update')
def soft_delete_protocol(request, protocol_id: int):
    """
    将协议标记为已删除（is_deleted=True），不出现在列表与查询中；非物理删除，可通过后台恢复。
    """
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    if not services.delete_protocol(protocol_id):
        return 404, {'code': 404, 'msg': '协议不存在'}
    return {'code': 200, 'msg': '已标记删除', 'data': {'id': protocol_id}}


@router.post('/{protocol_id}/parse', summary='触发 AI 解析')
@require_permission('protocol.protocol.update')
def trigger_parse(request, protocol_id: int):
    """触发协议 AI 解析；按数据权限校验可见性"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    account_id = account.id if account else None
    parse_log = services.trigger_parse(protocol_id, account_id=account_id)
    if not parse_log:
        return 400, {'code': 400, 'msg': '无法解析：协议不存在或未上传文件'}
    return {
        'code': 200,
        'msg': 'OK',
        'data': {'parse_log_id': parse_log.id, 'status': parse_log.status},
    }


@router.post('/{protocol_id}/accept-parsed', summary='采纳 AI 解析结果写入协议')
@require_permission('protocol.protocol.update')
def accept_parsed(request, protocol_id: int):
    """
    数字员工流程内嵌：把 AI 解析或编排产出的结构化数据写入协议 parsed_data。
    前端动作卡片点击"采纳写入协议"时调用。
    """
    import json
    account = _get_account_from_request(request)
    protocol = get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}

    body = {}
    try:
        body = json.loads(request.body) if request.body else {}
    except Exception:
        pass

    parsed_data = body.get('parsed_data') or protocol.parsed_data
    if not parsed_data or not isinstance(parsed_data, dict):
        return 400, {'code': 400, 'msg': 'parsed_data 不可为空'}

    try:
        result = services.set_parsed_data(protocol_id, parsed_data)
        if not result:
            return 400, {'code': 400, 'msg': '写入失败'}
    except ValueError as exc:
        return 400, {'code': 400, 'msg': str(exc)}

    try:
        from apps.secretary.runtime_plane import create_execution_task, finalize_execution_task
        task_id = create_execution_task(
            runtime_type='service',
            name='accept-parsed-result',
            target='protocol.accept_parsed',
            account_id=getattr(account, 'id', None),
            input_payload={'protocol_id': protocol_id},
            role_code='solution_designer',
            workstation_key='research',
            business_object_type='protocol',
            business_object_id=str(protocol_id),
        )
        finalize_execution_task(task_id, ok=True, output={'protocol_id': protocol_id, 'status': 'parsed'})
    except Exception:
        pass

    return {'code': 200, 'msg': 'OK', 'data': {'protocol_id': protocol_id, 'status': 'parsed'}}


@router.get('/{protocol_id}/dashboard', summary='项目级仪表板')
@require_permission('protocol.protocol.read')
def protocol_dashboard(request, protocol_id: int):
    """项目级聚合仪表板：入组/工单/访视/偏差/CAPA/财务"""
    from datetime import date
    from django.db.models import Count, Q

    account = _get_account_from_request(request)
    protocol = get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}

    today = date.today()
    result = {'protocol': _protocol_to_dict(protocol)}

    # Enrollment stats
    try:
        from apps.subject.models import Enrollment
        enrollments = Enrollment.objects.filter(protocol_id=protocol_id)
        enrollment_stats = list(enrollments.values('status').annotate(count=Count('id')))
        enrolled_count = enrollments.filter(status='enrolled').count()
        result['enrollment'] = {
            'by_status': enrollment_stats,
            'enrolled': enrolled_count,
            'total': enrollments.count(),
            'rate': round(enrolled_count / protocol.sample_size * 100, 1) if protocol.sample_size else 0,
        }
    except Exception:
        result['enrollment'] = {'by_status': [], 'enrolled': 0, 'total': 0, 'rate': 0}

    # WorkOrder stats
    try:
        from apps.workorder.models import WorkOrder
        wo_qs = WorkOrder.objects.filter(enrollment__protocol_id=protocol_id, is_deleted=False)
        wo_status = list(wo_qs.values('status').annotate(count=Count('id')))
        wo_total = wo_qs.count()
        wo_done = wo_qs.filter(status__in=['completed', 'approved']).count()
        wo_overdue = wo_qs.filter(due_date__lt=today).exclude(
            status__in=['completed', 'approved', 'cancelled'],
        ).count()

        by_assignee = list(
            wo_qs.exclude(assigned_to__isnull=True)
            .values('assigned_to')
            .annotate(
                total=Count('id'),
                completed=Count('id', filter=Q(status__in=['completed', 'approved'])),
            )
            .order_by('-total')[:10]
        )

        result['workorders'] = {
            'by_status': wo_status,
            'total': wo_total,
            'completed': wo_done,
            'completion_rate': round(wo_done / wo_total * 100, 1) if wo_total else 0,
            'overdue': wo_overdue,
            'by_assignee': by_assignee,
        }
    except Exception:
        result['workorders'] = {'by_status': [], 'total': 0, 'completed': 0, 'completion_rate': 0, 'overdue': 0, 'by_assignee': []}

    # Visit compliance
    try:
        from apps.visit.models import VisitPlan
        plans = VisitPlan.objects.filter(protocol_id=protocol_id)
        if plans.exists():
            plan = plans.first()
            from apps.visit.services.compliance_service import ComplianceAnalysisService
            compliance = ComplianceAnalysisService.analyze_visit_completeness(plan.id)
            result['visit_compliance'] = compliance
        else:
            result['visit_compliance'] = None
    except Exception:
        result['visit_compliance'] = None

    # Quality: deviations + CAPA
    try:
        from apps.quality.models import Deviation, CAPA
        dev_qs = Deviation.objects.filter(project_id=protocol_id)
        dev_stats = list(dev_qs.values('status').annotate(count=Count('id')))
        dev_severity = list(dev_qs.values('severity').annotate(count=Count('id')))
        capa_qs = CAPA.objects.filter(deviation__project_id=protocol_id)
        capa_stats = list(capa_qs.values('status').annotate(count=Count('id')))
        result['quality'] = {
            'deviation_by_status': dev_stats,
            'deviation_by_severity': dev_severity,
            'deviation_total': dev_qs.count(),
            'capa_by_status': capa_stats,
            'capa_total': capa_qs.count(),
        }
    except Exception:
        result['quality'] = {'deviation_by_status': [], 'deviation_by_severity': [], 'deviation_total': 0, 'capa_by_status': [], 'capa_total': 0}

    # Finance summary
    try:
        from apps.finance.models import Contract, Invoice, Payment
        contracts = Contract.objects.filter(protocol_id=protocol_id)
        contract_amount = sum(float(c.amount or 0) for c in contracts)
        invoices = Invoice.objects.filter(contract__protocol_id=protocol_id)
        invoiced = sum(float(i.total or 0) for i in invoices)
        payments = Payment.objects.filter(invoice__contract__protocol_id=protocol_id)
        received = sum(float(p.actual_amount or 0) for p in payments)
        result['finance'] = {
            'contract_amount': contract_amount,
            'invoiced': invoiced,
            'received': received,
            'outstanding': round(invoiced - received, 2),
        }
    except Exception:
        result['finance'] = {'contract_amount': 0, 'invoiced': 0, 'received': 0, 'outstanding': 0}

    return {'code': 200, 'msg': 'OK', 'data': result}


@router.post('/{protocol_id}/startup-package', summary='生成启动包')
@require_permission('protocol.protocol.update')
def generate_startup_package(request, protocol_id: int):
    """B3：一键生成项目启动包（访视计划/资源需求/eTMF/CRF/伦理/预算/里程碑/飞书群）"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    from .services.startup_package_service import generate_startup_package as gen_pkg
    result = gen_pkg(protocol_id)
    return {'code': 200, 'msg': 'OK', 'data': result}


@router.post('/{protocol_id}/publish-status', summary='发布项目状态通报')
@require_permission('protocol.protocol.read')
def publish_status(request, protocol_id: int):
    """E2：一键发布项目状态到飞书群（含数字员工门禁）"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    try:
        from apps.secretary.evidence_gate_service import check_business_gate
        passed, reason, _ = check_business_gate(
            'release_digital_worker',
            {'skill_id': 'efficacy-report-generator', 'role_code': 'report_generator'},
        )
        if not passed:
            return 400, {'code': 400, 'msg': f'数字员工门禁未通过，禁止发布：{reason}'}
    except ImportError:
        pass
    from apps.notification.card_template_service import publish_status_report
    result = publish_status_report(protocol_id)
    return {'code': 200, 'msg': 'OK', 'data': result}


@router.get('/{protocol_id}/logs', summary='解析日志')
@require_permission('protocol.protocol.read')
def get_parse_logs(request, protocol_id: int):
    """获取协议的解析日志；按数据权限校验协议可见性"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    logs = services.get_parse_logs(protocol_id)
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': [
                {
                    'id': log.id,
                    'status': log.status,
                    'error_message': log.error_message,
                    'create_time': log.create_time.isoformat(),
                    'finish_time': log.finish_time.isoformat() if log.finish_time else None,
                }
                for log in logs
            ],
        },
    }


# ============================================================================
# 协议状态变更
# ============================================================================
@router.post('/{protocol_id}/activate', summary='激活协议', response={200: dict, 400: dict, 404: dict})
@require_permission('protocol.protocol.update')
def activate_protocol(request, protocol_id: int):
    """将协议状态变更为 active（生效中）。草稿/已上传状态可激活。"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    p = Protocol.objects.filter(id=protocol_id, is_deleted=False).first()
    if p.status == 'active':
        return {'code': 200, 'msg': '协议已处于生效状态', 'data': {'id': p.id, 'status': p.status}}

    # 数字员工启动门禁：激活协议前通过 EvidenceGate 检查
    try:
        from apps.secretary.evidence_gate_service import check_business_gate
        passed, reason, gate_run_id = check_business_gate(
            'release_digital_worker',
            {'protocol_id': protocol_id, 'role_code': 'startup_gate_assistant'},
        )
        if not passed:
            return 400, {
                'code': 400,
                'msg': f'数字员工启动门禁未通过，无法激活协议：{reason}',
                'data': {'gate_run_id': gate_run_id, 'gate_blocked': True},
            }
    except ImportError:
        pass
    except Exception as exc:
        import logging as _log
        _log.getLogger(__name__).debug('activate_protocol gate check skipped: %s', exc)

    updated = services.update_protocol(protocol_id, status='active')
    if not updated:
        return 400, {'code': 400, 'msg': '激活失败'}
    return {'code': 200, 'msg': '协议已激活', 'data': {'id': updated.id, 'status': updated.status}}


@router.post('/{protocol_id}/deactivate', summary='归档协议', response={200: dict, 400: dict, 404: dict})
@require_permission('protocol.protocol.update')
def deactivate_protocol(request, protocol_id: int):
    """将协议状态变更为 archived（已归档）。"""
    account = _get_account_from_request(request)
    if not get_visible_object(Protocol.objects.filter(id=protocol_id, is_deleted=False), account):
        return 404, {'code': 404, 'msg': '协议不存在'}
    readiness = services.evaluate_archive_readiness(protocol_id)
    if not readiness.get('passed'):
        return 400, {
            'code': 400,
            'msg': '协议未满足归档前置条件，请先完成结项链路',
            'data': readiness,
        }
    updated = services.update_protocol(protocol_id, status='archived')
    if not updated:
        return 400, {'code': 400, 'msg': '归档失败'}
    return {'code': 200, 'msg': '协议已归档', 'data': {'id': updated.id, 'status': updated.status}}


# ============================================================================
# 双签工作人员档案 + 邮件身份验证（执行台知情管理）
# ============================================================================


class WitnessStaffCreateIn(Schema):
    """从鹿鸣·治理台（3008）已存在的账号建档；须具备 qa（QA质量管理）全局角色。"""
    account_id: int


class WitnessStaffCreatePartTimeIn(Schema):
    """双签工作人员（无治理台账号）：先录姓名与工作邮箱；身份证与手机号可在人脸核验时由本人补全。"""

    name: str
    email: str
    phone: Optional[str] = ''
    id_card_no: Optional[str] = None
    gender: Optional[str] = None


class WitnessStaffUpdateIn(Schema):
    name: Optional[str] = None
    email: Optional[str] = None
    gender: Optional[str] = None
    id_card_no: Optional[str] = None
    phone: Optional[str] = None
    priority: Optional[int] = None


class DualSignAuthRequestIn(Schema):
    witness_staff_id: int
    icf_version_id: int
    notify_email: Optional[str] = None


class WitnessStaffNotifyIn(Schema):
    """从工作人员列表发起邮件：路径含 staff_id，body 含协议与节点"""

    protocol_id: int
    icf_version_id: int
    notify_email: Optional[str] = None


class WitnessStaffProfileNotifyIn(Schema):
    """双签名单「核验」：仅档案人脸+签名登记，不绑定协议"""

    notify_email: Optional[str] = None


class WitnessFaceSubmitIn(Schema):
    """【已废弃】原占位提交；请改用 face-start + face-result。"""

    token: str
    face_order_id: str = ''
    id_card_no: Optional[str] = None
    phone: Optional[str] = None


class WitnessFaceStartIn(Schema):
    """公开人脸核验：发起火山引擎 H5；档案缺身份证与手机号时须填。"""

    token: str
    id_card_no: Optional[str] = None
    phone: Optional[str] = None


class IcfVersionAnswersIn(Schema):
    """联调：单节点勾选结果（与 icf_checkbox_answers 结构一致）。"""

    icf_version_id: int
    answers: List[Any] = []


class WitnessDevConsentSubmitIn(Schema):
    """联调：将 ICF 节点签署写入 SubjectConsent（测试类型）。"""

    token: str
    icf_version_ids: List[int]
    icf_version_answers: Optional[List[IcfVersionAnswersIn]] = None


class IcfVersionSignaturesIn(Schema):
    """核验测试 H5：各节点手写签名（data URL 或纯 base64）。"""

    icf_version_id: int
    signature_images: List[str] = []


class ConsentTestScanSubmitIn(Schema):
    """核验测试 H5：凭 p/t 写入测试类型 SubjectConsent。"""

    p: int
    t: str
    icf_version_ids: List[int]
    icf_version_answers: Optional[List[IcfVersionAnswersIn]] = None
    icf_version_signatures: Optional[List[IcfVersionSignaturesIn]] = None
    subject_name: Optional[str] = None
    id_card_no: Optional[str] = None
    phone: Optional[str] = None
    screening_number: Optional[str] = None


def _upsert_dual_sign_staff_row(protocol: Protocol, ws: WitnessStaff) -> None:
    settings_data = dict(_get_effective_consent_settings(protocol))
    staffs = list(settings_data.get('dual_sign_staffs') or [])
    sid = str(ws.id)
    row = {
        'staff_id': sid,
        'name': ws.name,
        'email': ws.email or '',
        'phone': ws.phone or '',
        'id_card_no': ws.id_card_no or '',
        'identity_verified': bool(ws.identity_verified),
    }
    replaced = False
    for i, s in enumerate(staffs):
        if str(s.get('staff_id') or '') == sid:
            staffs[i] = {**dict(s), **row}
            replaced = True
            break
    if not replaced:
        staffs.append(row)
    settings_data['dual_sign_staffs'] = staffs
    settings_data['require_dual_sign'] = True
    _save_consent_settings(protocol, settings_data)


def _perform_dual_sign_auth(protocol_id: int, data: DualSignAuthRequestIn):
    """发送授权邮件并写入双签名单；成功返回 dict，失败返回 (status_code, err_body)。"""
    from apps.protocol.services import witness_staff_service as ws_svc
    from apps.subject.models import ICFVersion

    protocol = Protocol.objects.filter(id=protocol_id, is_deleted=False).first()
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 400, {
            'code': 400,
            'msg': '知情已发布，请先取消发布（下架）后再进行授权核验测试',
        }
    ws = WitnessStaff.objects.filter(id=data.witness_staff_id, is_deleted=False).first()
    if not ws:
        return 404, {'code': 404, 'msg': '工作人员不存在'}
    if not ws_svc.witness_staff_can_receive_auth_emails(ws):
        return 400, {
            'code': 400,
            'msg': '该人员无法发送授权邮件：须为具备 QA质量管理 全局角色的治理台账号，或已填写姓名与工作邮箱的无账号档案',
        }
    icf = ICFVersion.objects.filter(id=data.icf_version_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': '签署节点不存在'}
    notify = (data.notify_email or '').strip() or (ws.email or '').strip()
    if not notify:
        return 400, {'code': 400, 'msg': '请填写通知邮箱'}
    try:
        ws_svc.send_witness_authorization_email(
            protocol=protocol,
            witness=ws,
            icf_version_id=icf.id,
            notify_email=notify,
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    _upsert_dual_sign_staff_row(protocol, ws)
    settings_after = dict(_get_effective_consent_settings(protocol))
    settings_after['consent_verify_test_staff_name'] = (ws.name or '').strip()
    _save_consent_settings(protocol, settings_after)
    return {'code': 200, 'msg': '授权邮件已发送', 'data': {'notify_email': notify}}


@router.get('/witness-staff/list', summary='双签工作人员列表（可搜索）')
@require_any_permission(['protocol.protocol.read', 'protocol.protocol.update', 'subject.subject.read'])
def witness_staff_list(
    request,
    search: str = None,
    page: int = 1,
    page_size: int = 20,
    focus_witness_staff_id: int = None,
):
    from apps.protocol.services import witness_staff_service as ws_svc

    result = ws_svc.list_witness_staff(
        search=search,
        page=page,
        page_size=min(page_size, 500),
        focus_witness_staff_id=focus_witness_staff_id,
    )
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'items': result['items'],
            'total': result['total'],
            'page': result['page'],
            'page_size': result['page_size'],
        },
    }


@router.get('/witness-staff/eligible-accounts', summary='可选：治理台账号（QA质量管理，用于建档）')
@require_permission('protocol.protocol.update')
def witness_staff_eligible_accounts(
    request,
    search: str = None,
    page: int = 1,
    page_size: int = 50,
    only_without_profile: bool = False,
):
    from apps.protocol.services import witness_staff_service as ws_svc

    result = ws_svc.list_eligible_accounts_for_picker(
        search=search,
        page=page,
        page_size=min(page_size, 100),
        only_without_profile=only_without_profile,
    )
    return {'code': 200, 'msg': 'OK', 'data': result}


@router.post('/witness-staff/sync-from-accounts', summary='从治理台账号批量同步双签档案')
@require_permission('protocol.protocol.update')
def witness_staff_sync_from_accounts(request):
    from apps.protocol.services import witness_staff_service as ws_svc

    result = ws_svc.sync_witness_staff_from_accounts()
    return {'code': 200, 'msg': 'OK', 'data': result}


@router.post('/witness-staff', summary='新增双签工作人员')
@require_permission('protocol.protocol.update')
def witness_staff_create(request, data: WitnessStaffCreateIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    try:
        ws = ws_svc.create_witness_staff_from_account(data.account_id)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': ws_svc.witness_staff_to_dict(ws)}


@router.post('/witness-staff/part-time', summary='新增双签工作人员（无治理台账号）')
@require_permission('protocol.protocol.update')
def witness_staff_create_part_time(request, data: WitnessStaffCreatePartTimeIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    try:
        ws = ws_svc.create_witness_staff_part_time(
            name=data.name,
            email=data.email,
            phone=data.phone,
            id_card_no=data.id_card_no,
            gender=data.gender,
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': ws_svc.witness_staff_to_dict(ws)}


@router.put('/witness-staff/{staff_id}', summary='编辑双签工作人员')
@require_permission('protocol.protocol.update')
def witness_staff_update(request, staff_id: int, data: WitnessStaffUpdateIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    ws = ws_svc.update_witness_staff(
        staff_id,
        name=data.name,
        email=data.email,
        gender=data.gender,
        id_card_no=data.id_card_no,
        phone=data.phone,
        priority=data.priority,
    )
    if not ws:
        return 404, {'code': 404, 'msg': '记录不存在'}
    return {'code': 200, 'msg': 'OK', 'data': ws_svc.witness_staff_to_dict(ws)}


@router.delete('/witness-staff/{staff_id}', summary='删除双签工作人员（软删除）')
@require_permission('protocol.protocol.update')
def witness_staff_delete(request, staff_id: int):
    from apps.protocol.services import witness_staff_service as ws_svc

    if not ws_svc.soft_delete_witness_staff(staff_id):
        return 404, {'code': 404, 'msg': '记录不存在'}
    return {'code': 200, 'msg': 'OK', 'data': None}


@router.get(
    '/{protocol_id}/dual-sign-staff-status',
    summary='双签工作人员在本协议+签署节点下的核验阶段（待发邮件/待核验/核验中/已核验）',
)
@require_any_permission(['protocol.protocol.read', 'protocol.protocol.update', 'subject.subject.read'])
def dual_sign_staff_status_list(
    request,
    protocol_id: int,
    icf_version_id: int = Query(...),
    staff_ids: Optional[str] = Query(None),
):
    """
    若传入 staff_ids（逗号分隔），按该列表计算；否则使用协议知情配置中已保存的双签名单。
    与执行台「已选工作人员」列表对齐时，请传入当前勾选顺序的 id。
    """
    from apps.subject.models import ICFVersion
    from apps.protocol.services import witness_staff_service as ws_svc

    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    icf = ICFVersion.objects.filter(id=icf_version_id, protocol_id=protocol_id).first()
    if not icf:
        return 404, {'code': 404, 'msg': '签署节点不存在'}
    ids: list[int] = []
    if staff_ids and str(staff_ids).strip():
        for part in str(staff_ids).split(','):
            part = part.strip()
            if not part:
                continue
            try:
                ids.append(int(part))
            except ValueError:
                continue
    else:
        settings = _get_effective_consent_settings(protocol)
        for s in settings.get('dual_sign_staffs') or []:
            sid = s.get('staff_id')
            if sid is None:
                continue
            try:
                ids.append(int(str(sid).strip()))
            except ValueError:
                continue
    items = ws_svc.dual_sign_staff_status_batch(protocol_id, icf_version_id, ids)
    return {'code': 200, 'msg': 'OK', 'data': {'items': items}}


@router.post(
    '/{protocol_id}/dual-sign-auth-request',
    summary='提交双签身份验证（发邮件）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def dual_sign_auth_request(request, protocol_id: int, data: DualSignAuthRequestIn):
    """工作人员 + 签署节点 + 通知邮箱就绪后发送授权邮件，并写入协议知情配置中的双签名单。"""
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    result = _perform_dual_sign_auth(protocol_id, data)
    if isinstance(result, tuple):
        return result
    return result


@router.post(
    '/{protocol_id}/dual-sign-verification-reset',
    summary='清空本协议双签全员人脸核验状态（未发布）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def dual_sign_verification_reset(request, protocol_id: int):
    """
    将 **已保存** 双签名单中对应工作人员的档案核验字段清空，并同步知情配置快照；
    清除本协议下待轮询的火山 byted_token。重做须重新「发送验证邮件」，每人走 witness-auth/face-start + face-result（真实火山核身，非 Mock）。
    """
    _, protocol = _check_protocol_visible(request, protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    from apps.protocol.services import witness_staff_service as ws_svc

    settings_data = _get_consent_settings(protocol)
    if settings_data.get('consent_launched'):
        return 400, {'code': 400, 'msg': '知情已发布，无法清空核验状态。请先取消发布后再操作。'}
    if not settings_data.get('require_dual_sign'):
        return 400, {'code': 400, 'msg': '未启用工作人员见证双签'}
    staffs = settings_data.get('dual_sign_staffs') or []
    if not staffs:
        return 400, {'code': 400, 'msg': '当前协议未保存双签名单，请先保存配置后再清空'}
    ids: List[int] = []
    for s in staffs:
        sid = str(s.get('staff_id') or '').strip()
        if sid.isdigit():
            ids.append(int(sid))
    if not ids:
        return 400, {'code': 400, 'msg': '双签名单无有效工作人员 ID'}
    cleared = ws_svc.clear_witness_staff_face_verification(ids)
    WitnessDualSignAuthToken.objects.filter(protocol_id=protocol.id, witness_staff_id__in=ids).update(
        face_byted_token=''
    )
    fresh = _get_consent_settings(protocol)
    fresh['consent_verify_signature_authorized'] = False
    _save_consent_settings(protocol, fresh)
    return {
        'code': 200,
        'msg': 'OK',
        'data': {'cleared_staff': cleared, 'staff_ids': ids},
    }


@router.post(
    '/witness-staff/{staff_id}/send-auth-email',
    summary='从工作人员列表发起身份验证邮件',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def witness_staff_send_auth_email(request, staff_id: int, data: WitnessStaffNotifyIn):
    _, protocol = _check_protocol_visible(request, data.protocol_id)
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    payload = DualSignAuthRequestIn(
        witness_staff_id=staff_id,
        icf_version_id=data.icf_version_id,
        notify_email=data.notify_email,
    )
    result = _perform_dual_sign_auth(data.protocol_id, payload)
    if isinstance(result, tuple):
        return result
    return result


@router.post(
    '/witness-staff/{staff_id}/send-profile-verify-email',
    summary='双签名单：发送档案核验邮件（人脸+手写签名登记，不绑定协议）',
    response={200: dict, 400: dict, 404: dict},
)
@require_permission('protocol.protocol.update')
def witness_staff_send_profile_verify_email(request, staff_id: int, data: WitnessStaffProfileNotifyIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    ws = WitnessStaff.objects.filter(id=staff_id, is_deleted=False).first()
    if not ws:
        return 404, {'code': 404, 'msg': '工作人员不存在'}
    if not ws_svc.witness_staff_can_receive_auth_emails(ws):
        return 400, {
            'code': 400,
            'msg': '该人员无法发送核验邮件：须为具备 QA质量管理 全局角色的治理台账号，或已填写姓名与工作邮箱的无账号档案',
        }
    notify = (data.notify_email or '').strip() or (ws.email or '').strip()
    if not notify:
        return 400, {'code': 400, 'msg': '请填写工作邮箱或在治理台维护邮箱后再试'}
    try:
        ws_svc.send_witness_profile_verification_email(witness=ws, notify_email=notify)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': '核验邮件已发送', 'data': {'notify_email': notify}}


@router.get('/witness-auth/resolve', summary='解析双签授权令牌（公开）', auth=None)
def witness_auth_resolve(request, token: str = Query(...)):
    from apps.protocol.services import witness_staff_service as ws_svc

    row = ws_svc.resolve_auth_token(token)
    if not row:
        return 400, {'code': 400, 'msg': '链接无效或已过期'}
    ws = row.witness_staff
    protocol = (
        Protocol.objects.filter(id=row.protocol_id, is_deleted=False).first()
        if row.protocol_id is not None
        else None
    )
    code = getattr(protocol, 'code', '') if protocol else ''
    title = getattr(protocol, 'title', '') if protocol else ''
    has_id = bool((ws.id_card_no or '').strip())
    has_phone = bool((ws.phone or '').strip())
    eff = ws_svc.witness_face_verification_effective(ws)
    legacy = ws_svc.witness_has_legacy_placeholder_face_record(ws)
    from apps.subject.services.identity_provider_service import get_identity_provider_config_state

    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'name': ws.name,
            'id_card_no': ws.id_card_no or '',
            'phone': ws.phone or '',
            'has_id_card_and_phone': has_id and has_phone,
            'email': ws.email or '',
            'token_scope': 'profile' if row.protocol_id is None else 'project',
            'protocol_id': row.protocol_id,
            'protocol_code': code,
            'protocol_title': title,
            'icf_version_id': row.icf_version_id,
            'staff_signature_registered': bool(getattr(row, 'staff_signature_registered_at', None)),
            'identity_verified': bool(ws.identity_verified),
            'face_verified_at': ws.face_verified_at.isoformat() if ws.face_verified_at else None,
            # 邮件页须以 effective 为准：identity_verified 曾为占位接口误标为 True
            'face_verification_effective': eff,
            'legacy_placeholder_face_record': legacy,
            # 与受试者 L2 共用：便于公开页在发起前提示缺哪项环境变量
            'identity_provider_state': get_identity_provider_config_state(),
            # 联调：为 true 时 face-start 将跳过火山并跳转知情联调页
            'witness_face_dev_bypass': bool(getattr(django_settings, 'WITNESS_FACE_DEV_BYPASS', False)),
            'signature_auth_decision': (row.signature_auth_decision or '').strip() or None,
            'signature_auth_at': (
                row.signature_auth_at.isoformat()
                if getattr(row, 'signature_auth_at', None)
                else None
            ),
            # 双签档案中是否已上传手写签名（同意授权前置条件）
            'staff_signature_on_file': bool((ws.signature_file or '').strip()),
            # 深链定位双签名单行（与 t_witness_staff.id 一致）
            'witness_staff_id': ws.id,
        },
    }


class WitnessSignatureAuthorizeIn(Schema):
    """人脸通过后：是否同意本项目使用签名信息（与知情签署测试入口分离）。"""

    token: str
    decision: str  # agree | refuse


class WitnessStaffSignatureRegisterIn(Schema):
    """档案核验邮件：人脸通过后提交手写签名图片（公开）。"""

    token: str
    image_base64: str


@router.post(
    '/witness-auth/signature-authorize',
    summary='记录签名授权决策（公开，须先完成有效人脸核验）',
    auth=None,
    response={200: dict, 400: dict},
)
def witness_auth_signature_authorize(request, data: WitnessSignatureAuthorizeIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    try:
        out = ws_svc.record_witness_signature_authorization(data.token, data.decision)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.post(
    '/witness-auth/register-staff-signature',
    summary='档案核验：人脸通过后提交手写签名图片（公开）',
    auth=None,
    response={200: dict, 400: dict},
)
def witness_auth_register_staff_signature(request, data: WitnessStaffSignatureRegisterIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    try:
        out = ws_svc.register_witness_staff_signature_from_token(data.token, data.image_base64)
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.get(
    '/witness-auth/dev-consent-queue',
    summary='联调：公开令牌下列出协议 ICF 签署节点（仅 WITNESS_FACE_DEV_BYPASS）',
    auth=None,
    response={200: dict, 400: dict, 403: dict},
)
def witness_auth_dev_consent_queue(request, token: str = Query(...)):
    """供执行台 /#/witness-consent-dev 列出 ICF；写入签署记录见 POST …/dev-consent-submit。"""
    from apps.subject.models import ICFVersion
    from apps.protocol.services import protocol_service as protocol_svc
    from apps.protocol.services import witness_staff_service as ws_svc

    if not getattr(django_settings, 'WITNESS_FACE_DEV_BYPASS', False):
        return 403, {
            'code': 403,
            'msg': '未启用联调模式（环境变量 WITNESS_FACE_DEV_BYPASS=true 并重启 Django）',
        }

    row = ws_svc.resolve_auth_token(token)
    if not row:
        return 400, {'code': 400, 'msg': '链接无效或已过期'}
    if row.protocol_id is None:
        return 400, {'code': 400, 'msg': '当前为档案核验链接，知情联调请使用项目授权邮件'}

    protocol = Protocol.objects.filter(id=row.protocol_id, is_deleted=False).first()
    if not protocol:
        return 400, {'code': 400, 'msg': '协议不存在'}

    qs = (
        ICFVersion.objects.filter(
            protocol_id=protocol.id,
            is_active=True,
            mini_sign_rules_saved=True,
        )
        .order_by('display_order', '-create_time')
    )
    icf_list = list(qs)
    if not icf_list:
        return 400, {
            'code': 400,
            'msg': '当前协议无已保存小程序签署规则的节点，请在执行台知情配置中维护各节点规则并至少保存一次后再试',
        }

    start_id = int(row.icf_version_id)
    ids = [x.id for x in icf_list]
    if start_id not in ids:
        return 400, {
            'code': 400,
            'msg': '授权链接对应的签署节点已不在当前有效节点列表中，请重新发送双签授权邮件',
        }
    icf_list = icf_list[ids.index(start_id) :]

    from apps.subject.services.consent_service import (
        effective_required_reading_seconds_for_icf,
        get_effective_mini_sign_rules,
    )

    items = []
    for icf in icf_list:
        sec = effective_required_reading_seconds_for_icf(icf, protocol)
        body_html = protocol_svc.resolve_icf_body_html_for_witness_dev(icf)
        rules = get_effective_mini_sign_rules(protocol, icf)
        enable_subj = bool(rules.get('enable_subject_signature', False))
        subj_times = _clamp_1_or_2(rules.get('subject_signature_times'), 1) if enable_subj else 0
        sup_labels = rules.get('supplemental_collect_labels')
        if not isinstance(sup_labels, list):
            sup_labels = []
        sup_labels = [str(x).strip() for x in sup_labels if str(x).strip()][:20]
        items.append(
            {
                'icf_version_id': icf.id,
                'node_title': (icf.node_title or '').strip() or f'签署节点 {icf.version}',
                'version': icf.version,
                'required_reading_duration_seconds': sec,
                'content': body_html,
                'enable_subject_signature': enable_subj,
                'subject_signature_times': subj_times,
                'enable_checkbox_recognition': bool(rules.get('enable_checkbox_recognition', False)),
                'supplemental_collect_labels': sup_labels,
                'collect_other_information': bool(rules.get('collect_other_information', False)),
            }
        )

    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'protocol_id': protocol.id,
            'protocol_code': getattr(protocol, 'code', '') or '',
            'protocol_title': getattr(protocol, 'title', '') or '',
            'items': items,
        },
    }


@router.post(
    '/witness-auth/dev-consent-submit',
    summary='联调：写入测试签署记录（公开，WITNESS_FACE_DEV_BYPASS）',
    auth=None,
    response={200: dict, 400: dict, 403: dict},
)
def witness_auth_dev_consent_submit(request, data: WitnessDevConsentSubmitIn):
    from apps.protocol.services import witness_staff_service as ws_svc

    if not getattr(django_settings, 'WITNESS_FACE_DEV_BYPASS', False):
        return 403, {
            'code': 403,
            'msg': '未启用联调模式（环境变量 WITNESS_FACE_DEV_BYPASS=true 并重启 Django）',
        }
    try:
        out = ws_svc.submit_witness_dev_consent_records(
            data.token,
            list(data.icf_version_ids or []),
            list(data.icf_version_answers) if data.icf_version_answers else None,
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    except RuntimeError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.post(
    '/witness-auth/face-start',
    summary='发起人脸核身 H5（公开，火山引擎）',
    auth=None,
    # 须声明非 200 状态码，否则 Ninja 抛 ConfigError，全局处理器会把正文替换成「操作失败」
    response={200: dict, 400: dict, 503: dict, 500: dict},
)
def witness_auth_face_start(request, data: WitnessFaceStartIn):
    from apps.protocol.services import witness_staff_service as ws_svc
    from apps.subject.services.identity_provider_service import get_identity_provider_config_state

    try:
        out = ws_svc.start_witness_face_verification(
            data.token,
            id_card_no=data.id_card_no,
            phone=data.phone,
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    except RuntimeError as e:
        err = str(e)
        if err == 'IDENTITY_PROVIDER_UNAVAILABLE':
            return 503, {
                'code': 503,
                'msg': '实名核身服务未配置，请联系管理员（与受试者 L2 相同的火山引擎身份认证环境变量）。',
                'error_code': 'IDENTITY_PROVIDER_UNAVAILABLE',
                'data': get_identity_provider_config_state(),
            }
        if err == 'WITNESS_FACE_TENCENT_NOT_IMPLEMENTED':
            return 503, {
                'code': 503,
                'msg': '腾讯云人脸核身尚未接入；请优先完成火山引擎配置，或参阅 docs/WITNESS_FACE_IDENTITY_PROVIDERS.md 扩展备选。',
                'error_code': 'WITNESS_FACE_TENCENT_NOT_IMPLEMENTED',
            }
        return 500, {'code': 500, 'msg': err}
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.get('/witness-auth/face-result', summary='查询人脸核身结果（公开，轮询）', auth=None)
def witness_auth_face_result(request, token: str = Query(...)):
    from apps.protocol.services import witness_staff_service as ws_svc

    out = ws_svc.poll_witness_face_verification(token)
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.post('/witness-auth/face-submit', summary='【已废弃】人脸识别占位提交', auth=None)
def witness_auth_face_submit(request, data: WitnessFaceSubmitIn):
    return 400, {
        'code': 400,
        'msg': '该流程已升级为火山引擎在线人脸核验，请刷新页面后使用「开始人脸核验」。',
    }


@router.get('/public/consent-test-queue', summary='核验测试 H5：公开口令下列出 ICF 节点', auth=None)
def consent_test_scan_queue(request, p: int = Query(...), t: str = Query(...)):
    """供执行台 /#/consent-test-scan 拉取与联调页一致的节点列表（不须小程序）。"""
    from .consent_test_tokens import unsign_consent_test_scan_token

    pid = unsign_consent_test_scan_token(t)
    if pid is None or int(pid) != int(p):
        return 400, {'code': 400, 'msg': '链接无效或已过期'}
    protocol = Protocol.objects.filter(id=p, is_deleted=False).first()
    if not protocol:
        return 404, {'code': 404, 'msg': '协议不存在'}
    if _is_consent_launched(protocol):
        return 403, {
            'code': 403,
            'msg': '知情已发布，预发布核验测试不可用，请使用正式入口或先下架知情。',
        }
    st = get_consent_config_status_for_protocol(protocol)
    if st not in ('已授权待测试', '已测试待开始', '核验测试中', '待测试'):
        return 403, {
            'code': 403,
            'msg': '请先完成配置与工作人员授权核验',
            'data': {'config_status': st},
        }
    items, err = _build_consent_test_scan_icf_items(protocol)
    if err:
        return 400, {'code': 400, 'msg': err}
    cs = _get_consent_settings(protocol)
    require_face = bool(cs.get('require_face_verify', False))
    return {
        'code': 200,
        'msg': 'OK',
        'data': {
            'protocol_id': protocol.id,
            'protocol_code': getattr(protocol, 'code', '') or '',
            'protocol_title': getattr(protocol, 'title', '') or '',
            'items': items,
            'require_face_verify': require_face,
            'enable_auto_sign_date': bool(cs.get('enable_auto_sign_date', False)),
        },
    }


@router.post('/public/consent-test-submit', summary='核验测试 H5：写入测试类型签署记录', auth=None)
def consent_test_scan_submit(request, data: ConsentTestScanSubmitIn):
    from .services import witness_staff_service as ws_svc

    try:
        out = ws_svc.submit_consent_test_scan_records(
            data.p,
            data.t,
            list(data.icf_version_ids or []),
            list(data.icf_version_answers) if data.icf_version_answers else None,
            subject_name=data.subject_name,
            id_card_no=data.id_card_no,
            phone=data.phone,
            screening_number=data.screening_number,
            icf_version_signatures=list(data.icf_version_signatures) if data.icf_version_signatures else None,
        )
    except ValueError as e:
        return 400, {'code': 400, 'msg': str(e)}
    except RuntimeError as e:
        return 400, {'code': 400, 'msg': str(e)}
    return {'code': 200, 'msg': 'OK', 'data': out}


@router.get('/public/consent-test-receipt', summary='核验测试 H5：凭口令下载签署回执 PDF', auth=None)
def consent_test_scan_receipt(
    request,
    p: int = Query(...),
    t: str = Query(...),
    consent_id: int = Query(...),
    batch_id: str = Query(...),
    download: int = Query(0),
):
    """
    与 submit 返回的 consent_test_scan_batch_id + consent_id 配套使用；
    校验 scan_token、协议、批次，避免仅凭 consent_id 枚举他人回执。
    """
    from .consent_test_tokens import unsign_consent_test_scan_token
    from apps.subject.models import SubjectConsent
    from apps.subject.services.consent_service import _ensure_receipt_pdf

    pid = unsign_consent_test_scan_token(t)
    if pid is None or int(pid) != int(p):
        return 400, {'code': 400, 'msg': '链接无效或已过期'}

    bid = (batch_id or '').strip()
    if not bid:
        return 400, {'code': 400, 'msg': '缺少 batch_id'}

    c = (
        SubjectConsent.objects.filter(id=consent_id)
        .select_related('icf_version', 'icf_version__protocol')
        .first()
    )
    if not c:
        return 404, {'code': 404, 'msg': '记录不存在'}

    icf = c.icf_version
    if not icf or int(icf.protocol_id) != int(p):
        return 403, {'code': 403, 'msg': '无权访问'}

    sig = dict(c.signature_data or {})
    if not sig.get('consent_test_scan_h5'):
        return 403, {'code': 403, 'msg': '无权访问'}
    if (sig.get('consent_test_scan_batch_id') or '').strip() != bid:
        return 403, {'code': 403, 'msg': '无权访问'}

    if not sig.get('receipt_pdf_path') or sig.get('receipt_stub') is not False:
        _ensure_receipt_pdf(c)
        c.save(update_fields=['signature_data', 'update_time'])
        sig = dict(c.signature_data or {})

    rel = (sig.get('receipt_pdf_path') or '').strip()
    if not rel or '..' in rel or rel.startswith('/'):
        return 404, {'code': 404, 'msg': '回执文件不存在'}

    media_root = os.path.abspath(os.path.normpath(django_settings.MEDIA_ROOT))
    abs_path = os.path.abspath(os.path.normpath(os.path.join(media_root, rel)))
    if not abs_path.startswith(media_root + os.sep) and abs_path != media_root:
        return 400, {'code': 400, 'msg': '非法文件路径'}
    if not os.path.isfile(abs_path):
        return 404, {'code': 404, 'msg': '文件不存在'}

    rno = (c.receipt_no or str(c.id)).replace('/', '_').replace('\\', '_')
    safe_name = f'icf_receipt_{rno}.pdf'
    fh = open(abs_path, 'rb')
    as_att = bool(int(download or 0))
    resp = FileResponse(fh, content_type='application/pdf', as_attachment=False, filename=safe_name)
    # 显式区分：预览 inline，下载 attachment（避免部分浏览器将 inline 也当成下载）
    if as_att:
        resp['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    else:
        resp['Content-Disposition'] = f'inline; filename="{safe_name}"'
    return resp


@router.get('/public/consent-test-landing', summary='知情签署测试：兼容旧链，302 至执行台 H5', auth=None)
def consent_test_landing(request, p: int = Query(...), t: str = Query(...)):
    """
    旧二维码指向本接口时：校验后与列表一致 **302** 至执行台
    `…/execution/#/consent-test-scan?p=&t=`（移动端 H5 核验，不拉起小程序）。
    """
    from django.http import HttpResponse
    import html as html_module

    from .consent_test_tokens import unsign_consent_test_scan_token

    pid = unsign_consent_test_scan_token(t)
    if pid is None or int(pid) != int(p):
        return HttpResponse(
            '<!DOCTYPE html><meta charset="utf-8"><title>链接无效</title>'
            '<body style="font-family:system-ui;padding:24px;">链接无效或已过期，请重新在执行台知情管理列表获取二维码。</body>',
            status=400,
            content_type='text/html; charset=utf-8',
        )
    protocol = Protocol.objects.filter(id=p, is_deleted=False).first()
    if not protocol:
        return HttpResponse(
            '<!DOCTYPE html><meta charset="utf-8"><title>协议不存在</title>'
            '<body style="font-family:system-ui;padding:24px;">协议不存在或已删除。</body>',
            status=404,
            content_type='text/html; charset=utf-8',
        )
    if _is_consent_launched(protocol):
        block_msg = '知情已发布，预发布核验测试扫码不可用，请使用正式入口或先下架知情。'
        safe = html_module.escape(block_msg)
        body = (
            f'<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>暂无法开始签署测试</title></head><body style="font-family:system-ui;padding:24px;line-height:1.6;">'
            f'<p style="font-size:18px;font-weight:600;">暂无法开始签署测试</p>'
            f'<p style="color:#334155;">{safe}</p>'
            f'</body></html>'
        )
        return HttpResponse(body, content_type='text/html; charset=utf-8')
    st = get_consent_config_status_for_protocol(protocol)
    block_msg = '请先完成配置与工作人员授权核验'
    if st not in ('已授权待测试', '已测试待开始', '核验测试中', '待测试'):
        safe = html_module.escape(block_msg)
        safe_status = html_module.escape(st)
        body = (
            f'<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>暂无法开始签署测试</title></head><body style="font-family:system-ui;padding:24px;line-height:1.6;">'
            f'<p style="font-size:18px;font-weight:600;">暂无法开始签署测试</p>'
            f'<p style="color:#334155;">{safe}</p>'
            f'<p style="color:#64748b;font-size:14px;">当前状态：{safe_status}</p></body></html>'
        )
        return HttpResponse(body, content_type='text/html; charset=utf-8')

    h5_url = _build_consent_test_scan_url(request, p, t)
    resp = HttpResponse(status=302)
    resp['Location'] = h5_url
    return resp

