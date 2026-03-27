"""
样品发放（产品发放）API — 与 KIS /api/v1/product/* 响应格式一致：{ success, data, message }。
数据存 cn_kis default 库，不再依赖 kis_test。
"""
import io
from typing import Optional, Any, List
from urllib.parse import quote

from django.http import JsonResponse, HttpResponse
from ninja import Router, Schema, Query
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.identity.decorators import _get_account_from_request
from . import services

router = Router()


def _kis_response(data: Any = None, message: str = "操作成功") -> dict:
    return {"success": True, "data": data, "message": message}


def _get_user(request):
    """返回 (account_id, 显示名)，显示名优先 display_name，用于操作人/创建人等展示."""
    account = _get_account_from_request(request)
    if not account:
        return None, None
    display = (getattr(account, "display_name", None) or "").strip() or None
    if not display:
        display = getattr(account, "name", None) or getattr(account, "username", None)
    return getattr(account, "id", None), display


# ---------- 工单 ----------

@router.get("/workorders")
def get_work_orders(
    request,
    keyword: Optional[str] = Query(None),
    project_no: Optional[str] = Query(None),
    project_start_date: Optional[str] = Query(None),
    project_end_date: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
):
    data = services.work_order_list(
        keyword=keyword,
        project_no=project_no,
        project_start_date=project_start_date,
        project_end_date=project_end_date,
        page=page,
        page_size=pageSize,
    )
    return JsonResponse(_kis_response(data=data))


@router.get("/workorders/{id}/executions")
def get_work_order_executions(
    request,
    id: int,
    page: int = Query(1, ge=1),
    pageSize: int = Query(10, ge=1, le=100),
):
    """工单下执行记录分页（摘要），与接待台分页拉取详情配合。"""
    data = services.work_order_executions_page(id, page=page, page_size=pageSize)
    if data is None:
        return JsonResponse({"success": False, "data": None, "message": "工单不存在"}, status=404)
    return JsonResponse(_kis_response(data=data))


@router.get("/workorders/{id}")
def get_work_order(
    request,
    id: int,
    include_executions: bool = Query(
        True,
        description="为 false 时不返回 executions 摘要列表（仅 executions_total），请用 /workorders/{id}/executions 分页",
    ),
):
    data = services.work_order_detail(id, include_executions=include_executions)
    if data is None:
        return JsonResponse({"success": False, "data": None, "message": "工单不存在"}, status=404)
    return JsonResponse(_kis_response(data=data))


class WorkOrderCreateIn(Schema):
    project_no: str
    project_name: str
    project_start_date: str
    project_end_date: str
    visit_count: int = 0
    researcher: Optional[str] = None
    supervisor: Optional[str] = None
    usage_method: Optional[str] = None
    usage_frequency: Optional[str] = None
    precautions: Optional[str] = None
    project_requirements: Optional[str] = None


@router.post("/workorders")
def create_work_order(request, data: WorkOrderCreateIn):
    user_id, user_name = _get_user(request)
    try:
        payload = data.dict()
        payload["project_start_date"] = payload.get("project_start_date")
        payload["project_end_date"] = payload.get("project_end_date")
        result = services.work_order_create(payload, user_id=user_id, user_name=user_name)
        return JsonResponse(_kis_response(data=result, message="创建成功"))
    except ValueError as e:
        return JsonResponse({"success": False, "data": None, "message": str(e)}, status=400)


class WorkOrderUpdateIn(Schema):
    project_no: Optional[str] = None
    project_name: Optional[str] = None
    project_start_date: Optional[str] = None
    project_end_date: Optional[str] = None
    visit_count: Optional[int] = None
    researcher: Optional[str] = None
    supervisor: Optional[str] = None
    usage_method: Optional[str] = None
    usage_frequency: Optional[str] = None
    precautions: Optional[str] = None
    project_requirements: Optional[str] = None


@router.put("/workorders/{id}")
def update_work_order(request, id: int, data: WorkOrderUpdateIn):
    user_id, _ = _get_user(request)
    payload = data.dict(exclude_unset=True)
    result = services.work_order_update(id, payload, user_id=user_id)
    if result is None:
        return JsonResponse({"success": False, "data": None, "message": "工单不存在"}, status=404)
    try:
        return JsonResponse(_kis_response(data=result, message="更新成功"))
    except ValueError as e:
        return JsonResponse({"success": False, "data": None, "message": str(e)}, status=400)


# ---------- 项目执行概览（接待台日历） ----------

@router.get("/execution-overview")
def get_execution_overview(request, date: str = Query(..., description="YYYY-MM-DD")):
    """按日期返回项目执行概览。"""
    items = services.execution_overview_by_date(date)
    return JsonResponse(_kis_response(data={"items": items}))


@router.get("/execution-overview/counts")
def get_execution_overview_counts(request, month: str = Query(..., description="YYYY-MM")):
    """按月份返回每日项目数量（月历角标）。"""
    counts = services.execution_overview_counts_by_month(month)
    return JsonResponse(_kis_response(data=counts))


# ---------- 枚举 ----------

@router.get("/enums/execution-stage")
def get_enums_execution_stage(request):
    return JsonResponse(_kis_response(data=services.enum_execution_stage()))


@router.get("/enums/exception-type")
def get_enums_exception_type(request):
    return JsonResponse(_kis_response(data=services.enum_exception_type()))


# ---------- 样品台账与记录 ----------

@router.get("/orders-ledger")
def get_orders_ledger(
    request,
    related_project_no: Optional[str] = Query(None),
    product_code: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
):
    data = services.sample_ledger_list(
        related_project_no=related_project_no,
        product_code=product_code,
        keyword=keyword,
        page=page,
        page_size=pageSize,
    )
    return JsonResponse(_kis_response(data=data))


@router.get("/orders")
def get_sample_orders(
    request,
    keyword: Optional[str] = Query(None),
    related_project_no: Optional[str] = Query(None),
    product_code: Optional[str] = Query(None),
    operation_type: Optional[str] = Query(None),
    operation_date_from: Optional[str] = Query(None),
    operation_date_to: Optional[str] = Query(None),
    purpose: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
):
    data = services.sample_order_list(
        keyword=keyword,
        related_project_no=related_project_no,
        product_code=product_code,
        operation_type=operation_type,
        operation_date_from=operation_date_from,
        operation_date_to=operation_date_to,
        purpose=purpose,
        page=page,
        page_size=pageSize,
    )
    return JsonResponse(_kis_response(data=data))


@router.get("/orders/{id}")
def get_sample_order(request, id: int):
    data = services.sample_order_detail(id)
    if data is None:
        return JsonResponse({"success": False, "data": None, "message": "记录不存在"}, status=404)
    return JsonResponse(_kis_response(data=data))


class SampleOrderCreateIn(Schema):
    operation_type: str = "receive"
    operation_date: str
    related_project_no: str
    project_name: Optional[str] = None
    project_start_date: Optional[str] = None
    project_end_date: Optional[str] = None
    researcher: Optional[str] = None
    supervisor: Optional[str] = None
    product_name: str
    product_code: str
    quantity: float
    unit: Optional[str] = None
    purpose: str
    remark: Optional[str] = None


@router.post("/orders")
def create_sample_order(request, data: SampleOrderCreateIn):
    user_id, user_name = _get_user(request)
    payload = data.dict()
    result = services.sample_order_create(payload, user_id=user_id, user_name=user_name)
    return JsonResponse(_kis_response(data=result, message="创建成功"))


class SampleOrderUpdateIn(Schema):
    operation_type: Optional[str] = None
    operation_date: Optional[str] = None
    related_project_no: Optional[str] = None
    project_name: Optional[str] = None
    project_start_date: Optional[str] = None
    project_end_date: Optional[str] = None
    researcher: Optional[str] = None
    supervisor: Optional[str] = None
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    purpose: Optional[str] = None
    remark: Optional[str] = None


@router.put("/orders/{id}")
def update_sample_order(request, id: int, data: SampleOrderUpdateIn):
    user_id, user_name = _get_user(request)
    payload = data.dict(exclude_unset=True)
    result = services.sample_order_update(id, payload, user_id=user_id, user_name=user_name)
    if result is None:
        return JsonResponse({"success": False, "data": None, "message": "记录不存在"}, status=404)
    return JsonResponse(_kis_response(data=result, message="更新成功"))


@router.get("/project-products")
def get_project_products(request, related_project_no: str = Query(...)):
    data = services.project_products(related_project_no)
    return JsonResponse(_kis_response(data=data))


# ---------- 执行记录 ----------

@router.get("/execution-orders")
def get_execution_orders(
    request,
    work_order_id: Optional[int] = Query(None),
    subject_rd: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    execution_date_from: Optional[str] = Query(None),
    execution_date_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
):
    data = services.execution_list(
        work_order_id=work_order_id,
        subject_rd=subject_rd,
        keyword=keyword,
        execution_date_from=execution_date_from,
        execution_date_to=execution_date_to,
        page=page,
        page_size=pageSize,
    )
    return JsonResponse(_kis_response(data=data))


@router.get("/execution-orders/pending-skips")
def get_execution_pending_skips(request, queue_date: str = Query(..., description="队列日期 YYYY-MM-DD")):
    """当日已「无需执行」结案的待执行项，用于前端从待执行列表排除。"""
    data = services.pending_execution_skips_for_queue_date(queue_date)
    return JsonResponse(_kis_response(data=data))


@router.get("/execution-orders/{id}")
def get_execution_order(request, id: int):
    data = services.execution_detail(id)
    if data is None:
        return JsonResponse({"success": False, "data": None, "message": "执行记录不存在"}, status=404)
    return JsonResponse(_kis_response(data=data))


class ProductOperationItemIn(Schema):
    id: Optional[int] = None  # 已有记录时传，用于仅更新有变更的行
    stage: str
    execution_cycle: Optional[str] = None
    product_code: str
    product_name: str
    bottle_sequence: Optional[str] = None
    is_selected: int = 1
    product_operation_type: Optional[str] = None
    product_distribution: bool = False
    product_inspection: bool = False
    product_recovery: bool = False
    product_site_use: bool = False
    distribution_weight: Optional[float] = None
    inspection_weight: Optional[float] = None
    recovery_weight: Optional[float] = None
    diary_distribution: bool = False
    diary_inspection: bool = False
    diary_recovery: bool = False
    usage_diagram_file_id: Optional[int] = None


class ExecutionCreateIn(Schema):
    work_order_id: int
    related_project_no: str
    subject_rd: Optional[str] = None  # 选填；空字符串表示未填
    subject_initials: str
    operator_name: Optional[str] = None  # 当前登录用户显示名，不传则用后端 account.name
    screening_no: str  # 受试者SC号（必填）
    execution_date: Optional[str] = None
    exception_type: Optional[str] = None
    exception_description: Optional[str] = None
    remark: Optional[str] = None
    products: List[ProductOperationItemIn] = []
    # 待执行工单「无需执行」：无产品行，remark 内写入结案标记
    skip_execution: bool = False
    pending_queue_date: Optional[str] = None
    pending_checkin_id: Optional[int] = None
    pending_subject_id: Optional[int] = None


@router.post("/execution-orders")
def create_execution_order(request, data: ExecutionCreateIn):
    user_id, user_name = _get_user(request)
    payload = data.dict() if hasattr(data, "dict") else data.model_dump()
    prods = payload.get("products") or []
    payload["products"] = [p.dict() if hasattr(p, "dict") else (p.model_dump() if hasattr(p, "model_dump") else p) for p in prods]
    try:
        result = services.execution_create(payload, user_id=user_id, user_name=user_name)
        return JsonResponse(_kis_response(data=result, message="创建成功"))
    except ValueError as e:
        return JsonResponse({"success": False, "data": None, "message": str(e)}, status=400)


class ExecutionUpdateIn(Schema):
    related_project_no: Optional[str] = None
    subject_rd: Optional[str] = None
    subject_initials: Optional[str] = None
    operator_name: Optional[str] = None  # 当前登录用户显示名
    screening_no: Optional[str] = None
    execution_date: Optional[str] = None
    exception_type: Optional[str] = None
    exception_description: Optional[str] = None
    remark: Optional[str] = None
    products: Optional[List[ProductOperationItemIn]] = None


@router.put("/execution-orders/{id}")
def update_execution_order(request, id: int, data: ExecutionUpdateIn):
    user_id, user_name = _get_user(request)
    payload = data.dict(exclude_unset=True) if hasattr(data, "dict") else data.model_dump(exclude_unset=True)
    if payload.get("operator_name") is None and user_name:
        payload["operator_name"] = user_name
    if payload.get("products") is not None:
        payload["products"] = [
            p if isinstance(p, dict) else (p.dict() if hasattr(p, "dict") else getattr(p, "model_dump", lambda: p)())
            for p in payload["products"]
        ]
    try:
        result = services.execution_update(id, payload, user_id=user_id, user_name=user_name)
        if result is None:
            return JsonResponse({"success": False, "data": None, "message": "执行记录不存在"}, status=404)
        return JsonResponse(_kis_response(data=result, message="更新成功"))
    except ValueError as e:
        return JsonResponse({"success": False, "data": None, "message": str(e)}, status=400)


@router.delete("/execution-orders/{id}")
def delete_execution_order(request, id: int):
    ok = services.execution_delete(id)
    if not ok:
        return JsonResponse({"success": False, "data": None, "message": "执行记录不存在"}, status=404)
    return JsonResponse(_kis_response(data=None, message="删除成功"))


def _build_excel_bytes(sheet_name: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportExcelIn(Schema):
    sheet_name: str = "Sheet1"
    filename: str = "export.xlsx"
    headers: List[str]
    rows: List[List[Any]]


@router.post("/export-excel")
def export_excel(request, data: ExportExcelIn):
    content = _build_excel_bytes(
        data.sheet_name,
        data.headers,
        data.rows,
    )
    filename_encoded = quote(data.filename, safe="")
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename_encoded}"
    return response
