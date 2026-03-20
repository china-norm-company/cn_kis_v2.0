"""
设备台账批量导入服务

支持 Excel（.xlsx）格式，字段映射参考 FS-RF-036 设备台账模板：
- 资产类别、货主组织、卡片编码、资产名称、计量单位、资产数量
- 开始使用日期、期初原值、对应lims上的编号、位置、组别、设备名称
"""
import logging
from datetime import datetime
from decimal import Decimal
from typing import Any, Optional
import openpyxl

from ..models import ResourceItem, ResourceCategory, ResourceType, ResourceStatus

logger = logging.getLogger(__name__)

# Excel 列名 -> 内部字段映射（支持附件模板及常见变体）
EXCEL_COLUMN_MAP = {
    '资产类别': 'asset_category',
    '货主组织': 'organization',
    '卡片编码': 'code',
    '资产名称': 'name',
    '计量单位（卡片）': 'unit',
    '计量单位': 'unit',
    '资产数量（卡片）': 'quantity',
    '资产数量': 'quantity',
    '开始使用日期': 'purchase_date',
    '购入日期': 'purchase_date',
    '期初原值': 'initial_value',
    '对应lims上的编号': 'lims_code',
    '对应lims上的编号 ': 'lims_code',
    '位置': 'location',
    '组别': 'group',
    '设备名称': 'device_name',
    '设备编号': 'code',
    '设备型号': 'model_number',
    '制造商': 'manufacturer',
    '序列号': 'serial_number',
    '保修到期': 'warranty_expiry',
}


def _normalize(val: Any) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    if s.upper() in ('#N/A', 'N/A', 'NA', '-'):
        return ''
    return s


def _parse_date(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date() if hasattr(val, 'date') else val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(val[:10], fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(val: Any) -> Optional[Decimal]:
    if val is None or val == '':
        return None
    try:
        return Decimal(str(val))
    except Exception:
        return None


def _parse_int(val: Any) -> Optional[int]:
    if val is None or val == '':
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _resolve_category(default_name: str = '仪器设备') -> Optional[ResourceCategory]:
    """解析设备类别，优先取第一个设备类型类别"""
    cat = ResourceCategory.objects.filter(
        resource_type=ResourceType.EQUIPMENT,
        is_active=True,
    ).first()
    if cat:
        return cat
    # 若无类别则尝试按名称匹配
    cat = ResourceCategory.objects.filter(
        resource_type=ResourceType.EQUIPMENT,
        name__icontains=default_name or '设备',
        is_active=True,
    ).first()
    return cat


def parse_excel_rows(file) -> list[dict]:
    """解析 Excel 文件，返回标准化行数据"""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        h = _normalize(cell.value) if cell.value else ''
        headers.append(h)
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {'_row': row_idx}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                std_key = EXCEL_COLUMN_MAP.get(headers[i])
                if std_key and val is not None and str(val).strip():
                    row_dict[std_key] = val
        # 至少需要 code 或 name
        if row_dict.get('code') or row_dict.get('name'):
            rows.append(row_dict)
    wb.close()
    return rows


def import_equipment_from_rows(
    rows: list[dict],
    created_by_id: Optional[int] = None,
    update_existing: bool = False,
) -> dict:
    """
    从解析后的行数据批量导入设备

    :param rows: parse_excel_rows 返回的数据
    :param created_by_id: 操作人 ID
    :param update_existing: 若 True，按 code 更新已存在设备；否则跳过
    :return: {total, success, failed, created_ids, updated_ids, errors}
    """
    errors = []
    created_ids = []
    updated_ids = []
    category = _resolve_category()

    if not category:
        return {
            'total': len(rows),
            'success': 0,
            'failed': len(rows),
            'created_ids': [],
            'updated_ids': [],
            'errors': [{'row': 0, 'code': '', 'message': '系统中无设备类别，请先创建资源类别'}],
        }

    for row in rows:
        row_num = row.get('_row', 0)
        code = _normalize(row.get('code'))
        name = _normalize(row.get('name'))
        if not code and not name:
            continue
        if not code:
            code = name or f'IMPORT-{row_num}'
        if not name:
            name = code

        # 扩展属性（存入 attributes）
        attrs = {}
        for k in ['organization', 'unit', 'quantity', 'initial_value', 'lims_code', 'group', 'device_name']:
            v = row.get(k)
            if v is not None and _normalize(v):
                if k == 'quantity':
                    attrs[k] = _parse_int(v)
                elif k == 'initial_value':
                    attrs[k] = float(_parse_decimal(v) or 0)
                else:
                    attrs[k] = _normalize(v)

        purchase_date = _parse_date(row.get('purchase_date'))
        warranty_expiry = _parse_date(row.get('warranty_expiry'))
        location = _normalize(row.get('location')) or ''
        manufacturer = _normalize(row.get('manufacturer')) or ''
        model_number = _normalize(row.get('model_number')) or attrs.get('device_name', '') or ''
        serial_number = _normalize(row.get('serial_number')) or ''

        existing = ResourceItem.objects.filter(code=code, is_deleted=False).first()
        if existing:
            if update_existing:
                existing.name = name
                existing.location = location
                existing.manufacturer = manufacturer
                existing.model_number = model_number or existing.model_number
                existing.serial_number = serial_number or existing.serial_number
                existing.purchase_date = purchase_date or existing.purchase_date
                existing.warranty_expiry = warranty_expiry or existing.warranty_expiry
                existing.attributes = {**existing.attributes, **attrs}
                existing.save()
                updated_ids.append(existing.id)
            else:
                errors.append({'row': row_num, 'code': code, 'message': '设备编号已存在，跳过'})
            continue

        try:
            item = ResourceItem.objects.create(
                name=name,
                code=code,
                category=category,
                status=ResourceStatus.ACTIVE,
                location=location,
                manufacturer=manufacturer,
                model_number=model_number,
                serial_number=serial_number,
                purchase_date=purchase_date,
                warranty_expiry=warranty_expiry,
                attributes=attrs,
            )
            created_ids.append(item.id)
        except Exception as e:
            errors.append({'row': row_num, 'code': code, 'message': str(e)})

    return {
        'total': len(rows),
        'success': len(created_ids) + len(updated_ids),
        'failed': len(errors),
        'created_ids': created_ids,
        'updated_ids': updated_ids,
        'errors': errors,
    }


def parse_and_import_equipment(file, created_by_id=None, update_existing=False) -> dict:
    """解析并导入设备 Excel"""
    rows = parse_excel_rows(file)
    return import_equipment_from_rows(rows, created_by_id=created_by_id, update_existing=update_existing)
