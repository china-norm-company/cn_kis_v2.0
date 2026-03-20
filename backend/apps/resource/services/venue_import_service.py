"""
场地批量导入服务

支持 CSV、Excel（.xlsx）格式。
模板列：场地名称、场地编码、所属中心、面积、场地功能、场地环境要求、场地状态
"""
import csv
import io
import logging
from typing import Any, Optional

import openpyxl

from ..models import ResourceItem, ResourceType
from .. import services_facility as svc

logger = logging.getLogger(__name__)


def _get_existing_venue_codes():
    """获取已存在的场地编码（含软删除），因 DB 唯一约束对全表生效，需全部纳入检查"""
    return set(
        ResourceItem.objects.filter(
            category__resource_type=ResourceType.ENVIRONMENT,
        ).values_list('code', flat=True)
    )

# 模板表头（支持中英文）
TEMPLATE_HEADERS = [
    '场地名称', '场地编码', '所属中心', '面积', '场地功能', '场地环境要求', '场地状态',
]

# 列名 -> 内部字段映射（支持自定义模板，多种列名自动识别）
# 每个字段可对应多种表头写法
COLUMN_ALIASES = {
    'name': ['场地名称', '名称', 'name', '场地名', 'venue_name', 'venue name', '场地', '房间名称', '房间', 'Venue Name'],
    'code': ['场地编码', '编码', 'code', '编号', 'venue_code', 'venue code', '场地编号', '房间编码'],
    'center': ['所属中心', '中心', 'center', '中心名称', '所属', '中心/区域'],
    'area': ['面积', 'area', '面积(m²)', '面积(m2)', '面积（平方米）', '面积㎡'],
    'venue_type': ['场地功能', '场地类型', '功能', 'venue_type', '类型', '用途', '场地用途'],
    'env_requirements': ['场地环境要求', '环境要求', 'env_requirements', '环境', '温湿度要求', '环境参数'],
    'status': ['场地状态', '状态', 'status', '场地状态', '使用状态'],
}

# 展开为 列名 -> 字段 的扁平映射（用于快速查找，支持自定义模板列名）
COLUMN_MAP = {}
for field, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        COLUMN_MAP[alias] = field
        COLUMN_MAP[alias.lower()] = field  # 英文列名不区分大小写

# 场地功能值映射（支持中文或英文）
VENUE_TYPE_MAP = {
    '恒温恒湿测试室': 'testing_room',
    'testing_room': 'testing_room',
    '等候区': 'waiting_area',
    'waiting_area': 'waiting_area',
    '洗漱区': 'washing_area',
    'washing_area': 'washing_area',
    '存储室': 'storage_room',
    'storage_room': 'storage_room',
    '办公室': 'office',
    'office': 'office',
    '功能间': 'utility_room',
    'utility_room': 'utility_room',
    '接待': 'reception',
    'reception': 'reception',
}

# 场地状态映射
STATUS_MAP = {
    '启用': 'active',
    'active': 'active',
    '使用中': 'reserved',
    'reserved': 'reserved',
    '维修中': 'maintenance',
    'maintenance': 'maintenance',
    '停用': 'retired',
    'retired': 'retired',
}


def _normalize(val: Any) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    # #N/A、N/A、NA、N A（空格）、- 视为空；单字母如 A/B/C 保留
    s_compact = s.replace(' ', '').replace('/', '').upper()
    if s_compact in ('#NA', 'NA', 'N/A', '-') or s.upper() in ('#N/A', 'N/A', '-'):
        return ''
    if s.upper() == 'NA' and len(s) == 2:
        return ''
    return s


def _parse_float(val: Any) -> Optional[float]:
    if val is None or val == '':
        return None
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return None


def _row_to_data(headers: list[str], row: list) -> dict:
    """将一行数据转为 create_venue 所需的 dict，支持自定义模板列名自动识别"""
    data = {}
    for i, h in enumerate(headers):
        if i >= len(row):
            break
        h_clean = _normalize(h)
        field = COLUMN_MAP.get(h_clean) or COLUMN_MAP.get(h_clean.lower())
        if not field:
            continue
        val = _normalize(row[i])
        if field == 'name':
            data['name'] = val
        elif field == 'code':
            data['code'] = val
        elif field == 'center':
            data['center'] = val
        elif field == 'area':
            area = _parse_float(row[i])
            data['area'] = area if area is not None else 0
        elif field == 'venue_type':
            data['venue_type'] = VENUE_TYPE_MAP.get(val, val) or 'testing_room'
        elif field == 'env_requirements':
            data['env_requirements'] = val
        elif field == 'status':
            data['status'] = STATUS_MAP.get(val, val) or 'active'
    return data


def _parse_csv(content: bytes) -> list[dict]:
    """解析 CSV 文件"""
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content.decode('gbk', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    result = []
    for idx, row in enumerate(rows[1:], start=2):
        if not any(_normalize(c) for c in row):
            continue
        data = _row_to_data(headers, row)
        if data.get('name') or data.get('code'):
            data['_row'] = idx
            result.append(data)
    return result


def _header_row_match_score(headers: list[str]) -> int:
    """计算表头与已知列名的匹配数量，用于自动识别表头行"""
    score = 0
    for h in headers:
        h_clean = _normalize(h)
        if COLUMN_MAP.get(h_clean) or COLUMN_MAP.get(h_clean.lower()):
            score += 1
    return score


def _parse_excel(content: bytes) -> list[dict]:
    """解析 Excel 文件，支持自动识别表头行（前5行内查找）"""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # 在前5行内查找最佳表头行（匹配列名最多的行）
    header_row_idx = 0
    best_score = 0
    for i in range(min(5, len(rows))):
        cand = [str(h or '').strip() for h in rows[i]]
        score = _header_row_match_score(cand)
        if score > best_score:
            best_score = score
            header_row_idx = i
    headers = [str(h or '').strip() for h in rows[header_row_idx]]
    data_start = header_row_idx + 1
    result = []
    for idx, row in enumerate(rows[data_start:], start=data_start + 1):
        row_list = [str(c) if c is not None else '' for c in row]
        if not any(_normalize(c) for c in row_list):
            continue
        data = _row_to_data(headers, row_list)
        if data.get('name') or data.get('code'):
            data['_row'] = idx
            result.append(data)
    return result


def parse_and_import(content: bytes, filename: str) -> dict:
    """
    解析并导入场地

    Returns:
        {
            'total': int,
            'success': int,
            'failed': int,
            'created_ids': list[int],
            'errors': list[{'row': int, 'code': str, 'message': str}],
        }
    """
    name_lower = (filename or '').lower()
    if name_lower.endswith('.csv'):
        rows_data = _parse_csv(content)
    elif name_lower.endswith('.xlsx'):
        rows_data = _parse_excel(content)
    elif name_lower.endswith('.xls'):
        return {
            'total': 0,
            'success': 0,
            'failed': 0,
            'created_ids': [],
            'errors': [{'row': 0, 'code': 'FORMAT', 'message': '请使用 .xlsx 格式，或改用 .csv 文件'}],
        }
    else:
        return {
            'total': 0,
            'success': 0,
            'failed': 0,
            'created_ids': [],
            'errors': [{'row': 0, 'code': 'FORMAT', 'message': '仅支持 .csv、.xlsx 文件'}],
        }

    result = {
        'total': len(rows_data),
        'success': 0,
        'failed': 0,
        'created_ids': [],
        'errors': [],
    }

    existing_codes = _get_existing_venue_codes()

    for item in rows_data:
        row = item.pop('_row', 0)
        name = item.get('name', '').strip()
        code = item.get('code', '').strip()
        if not name and not code:
            continue
        if not name:
            result['errors'].append({'row': row, 'code': 'MISSING', 'message': '场地名称不能为空'})
            result['failed'] += 1
            continue
        if not code:
            result['errors'].append({'row': row, 'code': 'MISSING', 'message': '场地编码不能为空'})
            result['failed'] += 1
            continue
        if code in existing_codes:
            result['errors'].append({'row': row, 'code': 'DUPLICATE', 'message': f'场地编码 {code} 已存在'})
            result['failed'] += 1
            continue
        try:
            created = svc.create_venue(item)
            result['created_ids'].append(created['id'])
            existing_codes.add(code)
            result['success'] += 1
        except Exception as e:
            result['errors'].append({'row': row, 'code': 'ERROR', 'message': str(e)})
            result['failed'] += 1

    return result


def build_template_excel() -> bytes:
    """生成 Excel 导入模板（含表头和示例行）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws:
        ws.title = '场地导入'
        for col, h in enumerate(TEMPLATE_HEADERS, 1):
            ws.cell(row=1, column=col, value=h)
        # 示例行
        ws.cell(row=2, column=1, value='恒温恒湿测试室 A')
        ws.cell(row=2, column=2, value='VNU-TH-A')
        ws.cell(row=2, column=3, value='上海中心')
        ws.cell(row=2, column=4, value=35)
        ws.cell(row=2, column=5, value='恒温恒湿测试室')
        ws.cell(row=2, column=6, value='22±2°C, 50±10%RH')
        ws.cell(row=2, column=7, value='启用')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    """生成 CSV 导入模板（含表头和示例行）"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(['恒温恒湿测试室 A', 'VNU-TH-A', '上海中心', '35', '恒温恒湿测试室', '22±2°C, 50±10%RH', '启用'])
    return buf.getvalue().encode('utf-8-sig')
