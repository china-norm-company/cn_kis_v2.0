"""
设备管理工作台（器监）API

端点（挂载到 /equipment/）：

仪表盘：
- GET  /dashboard                       设备管理总览面板

设备台账：
- GET  /ledger                          设备列表（增强筛选）
- GET  /index                           设备列表（与 /ledger 相同，兼容旧客户端）
- GET  /ledger/{id}                     设备详情
- POST /ledger/create                   新增设备
- PUT  /ledger/{id}                     更新设备
- POST /ledger/{id}/retire              设备报废
- POST /ledger/{id}/change-status       变更状态

校准管理：
- GET  /calibrations/plan               校准计划视图
- GET  /calibrations/list               校准记录列表
- POST /calibrations/create             新增校准记录
- GET  /calibrations/{id}               校准详情

维护工单：
- GET  /maintenance/list                维护工单列表
- GET  /maintenance/stats               维护统计
- POST /maintenance/create              创建维护工单
- GET  /maintenance/{id}                维护工单详情
- PUT  /maintenance/{id}                更新维护工单
- POST /maintenance/{id}/assign         分配维护
- POST /maintenance/{id}/start          开始维护
- POST /maintenance/{id}/complete       完成维护
- POST /maintenance/{id}/cancel         取消维护

使用记录：
- GET  /usage/list                      使用记录列表
- GET  /usage/stats                     使用统计
- POST /usage/register                  手动登记使用
- POST /usage/{id}/end                  结束使用

操作授权：
- GET  /authorizations/list             授权列表
- POST /authorizations/grant            授予授权
- POST /authorizations/{id}/revoke      撤销授权
- GET  /authorizations/check            检查授权

检测方法：
- GET  /detection-methods/list          检测方法列表
- GET  /detection-methods/{id}          检测方法详情
- POST /detection-methods/sop-upload      上传 SOP 附件（返回 url）
- POST /detection-methods/create        创建检测方法
- PUT  /detection-methods/{id}          更新检测方法
- POST /detection-methods/{id}/resources/add      添加资源需求
- POST /detection-methods/{id}/personnel/add      添加人员要求
- DELETE /detection-methods/resources/{res_id}     删除资源需求
- DELETE /detection-methods/personnel/{per_id}     删除人员要求
"""
from ninja import Router, Schema, Query, File
from ninja.files import UploadedFile
from typing import Optional, List
from pydantic import ConfigDict
from pathlib import Path
import os
import uuid
from django.conf import settings

from apps.identity.decorators import (
    _get_account_from_request,
    require_permission,
    require_any_permission,
)

router = Router()


# ============================================================================
# Schema 定义
# ============================================================================

# --- 设备台账 ---
class LedgerQueryIn(Schema):
    keyword: Optional[str] = None
    category_id: Optional[int] = None
    status: Optional[str] = None
    calibration_status: Optional[str] = None
    location: Optional[str] = None
    page: int = 1
    page_size: int = 20
    sort_by: Optional[str] = '-create_time'
    # LIMS 来源过滤
    lims_only: Optional[bool] = None       # True = 只显示 LIMS 导入；False = 只显示手工录入


class LedgerSummaryQueryIn(Schema):
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 20


class EquipmentCreateIn(Schema):
    model_config = ConfigDict(protected_namespaces=())

    name: str
    code: str
    category_id: int
    name_classification: Optional[str] = ''
    status: Optional[str] = 'active'
    location: Optional[str] = ''
    manufacturer: Optional[str] = ''
    model_number: Optional[str] = ''
    serial_number: Optional[str] = ''
    purchase_date: Optional[str] = None
    warranty_expiry: Optional[str] = None
    next_calibration_date: Optional[str] = None
    next_verification_date: Optional[str] = None
    next_maintenance_date: Optional[str] = None
    calibration_cycle_days: Optional[int] = None
    verification_cycle_days: Optional[int] = None
    maintenance_cycle_days: Optional[int] = None
    manager_id: Optional[int] = None


class EquipmentUpdateIn(Schema):
    model_config = ConfigDict(protected_namespaces=())

    name: Optional[str] = None
    name_classification: Optional[str] = None
    location: Optional[str] = None
    manufacturer: Optional[str] = None
    model_number: Optional[str] = None
    serial_number: Optional[str] = None
    purchase_date: Optional[str] = None
    warranty_expiry: Optional[str] = None
    next_calibration_date: Optional[str] = None
    next_verification_date: Optional[str] = None
    next_maintenance_date: Optional[str] = None
    calibration_cycle_days: Optional[int] = None
    verification_cycle_days: Optional[int] = None
    maintenance_cycle_days: Optional[int] = None
    manager_id: Optional[int] = None


class StatusChangeIn(Schema):
    status: str
    reason: Optional[str] = ''


# --- 校准 ---
class CalibrationQueryIn(Schema):
    equipment_id: Optional[int] = None
    result: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    page: int = 1
    page_size: int = 20


class CalibrationPlanListQueryIn(Schema):
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 50


class CalibrationWorkOrdersIn(Schema):
    equipment_ids: List[int]


class CalibrationCreateIn(Schema):
    equipment_id: int
    calibration_date: str
    next_due_date: str
    calibration_type: Optional[str] = 'internal'
    calibrator: Optional[str] = ''
    certificate_no: Optional[str] = ''
    certificate_file_url: Optional[str] = ''
    result: Optional[str] = 'pass'
    notes: Optional[str] = ''


# --- 核查计划 ---
class VerificationPlanListQueryIn(Schema):
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 50


class VerificationWorkOrdersIn(Schema):
    equipment_ids: List[int]


class VerificationQueryIn(Schema):
    equipment_id: Optional[int] = None
    result: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    page: int = 1
    page_size: int = 20


class VerificationCreateIn(Schema):
    equipment_id: int
    verification_date: str
    next_due_date: str
    verifier: Optional[str] = ''
    result: Optional[str] = 'pass'
    method_notes: Optional[str] = ''
    notes: Optional[str] = ''


# --- 维护计划 ---
class MaintenancePlanListQueryIn(Schema):
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 50


class MaintenanceWorkOrdersIn(Schema):
    equipment_ids: List[int]


# --- 维护工单 ---
class MaintenanceQueryIn(Schema):
    equipment_id: Optional[int] = None
    status: Optional[str] = None
    maintenance_type: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    assigned_to_id: Optional[int] = None
    page: int = 1
    page_size: int = 20


class MaintenanceCreateIn(Schema):
    equipment_id: int
    maintenance_type: str
    title: str
    description: str
    maintenance_date: Optional[str] = None
    assigned_to_id: Optional[int] = None


class MaintenanceUpdateIn(Schema):
    title: Optional[str] = None
    description: Optional[str] = None
    maintenance_date: Optional[str] = None


class MaintenanceAssignIn(Schema):
    assigned_to_id: int


class MaintenanceCompleteIn(Schema):
    result_notes: str
    cost: Optional[float] = None
    requires_recalibration: bool = False
    next_maintenance_date: Optional[str] = None
    performed_by: Optional[str] = ''


class MaintenanceCancelIn(Schema):
    reason: Optional[str] = ''


# --- 使用记录 ---
class UsageQueryIn(Schema):
    equipment_id: Optional[int] = None
    operator_id: Optional[int] = None
    usage_type: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    page: int = 1
    page_size: int = 20


class UsageRegisterIn(Schema):
    equipment_id: int
    usage_type: Optional[str] = 'manual'
    notes: Optional[str] = ''


# --- 授权 ---
class AuthGrantIn(Schema):
    equipment_id: int
    operator_id: int
    operator_name: Optional[str] = ''
    authorized_at: Optional[str] = None
    expires_at: Optional[str] = None
    training_record: Optional[str] = ''
    notes: Optional[str] = ''


class AuthCheckIn(Schema):
    equipment_id: int
    operator_id: int


# --- 检测方法 ---
class MethodQueryIn(Schema):
    category: Optional[str] = None
    status: Optional[str] = None
    keyword: Optional[str] = None
    page: int = 1
    page_size: int = 20


class MethodCreateIn(Schema):
    code: str
    name: str
    name_en: Optional[str] = ''
    equipment_name_classification: Optional[str] = ''
    category: str
    description: Optional[str] = ''
    qc_requirements: Optional[str] = ''
    sop_attachment_url: Optional[str] = ''
    estimated_duration_minutes: Optional[int] = 30
    preparation_time_minutes: Optional[int] = 10
    temperature_min: Optional[float] = None
    temperature_max: Optional[float] = None
    humidity_min: Optional[float] = None
    humidity_max: Optional[float] = None
    environment_notes: Optional[str] = ''
    sop_reference: Optional[str] = ''
    keywords: Optional[List[str]] = None
    status: Optional[str] = 'draft'


class MethodUpdateIn(Schema):
    name: Optional[str] = None
    name_en: Optional[str] = None
    equipment_name_classification: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    qc_requirements: Optional[str] = None
    sop_attachment_url: Optional[str] = None
    estimated_duration_minutes: Optional[int] = None
    preparation_time_minutes: Optional[int] = None
    temperature_min: Optional[float] = None
    temperature_max: Optional[float] = None
    humidity_min: Optional[float] = None
    humidity_max: Optional[float] = None
    environment_notes: Optional[str] = None
    status: Optional[str] = None
    keywords: Optional[List[str]] = None


class MethodResourceIn(Schema):
    resource_type: str
    resource_category_id: Optional[int] = None
    quantity: Optional[int] = 1
    is_mandatory: Optional[bool] = True
    recommended_models: Optional[List[str]] = None
    usage_notes: Optional[str] = ''


class MethodPersonnelIn(Schema):
    qualification_name: str
    qualification_code: Optional[str] = ''
    level: Optional[str] = 'required'
    min_experience_months: Optional[int] = 0
    notes: Optional[str] = ''


# ============================================================================
# 仪表盘
# ============================================================================
@router.get('/dashboard', summary='设备管理总览')
@require_permission('resource.equipment.read')
def dashboard(request):
    """设备管理员工作面板 — 设备总况、校准预警、维护概览、今日使用"""
    from .services.equipment_service import get_dashboard
    data = get_dashboard()
    return {'code': 0, 'msg': 'ok', 'data': data}


# ============================================================================
# 设备台账
# ============================================================================
@router.get('/ledger', summary='设备列表')
@require_permission('resource.equipment.read')
def list_ledger(request, params: Query[LedgerQueryIn]):
    """设备列表（增强筛选+统计），支持 lims_only 过滤 LIMS 导入数据"""
    from .services.equipment_service import list_equipment
    data = list_equipment(
        keyword=params.keyword, category_id=params.category_id,
        status=params.status, calibration_status=params.calibration_status,
        location=params.location, page=params.page,
        page_size=params.page_size, sort_by=params.sort_by,
        lims_only=params.lims_only,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/index', summary='设备列表（别名，同 GET /ledger）')
@require_permission('resource.equipment.read')
def list_equipment_index_alias(request, params: Query[LedgerQueryIn]):
    """与 list_ledger 一致，兼容请求 /equipment/index 的客户端"""
    from .services.equipment_service import list_equipment
    data = list_equipment(
        keyword=params.keyword, category_id=params.category_id,
        status=params.status, calibration_status=params.calibration_status,
        location=params.location, page=params.page,
        page_size=params.page_size, sort_by=params.sort_by,
        lims_only=params.lims_only,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/ledger-categories', summary='设备类别台账')
@require_permission('resource.equipment.read')
def list_ledger_categories(request, params: Query[LedgerSummaryQueryIn]):
    """按设备类别聚合设备数量，支持模糊检索类别名/编码。"""
    from .services.equipment_service import list_equipment_category_ledger
    data = list_equipment_category_ledger(
        keyword=params.keyword,
        page=params.page,
        page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/ledger-name-classifications', summary='设备细分类别台账')
@require_permission('resource.equipment.read')
def list_ledger_name_classifications(request, params: Query[LedgerSummaryQueryIn]):
    """按名称分类 + 设备类别聚合设备数量，支持模糊检索。"""
    from .services.equipment_service import list_equipment_name_classification_ledger
    data = list_equipment_name_classification_ledger(
        keyword=params.keyword,
        page=params.page,
        page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/ledger/{equipment_id}', summary='设备详情')
@require_permission('resource.equipment.read')
def get_ledger_detail(request, equipment_id: int):
    """设备完整档案（含校准/维护/使用/授权）"""
    from .services.equipment_service import get_equipment_detail
    data = get_equipment_detail(equipment_id)
    if not data:
        return {'code': 404, 'msg': '设备不存在', 'data': None}
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/ledger/import', summary='批量导入设备')
@require_permission('resource.equipment.write')
def import_ledger(request, file: File[UploadedFile] = File(...)):
    """从 Excel 批量导入设备台账，支持 FS-RF-036 模板格式"""
    from .services.equipment_import_service import parse_and_import_equipment
    name_lower = (file.name or '').lower()
    if not any(name_lower.endswith(ext) for ext in ('.xlsx', '.xls')):
        return {'code': 400, 'msg': '仅支持 .xlsx、.xls 文件', 'data': None}
    account = _get_account_from_request(request)
    try:
        from io import BytesIO
        content = file.read()
        result = parse_and_import_equipment(BytesIO(content), created_by_id=account.id if account else None)
        return {'code': 0, 'msg': '导入完成', 'data': result}
    except Exception as e:
        return {'code': 500, 'msg': f'导入失败: {str(e)}', 'data': None}


@router.post('/ledger/create', summary='新增设备')
@require_permission('resource.equipment.write')
def create_ledger(request, payload: EquipmentCreateIn):
    """录入新设备"""
    from .services.equipment_service import create_equipment
    from datetime import date as dt_date
    kwargs = payload.dict(exclude_unset=True)
    kwargs.pop('name', None)
    kwargs.pop('code', None)
    kwargs.pop('category_id', None)
    name_classification = (kwargs.pop('name_classification', '') or '').strip()
    for field_name in (
        'purchase_date', 'warranty_expiry',
        'next_calibration_date', 'next_verification_date', 'next_maintenance_date',
    ):
        if field_name in kwargs and kwargs[field_name]:
            kwargs[field_name] = dt_date.fromisoformat(kwargs[field_name])
    if name_classification:
        attrs = dict(kwargs.get('attributes') or {})
        attrs['name_classification'] = name_classification
        kwargs['attributes'] = attrs
    item = create_equipment(
        name=payload.name, code=payload.code,
        category_id=payload.category_id, **kwargs,
    )
    return {'code': 0, 'msg': '设备创建成功', 'data': {'id': item.id, 'code': item.code}}


@router.put('/ledger/{equipment_id}', summary='更新设备')
@require_permission('resource.equipment.write')
def update_ledger(request, equipment_id: int, payload: EquipmentUpdateIn):
    """更新设备信息"""
    from .services.equipment_service import update_equipment
    from datetime import date as dt_date
    kwargs = payload.dict(exclude_unset=True)
    name_classification = kwargs.pop('name_classification', None)
    for field_name in (
        'purchase_date', 'warranty_expiry',
        'next_calibration_date', 'next_verification_date', 'next_maintenance_date',
    ):
        if field_name in kwargs and kwargs[field_name]:
            kwargs[field_name] = dt_date.fromisoformat(kwargs[field_name])
    if name_classification is not None:
        item = update_equipment(equipment_id)
        if not item:
            return {'code': 404, 'msg': '设备不存在', 'data': None}
        attrs = dict(item.attributes or {})
        attrs['name_classification'] = (name_classification or '').strip()
        kwargs['attributes'] = attrs
    item = update_equipment(equipment_id, **kwargs)
    if not item:
        return {'code': 404, 'msg': '设备不存在', 'data': None}
    return {'code': 0, 'msg': '更新成功', 'data': {'id': item.id}}


@router.post('/ledger/{equipment_id}/retire', summary='设备报废')
@require_permission('resource.equipment.write')
def retire_ledger(request, equipment_id: int):
    """设备报废"""
    from .services.equipment_service import retire_equipment
    result = retire_equipment(equipment_id)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '设备已报废', 'data': result}


@router.post('/ledger/{equipment_id}/change-status', summary='变更设备状态')
@require_permission('resource.equipment.write')
def change_status(request, equipment_id: int, payload: StatusChangeIn):
    """变更设备状态"""
    from .services.equipment_service import change_equipment_status
    result = change_equipment_status(equipment_id, payload.status, payload.reason or '')
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '状态已变更', 'data': result}


# ============================================================================
# 校准管理
# ============================================================================
@router.get('/calibrations/plan', summary='校准计划视图')
@require_permission('resource.calibration.read')
def calibration_plan(request):
    """校准计划 — 逾期/7日内/本月到期 + 待办校准工单"""
    from .services.equipment_service import get_calibration_plan
    return {'code': 0, 'msg': 'ok', 'data': get_calibration_plan()}


@router.get('/calibrations/plan/template', summary='下载校准计划导入模板')
@require_permission('resource.calibration.read')
def download_calibration_plan_template(request):
    """返回校准计划 Excel 导入模板（含表头）"""
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '校准计划'
    headers = [
        '设备编号', '设备名称', '设备状态', '设备规格/型号', '出厂编号',
        '溯源方式', '校准方式', '校准机构', '校准方法', '校准周期(天)',
        '上次校准时间', '下次校准时间', '校准提前提醒(天)', '校准提醒人员', '量值溯源参数',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    # 示例行
    ws.cell(row=2, column=1, value='FSD0405093')
    ws.cell(row=2, column=2, value='D01高温高湿间')
    ws.cell(row=2, column=3, value='启用')
    ws.cell(row=2, column=6, value='校准')
    ws.cell(row=2, column=7, value='上门检测')
    ws.cell(row=2, column=10, value=365)
    ws.cell(row=2, column=11, value='2025-06-20')
    ws.cell(row=2, column=12, value='2026-06-19')
    ws.cell(row=2, column=13, value=30)
    ws.cell(row=2, column=14, value='张造')
    ws.cell(row=2, column=15, value='明细')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="calibration_plan_template.xlsx"'
    return resp


@router.post('/calibrations/plan/import', summary='批量导入校准计划')
@require_permission('resource.calibration.write')
def import_calibration_plan(request, file: File[UploadedFile] = File(...)):
    """从 Excel 批量导入校准计划（设备编号、校准日期、下次到期日）"""
    from .services.calibration_plan_import_service import parse_and_import_calibration_plan
    name_lower = (file.name or '').lower()
    if not any(name_lower.endswith(ext) for ext in ('.xlsx', '.xls')):
        return {'code': 400, 'msg': '仅支持 .xlsx、.xls 文件', 'data': None}
    account = _get_account_from_request(request)
    try:
        from io import BytesIO
        content = file.read()
        result = parse_and_import_calibration_plan(BytesIO(content), created_by_id=account.id if account else None)
        return {'code': 0, 'msg': '导入完成', 'data': result}
    except Exception as e:
        return {'code': 500, 'msg': f'导入失败: {str(e)}', 'data': None}


@router.get('/calibrations/plan/list', summary='校准计划列表')
@require_permission('resource.calibration.read')
def list_calibration_plans_api(request, params: Query[CalibrationPlanListQueryIn]):
    """校准计划列表 — 已导入的计划（含设备编号、溯源方式、校准机构、周期、提醒等）"""
    from .services.equipment_service import list_calibration_plans
    data = list_calibration_plans(
        keyword=params.keyword,
        page=params.page,
        page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/calibrations/pending-work-orders', summary='待发起校准工单列表')
@require_permission('resource.calibration.read')
def pending_calibration_work_orders(request):
    """获取待发起校准工单的设备（下次到期日<=下月末，且尚未发起工单）"""
    from .services.equipment_service import get_pending_calibration_work_orders
    items = get_pending_calibration_work_orders()
    return {'code': 0, 'msg': 'ok', 'data': {'items': items, 'count': len(items)}}


@router.post('/calibrations/create-work-orders', summary='批量发起校准工单')
@require_permission('resource.calibration.write')
def create_calibration_work_orders_api(request, payload: CalibrationWorkOrdersIn):
    """为指定设备批量发起校准工单"""
    from .services.equipment_service import create_calibration_work_orders
    ids = payload.equipment_ids or []
    if not ids:
        return {'code': 400, 'msg': '请提供 equipment_ids', 'data': None}
    result = create_calibration_work_orders(ids)
    return {'code': 0, 'msg': f'已发起 {len(result["created"])} 个校准工单', 'data': result}


@router.get('/calibrations/list', summary='校准记录列表')
@require_permission('resource.calibration.read')
def list_calibrations_api(request, params: Query[CalibrationQueryIn]):
    """校准记录列表（分页+筛选）"""
    from .services.equipment_service import list_calibrations
    data = list_calibrations(
        equipment_id=params.equipment_id, result=params.result,
        date_from=params.date_from, date_to=params.date_to,
        page=params.page, page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/calibrations/create', summary='新增校准记录')
@require_permission('resource.calibration.write')
def create_calibration_api(request, payload: CalibrationCreateIn):
    """新增校准记录（自动更新设备状态）"""
    from .services.equipment_service import create_calibration
    result = create_calibration(
        equipment_id=payload.equipment_id,
        calibration_date=payload.calibration_date,
        next_due_date=payload.next_due_date,
        calibration_type=payload.calibration_type or 'internal',
        calibrator=payload.calibrator or '',
        certificate_no=payload.certificate_no or '',
        certificate_file_url=payload.certificate_file_url or '',
        result=payload.result or 'pass',
        notes=payload.notes or '',
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '校准记录已添加', 'data': result}


@router.get('/calibrations/{calibration_id}', summary='校准详情')
@require_permission('resource.calibration.read')
def get_calibration_api(request, calibration_id: int):
    """校准记录详情"""
    from .services.equipment_service import get_calibration_detail
    data = get_calibration_detail(calibration_id)
    if not data:
        return {'code': 404, 'msg': '校准记录不存在', 'data': None}
    return {'code': 0, 'msg': 'ok', 'data': data}


# ============================================================================
# 核查计划
# ============================================================================
@router.get('/verifications/plan', summary='核查计划视图')
@require_permission('resource.verification.read')
def verification_plan(request):
    """核查计划 — 逾期/7日内/本月到期 + 待办核查工单"""
    from .services.equipment_service import get_verification_plan
    return {'code': 0, 'msg': 'ok', 'data': get_verification_plan()}


@router.get('/verifications/plan/template', summary='下载核查计划导入模板')
@require_permission('resource.verification.read')
def download_verification_plan_template(request):
    """返回核查计划 Excel 导入模板"""
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '核查计划'
    headers = ['设备编号', '设备名称', '设备规格/型号', '设备状态', '核查周期(天)', '上次核查时间', '下次核查时间', '核查提前提醒(天)', '核查提醒人员', '核查方法']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value='FSD0405004-02')
    ws.cell(row=2, column=2, value='VISIA-CR 校准色卡')
    ws.cell(row=2, column=4, value='启用')
    ws.cell(row=2, column=5, value=180)
    ws.cell(row=2, column=6, value='2025-10-13')
    ws.cell(row=2, column=7, value='2026-04-11')
    ws.cell(row=2, column=8, value=30)
    ws.cell(row=2, column=9, value='谷勤秀')
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="verification_plan_template.xlsx"'
    return resp


@router.post('/verifications/plan/import', summary='批量导入核查计划')
@require_permission('resource.verification.write')
def import_verification_plan(request, file: File[UploadedFile] = File(...)):
    """从 Excel 批量导入核查计划"""
    from .services.verification_plan_import_service import parse_and_import_verification_plan
    name_lower = (file.name or '').lower()
    if not any(name_lower.endswith(ext) for ext in ('.xlsx', '.xls')):
        return {'code': 400, 'msg': '仅支持 .xlsx、.xls 文件', 'data': None}
    account = _get_account_from_request(request)
    try:
        from io import BytesIO
        result = parse_and_import_verification_plan(BytesIO(file.read()), created_by_id=account.id if account else None)
        return {'code': 0, 'msg': '导入完成', 'data': result}
    except Exception as e:
        return {'code': 500, 'msg': f'导入失败: {str(e)}', 'data': None}


@router.get('/verifications/plan/list', summary='核查计划列表')
@require_permission('resource.verification.read')
def list_verification_plans_api(request, params: Query[VerificationPlanListQueryIn]):
    """核查计划列表"""
    from .services.equipment_service import list_verification_plans
    data = list_verification_plans(keyword=params.keyword, page=params.page, page_size=params.page_size)
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/verifications/pending-work-orders', summary='待发起核查工单列表')
@require_permission('resource.verification.read')
def pending_verification_work_orders(request):
    """获取待发起核查工单的设备"""
    from .services.equipment_service import get_pending_verification_work_orders
    items = get_pending_verification_work_orders()
    return {'code': 0, 'msg': 'ok', 'data': {'items': items, 'count': len(items)}}


@router.post('/verifications/create-work-orders', summary='批量发起核查工单')
@require_permission('resource.verification.write')
def create_verification_work_orders_api(request, payload: VerificationWorkOrdersIn):
    """为指定设备批量发起核查工单"""
    from .services.equipment_service import create_verification_work_orders
    ids = payload.equipment_ids or []
    if not ids:
        return {'code': 400, 'msg': '请提供 equipment_ids', 'data': None}
    result = create_verification_work_orders(ids)
    return {'code': 0, 'msg': f'已发起 {len(result["created"])} 个核查工单', 'data': result}


@router.get('/verifications/list', summary='核查记录列表')
@require_permission('resource.verification.read')
def list_verifications_api(request, params: Query[VerificationQueryIn]):
    """核查记录列表"""
    from .services.equipment_service import list_verifications
    data = list_verifications(
        equipment_id=params.equipment_id, result=params.result,
        date_from=params.date_from, date_to=params.date_to,
        page=params.page, page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/verifications/create', summary='新增核查记录')
@require_permission('resource.verification.write')
def create_verification_api(request, payload: VerificationCreateIn):
    """新增核查记录"""
    from .services.equipment_service import create_verification
    result = create_verification(
        equipment_id=payload.equipment_id,
        verification_date=payload.verification_date,
        next_due_date=payload.next_due_date,
        verifier=payload.verifier or '',
        result=payload.result or 'pass',
        method_notes=payload.method_notes or '',
        notes=payload.notes or '',
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '核查记录已添加', 'data': result}


@router.get('/verifications/{verification_id}', summary='核查详情')
@require_permission('resource.verification.read')
def get_verification_api(request, verification_id: int):
    """核查记录详情"""
    from .services.equipment_service import get_verification_detail
    data = get_verification_detail(verification_id)
    if not data:
        return {'code': 404, 'msg': '核查记录不存在', 'data': None}
    return {'code': 0, 'msg': 'ok', 'data': data}


# ============================================================================
# 维护计划
# ============================================================================
@router.get('/maintenance-plans/plan', summary='维护计划视图')
@require_permission('resource.maintenance.read')
def maintenance_plan(request):
    """维护计划 — 逾期/7日内/本月到期 + 待办维护工单"""
    from .services.equipment_service import get_maintenance_plan
    return {'code': 0, 'msg': 'ok', 'data': get_maintenance_plan()}


@router.get('/maintenance-plans/plan/template', summary='下载维护计划导入模板')
@require_permission('resource.maintenance.read')
def download_maintenance_plan_template(request):
    """返回维护计划 Excel 导入模板"""
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '维护计划'
    headers = ['设备编号', '设备名称', '设备规格/型号', '设备状态', '维护周期(天)', '上次维护时间', '下次维护时间', '维护提前提醒(天)', '维护提醒人员', '维护方法']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value='FS-MC-E036')
    ws.cell(row=2, column=2, value='多参数监护仪')
    ws.cell(row=2, column=3, value='KB12A')
    ws.cell(row=2, column=4, value='启用')
    ws.cell(row=2, column=5, value=180)
    ws.cell(row=2, column=6, value='2026-01-26')
    ws.cell(row=2, column=7, value='2026-07-25')
    ws.cell(row=2, column=8, value=30)
    ws.cell(row=2, column=9, value='刘利伟')
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="maintenance_plan_template.xlsx"'
    return resp


@router.post('/maintenance-plans/plan/import', summary='批量导入维护计划')
@require_permission('resource.maintenance.write')
def import_maintenance_plan(request, file: File[UploadedFile] = File(...)):
    """从 Excel 批量导入维护计划"""
    from .services.maintenance_plan_import_service import parse_and_import_maintenance_plan
    name_lower = (file.name or '').lower()
    if not any(name_lower.endswith(ext) for ext in ('.xlsx', '.xls')):
        return {'code': 400, 'msg': '仅支持 .xlsx、.xls 文件', 'data': None}
    account = _get_account_from_request(request)
    try:
        from io import BytesIO
        result = parse_and_import_maintenance_plan(BytesIO(file.read()), created_by_id=account.id if account else None)
        return {'code': 0, 'msg': '导入完成', 'data': result}
    except Exception as e:
        return {'code': 500, 'msg': f'导入失败: {str(e)}', 'data': None}


@router.get('/maintenance-plans/plan/list', summary='维护计划列表')
@require_permission('resource.maintenance.read')
def list_maintenance_plans_api(request, params: Query[MaintenancePlanListQueryIn]):
    """维护计划列表"""
    from .services.equipment_service import list_maintenance_plans
    data = list_maintenance_plans(keyword=params.keyword, page=params.page, page_size=params.page_size)
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/maintenance-plans/pending-work-orders', summary='待发起维护工单列表')
@require_permission('resource.maintenance.read')
def pending_maintenance_work_orders(request):
    """获取待发起维护工单的设备"""
    from .services.equipment_service import get_pending_maintenance_work_orders
    items = get_pending_maintenance_work_orders()
    return {'code': 0, 'msg': 'ok', 'data': {'items': items, 'count': len(items)}}


@router.post('/maintenance-plans/create-work-orders', summary='批量发起维护工单')
@require_permission('resource.maintenance.write')
def create_maintenance_work_orders_api(request, payload: MaintenanceWorkOrdersIn):
    """为指定设备批量发起维护工单"""
    from .services.equipment_service import create_maintenance_work_orders
    ids = payload.equipment_ids or []
    if not ids:
        return {'code': 400, 'msg': '请提供 equipment_ids', 'data': None}
    result = create_maintenance_work_orders(ids)
    return {'code': 0, 'msg': f'已发起 {len(result["created"])} 个维护工单', 'data': result}


# ============================================================================
# 维护工单
# ============================================================================
@router.get('/maintenance/list', summary='维护工单列表')
@require_permission('resource.maintenance.read')
def list_maintenance_api(request, params: Query[MaintenanceQueryIn]):
    """维护工单列表（分页+筛选）"""
    from .services.equipment_service import list_maintenance
    data = list_maintenance(
        equipment_id=params.equipment_id, status=params.status,
        maintenance_type=params.maintenance_type,
        date_from=params.date_from, date_to=params.date_to,
        assigned_to_id=params.assigned_to_id,
        page=params.page, page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/maintenance/stats', summary='维护统计')
@require_permission('resource.maintenance.read')
def maintenance_stats_api(request):
    """维护工单统计"""
    from .services.equipment_service import get_maintenance_stats
    return {'code': 0, 'msg': 'ok', 'data': get_maintenance_stats()}


@router.post('/maintenance/create', summary='创建维护工单')
@require_permission('resource.maintenance.write')
def create_maintenance_api(request, payload: MaintenanceCreateIn):
    """创建维护工单"""
    account = _get_account_from_request(request)
    from .services.equipment_service import create_maintenance
    result = create_maintenance(
        equipment_id=payload.equipment_id,
        maintenance_type=payload.maintenance_type,
        title=payload.title,
        description=payload.description,
        maintenance_date=payload.maintenance_date,
        assigned_to_id=payload.assigned_to_id,
        reported_by_id=account.id if account else None,
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '维护工单已创建', 'data': result}


@router.get('/maintenance/{maintenance_id}', summary='维护工单详情')
@require_permission('resource.maintenance.read')
def get_maintenance_api(request, maintenance_id: int):
    """维护工单详情"""
    from .services.equipment_service import get_maintenance_detail
    data = get_maintenance_detail(maintenance_id)
    if not data:
        return {'code': 404, 'msg': '维护工单不存在', 'data': None}
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.put('/maintenance/{maintenance_id}', summary='更新维护工单')
@require_permission('resource.maintenance.write')
def update_maintenance_api(request, maintenance_id: int, payload: MaintenanceUpdateIn):
    """更新维护工单"""
    from .services.equipment_service import update_maintenance
    from datetime import date as dt_date
    kwargs = payload.dict(exclude_unset=True)
    if 'maintenance_date' in kwargs and kwargs['maintenance_date']:
        kwargs['maintenance_date'] = dt_date.fromisoformat(kwargs['maintenance_date'])
    result = update_maintenance(maintenance_id, **kwargs)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '更新成功', 'data': result}


@router.post('/maintenance/{maintenance_id}/assign', summary='分配维护')
@require_permission('resource.maintenance.write')
def assign_maintenance_api(request, maintenance_id: int, payload: MaintenanceAssignIn):
    """分配维护任务"""
    from .services.equipment_service import assign_maintenance
    result = assign_maintenance(maintenance_id, payload.assigned_to_id)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '已分配', 'data': result}


@router.post('/maintenance/{maintenance_id}/start', summary='开始维护')
@require_permission('resource.maintenance.write')
def start_maintenance_api(request, maintenance_id: int):
    """开始维护"""
    from .services.equipment_service import start_maintenance
    result = start_maintenance(maintenance_id)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '维护已开始', 'data': result}


@router.post('/maintenance/{maintenance_id}/complete', summary='完成维护')
@require_permission('resource.maintenance.write')
def complete_maintenance_api(request, maintenance_id: int, payload: MaintenanceCompleteIn):
    """完成维护"""
    from .services.equipment_service import complete_maintenance
    result = complete_maintenance(
        maintenance_id,
        result_notes=payload.result_notes,
        cost=payload.cost,
        requires_recalibration=payload.requires_recalibration,
        next_maintenance_date=payload.next_maintenance_date,
        performed_by=payload.performed_by or '',
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '维护已完成', 'data': result}


@router.post('/maintenance/{maintenance_id}/cancel', summary='取消维护')
@require_permission('resource.maintenance.write')
def cancel_maintenance_api(request, maintenance_id: int, payload: MaintenanceCancelIn):
    """取消维护工单"""
    from .services.equipment_service import cancel_maintenance
    result = cancel_maintenance(maintenance_id, payload.reason or '')
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '维护已取消', 'data': result}


# ============================================================================
# 使用记录
# ============================================================================
@router.get('/usage/list', summary='使用记录列表')
@require_any_permission(['resource.equipment.read', 'resource.usage.read'])
def list_usage_api(request, params: Query[UsageQueryIn]):
    """使用记录列表（分页+筛选）"""
    from .services.equipment_service import list_usage
    data = list_usage(
        equipment_id=params.equipment_id, operator_id=params.operator_id,
        usage_type=params.usage_type,
        date_from=params.date_from, date_to=params.date_to,
        page=params.page, page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.get('/usage/stats', summary='使用统计')
@require_any_permission(['resource.equipment.read', 'resource.usage.read'])
def usage_stats_api(request):
    """设备使用统计（30天）"""
    from .services.equipment_service import get_usage_stats
    return {'code': 0, 'msg': 'ok', 'data': get_usage_stats()}


@router.post('/usage/register', summary='登记使用')
@require_any_permission(['resource.equipment.write', 'resource.usage.write'])
def register_usage_api(request, payload: UsageRegisterIn):
    """手动登记设备使用（开始使用）"""
    account = _get_account_from_request(request)
    from .services.equipment_service import register_usage
    result = register_usage(
        equipment_id=payload.equipment_id,
        operator_id=account.id if account else None,
        operator_name=account.display_name if account and hasattr(account, 'display_name') else '',
        usage_type=payload.usage_type or 'manual',
        notes=payload.notes or '',
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '使用已登记', 'data': result}


@router.post('/usage/{usage_id}/end', summary='结束使用')
@require_any_permission(['resource.equipment.write', 'resource.usage.write'])
def end_usage_api(request, usage_id: int):
    """结束设备使用"""
    from .services.equipment_service import end_usage
    result = end_usage(usage_id)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '使用已结束', 'data': result}


# ============================================================================
# 操作授权
# ============================================================================
@router.get('/authorizations/list', summary='授权列表')
@require_permission('resource.authorization.read')
def list_authorizations_api(
    request,
    equipment_id: int = None,
    operator_id: int = None,
    is_active: bool = None,
):
    """设备操作授权列表"""
    from .services.equipment_service import list_authorizations
    data = list_authorizations(
        equipment_id=equipment_id, operator_id=operator_id, is_active=is_active,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/authorizations/grant', summary='授予授权')
@require_permission('resource.authorization.write')
def grant_authorization_api(request, payload: AuthGrantIn):
    """授予设备操作授权"""
    account = _get_account_from_request(request)
    from .services.equipment_service import grant_authorization
    result = grant_authorization(
        equipment_id=payload.equipment_id,
        operator_id=payload.operator_id,
        operator_name=payload.operator_name or '',
        authorized_at=payload.authorized_at,
        expires_at=payload.expires_at,
        training_record=payload.training_record or '',
        authorized_by_id=account.id if account else None,
        notes=payload.notes or '',
    )
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '授权成功', 'data': result}


@router.post('/authorizations/{authorization_id}/revoke', summary='撤销授权')
@require_permission('resource.authorization.write')
def revoke_authorization_api(request, authorization_id: int):
    """撤销操作授权"""
    from .services.equipment_service import revoke_authorization
    result = revoke_authorization(authorization_id)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '授权已撤销', 'data': result}


@router.get('/authorizations/check', summary='检查授权')
@require_permission('resource.authorization.read')
def check_authorization_api(request, params: Query[AuthCheckIn]):
    """检查操作人员对设备的授权状态"""
    from .services.equipment_service import check_authorization
    return {'code': 0, 'msg': 'ok', 'data': check_authorization(
        params.equipment_id, params.operator_id,
    )}


# ============================================================================
# 检测方法
# ============================================================================
@router.get('/detection-methods/list', summary='检测方法列表')
@require_permission('resource.method.read')
def list_methods_api(request, params: Query[MethodQueryIn]):
    """检测方法模板列表"""
    from .services.equipment_service import list_detection_methods
    data = list_detection_methods(
        category=params.category, status=params.status,
        keyword=params.keyword, page=params.page, page_size=params.page_size,
    )
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/detection-methods/sop-upload', summary='上传检测方法 SOP 附件')
@require_permission('resource.method.write')
def upload_detection_method_sop(request, file: File[UploadedFile] = File(...)):
    """保存到 MEDIA_ROOT/detection_methods/sop/，返回可访问的 url（相对站点根路径）"""
    suffix = Path(file.name or '').suffix.lower()
    allowed = ('.pdf', '.doc', '.docx', '.xlsx', '.xls', '.ppt', '.pptx', '.png', '.jpg', '.jpeg', '.zip')
    if suffix not in allowed:
        return {'code': 400, 'msg': f'不支持的文件类型，允许: {", ".join(allowed)}', 'data': None}
    rel_dir = 'detection_methods/sop'
    media_root = Path(str(settings.MEDIA_ROOT))
    dest_dir = media_root / rel_dir
    dest_dir.mkdir(parents=True, exist_ok=True)
    stored = f'{uuid.uuid4().hex}{suffix}'
    dest = dest_dir / stored
    with open(dest, 'wb') as out:
        if hasattr(file, 'chunks'):
            for chunk in file.chunks():
                out.write(chunk)
        else:
            out.write(file.read())
    media_url = (settings.MEDIA_URL or '/media/').rstrip('/') + '/'
    url = f'{media_url}{rel_dir}/{stored}'
    return {
        'code': 0,
        'msg': 'ok',
        'data': {'url': url, 'original_filename': os.path.basename(file.name or '')},
    }


@router.get('/detection-methods/{method_id}', summary='检测方法详情')
@require_permission('resource.method.read')
def get_method_api(request, method_id: int):
    """检测方法详情（含资源需求和人员要求）"""
    from .services.equipment_service import get_detection_method_detail
    data = get_detection_method_detail(method_id)
    if not data:
        return {'code': 404, 'msg': '检测方法不存在', 'data': None}
    return {'code': 0, 'msg': 'ok', 'data': data}


@router.post('/detection-methods/create', summary='创建检测方法')
@require_permission('resource.method.write')
def create_method_api(request, payload: MethodCreateIn):
    """创建检测方法模板"""
    from .services.equipment_service import create_detection_method
    kwargs = payload.dict(exclude_unset=True)
    result = create_detection_method(**kwargs)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '创建成功', 'data': result}


@router.put('/detection-methods/{method_id}', summary='更新检测方法')
@require_permission('resource.method.write')
def update_method_api(request, method_id: int, payload: MethodUpdateIn):
    """更新检测方法"""
    from .services.equipment_service import update_detection_method
    kwargs = payload.dict(exclude_unset=True)
    result = update_detection_method(method_id, **kwargs)
    if not result:
        return {'code': 404, 'msg': '检测方法不存在', 'data': None}
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '更新成功', 'data': result}


@router.post('/detection-methods/{method_id}/resources/add', summary='添加资源需求')
@require_permission('resource.method.write')
def add_resource_api(request, method_id: int, payload: MethodResourceIn):
    """添加检测方法资源需求"""
    from .services.equipment_service import add_method_resource
    kwargs = payload.dict(exclude_unset=True)
    result = add_method_resource(method_id, **kwargs)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '添加成功', 'data': result}


@router.post('/detection-methods/{method_id}/personnel/add', summary='添加人员要求')
@require_permission('resource.method.write')
def add_personnel_api(request, method_id: int, payload: MethodPersonnelIn):
    """添加检测方法人员要求"""
    from .services.equipment_service import add_method_personnel
    kwargs = payload.dict(exclude_unset=True)
    result = add_method_personnel(method_id, **kwargs)
    if 'error' in result:
        return {'code': 400, 'msg': result['error'], 'data': None}
    return {'code': 0, 'msg': '添加成功', 'data': result}


@router.delete('/detection-methods/resources/{resource_id}', summary='删除资源需求')
@require_permission('resource.method.write')
def remove_resource_api(request, resource_id: int):
    """删除检测方法资源需求"""
    from .services.equipment_service import remove_method_resource
    ok = remove_method_resource(resource_id)
    if not ok:
        return {'code': 404, 'msg': '资源需求不存在', 'data': None}
    return {'code': 0, 'msg': '删除成功', 'data': None}


@router.delete('/detection-methods/personnel/{personnel_id}', summary='删除人员要求')
@require_permission('resource.method.write')
def remove_personnel_api(request, personnel_id: int):
    """删除检测方法人员要求"""
    from .services.equipment_service import remove_method_personnel
    ok = remove_method_personnel(personnel_id)
    if not ok:
        return {'code': 404, 'msg': '人员要求不存在', 'data': None}
    return {'code': 0, 'msg': '删除成功', 'data': None}
