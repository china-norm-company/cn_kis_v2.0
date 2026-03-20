"""
受试者全生命周期知识库构建
以董彦吟+张煜佼的结算数据为锚点，逆向追溯每个受试者的完整档案：

  结算数据（谁、哪个项目、结算多少）
    ↓ 按项目编号关联
  项目信息（Protocol: 品类/客户/周期/测试方法）
    ↓ 按入组编号关联
  测试数据（Raw Data Excel: 仪器测量）
    ↓ 按受试者编号关联
  问卷数据（VAS/Likert/DLQI）
    ↓ 按IM群聊关联
  访视记录、初筛合格情况、脱落情况
    ↓ 综合
  受试者全景美丽档案（KnowledgeEntry）

Usage:
  python manage.py build_subject_knowledge --phase all
  python manage.py build_subject_knowledge --phase extract    # 从结算提取受试者
  python manage.py build_subject_knowledge --phase enrich     # 用项目数据丰富档案
  python manage.py build_subject_knowledge --phase im_screen  # 从IM提取初筛信号
  python manage.py build_subject_knowledge --phase profile    # 生成全景档案
  python manage.py build_subject_knowledge --phase vectorize  # 向量化
  python manage.py build_subject_knowledge --stats
"""
import os
import re
import logging
from collections import defaultdict
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import connection, transaction
from django.utils import timezone

logger = logging.getLogger(__name__)
MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')
PROJ_P = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)
NAME_P = re.compile(r'^[\u4e00-\u9fff·]{2,5}$')
PHONE_P = re.compile(r'1[3-9]\d{9}')
ID_P = re.compile(r'\d{17}[\dXx]')
SC_RD_P = re.compile(r'\b(SC|RD)(\d{3,6})\b', re.IGNORECASE)

INVALID_NAMES = {'合计','小计','总计','序号','姓名','备注','兼职','项目','结算',
                 '金额','实际','应发','发放','签名','联系','银行','证号','明细',
                 '汇总','负责','提交','确认','审核','当月','上月','本月','年度'}

DONG_ID = 'ou_1669ffd7dc086fc52be6cc848c7431ab'   # 董彦吟
ZHANG_ID = 'ou_a3a9c72e3a78dfb64d29d4483352acd2'  # 张煜佼


def extract_proj(text):
    m = PROJ_P.search(str(text))
    return m.group(0).upper() if m else ''


def mask_id(s):
    s = str(s).strip()
    return s[:6] + '****' + s[-4:] if len(s) == 18 else s


def safe_amount(v):
    try:
        return Decimal(re.sub(r'[^\d.]', '', str(v)) or '0')
    except (InvalidOperation, ValueError):
        return Decimal('0')


class Command(BaseCommand):
    help = '受试者全生命周期知识库构建（以结算数据为锚点）'

    def add_arguments(self, parser):
        parser.add_argument('--phase', default='all',
                            choices=['all', 'extract', 'enrich', 'im_screen', 'profile', 'vectorize'])
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--stats', action='store_true')

    def handle(self, *args, **options):
        if options['stats']:
            self._show_stats()
            return

        self.dry_run = options['dry_run']
        self.counters = defaultdict(int)

        phase = options['phase']
        phases = ['extract', 'enrich', 'im_screen', 'profile', 'vectorize'] \
            if phase == 'all' else [phase]

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}')()

        self.stdout.write(self.style.SUCCESS(
            f'\n总结: 新建{self.counters["new"]} 更新{self.counters["updated"]} '
            f'档案{self.counters["profiles"]} 向量化{self.counters["vectorized"]}'
        ))

    # =========================================================
    # Phase 1: 从结算Excel提取受试者（按项目编号分组）
    # =========================================================
    def phase_extract(self):
        import openpyxl
        try:
            import xlrd
            has_xlrd = True
        except ImportError:
            has_xlrd = False

        self.stdout.write('扫描所有结算/发放Excel...')
        files = self._find_settlement_files()
        self.stdout.write(f'找到 {len(files)} 个文件')

        # subject_map: name/phone -> {name, phone, id_card, projects: {proj_code: {role, amounts}}}
        subject_map = defaultdict(lambda: {
            'name': '', 'phone': '', 'id_card': '', 'projects': defaultdict(lambda: {'roles': set(), 'amounts': []})
        })

        for fpath, fname, file_proj in files:
            try:
                if fname.endswith('.xls') and has_xlrd:
                    sheets = self._load_xls_sheets(fpath, xlrd)
                elif fname.endswith('.xlsx'):
                    sheets = self._load_xlsx_sheets(fpath, openpyxl)
                else:
                    continue
            except Exception:
                continue

            for sheet_name, rows in sheets:
                if len(rows) < 2:
                    continue

                sheet_proj = file_proj or self._find_proj_in_rows(rows[:8]) or extract_proj(sheet_name)
                col = self._detect_columns(rows)
                if col.get('name') is None:
                    continue

                data_start = col.pop('_data_start', 1)
                for row in rows[data_start:]:
                    rec = self._extract_row(row, col)
                    if not rec:
                        continue

                    # 项目从行内找或用文件/sheet的项目
                    row_proj = (
                        extract_proj(rec.get('proj_raw', '')) or
                        extract_proj(' '.join(str(v) for v in row[:15] if v)) or
                        sheet_proj
                    )

                    key = rec['phone'] if rec['phone'] else rec['name']
                    subject_map[key]['name'] = rec['name']
                    if rec['phone']:
                        subject_map[key]['phone'] = rec['phone']
                    if rec['id_card']:
                        subject_map[key]['id_card'] = rec['id_card']

                    if row_proj:
                        subject_map[key]['projects'][row_proj]['roles'].add(rec.get('role', ''))
                        if rec.get('amount', 0) > 0:
                            subject_map[key]['projects'][row_proj]['amounts'].append(rec['amount'])

        self.stdout.write(f'提取唯一受试者: {len(subject_map):,}')

        if not self.dry_run:
            saved = self._upsert_subjects(subject_map)
            self.stdout.write(self.style.SUCCESS(f'写入: {saved} 条'))

    def _find_settlement_files(self):
        results = []
        kws = ['发放明细', '兼职劳务费', '兼职人员明细', '员工受试者结算',
               '受试者银行卡', '月度发放', '复硕正态众包', '灵工结算',
               '劳务报酬', '境内人员信息', '新增银行卡', '兼职劳务费表格2024',
               '兼职劳务费-202', 'SPF测试项目结算', '日照云库']
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not (f.endswith('.xlsx') or f.endswith('.xls')):
                    continue
                if any(kw in f for kw in kws):
                    proj = extract_proj(f) or extract_proj(root)
                    results.append((os.path.join(root, f), f, proj))
        return results

    def _load_xlsx_sheets(self, fpath, openpyxl):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 5000), values_only=True))
            sheets.append((ws.title, rows))
        wb.close()
        return sheets

    def _load_xls_sheets(self, fpath, xlrd):
        wb = xlrd.open_workbook(fpath)
        return [(ws.name, [ws.row_values(i) for i in range(min(ws.nrows, 5000))]) for ws in wb.sheets()]

    def _find_proj_in_rows(self, rows):
        for row in rows:
            row_text = ' '.join(str(v) for v in row if v)
            m = PROJ_P.search(row_text)
            if m:
                return m.group(0).upper()
        return ''

    def _detect_columns(self, rows):
        col = {}
        for hi in range(min(5, len(rows))):
            header = [str(c).strip() if c else '' for c in rows[hi][:20]]
            for i, h in enumerate(header):
                if h == '姓名' and 'name' not in col:
                    col['name'] = i
                    col['_data_start'] = hi + 1
                elif '手机号' in h and 'phone' not in col:
                    col['phone'] = i
                elif '身份证' in h and 'id_card' not in col:
                    col['id_card'] = i
                elif ('岗位' in h or '工作岗位' in h) and 'role' not in col:
                    col['role'] = i
                elif ('应发' in h or '实际发放' in h or '结算金额' in h) and 'amount' not in col:
                    col['amount'] = i
                elif '项目编号' in h and 'proj_raw' not in col:
                    col['proj_raw'] = i
            if 'name' in col:
                break
        return col

    def _extract_row(self, row, col):
        def get(field):
            ci = col.get(field)
            return str(row[ci]).strip() if ci is not None and ci < len(row) and row[ci] is not None else ''

        name = get('name')
        if not NAME_P.match(name) or name in INVALID_NAMES:
            return None

        phone = ''
        p = re.sub(r'[^\d]', '', get('phone'))
        if len(p) == 11 and p.startswith('1'):
            phone = p

        id_card = ''
        id_raw = re.sub(r'[\s\-]', '', get('id_card'))
        if re.match(r'^\d{17}[\dXx]$', id_raw):
            id_card = mask_id(id_raw)

        amount = safe_amount(get('amount'))

        return {
            'name': name,
            'phone': phone,
            'id_card': id_card,
            'role': get('role')[:20],
            'amount': float(amount),
            'proj_raw': get('proj_raw'),
        }

    def _upsert_subjects(self, subject_map):
        from apps.subject.models import Subject, Enrollment, EnrollmentStatus
        from apps.protocol.models import Protocol
        from apps.subject.models_execution import SubjectPayment, PaymentType

        proto_cache = {}
        saved = 0

        for key, info in subject_map.items():
            name = info['name']
            phone = info['phone']
            id_card = info['id_card']
            projects = info['projects']

            if not name:
                continue

            try:
                subject = None
                if phone:
                    subject = Subject.objects.filter(phone=phone[:20]).first()
                if subject is None:
                    subject = Subject.objects.filter(name=name).first()

                if subject is None:
                    subject = Subject(
                        name=name[:100], phone=phone[:20] if phone else '',
                        id_card_encrypted=id_card[:50] if id_card else '',
                        source_channel='database', status='completed',
                    )
                    subject.save()
                    self.counters['new'] += 1
                else:
                    changed = False
                    if not subject.phone and phone:
                        subject.phone = phone[:20]
                        changed = True
                    if not subject.id_card_encrypted and id_card:
                        subject.id_card_encrypted = id_card[:50]
                        changed = True
                    if changed:
                        subject.save(update_fields=['phone', 'id_card_encrypted', 'update_time'])
                    self.counters['updated'] += 1

                saved += 1

                # 入组关联 + 支付记录
                for proj_code, proj_info in projects.items():
                    if not proj_code or proj_code == '未知':
                        continue
                    if proj_code not in proto_cache:
                        proto_cache[proj_code] = Protocol.objects.filter(
                            code__iexact=proj_code
                        ).first()
                    proto = proto_cache.get(proj_code)
                    if proto and subject.pk:
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=proto,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )
                    # 支付记录
                    amounts = proj_info.get('amounts', [])
                    if amounts and subject.pk:
                        total = sum(amounts)
                        pay_no = f'settle_{subject.id}_{proj_code}'[:50]
                        SubjectPayment.objects.get_or_create(
                            payment_no=pay_no,
                            defaults={
                                'subject': subject,
                                'payment_type': PaymentType.VISIT_COMPENSATION,
                                'amount': Decimal(str(round(total, 2))),
                                'status': 'paid',
                                'notes': f'项目:{proj_code} 来源:结算台账',
                            }
                        )

            except Exception as e:
                logger.debug('upsert 跳过 %s: %s', name, e)

        return saved

    # =========================================================
    # Phase 2: 用项目数据丰富受试者档案
    # =========================================================
    def phase_enrich(self):
        """按项目编号关联测试数据（Raw Data / EDC / 问卷）"""
        from apps.subject.models import Subject, Enrollment
        from apps.protocol.models import Protocol
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire

        self.stdout.write('关联项目测试数据...')

        # 已有入组记录的受试者
        enrollments = Enrollment.objects.select_related('subject', 'protocol').filter(
            status='completed'
        )
        total = enrollments.count()
        self.stdout.write(f'有入组记录的受试者: {total:,}')

        # 按项目汇总
        proj_subjects = defaultdict(list)
        for enr in enrollments:
            if enr.protocol:
                proj_subjects[enr.protocol.code].append(enr.subject)

        self.stdout.write(f'涉及项目: {len(proj_subjects)}')

        enriched = 0
        for proj_code, subjects in proj_subjects.items():
            m_count = SkinMeasurementRecord.objects.filter(
                enrollment__protocol__code=proj_code
            ).values('subject_id').distinct().count()

            q_count = SubjectQuestionnaire.objects.filter(
                enrollment__protocol__code=proj_code
            ).values('subject_id').distinct().count()

            if m_count > 0 or q_count > 0:
                enriched += len(subjects)
                logger.debug('项目 %s: %d名受试者有测量数据', proj_code, m_count)

        self.stdout.write(f'有测试数据的受试者: {enriched}')

    # =========================================================
    # Phase 3: 从IM提取初筛/合格/脱落信号
    # =========================================================
    def phase_im_screen(self):
        """从招募IM群聊提取每日初筛/合格/入组/脱落数字和受试者编号"""
        from apps.knowledge.models import KnowledgeEntry

        RECRUIT_CHATS = [
            '特化招募', '组3招募', '组2招募群', '组1 招募沟通', '组7&组9 招募',
            '华山植发-招募重点跟踪群组', '组4-招募沟通群', 'C06招募沟通群',
            '忍者-组3', '头发测试研究团队', '组3运营', '招募天团🤙🏼🤙🏼🤙🏼',
        ]

        c = connection.cursor()
        chat_placeholders = ','.join(['%s'] * len(RECRUIT_CHATS))
        c.execute(f"""
        SELECT raw_content, metadata->>'chat_name', created_at::date
        FROM t_personal_context
        WHERE source_type='im'
        AND metadata->>'chat_name' IN ({chat_placeholders})
        AND (raw_content LIKE '%%到访%%' OR raw_content LIKE '%%入组%%'
             OR raw_content LIKE '%%合格%%' OR raw_content LIKE '%%初筛%%'
             OR raw_content LIKE '%%不合格%%' OR raw_content LIKE '%%脱落%%'
             OR raw_content LIKE '%%SC0%%' OR raw_content LIKE '%%RD0%%')
        ORDER BY created_at DESC
        LIMIT 5000
        """, RECRUIT_CHATS)
        msgs = c.fetchall()
        self.stdout.write(f'招募IM消息: {len(msgs):,}')

        # 解析规则
        patterns = {
            'visited': re.compile(r'(?:到访|来访|约访)\D{0,5}(\d+)'),
            'screened': re.compile(r'(?:初筛|筛选了?)\D{0,5}(\d+)'),
            'qualified': re.compile(r'(?:初筛合格|合格受试者?)\D{0,5}(\d+)'),
            'enrolled': re.compile(r'(?:入组了?|已入组)\D{0,5}(\d+)'),
            'dropout': re.compile(r'(?:脱落了?|不合格)\D{0,5}(\d+)'),
            'unqualified': re.compile(r'(?:不合格|未通过|筛选失败)\D{0,5}(\d+)'),
        }

        proj_daily = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        sc_rd_signals = defaultdict(lambda: {'projects': set(), 'dates': [], 'events': []})

        for raw, chat, dt in msgs:
            content = str(raw or '')
            if content.startswith('{') or content.startswith('['):
                continue

            projs = [p.upper() for p in PROJ_P.findall(content)]
            date_str = str(dt)

            # 提取招募数字
            for field, pat in patterns.items():
                m = pat.search(content)
                if m:
                    try:
                        val = int(m.group(1))
                        if 1 <= val <= 500:
                            for proj in (projs or ['未知']):
                                proj_daily[proj][date_str][field] += val
                    except Exception:
                        pass

            # 提取受试者编号
            for m in SC_RD_P.finditer(content):
                subj_id = f'{m.group(1).upper()}{m.group(2)}'
                sc_rd_signals[subj_id]['projects'].update(projs)
                sc_rd_signals[subj_id]['dates'].append(date_str)

        # 汇总统计
        total_visited = sum(
            sum(s.get('visited', 0) for s in dates.values())
            for dates in proj_daily.values()
        )
        total_enrolled = sum(
            sum(s.get('enrolled', 0) for s in dates.values())
            for dates in proj_daily.values()
        )
        total_qualified = sum(
            sum(s.get('qualified', 0) for s in dates.values())
            for dates in proj_daily.values()
        )
        total_unqualified = sum(
            sum(s.get('unqualified', 0) for s in dates.values())
            for dates in proj_daily.values()
        )
        total_dropout = sum(
            sum(s.get('dropout', 0) for s in dates.values())
            for dates in proj_daily.values()
        )

        self.stdout.write(f'\n  累计到访: {total_visited:,}')
        self.stdout.write(f'  累计合格: {total_qualified:,}')
        self.stdout.write(f'  累计不合格: {total_unqualified:,}')
        self.stdout.write(f'  累计入组: {total_enrolled:,}')
        self.stdout.write(f'  累计脱落: {total_dropout:,}')
        self.stdout.write(f'  SC/RD信号: {len(sc_rd_signals)} 个')

        # 按项目汇总写入知识库
        if not self.dry_run:
            proj_lines = []
            for proj, dates in sorted(proj_daily.items()):
                if proj == '未知':
                    continue
                totals = defaultdict(int)
                for date_stats in dates.values():
                    for k, v in date_stats.items():
                        totals[k] += v
                if sum(totals.values()) > 5:
                    proj_lines.append(
                        f"{proj}: 到访{totals['visited']} 初筛{totals['screened']} "
                        f"合格{totals['qualified']} 不合格{totals['unqualified']} "
                        f"入组{totals['enrolled']} 脱落{totals['dropout']}"
                    )

            content = (
                f"招募IM群聊 — 受试者全生命周期信号汇总\n"
                f"分析消息: {len(msgs):,}\n"
                f"累计到访: {total_visited:,}\n"
                f"累计合格: {total_qualified:,}\n"
                f"累计不合格: {total_unqualified:,}\n"
                f"累计入组: {total_enrolled:,}\n"
                f"累计脱落: {total_dropout:,}\n"
                f"识别受试者编号: {len(sc_rd_signals)}\n\n"
                f"=== 按项目统计 ===\n" + '\n'.join(proj_lines[:100])
            )

            KnowledgeEntry.objects.update_or_create(
                source_type='recruit_lifecycle_stats',
                source_id=0,
                source_key='recruit_lifecycle_dong_zhang',
                defaults={
                    'entry_type': 'lesson_learned',
                    'title': '招募团队 IM 受试者全生命周期统计',
                    'content': content,
                    'status': 'published', 'is_published': True,
                }
            )
            self.counters['profiles'] += 1
            self.stdout.write(self.style.SUCCESS('IM招募统计已写入知识库'))

    # =========================================================
    # Phase 4: 生成受试者全景知识档案
    # =========================================================
    def phase_profile(self):
        """为每位受试者生成融合所有来源的全景档案"""
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord, SubjectPayment
        from apps.subject.models_domain import SkinProfile
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'生成全景档案: {total:,} 名受试者')

        created = updated = 0
        for i in range(0, total, 500):
            batch = subjects[i:i + 500]
            for subject in batch:
                text = self._build_profile(subject)
                if not text:
                    continue
                if self.dry_run:
                    created += 1
                    continue
                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_full_lifecycle',
                    source_id=subject.id,
                    source_key=f'lifecycle_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 受试者全档案',
                        'content': text,
                        'status': 'published', 'is_published': True,
                        'index_status': 'pending',
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1

            if (i + 500) % 5000 == 0:
                self.stdout.write(f'  进度: {min(i+500, total)}/{total}')

        self.counters['profiles'] = created + updated
        self.stdout.write(self.style.SUCCESS(f'全景档案: 新建{created} 更新{updated}'))

    def _build_profile(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord, SubjectPayment
        from apps.subject.models_domain import SkinProfile

        lines = [f'受试者: {subject.name}']

        if subject.subject_no:
            lines.append(f'系统编号: {subject.subject_no}')

        demo = []
        if subject.gender:
            demo.append(subject.get_gender_display())
        if subject.age:
            demo.append(f'{subject.age}岁')
        if demo:
            lines.append(f'基本信息: {", ".join(demo)}')

        if subject.phone:
            lines.append(f'手机尾号: {subject.phone[-4:] if len(subject.phone) >= 4 else ""}')

        # 皮肤档案
        skin = SkinProfile.objects.filter(subject=subject).first()
        if skin:
            parts = []
            if skin.fitzpatrick_type:
                parts.append(f'Fitzpatrick {skin.fitzpatrick_type}型')
            if skin.skin_type_u_zone:
                parts.append(f'{skin.skin_type_u_zone}肤质')
            if skin.moisture_baseline:
                parts.append(f'水分基线{skin.moisture_baseline}')
            if parts:
                lines.append(f'皮肤特征: {", ".join(parts)}')

        # 项目参与（入组记录）
        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        n_enr = enrollments.count()
        if n_enr:
            projs = [e.protocol.code for e in enrollments[:8] if e.protocol]
            lines.append(f'参与项目({n_enr}个): {", ".join(projs)}')

        # 仪器测量
        m_count = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m_count:
            latest = SkinMeasurementRecord.objects.filter(subject=subject).order_by('-measured_at').first()
            m_parts = []
            for field, label in [('moisture','水分'), ('tewl','TEWL'), ('sebum','皮脂'),
                                  ('elasticity','弹性'), ('melanin','黑色素'), ('erythema','红斑')]:
                val = getattr(latest, field, None)
                if val:
                    m_parts.append(f'{label}{val}')
            if m_parts:
                lines.append(f'测量数据({m_count}次): {", ".join(m_parts)}')

        # 问卷
        q_count = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q_count:
            # 感官评价摘要
            sensory = SubjectQuestionnaire.objects.filter(
                subject=subject, questionnaire_type='sensory_questionnaire'
            ).first()
            if sensory and sensory.answers:
                s_parts = []
                for key, label in [('itch_nrs','瘙痒'), ('dry_skin_overall','干燥'),
                                    ('burning_nrs','灼烧'), ('acne_severity','痤疮')]:
                    val = sensory.answers.get(key)
                    if val is not None:
                        s_parts.append(f'{label}{val}分')
                if s_parts:
                    lines.append(f'感官评价: {", ".join(s_parts)}')
            else:
                lines.append(f'问卷记录: {q_count}份')

        # 依从性
        comp = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if comp:
            lines.append(f'依从性: {comp.get_level_display()}, 到访率{comp.visit_attendance_rate}%')

        # 礼金结算
        pays = SubjectPayment.objects.filter(subject=subject, status='paid')
        if pays.exists():
            total_pay = sum(p.amount for p in pays if p.amount)
            proj_pays = []
            for p in pays[:5]:
                if p.notes:
                    parts = p.notes.replace('项目:', '').split()
                    if parts:
                        proj_pays.append(parts[0])
            lines.append(f'礼金结算: {pays.count()}次 ¥{total_pay} 项目:{",".join(filter(None, proj_pays))}')

        return '\n'.join(lines) if len(lines) >= 2 else None

    # =========================================================
    # Phase 5: 向量化（更新 index_status）
    # =========================================================
    def phase_vectorize(self):
        from apps.knowledge.models import KnowledgeEntry

        c = connection.cursor()
        c.execute("""
        UPDATE t_knowledge_entry 
        SET index_status='pending'
        WHERE source_type IN ('subject_full_lifecycle', 'subject_beauty_profile',
                              'recruit_lifecycle_stats')
        AND index_status != 'indexed'
        """)
        updated = c.rowcount
        self.counters['vectorized'] = updated
        self.stdout.write(self.style.SUCCESS(f'标记向量化: {updated:,} 条'))

        # 触发 process_pending_contexts 如果有 vectorize worker
        self.stdout.write('向量化任务已加入队列（由 process_pending_contexts 处理）')

    # =========================================================
    # 统计
    # =========================================================
    def _show_stats(self):
        c = connection.cursor()
        print('\n=== 受试者知识库统计 ===')
        for table, label in [
            ('t_subject', '受试者总数'),
            ('t_enrollment', '入组记录'),
            ('t_subject_payment', '礼金支付'),
            ('t_subject_project_sc', 'SC/RD映射'),
            ('t_skin_measurement_record', '仪器测量'),
            ('t_subject_questionnaire', '问卷数据'),
            ('t_subject_compliance', '依从性'),
            ('t_subject_skin_profile', '皮肤档案'),
        ]:
            c.execute(f'SELECT COUNT(*) FROM {table}')
            print(f'  {label:20s}: {c.fetchone()[0]:>10,}')

        print('\n=== 知识库统计 ===')
        c.execute("""
        SELECT source_type, COUNT(*), 
               SUM(CASE WHEN index_status='indexed' THEN 1 ELSE 0 END) as indexed
        FROM t_knowledge_entry
        WHERE source_type LIKE 'subject%' OR source_type LIKE 'recruit%'
        GROUP BY source_type ORDER BY COUNT(*) DESC
        """)
        for r in c.fetchall():
            print(f'  {r[0]:30s}: {r[1]:>8,} (已索引:{r[2]:,})')

        c.execute("SELECT COUNT(*) FROM t_knowledge_entity")
        c2 = connection.cursor()
        c2.execute("SELECT COUNT(*) FROM t_knowledge_relation")
        print(f'\n  知识图谱实体: {c.fetchone()[0]:,}')
        print(f'  知识图谱关系: {c2.fetchone()[0]:,}')
