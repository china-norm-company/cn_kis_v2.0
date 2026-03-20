"""
受试者全生命周期数据挖掘

以张煜佼（财务）+ 董彦吟（项目管理）为核心，从结算单、知情同意书、
访视记录、样品发放、不良反应事件等全面重建受试者档案。

核心数据来源：
  1. 员工受试者结算模板（姓名+身份证+手机+项目+金额）
  2. 兼职劳务费发放表（姓名+兼职编号+项目+金额）
  3. 灵工结算模板（同上）
  4. 受试者测试项目信息周报（已有）
  5. China-Norm数据库（已有）
  6. IM群聊招募统计（已有）

目标：受试者库 > 50,000

Usage:
  python manage.py mine_lifecycle_subjects --phase all
  python manage.py mine_lifecycle_subjects --phase settlement    # 结算单（最高价值）
  python manage.py mine_lifecycle_subjects --phase freelance     # 兼职劳务费
  python manage.py mine_lifecycle_subjects --phase weekly_report # 周报
  python manage.py mine_lifecycle_subjects --phase icf           # 知情同意
  python manage.py mine_lifecycle_subjects --phase visit         # 访视记录
  python manage.py mine_lifecycle_subjects --phase vectorize     # 向量化
  python manage.py mine_lifecycle_subjects --stats
"""
import os
import re
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.db import connection, transaction
from django.utils import timezone

logger = logging.getLogger(__name__)
MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')
PROJECT_P = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)
ID_CARD_P = re.compile(r'\d{17}[\dXx]')
PHONE_P = re.compile(r'1[3-9]\d{9}')


def norm(v):
    return str(v).strip() if v is not None else ''


def safe_dec(v):
    try:
        s = re.sub(r'[^\d.]', '', str(v))
        return Decimal(s) if s else None
    except (InvalidOperation, ValueError):
        return None


def mask_id(id_card):
    """身份证脱敏：保留前6后4"""
    s = str(id_card).strip()
    if len(s) == 18:
        return s[:6] + '********' + s[-4:]
    return s[:4] + '****' + s[-2:] if len(s) > 6 else s


def mask_phone(phone):
    """手机号脱敏：保留前3后4"""
    s = re.sub(r'[^\d]', '', str(phone))
    if len(s) == 11:
        return s[:3] + '****' + s[-4:]
    return s


def extract_project(text):
    m = PROJECT_P.search(str(text))
    return m.group(0).upper() if m else ''


class Command(BaseCommand):
    help = '受试者全生命周期数据挖掘（以财务结算为核心）'

    def add_arguments(self, parser):
        parser.add_argument('--phase', default='all',
                            choices=['all', 'settlement', 'freelance', 'weekly_report',
                                     'icf', 'visit', 'vectorize', 'kg'])
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--limit', type=int, default=0)
        parser.add_argument('--stats', action='store_true')

    def handle(self, *args, **options):
        if options['stats']:
            self._show_stats()
            return

        self.dry_run = options['dry_run']
        self.limit = options['limit']
        self.counters = defaultdict(int)

        phase = options['phase']
        phases = ['settlement', 'freelance', 'weekly_report', 'icf', 'visit', 'kg', 'vectorize'] \
            if phase == 'all' else [phase]

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}', lambda: None)()

        self.stdout.write(self.style.SUCCESS(
            f'\n=== 汇总 ===\n'
            f'  新建受试者: {self.counters["new"]}\n'
            f'  更新受试者: {self.counters["updated"]}\n'
            f'  知识档案: {self.counters["profiles"]}\n'
            f'  向量化: {self.counters["vectorized"]}'
        ))

    # ==================================================================
    # 1. 员工受试者结算模板（最高价值 — 含真实身份信息）
    # ==================================================================
    def phase_settlement(self):
        import openpyxl, xlrd

        files = self._find_settlement_files()
        self.stdout.write(f'找到 {len(files)} 个结算相关文件')

        all_records = []
        for fpath in files:
            fname = os.path.basename(fpath)
            try:
                if fname.endswith('.xls'):
                    records = self._parse_xls_settlement(fpath)
                else:
                    records = self._parse_xlsx_settlement(fpath, openpyxl)
                if records:
                    self.stdout.write(f'  {fname}: {len(records)} 条')
                    all_records.extend(records)
            except Exception as e:
                logger.warning('结算文件解析失败 %s: %s', fname, e)

        # 去重（按姓名+手机号）
        seen = set()
        unique = []
        for rec in all_records:
            key = f"{rec.get('name','')}{rec.get('phone','')[-4:]}"
            if key and key != '' and key not in seen:
                seen.add(key)
                unique.append(rec)

        self.stdout.write(f'\n解析 {len(all_records)} 条, 去重后 {len(unique)} 名唯一受试者')

        if not self.dry_run:
            saved = self._upsert_settlement_subjects(unique)
            self.stdout.write(self.style.SUCCESS(f'结算: 写入 {saved} 条'))

    def _find_settlement_files(self):
        results = set()
        kws_path = ['员工受试者结算', '结算模板', '礼金', '受试者结算']
        kws_email = ['员工受试者', '礼金', '兼职劳务', '受试者结算']

        # 从数据库查张煜佼的结算附件
        c = connection.cursor()
        c.execute("""
        SELECT DISTINCT metadata->>'local_path'
        FROM t_personal_context
        WHERE source_type='mail_attachment'
        AND user_id='ou_a3a9c72e3a78dfb64d29d4483352acd2'
        AND (metadata->>'local_path' LIKE '%.xlsx' OR metadata->>'local_path' LIKE '%.xls')
        AND (
            metadata->>'local_path' ILIKE '%结算%'
            OR metadata->>'local_path' ILIKE '%礼金%'
            OR metadata->>'local_path' ILIKE '%受试者%'
            OR metadata->>'local_path' ILIKE '%劳务%'
            OR metadata->>'local_path' ILIKE '%发放%'
            OR metadata->>'local_path' ILIKE '%银行%'
            OR metadata->>'subject' ILIKE '%员工受试者%'
            OR metadata->>'subject' ILIKE '%礼金%'
            OR metadata->>'subject' ILIKE '%兼职劳务%'
        )
        """)
        for r in c.fetchall():
            if r[0]:
                fpath = os.path.join(MEDIA_ROOT, r[0].lstrip('/'))
                if os.path.exists(fpath):
                    results.add(fpath)

        # 磁盘扫描
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~'):
                    continue
                if not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f for kw in kws_path):
                    results.add(os.path.join(root, f))

        if self.limit:
            return list(results)[:self.limit]
        return list(results)

    def _parse_xlsx_settlement(self, fpath, openpyxl):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        records = []
        fname = os.path.basename(fpath)
        project_code = extract_project(fname)

        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 5000), values_only=True))
            if not rows:
                continue

            # 找表头
            header_idx = 0
            for i, row in enumerate(rows[:5]):
                row_text = ' '.join(norm(c) for c in row if c)
                if any(k in row_text for k in ['姓名', '手机', '身份证', '项目编号', '结算金额']):
                    header_idx = i
                    break

            header = [norm(c) for c in rows[header_idx]]
            col = self._detect_settlement_columns(header)
            if not col.get('name') and not col.get('phone'):
                continue

            sheet_project = extract_project(sname) or project_code

            for row in rows[header_idx + 1:]:
                if not row:
                    continue
                rec = self._extract_settlement_row(row, col, sheet_project, fname, sname)
                if rec:
                    records.append(rec)

        wb.close()
        return records

    def _parse_xls_settlement(self, fpath):
        try:
            import xlrd
        except ImportError:
            return []

        try:
            wb = xlrd.open_workbook(fpath)
        except Exception:
            return []

        records = []
        fname = os.path.basename(fpath)
        project_code = extract_project(fname)

        for ws in wb.sheets():
            rows = [ws.row_values(i) for i in range(min(ws.nrows, 5000))]
            if not rows:
                continue

            header_idx = 0
            for i, row in enumerate(rows[:5]):
                row_text = ' '.join(str(c) for c in row if c)
                if any(k in row_text for k in ['姓名', '手机', '身份证', '项目编号']):
                    header_idx = i
                    break

            header = [norm(c) for c in rows[header_idx]]
            col = self._detect_settlement_columns(header)
            if not col.get('name') and not col.get('phone'):
                continue

            sheet_project = extract_project(ws.name) or project_code

            for row in rows[header_idx + 1:]:
                rec = self._extract_settlement_row(row, col, sheet_project, fname, ws.name)
                if rec:
                    records.append(rec)

        return records

    def _detect_settlement_columns(self, header):
        col = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if '项目编号' in h:
                col['project'] = i
            elif '结算金额' in h or '应发' in h or '实际发放' in h:
                col['amount'] = i
            elif h == '姓名' or (('姓名' in h or 'name' in hl) and 'name' not in col):
                col['name'] = i
            elif '身份证' in h:
                col['id_card'] = i
            elif '手机号' in h or '手机' in h:
                col['phone'] = i
            elif '银行账号' in h or '卡号' in h:
                col['bank_card'] = i
            elif '开户银行' in h:
                col['bank'] = i
            elif '兼职编号' in h or '编号' == h:
                col['member_id'] = i
            elif '工作岗位' in h or '岗位' == h:
                col['role'] = i
            elif '备注' in h:
                col['notes'] = i
        return col

    def _extract_settlement_row(self, row, col, project_code, fname, sname):
        def get(field):
            ci = col.get(field)
            if ci is not None and ci < len(row) and row[ci] is not None:
                return norm(row[ci])
            return ''

        name = get('name')
        if not name or name in ('姓名', '合计', '小计', '总计', ''):
            return None

        phone = get('phone')
        id_card = get('id_card')
        amount = safe_dec(get('amount'))

        # 身份证格式检查
        id_clean = re.sub(r'[\s\-]', '', id_card)
        if id_clean and not re.match(r'^\d{17}[\dXx]$', id_clean):
            id_clean = ''

        # 手机号格式检查
        phone_clean = re.sub(r'[^\d]', '', phone)
        if phone_clean and (len(phone_clean) != 11 or not phone_clean.startswith('1')):
            phone_clean = ''

        return {
            'name': name[:50],
            'phone': phone_clean,
            'phone_masked': mask_phone(phone_clean) if phone_clean else '',
            'id_card': id_clean,
            'id_card_masked': mask_id(id_clean) if id_clean else '',
            'bank_card': get('bank_card'),
            'bank': get('bank'),
            'amount': amount,
            'project_code': get('project') or project_code,
            'member_id': get('member_id'),
            'role': get('role'),
            'notes': get('notes'),
            'source': 'settlement',
            'source_file': f'{fname}/{sname}',
        }

    def _upsert_settlement_subjects(self, records):
        from apps.subject.models import Subject, Enrollment, EnrollmentStatus
        from apps.protocol.models import Protocol
        from apps.knowledge.models import KnowledgeEntry

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name', '').strip()
            phone = rec.get('phone', '')
            id_card = rec.get('id_card', '')
            project_code = rec.get('project_code', '')

            if not name:
                continue

            try:
                subject = None

                # 精准匹配：优先用手机号
                if phone:
                    subject = Subject.objects.filter(phone=phone[:20]).first()

                # 其次按姓名+手机尾号
                if subject is None and phone and len(phone) >= 4:
                    subject = Subject.objects.filter(
                        name=name, phone__endswith=phone[-4:]
                    ).first()

                # 按身份证（加密存储）
                if subject is None and id_card:
                    subject = Subject.objects.filter(
                        id_card_encrypted__contains=mask_id(id_card)
                    ).first()

                # 最后按姓名
                if subject is None:
                    subject = Subject.objects.filter(name=name).first()

                if subject is None:
                    subject = Subject(
                        name=name, phone=phone[:20] if phone else '',
                        id_card_encrypted=mask_id(id_card) if id_card else '',
                        source_channel='database', status='completed',
                    )
                    subject.save()
                    self.counters['new'] += 1
                    saved += 1
                else:
                    changed = False
                    if not subject.phone and phone:
                        subject.phone = phone[:20]
                        changed = True
                    if not subject.id_card_encrypted and id_card:
                        subject.id_card_encrypted = mask_id(id_card)
                        changed = True
                    if changed:
                        subject.save(update_fields=['phone', 'id_card_encrypted', 'update_time'])
                    self.counters['updated'] += 1
                    saved += 1

                # 关联项目
                if project_code and subject.pk:
                    if project_code not in proto_cache:
                        proto_cache[project_code] = Protocol.objects.filter(
                            code__iexact=project_code
                        ).first()
                    proto = proto_cache.get(project_code)
                    if proto:
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=proto,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )

                # 礼金支付记录
                if subject.pk and rec.get('amount'):
                    from apps.subject.models_execution import SubjectPayment, PaymentType
                    SubjectPayment.objects.get_or_create(
                        subject=subject,
                        payment_no=f"settle_{subject.id}_{project_code}_{rec.get('source_file','')[:20]}",
                        defaults={
                            'payment_type': PaymentType.VISIT_COMPENSATION,
                            'amount': rec['amount'],
                            'status': 'paid',
                            'notes': f"来源:{rec.get('source_file','')[:50]}",
                        }
                    )

                self.counters['profiles'] += 1

            except Exception as e:
                logger.debug('settlement upsert 跳过: %s, name=%s', e, name)

        return saved

    # ==================================================================
    # 2. 兼职劳务费发放表（含兼职编号+岗位）
    # ==================================================================
    def phase_freelance(self):
        import openpyxl

        kws = ['兼职劳务费', '兼职联络员', '兼职人员', '访问员费用', '劳务报酬']
        files = self._find_by_keywords(kws)
        self.stdout.write(f'找到 {len(files)} 个兼职劳务文件')

        all_records = []
        for fpath in files:
            fname = os.path.basename(fpath)
            if not fname.endswith('.xlsx'):
                continue
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 2000), values_only=True))
                    if not rows:
                        continue
                    project_code = extract_project(sname) or extract_project(fname)
                    records = self._parse_freelance_sheet(rows, project_code, fname, sname)
                    if records:
                        all_records.extend(records)
                wb.close()
            except Exception as e:
                logger.debug('兼职解析失败 %s: %s', fname, e)

        # 去重
        seen = set()
        unique = [r for r in all_records if (k := f"{r.get('name','')}{r.get('member_id','')}") and k not in seen and not seen.add(k)]

        self.stdout.write(f'解析 {len(all_records)} 条, 去重后 {len(unique)} 条')

        if not self.dry_run:
            saved = self._upsert_freelance_subjects(unique)
            self.stdout.write(self.style.SUCCESS(f'兼职: 写入 {saved} 条'))

    def _parse_freelance_sheet(self, rows, project_code, fname, sname):
        if not rows:
            return []

        header_idx = 0
        for i, row in enumerate(rows[:5]):
            row_text = ' '.join(norm(c) for c in row if c)
            if any(k in row_text for k in ['姓名', '兼职编号', '工作岗位', '序号']):
                header_idx = i
                break

        header = [norm(c) for c in rows[header_idx]]
        col = self._detect_settlement_columns(header)

        records = []
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            name = norm(row[col['name']]) if col.get('name') and col['name'] < len(row) else ''
            if not name or name in ('姓名', '合计', ''):
                continue

            member_id = ''
            if col.get('member_id') and col['member_id'] < len(row):
                mid_raw = norm(row[col['member_id']])
                if mid_raw and re.match(r'^\d{3,6}$', mid_raw):
                    member_id = mid_raw

            role = norm(row[col['role']]) if col.get('role') and col['role'] < len(row) else ''
            amount = safe_dec(row[col['amount']]) if col.get('amount') and col['amount'] < len(row) else None
            proj = norm(row[col['project']]) if col.get('project') and col['project'] < len(row) else project_code

            records.append({
                'name': name[:50],
                'member_id': member_id,
                'role': role,
                'amount': amount,
                'project_code': proj or project_code,
                'source': 'freelance',
                'source_file': f'{fname}/{sname}',
            })
        return records

    def _upsert_freelance_subjects(self, records):
        from apps.subject.models import Subject, Enrollment, EnrollmentStatus
        from apps.protocol.models import Protocol
        from apps.subject.models_execution import SubjectProjectSC

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name', '').strip()
            member_id = rec.get('member_id', '')
            project_code = rec.get('project_code', '')

            if not name:
                continue

            try:
                subject = Subject.objects.filter(name=name).first()
                if subject is None:
                    subject = Subject(
                        name=name, source_channel='database', status='completed',
                    )
                    subject.save()
                    self.counters['new'] += 1
                    saved += 1
                else:
                    self.counters['updated'] += 1
                    saved += 1

                # 兼职编号记录 → SC号
                if member_id and project_code and subject.pk:
                    SubjectProjectSC.objects.get_or_create(
                        subject=subject, project_code=project_code,
                        defaults={'sc_number': member_id, 'rd_number': ''},
                    )

                if project_code and subject.pk:
                    if project_code not in proto_cache:
                        proto_cache[project_code] = Protocol.objects.filter(
                            code__iexact=project_code
                        ).first()
                    proto = proto_cache.get(project_code)
                    if proto:
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=proto,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )

            except Exception as e:
                logger.debug('freelance upsert 跳过: %s', e)

        return saved

    # ==================================================================
    # 3. 受试者测试项目信息周报
    # ==================================================================
    def phase_weekly_report(self):
        import openpyxl

        files = self._find_by_keywords(['受试者测试项目信息周报'])
        self.stdout.write(f'找到 {len(files)} 个周报文件')

        all_records = []
        for fpath in files:
            fname = os.path.basename(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    wb.close()
                    continue

                header = [norm(c) for c in rows[0]]
                col = {}
                for i, h in enumerate(header):
                    if '受试者名称' in h or '姓名' in h:
                        col['name'] = i
                    elif '手机号' in h:
                        col['phone_masked'] = i
                    elif '性别' in h:
                        col['gender'] = i
                    elif '项目编号' in h:
                        col['project'] = i
                    elif '开始' in h:
                        col['start_date'] = i
                    elif '结束' in h:
                        col['end_date'] = i
                    elif '部位' in h:
                        col['test_site'] = i

                if 'name' not in col:
                    wb.close()
                    continue

                for row in rows[1:]:
                    if not row:
                        continue
                    rec = {f: norm(row[ci]) for f, ci in col.items() if ci < len(row) and row[ci] is not None}
                    name = rec.get('name', '')
                    project_code = rec.get('project', '')
                    if not name or not project_code:
                        continue

                    pm = rec.get('phone_masked', '')
                    last4 = re.search(r'(\d{4})\s*$', pm)
                    rec['phone_last4'] = last4.group(1) if last4 else ''
                    rec['source'] = 'weekly_report'
                    rec['source_file'] = fname
                    all_records.append(rec)

                wb.close()
            except Exception as e:
                logger.debug('周报解析失败 %s: %s', fname, e)

        # 去重
        seen = set()
        unique = []
        for rec in all_records:
            key = f"{rec.get('name','')[:20]}_{rec.get('project','')}"
            if key not in seen and key != '_':
                seen.add(key)
                unique.append(rec)

        self.stdout.write(f'解析 {len(all_records)} 条, 去重后 {len(unique)} 条')

        if not self.dry_run:
            saved = self._upsert_weekly(unique)
            self.stdout.write(self.style.SUCCESS(f'周报: 写入 {saved} 条'))

    def _upsert_weekly(self, records):
        from apps.subject.models import Subject, Enrollment, EnrollmentStatus
        from apps.protocol.models import Protocol
        from apps.knowledge.models import KnowledgeEntry

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name', '').strip()
            project_code = rec.get('project', '').strip()
            if not name:
                continue

            gender_raw = rec.get('gender', '')
            gender = 'female' if '女' in gender_raw else ('male' if '男' in gender_raw else '')
            phone4 = rec.get('phone_last4', '')

            try:
                subject = Subject.objects.filter(name=name).first()
                if subject is None:
                    subject = Subject(
                        name=name[:100], gender=gender,
                        source_channel='database', status='completed',
                    )
                    subject.save()
                    self.counters['new'] += 1
                    saved += 1
                else:
                    if not subject.gender and gender:
                        subject.gender = gender
                        subject.save(update_fields=['gender', 'update_time'])
                    self.counters['updated'] += 1
                    saved += 1

                if project_code and subject.pk:
                    if project_code not in proto_cache:
                        proto_cache[project_code] = Protocol.objects.filter(
                            code__iexact=project_code
                        ).first()
                    proto = proto_cache.get(project_code)
                    if proto:
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=proto,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )

            except Exception as e:
                logger.debug('weekly upsert 跳过: %s', e)

        return saved

    # ==================================================================
    # 4. ICF（知情同意书）数据
    # ==================================================================
    def phase_icf(self):
        """从邮件和附件中提取知情同意相关信息"""
        c = connection.cursor()

        # 从IM中提取知情同意相关消息
        c.execute("""
        SELECT raw_content, metadata->>'chat_name', created_at::date
        FROM t_personal_context
        WHERE source_type='im'
        AND (raw_content ILIKE '%ICF%' OR raw_content ILIKE '%知情同意%'
             OR raw_content ILIKE '%知情书%')
        ORDER BY created_at DESC
        LIMIT 500
        """)
        msgs = c.fetchall()
        self.stdout.write(f'ICF相关IM消息: {len(msgs)}')

        icf_subjects = []
        sc_rd_p = re.compile(r'(SC|RD)(\d{3,6})', re.IGNORECASE)
        proj_p = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)

        for raw, chat, dt in msgs:
            content = str(raw or '')
            projs = [p.upper() for p in proj_p.findall(content)]
            for m in sc_rd_p.finditer(content):
                rec = {
                    'sc_or_rd': f'{m.group(1).upper()}{m.group(2)}',
                    'projects': projs,
                    'date': str(dt),
                    'chat': chat or '',
                    'event_type': 'icf_signed',
                }
                icf_subjects.append(rec)

        self.stdout.write(f'从ICF消息提取受试者信号: {len(icf_subjects)}')

        if not self.dry_run and icf_subjects:
            self._write_icf_knowledge(icf_subjects)

    def _write_icf_knowledge(self, signals):
        from apps.knowledge.models import KnowledgeEntry

        summary = defaultdict(lambda: {'projects': set(), 'dates': []})
        for sig in signals:
            k = sig['sc_or_rd']
            for p in sig['projects']:
                summary[k]['projects'].add(p)
            summary[k]['dates'].append(sig['date'])

        content = f"ICF知情同意签署信号 (共{len(summary)}名受试者)\n\n"
        for sc_rd, info in list(summary.items())[:100]:
            content += f"{sc_rd}: 项目={list(info['projects'])[:3]}, 日期={info['dates'][-1] if info['dates'] else '?'}\n"

        KnowledgeEntry.objects.update_or_create(
            source_type='icf_signals',
            source_id=0,
            source_key='icf_signals_all',
            defaults={
                'entry_type': 'lesson_learned',
                'title': f'知情同意书签署信号汇总({len(summary)}名受试者)',
                'content': content,
                'status': 'published', 'is_published': True,
            }
        )
        self.counters['profiles'] += 1

    # ==================================================================
    # 5. 访视记录
    # ==================================================================
    def phase_visit(self):
        c = connection.cursor()

        # 从IM中提取访视数字
        c.execute("""
        SELECT raw_content, metadata->>'chat_name', created_at::date
        FROM t_personal_context
        WHERE source_type='im'
        AND (raw_content ILIKE '%T0%' OR raw_content ILIKE '%T4W%' OR raw_content ILIKE '%T8W%'
             OR raw_content ILIKE '%T12W%' OR raw_content ILIKE '%访视%' OR raw_content ILIKE '%V0%'
             OR raw_content ILIKE '%V1%' OR raw_content ILIKE '%基线%' OR raw_content ILIKE '%Baseline%')
        AND (raw_content ILIKE '%SC%' OR raw_content ILIKE '%RD%' OR raw_content ILIKE '%入组%')
        ORDER BY created_at DESC
        LIMIT 1000
        """)
        msgs = c.fetchall()
        self.stdout.write(f'访视相关IM消息: {len(msgs)}')

        visit_pattern = re.compile(r'(T\d+\.?\d*[WwMm]?|V\d+|基线|Baseline)', re.IGNORECASE)
        sc_rd_p = re.compile(r'(SC|RD)(\d{3,6})', re.IGNORECASE)
        proj_p = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)

        visit_records = []
        for raw, chat, dt in msgs:
            content = str(raw or '')
            projs = [p.upper() for p in proj_p.findall(content)]
            for m in sc_rd_p.finditer(content):
                timepoints = visit_pattern.findall(content)
                if timepoints:
                    visit_records.append({
                        'subject_id': f'{m.group(1).upper()}{m.group(2)}',
                        'timepoints': list(set(tp.upper() for tp in timepoints[:5])),
                        'projects': projs[:3],
                        'date': str(dt),
                    })

        self.stdout.write(f'提取访视记录: {len(visit_records)}')

        if not self.dry_run and visit_records:
            from apps.knowledge.models import KnowledgeEntry
            # 按项目汇总
            proj_visits = defaultdict(lambda: defaultdict(set))
            for vr in visit_records:
                for proj in (vr['projects'] or ['未知']):
                    for tp in vr['timepoints']:
                        proj_visits[proj][tp].add(vr['subject_id'])

            content = f"访视记录信号汇总 ({len(visit_records)}条)\n\n"
            for proj, tps in sorted(proj_visits.items())[:50]:
                content += f"\n{proj}:\n"
                for tp, subjects in sorted(tps.items()):
                    content += f"  {tp}: {len(subjects)}名受试者 (如: {', '.join(list(subjects)[:3])})\n"

            KnowledgeEntry.objects.update_or_create(
                source_type='visit_signals',
                source_id=0,
                source_key='visit_signals_all',
                defaults={
                    'entry_type': 'lesson_learned',
                    'title': f'访视记录信号汇总',
                    'content': content,
                    'status': 'published', 'is_published': True,
                }
            )
            self.counters['profiles'] += 1

    # ==================================================================
    # 6. 知识图谱 — 受试者全生命周期
    # ==================================================================
    def phase_kg(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_execution import SubjectPayment
        from apps.knowledge.models import KnowledgeEntity, KnowledgeRelation

        self.stdout.write('构建全生命周期知识图谱...')

        # 枚举生命周期阶段实体
        stages = ['报名', '初筛', '知情同意', '基线测量', '产品使用', '随访', '完成', '结算', '脱落']
        for stage in stages:
            KnowledgeEntity.objects.get_or_create(
                uri=f'cnkis:lifecycle_stage/{stage}',
                namespace='cnkis',
                defaults={
                    'entity_type': 'concept',
                    'label': stage,
                    'label_en': stage,
                    'definition': f'受试者生命周期阶段: {stage}',
                }
            )

        # 为有支付记录的受试者建立结算关系
        paid_enrollments = Enrollment.objects.select_related(
            'subject', 'protocol'
        ).filter(
            subject__payments__isnull=False
        ).distinct()

        self.stdout.write(f'有结算记录的入组: {paid_enrollments.count()}')

        r_created = e_created = 0
        for enr in paid_enrollments[:2000]:
            subject = enr.subject
            proto = enr.protocol
            if not proto:
                continue

            # 受试者实体
            uri = f'cnkis:subject/{subject.subject_no or subject.id}'
            subj_entity, s_new = KnowledgeEntity.objects.get_or_create(
                uri=uri, namespace='cnkis',
                defaults={
                    'entity_type': 'instance',
                    'label': subject.subject_no or str(subject.id),
                    'definition': f'{subject.name}, {subject.get_gender_display() or ""}, {subject.age or "?"}岁',
                    'properties': {'subject_id': subject.id},
                }
            )
            if s_new:
                e_created += 1

            # 完成阶段关系
            completed_entity = KnowledgeEntity.objects.filter(
                uri='cnkis:lifecycle_stage/完成'
            ).first()
            if completed_entity:
                _, r_new = KnowledgeRelation.objects.get_or_create(
                    subject=subj_entity, object=completed_entity,
                    predicate_uri='cnkis:reached_stage',
                    relation_type='custom',
                    defaults={'source': 'mine_lifecycle'},
                )
                if r_new:
                    r_created += 1

        self.counters['kg_entities'] = e_created
        self.counters['kg_relations'] = r_created
        self.stdout.write(self.style.SUCCESS(f'KG: 新增 {e_created} 实体, {r_created} 关系'))

    # ==================================================================
    # 7. 向量化（批量生成全景文本并索引）
    # ==================================================================
    def phase_vectorize(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord, SubjectPayment
        from apps.subject.models_domain import SkinProfile
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'向量化受试者档案: {total:,}')

        created = updated = 0
        for i in range(0, total, 500):
            batch = subjects[i:i + 500]
            for subject in batch:
                text = self._build_vector_text(subject)
                if not text:
                    continue

                if self.dry_run:
                    created += 1
                    continue

                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_full_lifecycle',
                    source_id=subject.id,
                    source_key=f'lifecycle_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 全档案',
                        'content': text,
                        'status': 'published',
                        'is_published': True,
                        'index_status': 'pending',  # 触发向量化
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1

            if (i + 500) % 2000 == 0:
                self.stdout.write(f'  进度: {min(i + 500, total)}/{total}')

        self.counters['vectorized'] = created + updated
        self.stdout.write(self.style.SUCCESS(f'向量化: 新建 {created}, 更新 {updated}'))

    def _build_vector_text(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord, SubjectPayment
        from apps.subject.models_domain import SkinProfile

        lines = [f'受试者: {subject.name}']

        if subject.subject_no:
            lines.append(f'编号: {subject.subject_no}')

        demo = []
        if subject.gender:
            demo.append(subject.get_gender_display())
        if subject.age:
            demo.append(f'{subject.age}岁')
        if demo:
            lines.append(f'基本信息: {", ".join(demo)}')

        if subject.phone:
            lines.append(f'手机尾号: {subject.phone[-4:]}')

        # 皮肤档案
        skin = SkinProfile.objects.filter(subject=subject).first()
        if skin:
            parts = []
            if skin.fitzpatrick_type:
                parts.append(f'Fitzpatrick {skin.fitzpatrick_type}型')
            if skin.skin_type_u_zone:
                parts.append(f'{skin.skin_type_u_zone}性肌')
            if skin.moisture_baseline:
                parts.append(f'水分基线{skin.moisture_baseline}')
            if parts:
                lines.append(f'皮肤特征: {", ".join(parts)}')

        # 项目参与
        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        enr_count = enrollments.count()
        if enr_count > 0:
            projs = [e.protocol.code for e in enrollments[:6] if e.protocol]
            lines.append(f'参与项目({enr_count}个): {", ".join(projs)}')

        # 测量数据
        m_count = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m_count:
            latest_m = SkinMeasurementRecord.objects.filter(subject=subject).order_by('-measured_at').first()
            m_parts = []
            for field, label in [('moisture', '水分'), ('tewl', 'TEWL'), ('sebum', '皮脂'), ('elasticity', '弹性')]:
                val = getattr(latest_m, field, None)
                if val:
                    m_parts.append(f'{label}{val}')
            if m_parts:
                lines.append(f'皮肤测量({m_count}次): {", ".join(m_parts)}')

        # 问卷
        q_count = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q_count:
            lines.append(f'问卷记录: {q_count}份')

        # 依从性
        compliance = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if compliance:
            lines.append(f'依从性: {compliance.get_level_display()}, 到访率{compliance.visit_attendance_rate}%')

        # 结算
        payments = SubjectPayment.objects.filter(subject=subject, status='paid')
        if payments.exists():
            total_amt = sum(p.amount for p in payments if p.amount)
            lines.append(f'礼金结算: {len(payments)}次, 累计¥{total_amt}')

        return '\n'.join(lines) if len(lines) >= 2 else None

    # ==================================================================
    # 工具方法
    # ==================================================================
    def _find_by_keywords(self, kws):
        results = set()
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f for kw in kws):
                    results.add(os.path.join(root, f))
        if self.limit:
            return list(results)[:self.limit]
        return list(results)

    def _show_stats(self):
        c = connection.cursor()
        print('\n=== 受试者全生命周期统计 ===')
        for table, label in [
            ('t_subject', '受试者总数'),
            ('t_enrollment', '入组记录'),
            ('t_subject_project_sc', 'SC/RD映射'),
            ('t_skin_measurement_record', '仪器测量'),
            ('t_subject_questionnaire', '问卷'),
            ('t_subject_compliance', '依从性'),
            ('t_subject_payment', '礼金支付'),
            ('t_subject_skin_profile', '皮肤档案'),
        ]:
            c.execute(f'SELECT COUNT(*) FROM {table}')
            print(f'  {label:20s}: {c.fetchone()[0]:>10,}')

        for src in ['subject_full_lifecycle', 'subject_beauty_profile',
                    'recruit_im_stats', 'icf_signals', 'visit_signals']:
            c.execute(f"SELECT COUNT(*) FROM t_knowledge_entry WHERE source_type='{src}'")
            print(f'  KE/{src[:22]:22s}: {c.fetchone()[0]:>10,}')

        c.execute("SELECT COUNT(*) FROM t_knowledge_entity")
        print(f'  {"知识图谱实体":20s}: {c.fetchone()[0]:>10,}')
        c.execute("SELECT COUNT(*) FROM t_knowledge_relation")
        print(f'  {"知识图谱关系":20s}: {c.fetchone()[0]:>10,}')
