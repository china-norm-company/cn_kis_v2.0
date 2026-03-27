"""
殷淑雯&王梦丹上下游受试者数据深度挖掘

从已采集的附件（特别是顾晶的7656个Raw Data Excel）
提取受试者信息，建立完整档案。

执行：
  python manage.py dig_downstream_subjects --phase all
  python manage.py dig_downstream_subjects --phase parse_attachments   # 解析已下载附件
  python manage.py dig_downstream_subjects --phase monitor             # 显示采集进度
  python manage.py dig_downstream_subjects --phase extract_subjects    # 提取受试者
"""
import os
import re
import logging
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.db import connection

logger = logging.getLogger(__name__)
MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')

# 核心招募相关人员
KEY_PERSONS = {
    '殷淑雯': {'open_id': 'ou_f4233d085de300ec5082439be0a71262', 'role': '招募负责人'},
    '王梦丹': {'open_id': 'ou_02f2a426b62fe050fd90817b19975fbb', 'role': '招募'},
    '顾晶': {'open_id': 'ou_30d27f8999d5285b26b6e167284f6b1e', 'role': '数据分析'},
    '段晨': {'open_id': 'ou_c48e17288348ca050aff9c64edbcafc9', 'role': '运营'},
    '童晓婷': {'open_id': 'ou_655f67046ef7a4b947521528a3f580d2', 'role': '招募执行'},
    '李思雨': {'open_id': 'ou_4fdb0aedd7a69f1bad00b4e8b214641e', 'role': '招募执行'},
    '卫婷婷': {'open_id': 'ou_44a7cba596e08dcbcca1389f57b057ae', 'role': '招募'},
    '茅晓珏': {'open_id': 'ou_ad72da4c356693f92608505890f6a5f3', 'role': '招募'},
    '姚志成': {'open_id': 'ou_e2db4f4935546ac6bddd2d4ebad6be60', 'role': '招募'},
}

PROJECT_P = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)


class Command(BaseCommand):
    help = '殷淑雯&王梦丹上下游受试者数据深度挖掘'

    def add_arguments(self, parser):
        parser.add_argument('--phase', default='all',
                            choices=['all', 'monitor', 'parse_attachments',
                                     'extract_subjects', 'im_recruit_stats',
                                     'build_profiles'])
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        phase = options['phase']
        phases = ['monitor', 'parse_attachments', 'extract_subjects',
                  'im_recruit_stats', 'build_profiles'] if phase == 'all' else [phase]

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}')()

    # ==================================================================
    # 监控采集进度
    # ==================================================================
    def phase_monitor(self):
        c = connection.cursor()
        c.execute("""
        SELECT cp.user_name, cp.source_type, cp.status,
               cp.total_fetched, cp.total_deposited, cp.total_errors,
               cp.updated_at
        FROM t_feishu_migration_checkpoint cp
        WHERE cp.user_name IN (
            '殷淑雯','王梦丹','童晓婷','李思雨','段晨','卫婷婷','顾晶','姚志成','徐蓓蕾','茅晓珏'
        )
        ORDER BY cp.user_name, cp.source_type
        """)
        rows = c.fetchall()
        self.stdout.write(f"\n{'用户':10s} {'类型':15s} {'状态':12s} {'已取':>8} {'已存':>8} {'错误':>6}")
        for r in rows:
            status_icon = '🔄' if r[2] == 'running' else ('✅' if r[2] == 'completed' else '⏳')
            self.stdout.write(
                f"  {status_icon} {r[0]:10s} {r[1]:15s} {r[2]:12s} "
                f"{r[3]:>8,} {r[4]:>8,} {r[5]:>6}"
            )

        # 实时 PersonalContext 统计
        c.execute("""
        SELECT a.display_name, pc.source_type, COUNT(*) as cnt,
               MAX(pc.created_at)::date as latest
        FROM t_personal_context pc
        JOIN t_account a ON a.feishu_open_id = pc.user_id
        WHERE a.display_name IN (
            '殷淑雯','王梦丹','童晓婷','李思雨','段晨','卫婷婷','顾晶','姚志成','徐蓓蕾','茅晓珏'
        )
        GROUP BY a.display_name, pc.source_type
        ORDER BY a.display_name, pc.source_type
        """)
        self.stdout.write('\n\n数据量统计:')
        for r in c.fetchall():
            self.stdout.write(f"  {r[0]:10s} {r[1]:20s} {r[2]:>10,} (至 {r[3]})")

    # ==================================================================
    # 解析已下载附件（特别是顾晶的Raw Data）
    # ==================================================================
    def phase_parse_attachments(self):
        import openpyxl

        # 找顾晶及相关人员的已下载Excel附件
        c = connection.cursor()
        c.execute("""
        SELECT DISTINCT metadata->>'local_path' as path, metadata->>'subject' as email_subj
        FROM t_personal_context
        WHERE source_type='mail_attachment'
        AND (metadata->>'local_path' LIKE '%.xlsx' OR metadata->>'local_path' LIKE '%.xls')
        AND metadata->>'local_path' IS NOT NULL
        AND metadata->>'local_path' != ''
        AND user_id IN (
            SELECT feishu_open_id FROM t_account WHERE display_name IN (
                '殷淑雯','王梦丹','童晓婷','李思雨','段晨','卫婷婷','顾晶'
            )
        )
        """)
        attachments = c.fetchall()
        self.stdout.write(f'找到 {len(attachments)} 个待解析 Excel 附件')

        parsed_records = []
        for local_path, email_subj in attachments:
            fpath = os.path.join(MEDIA_ROOT, local_path.lstrip('/'))
            if not os.path.exists(fpath):
                continue

            project_code = self._extract_project(os.path.basename(fpath)) or \
                           self._extract_project(email_subj or '')
            fname = os.path.basename(fpath)

            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            except Exception:
                continue

            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(max_row=min(ws.max_row, 500), values_only=True))
                if not rows:
                    continue

                records = self._try_parse_subject_sheet(rows, project_code, fname, sname)
                if records:
                    self.stdout.write(f'  {fname}/{sname}: {len(records)} 条受试者记录')
                    parsed_records.extend(records)

            wb.close()

        self.stdout.write(f'\n总解析受试者记录: {len(parsed_records)}')

        if not self.dry_run and parsed_records:
            saved = self._bulk_upsert_subjects(parsed_records)
            self.stdout.write(self.style.SUCCESS(f'写入 {saved} 条受试者记录'))

    def _try_parse_subject_sheet(self, rows, project_code, fname, sname):
        """智能识别并解析受试者相关sheet"""
        if not rows:
            return []

        # 收集前几行的文本特征
        header = [str(c).strip() if c else '' for c in rows[0]]
        preview_text = ' '.join(header)

        # 模式1: China-Norm 格式（含 Subject #, Initials, Age, Gender）
        if any(k in preview_text for k in ['Subject #', 'Subject Initials', '受试者编号', 'RD']):
            return self._parse_chinanorm_style(rows, project_code, fname)

        # 模式2: subinfo 格式（含 SC, RD, SEX, BRTHDAT）
        if len(rows) > 1:
            row1_text = ' '.join(str(c) for c in rows[1] if c)
            if any(k in row1_text for k in ['SUBJECT_CODE', 'FIELD1', 'SEX', 'BRTHDAT']):
                return self._parse_edc_subinfo(rows, project_code, fname)

        # 模式3: 受试者测试项目信息周报格式
        if '受试者名称' in preview_text and '项目编号' in preview_text:
            return self._parse_weekly_report_style(rows, project_code, fname)

        # 模式4: 合格受试者基础信息（SC, Age, Gender）
        if 'SC' in preview_text and ('Age' in preview_text or 'Gender' in preview_text):
            return self._parse_qualified_style(rows, project_code, fname)

        return []

    def _parse_chinanorm_style(self, rows, project_code, fname):
        header = [str(c).strip() if c else '' for c in rows[0]]
        col = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if 'cro study' in hl or '测试机构项目' in hl:
                col['project'] = i
            elif 'subject initials' in hl or '姓名缩写' in hl:
                col['initials'] = i
            elif 'subject #' in hl or '受试者编号' in hl:
                col['rd'] = i
            elif 'age' == hl or '年龄' in h:
                col['age'] = i
            elif 'gender' == hl or '性别' in h:
                col['gender'] = i
            elif 'skin type' in hl:
                col['skin_type'] = i
            elif 'fitzpatrick' in hl:
                col['fitzpatrick'] = i

        records = []
        for row in rows[2:]:
            if not row:
                continue
            rec = {}
            for f, ci in col.items():
                if ci < len(row) and row[ci] is not None:
                    rec[f] = str(row[ci]).strip()

            rd = rec.get('rd', '')
            if not rd or rd.lower() in ('e.g. rd001', 'subject #', ''):
                continue

            proj = rec.get('project', '') or project_code
            records.append({
                'name': rec.get('initials', ''),
                'rd_number': rd,
                'project_code': proj,
                'age': self._safe_int(rec.get('age', '')),
                'gender': self._norm_gender(rec.get('gender', '')),
                'skin_type': rec.get('skin_type', ''),
                'fitzpatrick': rec.get('fitzpatrick', ''),
                'source': 'chinanorm_attachment',
                'source_file': fname,
            })
        return records

    def _parse_edc_subinfo(self, rows, project_code, fname):
        if len(rows) < 3:
            return []
        header = [str(c).strip() if c else '' for c in rows[0]]
        field_row = [str(c).strip() if c else '' for c in rows[1]]

        sc_col = rd_col = sex_col = birth_col = None
        for i, (h, f) in enumerate(zip(header, field_row)):
            hl, fl = h.lower(), f.upper()
            if '筛选编号' in h or fl == 'SUBJECT_CODE':
                sc_col = i
            elif '入组编号' in h or fl == 'FIELD1':
                rd_col = i
            elif '性别' in h or fl == 'SEX':
                sex_col = i
            elif '出生日期' in h or fl == 'BRTHDAT':
                birth_col = i

        if sc_col is None:
            return []

        records = []
        for row in rows[2:]:
            if not row or len(row) <= sc_col:
                continue
            sc = str(row[sc_col]).strip() if row[sc_col] else ''
            rd = str(row[rd_col]).strip() if rd_col is not None and rd_col < len(row) and row[rd_col] else ''
            if not sc or sc in ('SUBJECT_CODE', '筛选编号SC', ''):
                continue

            gender = ''
            if sex_col is not None and sex_col < len(row):
                sex_val = str(row[sex_col]).strip()
                gender = 'female' if sex_val in ('2', '女', 'F', 'Female') else \
                         ('male' if sex_val in ('1', '男', 'M', 'Male') else '')

            age = None
            if birth_col is not None and birth_col < len(row) and row[birth_col]:
                try:
                    from datetime import datetime
                    bdate = str(row[birth_col]).strip()
                    if '-' in bdate:
                        birth = datetime.strptime(bdate[:10], '%Y-%m-%d')
                        age = (datetime.now() - birth).days // 365
                except Exception:
                    pass

            records.append({
                'name': f'{project_code}-RD{rd}' if rd else f'{project_code}-SC{sc}',
                'sc_number': sc,
                'rd_number': rd,
                'project_code': project_code,
                'age': age,
                'gender': gender,
                'source': 'edc_subinfo',
                'source_file': fname,
            })
        return records

    def _parse_weekly_report_style(self, rows, project_code, fname):
        header = [str(c).strip() if c else '' for c in rows[0]]
        col = {}
        for i, h in enumerate(header):
            if '受试者名称' in h or '姓名' in h:
                col['name'] = i
            elif '手机' in h:
                col['phone_masked'] = i
            elif '性别' in h:
                col['gender'] = i
            elif '项目编号' in h:
                col['project'] = i

        if 'name' not in col:
            return []

        records = []
        for row in rows[1:]:
            if not row:
                continue
            rec = {}
            for f, ci in col.items():
                if ci < len(row) and row[ci] is not None:
                    rec[f] = str(row[ci]).strip()

            name = rec.get('name', '')
            if not name:
                continue

            pm = rec.get('phone_masked', '')
            last4 = re.search(r'(\d{4})\s*$', pm)
            records.append({
                'name': name,
                'phone_last4': last4.group(1) if last4 else '',
                'gender': self._norm_gender(rec.get('gender', '')),
                'project_code': rec.get('project', '') or project_code,
                'source': 'weekly_report_attachment',
                'source_file': fname,
            })
        return records

    def _parse_qualified_style(self, rows, project_code, fname):
        header = [str(c).strip() if c else '' for c in rows[0]]
        col = {}
        for i, h in enumerate(header):
            if h.upper() == 'SC':
                col['sc'] = i
            elif h.upper() in ('RD', 'FIELD1'):
                col['rd'] = i
            elif h.lower() == 'age' or '年龄' in h:
                col['age'] = i
            elif h.lower() == 'gender' or '性别' in h:
                col['gender'] = i

        records = []
        for row in rows[1:]:
            if not row:
                continue
            rec = {f: str(row[ci]).strip() for f, ci in col.items() if ci < len(row) and row[ci]}
            sc = rec.get('sc', '')
            rd = rec.get('rd', '')
            if not sc and not rd:
                continue

            records.append({
                'name': f'{project_code}-RD{rd}' if rd else f'{project_code}-SC{sc}',
                'sc_number': sc,
                'rd_number': rd,
                'project_code': project_code,
                'age': self._safe_int(rec.get('age', '')),
                'gender': self._norm_gender(rec.get('gender', '')),
                'source': 'qualified_list',
                'source_file': fname,
            })
        return records

    def _bulk_upsert_subjects(self, records):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_domain import SkinProfile
        from apps.protocol.models import Protocol
        from django.utils import timezone

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name', '').strip()
            rd = rec.get('rd_number', '')
            sc = rec.get('sc_number', '')
            project_code = rec.get('project_code', '').strip()

            if not name:
                continue

            try:
                cand_no = f'{project_code}-{rd}'[:20] if rd and project_code else ''
                subject = None

                if cand_no:
                    subject = Subject.objects.filter(subject_no=cand_no).first()

                if subject is None and rec.get('phone_last4'):
                    subject = Subject.objects.filter(
                        name=name, phone__endswith=rec['phone_last4']
                    ).first()

                if subject is None:
                    subject = Subject.objects.filter(name=name).first()

                gender = rec.get('gender', '')
                age = rec.get('age')

                if subject is None:
                    subject = Subject(
                        name=name[:100], gender=gender or '', age=age,
                        source_channel='database', status='completed',
                    )
                    if cand_no:
                        subject.subject_no = cand_no
                    if not self.dry_run:
                        subject.save()
                    saved += 1
                else:
                    changed = False
                    if not subject.gender and gender:
                        subject.gender = gender
                        changed = True
                    if not subject.age and age:
                        subject.age = age
                        changed = True
                    if not self.dry_run and changed:
                        subject.save(update_fields=['gender', 'age', 'update_time'])
                    saved += 1

                if self.dry_run:
                    continue

                # SkinProfile
                if rec.get('skin_type') or rec.get('fitzpatrick'):
                    if subject.pk:
                        sp, _ = SkinProfile.objects.get_or_create(subject=subject)
                        changed = False
                        if rec.get('fitzpatrick') and not sp.fitzpatrick_type:
                            sp.fitzpatrick_type = rec['fitzpatrick']
                            changed = True
                        if rec.get('skin_type') and not sp.skin_type_u_zone:
                            sp.skin_type_u_zone = rec['skin_type']
                            changed = True
                        if changed:
                            sp.save(update_fields=['fitzpatrick_type', 'skin_type_u_zone', 'update_time'])

                # 入组关联
                if project_code and subject.pk:
                    if project_code not in proto_cache:
                        proto_cache[project_code] = Protocol.objects.filter(
                            code__iexact=project_code
                        ).first()
                    proto = proto_cache.get(project_code)
                    if proto:
                        from apps.subject.models import EnrollmentStatus
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=proto,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )

                # SC号记录
                if (sc or rd) and project_code and subject.pk:
                    from apps.subject.models_execution import SubjectProjectSC
                    SubjectProjectSC.objects.get_or_create(
                        subject=subject, project_code=project_code,
                        defaults={'sc_number': sc, 'rd_number': rd}
                    )

            except Exception as e:
                logger.debug('upsert 跳过: %s', e)

        return saved

    # ==================================================================
    # IM 招募统计
    # ==================================================================
    def phase_im_recruit_stats(self):
        from apps.secretary.models import PersonalContext
        from apps.knowledge.models import KnowledgeEntry
        import json

        recruit_chats = [
            '特化招募', '组3招募', '组2招募群', '组1 招募沟通', '组7&组9 招募',
            '华山植发-招募重点跟踪群组', '组4-招募沟通群', 'C06招募沟通群',
            '忍者-组3', '头发测试研究团队', '组3运营', '招募天团🤙🏼🤙🏼🤙🏼',
            '压力防脱测试招募渠道扩展沟通',
        ]

        # 数字提取模式
        patterns = {
            'visited': re.compile(r'(?:到访|来访|约访|今天来了?)\D{0,5}(\d+)'),
            'screened': re.compile(r'(?:初筛了?|粗筛了?|筛选了?)\D{0,5}(\d+)'),
            'qualified': re.compile(r'(?:初筛合格|合格了?)\D{0,5}(\d+)'),
            'enrolled': re.compile(r'(?:入组了?|已入组)\D{0,5}(\d+)'),
            'dropout': re.compile(r'(?:脱落了?|退出了?)\D{0,5}(\d+)'),
        }

        ctxs = PersonalContext.objects.filter(
            source_type='im',
            metadata__chat_name__in=recruit_chats,
        )

        daily_stats = []
        project_stats = defaultdict(lambda: defaultdict(int))

        for ctx in ctxs.iterator(chunk_size=1000):
            content = str(ctx.raw_content or '')
            if content.startswith('{') or content.startswith('['):
                try:
                    obj = json.loads(content)
                    content = obj.get('text', '') if isinstance(obj, dict) else ''
                except Exception:
                    pass

            if not any(k in content for k in ['入组', '到访', '合格', '初筛']):
                continue

            projs = [p.upper() for p in PROJECT_P.findall(content)]
            rec = {
                'date': str(ctx.created_at.date()) if ctx.created_at else '',
                'chat': (ctx.metadata or {}).get('chat_name', ''),
            }

            has_data = False
            for field, pat in patterns.items():
                m = pat.search(content)
                if m:
                    try:
                        val = int(m.group(1))
                        if 1 <= val <= 500:
                            rec[field] = val
                            has_data = True
                    except Exception:
                        pass

            if has_data:
                rec['projects'] = projs
                rec['snippet'] = content[:200]
                daily_stats.append(rec)
                for proj in (projs or ['未知项目']):
                    for f in ['visited', 'screened', 'qualified', 'enrolled', 'dropout']:
                        project_stats[proj][f] += rec.get(f, 0)

        total_visited = sum(r.get('visited', 0) for r in daily_stats)
        total_enrolled = sum(r.get('enrolled', 0) for r in daily_stats)
        total_qualified = sum(r.get('qualified', 0) for r in daily_stats)
        total_screened = sum(r.get('screened', 0) for r in daily_stats)
        total_dropout = sum(r.get('dropout', 0) for r in daily_stats)

        self.stdout.write(f'\n有效招募消息: {len(daily_stats)} 条')
        self.stdout.write(f'累计到访: {total_visited:,}')
        self.stdout.write(f'累计初筛: {total_screened:,}')
        self.stdout.write(f'累计合格: {total_qualified:,}')
        self.stdout.write(f'累计入组: {total_enrolled:,}')
        self.stdout.write(f'累计脱落: {total_dropout:,}')

        self.stdout.write('\n按项目汇总 (入组前25):')
        pl = sorted(project_stats.items(), key=lambda x: x[1].get('enrolled', 0), reverse=True)
        for proj, s in pl[:25]:
            if sum(s.values()) > 2:
                self.stdout.write(
                    f"  {proj:15s} 到访:{s.get('visited',0):>5} "
                    f"初筛:{s.get('screened',0):>5} 合格:{s.get('qualified',0):>5} "
                    f"入组:{s.get('enrolled',0):>5} 脱落:{s.get('dropout',0):>5}"
                )

        if not self.dry_run and daily_stats:
            proj_summary = '\n'.join([
                f"{proj}: 到访{s.get('visited',0)} 初筛{s.get('screened',0)} "
                f"合格{s.get('qualified',0)} 入组{s.get('enrolled',0)} 脱落{s.get('dropout',0)}"
                for proj, s in pl[:50] if sum(s.values()) > 2
            ])
            content = (
                f"殷淑雯&王梦丹招募团队 IM 群聊招募统计\n"
                f"分析记录: {len(daily_stats)} 条\n"
                f"累计到访: {total_visited:,}\n"
                f"累计初筛: {total_screened:,}\n"
                f"累计合格: {total_qualified:,}\n"
                f"累计入组: {total_enrolled:,}\n"
                f"累计脱落: {total_dropout:,}\n\n"
                f"=== 按项目统计 ===\n{proj_summary}"
            )
            KnowledgeEntry.objects.update_or_create(
                source_type='recruit_im_stats',
                source_id=1,
                source_key='yin_wang_recruit_stats',
                defaults={
                    'entry_type': 'lesson_learned',
                    'title': '殷淑雯&王梦丹招募团队 IM 统计',
                    'content': content,
                    'status': 'published',
                    'is_published': True,
                }
            )

    # ==================================================================
    # 提取受试者（从已解析数据中）
    # ==================================================================
    def phase_extract_subjects(self):
        import openpyxl

        # 从所有已下载的 Excel 中批量提取受试者
        all_records = []
        count = 0
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith('.xlsx'):
                    continue
                fpath = os.path.join(root, f)
                project_code = self._extract_project(f) or self._extract_project(root)
                try:
                    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                    for sname in wb.sheetnames:
                        ws = wb[sname]
                        rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 300), values_only=True))
                        if not rows:
                            continue
                        recs = self._try_parse_subject_sheet(rows, project_code, f, sname)
                        if recs:
                            all_records.extend(recs)
                    wb.close()
                    count += 1
                except Exception:
                    pass

                if count % 100 == 0 and count > 0:
                    self.stdout.write(f'  已扫描 {count} 个文件，提取 {len(all_records)} 条...')

        self.stdout.write(f'\n总扫描: {count} 个文件，提取 {len(all_records)} 条记录')

        # 去重
        seen = set()
        unique = []
        for rec in all_records:
            key = f"{rec.get('name','')[:20]}_{rec.get('project_code','')}"
            if key not in seen and key != '_':
                seen.add(key)
                unique.append(rec)

        self.stdout.write(f'去重后: {len(unique)} 条')

        if not self.dry_run and unique:
            saved = self._bulk_upsert_subjects(unique)
            self.stdout.write(self.style.SUCCESS(f'写入 {saved} 条受试者记录'))

    # ==================================================================
    # 生成知识档案
    # ==================================================================
    def phase_build_profiles(self):
        from apps.subject.models import Subject
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'受试者总数: {total:,}')

        created = updated = 0
        for i in range(0, total, 500):
            for subject in subjects[i:i + 500]:
                text = self._build_profile_text(subject)
                if not text:
                    continue
                if self.dry_run:
                    created += 1
                    continue
                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_full_lifecycle',
                    source_id=subject.id,
                    source_key=f'lifecycle_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 受试者全档案',
                        'content': text,
                        'status': 'published',
                        'is_published': True,
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            self.stdout.write(f'  进度: {min(i + 500, total)}/{total}')

        self.stdout.write(self.style.SUCCESS(f'档案: 新建 {created}, 更新 {updated}'))

    def _build_profile_text(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord
        from apps.subject.models_domain import SkinProfile

        lines = [
            f'受试者: {subject.name}',
            f'编号: {subject.subject_no or subject.id}',
        ]

        if subject.gender or subject.age:
            lines.append(f'基本: {subject.get_gender_display() or "?"}, {subject.age or "?"}岁')

        skin = SkinProfile.objects.filter(subject=subject).first()
        if skin:
            parts = []
            if skin.fitzpatrick_type:
                parts.append(f'Fitzpatrick {skin.fitzpatrick_type}')
            if skin.skin_type_u_zone:
                parts.append(f'{skin.skin_type_u_zone}肤质')
            if parts:
                lines.append(f'皮肤: {", ".join(parts)}')

        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        if enrollments.exists():
            projs = [e.protocol.code for e in enrollments[:6] if e.protocol]
            lines.append(f'项目({enrollments.count()}): {", ".join(projs)}')

        m = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m:
            lines.append(f'仪器测量: {m}次')

        q = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q:
            lines.append(f'问卷: {q}份')

        c = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if c:
            lines.append(f'依从性: {c.get_level_display()}')

        return '\n'.join(lines) if len(lines) > 2 else None

    # ==================================================================
    # 工具方法
    # ==================================================================
    def _extract_project(self, text):
        m = PROJECT_P.search(str(text))
        return m.group(0).upper() if m else ''

    def _safe_int(self, v):
        try:
            return int(re.sub(r'[^\d]', '', str(v))) if v else None
        except Exception:
            return None

    def _norm_gender(self, v):
        s = str(v).strip().lower()
        if '女' in s or s in ('f', 'female', '2'):
            return 'female'
        if '男' in s or s in ('m', 'male', '1'):
            return 'male'
        return ''
