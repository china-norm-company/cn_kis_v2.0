"""
大规模受试者名称挖掘 — 从所有兼职人员明细提取唯一姓名

策略：从磁盘上的所有结算/劳务文件提取中文姓名，
建立去重的受试者记录，目标总量 50,000+

Usage:
  python manage.py bulk_create_subjects_from_excel --phase all
  python manage.py bulk_create_subjects_from_excel --dry-run
"""
import os
import re
import logging

from django.core.management.base import BaseCommand
from django.db import connection

logger = logging.getLogger(__name__)
MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')

# 中文姓名正则（2-5个汉字）
NAME_P = re.compile(r'^[\u4e00-\u9fff·]{2,5}$')
# 有效姓名：不是常见词汇
INVALID_NAMES = {
    '合计', '小计', '总计', '序号', '姓名', '备注', '合同', '兼职', '项目', '结算', '金额',
    '实际', '应发', '发放', '签名', '联系', '银行', '卡号', '证号', '明细', '汇总',
    '负责', '提交', '确认', '审核', '备注', '工作', '正式', '兼职', '全职', '外包',
    '当月', '上月', '本月', '年度', '季度', '周期', '总计', '小计', '合计',
}
PROJECT_P = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)


class Command(BaseCommand):
    help = '从所有结算/劳务Excel批量创建受试者记录（目标50000+）'

    def add_arguments(self, parser):
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--phase', default='all',
                            choices=['all', 'scan', 'insert', 'vectorize'])
        parser.add_argument('--batch-size', type=int, default=500)

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.batch_size = options['batch_size']
        phase = options['phase']

        phases = ['scan', 'insert', 'vectorize'] if phase == 'all' else [phase]
        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}')()

    def phase_scan(self):
        """扫描所有Excel，收集唯一(姓名, 项目, 兼职编号)三元组"""
        import openpyxl

        self.stdout.write('扫描所有相关Excel文件...')
        records = {}  # key: 姓名, value: {projects, member_ids, amounts}

        scan_kws = [
            '兼职人员明细', '兼职劳务费', '劳务报酬', '员工受试者结算',
            '受试者银行卡', '境内人员信息', '复硕正态众包', '灵工结算',
            '发放明细', '月度发放', '银行卡信息', '新增银行卡',
        ]

        file_count = 0
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if not any(kw in f for kw in scan_kws):
                    continue
                fpath = os.path.join(root, f)
                project_code = self._extract_project(f) or self._extract_project(root)

                try:
                    if f.endswith('.xls'):
                        extracted = self._scan_xls(fpath, project_code)
                    else:
                        extracted = self._scan_xlsx(fpath, project_code, openpyxl)

                    for rec in extracted:
                        name = rec['name']
                        if name not in records:
                            records[name] = {
                                'name': name,
                                'projects': set(),
                                'member_ids': set(),
                                'phones': set(),
                                'id_cards': set(),
                            }
                        if rec.get('project'):
                            records[name]['projects'].add(rec['project'])
                        if rec.get('member_id'):
                            records[name]['member_ids'].add(rec['member_id'])
                        if rec.get('phone'):
                            records[name]['phones'].add(rec['phone'])
                        if rec.get('id_card'):
                            records[name]['id_cards'].add(rec['id_card'])

                    file_count += 1
                    if file_count % 50 == 0:
                        self.stdout.write(f'  已扫描 {file_count} 个文件, 唯一姓名: {len(records):,}')

                except Exception as e:
                    logger.debug('扫描失败 %s: %s', f, e)

        self.stdout.write(f'\n扫描完成: {file_count} 个文件, 唯一姓名: {len(records):,}')

        # 缓存到临时表或内存
        self._cached_records = list(records.values())

        c = connection.cursor()
        c.execute("SELECT COUNT(*) FROM t_subject WHERE is_deleted=FALSE")
        current = c.fetchone()[0]
        self.stdout.write(f'当前受试者: {current:,}, 新候选: {len(self._cached_records):,}')
        self.stdout.write(f'写入后预计: {current + max(0, len(self._cached_records) - current):,}')

    def phase_insert(self):
        """批量写入受试者记录"""

        if not hasattr(self, '_cached_records'):
            self.phase_scan()

        from apps.subject.models import Subject
        from apps.protocol.models import Protocol

        records = self._cached_records
        self.stdout.write(f'准备写入 {len(records):,} 名受试者...')

        proto_cache = {}
        created = updated = 0

        for i in range(0, len(records), self.batch_size):
            batch = records[i:i + self.batch_size]

            for rec in batch:
                name = rec['name']
                phones = rec.get('phones', set())
                id_cards = rec.get('id_cards', set())
                projects = rec.get('projects', set())
                member_ids = rec.get('member_ids', set())

                if self.dry_run:
                    created += 1
                    continue

                try:
                    subject = None

                    # 匹配顺序: 手机 → 姓名
                    for phone in phones:
                        if phone:
                            subject = Subject.objects.filter(phone=phone[:20]).first()
                            if subject:
                                break

                    if subject is None:
                        subject = Subject.objects.filter(name=name).first()

                    if subject is None:
                        first_phone = next(iter(phones), '') if phones else ''
                        first_id = next(iter(id_cards), '') if id_cards else ''
                        subject = Subject(
                            name=name[:100],
                            phone=first_phone[:20] if first_phone else '',
                            id_card_encrypted=first_id[:50] if first_id else '',
                            source_channel='database',
                            status='completed',
                        )
                        subject.save()
                        created += 1
                    else:
                        changed = False
                        if not subject.phone and phones:
                            subject.phone = next(iter(phones), '')[:20]
                            changed = True
                        if not subject.id_card_encrypted and id_cards:
                            subject.id_card_encrypted = next(iter(id_cards), '')[:50]
                            changed = True
                        if changed:
                            subject.save(update_fields=['phone', 'id_card_encrypted', 'update_time'])
                        updated += 1

                    # 关联项目
                    for proj_code in list(projects)[:3]:
                        if not proj_code:
                            continue
                        if proj_code not in proto_cache:
                            proto_cache[proj_code] = Protocol.objects.filter(
                                code__iexact=proj_code
                            ).first()
                        proto = proto_cache.get(proj_code)
                        if proto and subject.pk:
                            from apps.subject.models import Enrollment, EnrollmentStatus
                            from django.utils import timezone
                            Enrollment.objects.get_or_create(
                                subject=subject, protocol=proto,
                                defaults={
                                    'status': EnrollmentStatus.COMPLETED,
                                    'enrolled_at': timezone.now(),
                                }
                            )

                except Exception as e:
                    logger.debug('insert 跳过 %s: %s', name, e)

            if (i + self.batch_size) % 5000 == 0:
                self.stdout.write(f'  进度: {i + self.batch_size}/{len(records)}, 新建:{created}, 更新:{updated}')

        c = connection.cursor()
        c.execute("SELECT COUNT(*) FROM t_subject WHERE is_deleted=FALSE")
        total_now = c.fetchone()[0]
        self.stdout.write(self.style.SUCCESS(
            f'\n新建: {created}, 更新: {updated}\n受试者总数: {total_now:,}'
        ))

    def phase_vectorize(self):
        """为所有受试者生成全景知识档案"""
        from apps.subject.models import Subject
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'向量化 {total:,} 名受试者...')

        created = updated = 0
        for i in range(0, total, 1000):
            batch = subjects[i:i + 1000]
            entries = []

            for subject in batch:
                text = self._build_text(subject)
                if not text:
                    continue
                entries.append((subject.id, subject.name, text))

            if not self.dry_run:
                for sid, sname, text in entries:
                    _, is_new = KnowledgeEntry.objects.update_or_create(
                        source_type='subject_full_lifecycle',
                        source_id=sid,
                        source_key=f'lifecycle_{sid}',
                        defaults={
                            'entry_type': 'lesson_learned',
                            'title': f'{sname} 全档案',
                            'content': text,
                            'status': 'published',
                            'is_published': True,
                            'index_status': 'pending',
                        }
                    )
                    if is_new:
                        created += 1
                    else:
                        updated += 1
            else:
                created += len(entries)

            if (i + 1000) % 10000 == 0:
                self.stdout.write(f'  进度: {i + 1000}/{total}')

        self.stdout.write(self.style.SUCCESS(f'向量化: 新建 {created}, 更新 {updated}'))

        # 触发向量化索引
        if not self.dry_run:
            c = connection.cursor()
            c.execute("""
            UPDATE t_knowledge_entry
            SET index_status='pending'
            WHERE source_type='subject_full_lifecycle'
            AND index_status != 'indexed'
            """)
            self.stdout.write('已标记向量化队列')

    def _build_text(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord, SubjectPayment
        from apps.subject.models_domain import SkinProfile

        lines = [f'受试者: {subject.name}']

        demo = []
        if subject.gender:
            demo.append(subject.get_gender_display())
        if subject.age:
            demo.append(f'{subject.age}岁')
        if demo:
            lines.append(f'基本: {", ".join(demo)}')

        if subject.phone:
            lines.append(f'手机尾号: {subject.phone[-4:]}')

        skin = SkinProfile.objects.filter(subject=subject).first()
        if skin:
            parts = []
            if skin.fitzpatrick_type:
                parts.append(f'Fitzpatrick {skin.fitzpatrick_type}型')
            if skin.skin_type_u_zone:
                parts.append(f'{skin.skin_type_u_zone}肤质')
            if parts:
                lines.append(f'皮肤: {", ".join(parts)}')

        enrs = Enrollment.objects.filter(subject=subject).select_related('protocol')
        n_enr = enrs.count()
        if n_enr:
            projs = [e.protocol.code for e in enrs[:5] if e.protocol]
            lines.append(f'参与项目({n_enr}): {", ".join(projs)}')

        m = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m:
            lines.append(f'测量记录: {m}次')

        q = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q:
            lines.append(f'问卷: {q}份')

        comp = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if comp:
            lines.append(f'依从性: {comp.get_level_display()}')

        pays = SubjectPayment.objects.filter(subject=subject, status='paid').count()
        if pays:
            lines.append(f'结算次数: {pays}')

        return '\n'.join(lines) if len(lines) >= 2 else None

    def _scan_xlsx(self, fpath, project_code, openpyxl):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        results = []
        for ws in wb.worksheets:
            sheet_project = self._extract_project(ws.title) or project_code
            rows = list(ws.iter_rows(max_row=min(ws.max_row or 0, 3000), values_only=True))
            if not rows:
                continue

            # 找姓名列（通常在前3列）
            name_col = phone_col = id_col = proj_col = mid_col = None
            for header_idx in range(min(3, len(rows))):
                header = [str(c).strip() if c else '' for c in rows[header_idx]]
                for i, h in enumerate(header[:15]):
                    if h == '姓名' or (('姓名' in h) and name_col is None):
                        name_col = i
                    elif '手机' in h and phone_col is None:
                        phone_col = i
                    elif '身份证' in h and id_col is None:
                        id_col = i
                    elif '项目编号' in h and proj_col is None:
                        proj_col = i
                    elif '兼职编号' in h and mid_col is None:
                        mid_col = i
                if name_col is not None:
                    break

            if name_col is None:
                # 尝试第二列作为姓名列
                for row in rows[2:5]:
                    if row and len(row) > 1 and row[1]:
                        val = str(row[1]).strip()
                        if NAME_P.match(val) and val not in INVALID_NAMES:
                            name_col = 1
                            break

            if name_col is None:
                continue

            header_idx_final = 0
            for i, row in enumerate(rows[:5]):
                if row and name_col < len(row) and row[name_col]:
                    val = str(row[name_col]).strip()
                    if val == '姓名' or NAME_P.match(val):
                        header_idx_final = i
                        break

            for row in rows[header_idx_final + 1:]:
                if not row or name_col >= len(row) or not row[name_col]:
                    continue
                name = str(row[name_col]).strip()
                if not NAME_P.match(name) or name in INVALID_NAMES:
                    continue

                rec = {'name': name}
                if phone_col is not None and phone_col < len(row) and row[phone_col]:
                    phone_raw = re.sub(r'[^\d]', '', str(row[phone_col]))
                    if len(phone_raw) == 11 and phone_raw.startswith('1'):
                        rec['phone'] = phone_raw
                if id_col is not None and id_col < len(row) and row[id_col]:
                    id_raw = re.sub(r'[\s\-]', '', str(row[id_col]))
                    if re.match(r'^\d{17}[\dXx]$', id_raw):
                        rec['id_card'] = id_raw[:6] + '****' + id_raw[-4:]
                if proj_col is not None and proj_col < len(row) and row[proj_col]:
                    proj = self._extract_project(str(row[proj_col]))
                    if proj:
                        rec['project'] = proj
                elif sheet_project:
                    rec['project'] = sheet_project
                if mid_col is not None and mid_col < len(row) and row[mid_col]:
                    mid = str(row[mid_col]).strip()
                    if re.match(r'^\d{3,6}$', mid):
                        rec['member_id'] = mid

                results.append(rec)

        wb.close()
        return results

    def _scan_xls(self, fpath, project_code):
        try:
            import xlrd
        except ImportError:
            return []
        try:
            wb = xlrd.open_workbook(fpath)
        except Exception:
            return []

        results = []
        for ws in wb.sheets():
            sheet_project = self._extract_project(ws.name) or project_code
            rows = [ws.row_values(i) for i in range(min(ws.nrows, 3000))]

            name_col = None
            for hi in range(min(3, len(rows))):
                header = [str(c).strip() for c in rows[hi]]
                for i, h in enumerate(header[:15]):
                    if h == '姓名' or '姓名' in h:
                        name_col = i
                        break
                if name_col is not None:
                    break
            if name_col is None:
                name_col = 1

            for row in rows[2:]:
                if not row or name_col >= len(row) or not row[name_col]:
                    continue
                name = str(row[name_col]).strip()
                if not NAME_P.match(name) or name in INVALID_NAMES:
                    continue
                rec = {'name': name, 'project': sheet_project}
                results.append(rec)

        return results

    def _extract_project(self, text):
        m = PROJECT_P.search(str(text))
        return m.group(0).upper() if m else ''
