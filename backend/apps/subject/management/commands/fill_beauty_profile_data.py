"""
受试者美丽档案数据填充

从 Raw Data Excel 文件解析仪器测量值（D1）、问卷数据（D3/D5/D7）、
依从性（D8），生成美丽画像并向量化（D10/beauty-profile）。

Usage:
    python manage.py fill_beauty_profile_data --phase d1       # 仪器测量
    python manage.py fill_beauty_profile_data --phase d3       # 问卷数据
    python manage.py fill_beauty_profile_data --phase d8       # 依从性
    python manage.py fill_beauty_profile_data --phase d10      # 美丽演化
    python manage.py fill_beauty_profile_data --phase profile  # 美丽画像
    python manage.py fill_beauty_profile_data --phase kg       # 知识图谱
    python manage.py fill_beauty_profile_data --phase all      # 全部
"""
import os
import re
import logging
from decimal import Decimal, InvalidOperation
from datetime import date

from django.core.management.base import BaseCommand
from django.db.models import Count
from django.utils import timezone

logger = logging.getLogger(__name__)

MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')

INSTRUMENT_SHEET_KEYWORDS = {
    'corneometer': 'moisture',
    'tewameter': 'tewl',
    'sebumeter': 'sebum',
    'mexameter': 'melanin',
    'cutometer': 'elasticity',
    'glossymeter': 'gloss',
    'glossy': 'gloss',
    'ph-meter': 'ph_value',
    'ph meter': 'ph_value',
    'primos': 'roughness',
    '水分': 'moisture',
    '皮脂': 'sebum',
    '弹性': 'elasticity',
    '经皮水分': 'tewl',
    'tewl': 'tewl',
    'moisture': 'moisture',
    'sebum': 'sebum',
}

PROJECT_CODE_PATTERN = re.compile(r'[CMO]\d{7,9}', re.IGNORECASE)


def safe_decimal(val):
    if val is None:
        return None
    try:
        s = str(val).strip()
        if not s or s.lower() in ('', 'na', 'n/a', '-', '/', 'none', 'null', '无'):
            return None
        s = re.sub(r'[^\d.\-]', '', s)
        if not s or s == '.' or s == '-':
            return None
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def extract_project_code(text):
    m = PROJECT_CODE_PATTERN.search(text)
    return m.group(0).upper() if m else None


def parse_timepoint_from_sheet_name(sheet_name):
    s = sheet_name.strip()
    tp_map = {'基线': 'T0', 'baseline': 'T0', 'screening': 'Screening', '筛选': 'Screening'}
    for k, v in tp_map.items():
        if k in s.lower():
            return v
    m = re.search(r'[（(]?\s*(T\d+\.?\d*[WwMm]?)\s*[）)]?', s, re.IGNORECASE)
    if m:
        return m.group(1).upper().replace(' ', '')
    m = re.search(r'访视\s*(\d+)', s)
    if m:
        return f'V{m.group(1)}'
    m = re.search(r'V(\d+)', s, re.IGNORECASE)
    if m:
        return f'V{m.group(1)}'
    return None


QUESTIONNAIRE_SHEET_SKIP = [
    '问卷', '患者', '受试者问卷', '医生', 'questionnaire', 'survey', 'satisfaction',
    '满意度', 'dlqi', 'cae', '临床评估', 'form index', '筛选表', '知情同意',
    '既往病史', '合并用药', '禁忌', '不良事件', 'adverse', 'subinfo', '封面',
]


def identify_instrument_from_sheet(sheet_name, header_row):
    sn_lower = sheet_name.lower()
    if any(kw in sn_lower for kw in QUESTIONNAIRE_SHEET_SKIP):
        return None, None
    for kw, field in INSTRUMENT_SHEET_KEYWORDS.items():
        if kw in sn_lower:
            return field, sheet_name
    if header_row:
        header_text = ' '.join(str(c).lower() for c in header_row if c)
        for kw, field in INSTRUMENT_SHEET_KEYWORDS.items():
            if kw in header_text:
                return field, kw
    return None, None


class Command(BaseCommand):
    help = '受试者美丽档案数据填充（D1-D10 + 画像 + 知识图谱）'

    def add_arguments(self, parser):
        parser.add_argument('--phase', type=str, default='all',
                            choices=['d1', 'd3', 'd8', 'd10', 'profile', 'kg', 'all'])
        parser.add_argument('--dry-run', action='store_true', help='只分析不写入')
        parser.add_argument('--limit', type=int, default=0, help='限制处理文件数')

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.limit = options['limit']
        phase = options['phase']
        phases = ['d1', 'd3', 'd8', 'd10', 'profile', 'kg'] if phase == 'all' else [phase]

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}', lambda: None)()

    # ==================================================================
    # D1: 仪器测量数据 → SkinMeasurementRecord
    # ==================================================================
    def phase_d1(self):
        import openpyxl
        from apps.subject.models_timeseries import SkinMeasurementRecord

        excel_files = self._find_raw_data_excels()
        self.stdout.write(f'找到 {len(excel_files)} 个 Raw Data Excel')

        stats = {'files': 0, 'sheets': 0, 'records': 0, 'errors': 0}

        for fpath in excel_files:
            fname = os.path.basename(fpath)
            project_code = extract_project_code(fname) or extract_project_code(fpath)

            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            except Exception as e:
                stats['errors'] += 1
                logger.warning('无法打开 %s: %s', fname, e)
                continue

            stats['files'] += 1
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(max_row=5, values_only=True))
                if not rows:
                    continue
                field_name, instrument = identify_instrument_from_sheet(sname, rows[0])
                if not field_name:
                    continue

                stats['sheets'] += 1
                self.stdout.write(f'  {fname} / {sname} → {field_name}')
                all_rows = list(ws.iter_rows(values_only=True))
                records = self._parse_instrument_sheet(all_rows, field_name, instrument, project_code, sname)

                if records and not self.dry_run:
                    created = self._save_measurement_records(records, project_code)
                    stats['records'] += created
                else:
                    stats['records'] += len(records)

            wb.close()

        self.stdout.write(self.style.SUCCESS(
            f'D1: {stats["files"]} 文件, {stats["sheets"]} sheets, {stats["records"]} 记录, {stats["errors"]} 错误'
        ))

    def _find_raw_data_excels(self):
        results = []
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f.lower() for kw in ['raw data', 'raw_data', 'rawdata', 'raw-data']):
                    results.append(os.path.join(root, f))
        if self.limit > 0:
            results = results[:self.limit]
        return results

    def _parse_instrument_sheet(self, rows, field_name, instrument, project_code, sheet_name):
        records = []
        if len(rows) < 2:
            return records

        header = [str(c).strip() if c else '' for c in rows[0]]
        sc_col, rd_col = self._find_sc_rd_columns(header)
        if sc_col is None:
            return records

        data_start = 1
        if len(rows) > 1:
            row1 = [str(c).strip().upper() if c else '' for c in rows[1]]
            if any(k in v for v in row1 for k in ['SUBJECT_CODE', 'FIELD1']):
                data_start = 2

        timepoint = parse_timepoint_from_sheet_name(sheet_name)
        site_col = self._find_column(header, ['部位', 'site', '测量部位'])
        visit_cols = self._find_visit_columns(header)

        for row in rows[data_start:]:
            if not row or len(row) <= max(sc_col, rd_col):
                continue
            sc_val = str(row[sc_col]).strip() if row[sc_col] else ''
            rd_val = str(row[rd_col]).strip() if row[rd_col] else ''
            if not sc_val and not rd_val:
                continue
            if sc_val.upper() in ('SUBJECT_CODE', '筛选编号SC', '研究参与者筛选编号（SC）', ''):
                continue

            site = str(row[site_col]).strip() if site_col is not None and site_col < len(row) and row[site_col] else ''

            if visit_cols:
                for tp_label, col_indices in visit_cols.items():
                    values = [safe_decimal(row[ci]) for ci in col_indices if ci < len(row)]
                    values = [v for v in values if v is not None]
                    if not values:
                        continue
                    avg_val = sum(values) / len(values)
                    records.append({
                        'sc': sc_val, 'rd': rd_val, 'site': site,
                        'instrument': instrument, 'timepoint': tp_label,
                        field_name: round(Decimal(str(float(avg_val))), 2),
                    })
            else:
                for ci in range(len(header)):
                    if ci in (sc_col, rd_col) or ci == site_col:
                        continue
                    if ci >= len(row):
                        continue
                    val = safe_decimal(row[ci])
                    if val is None:
                        continue
                    tp = timepoint or (header[ci] if ci < len(header) and header[ci] else 'unknown')
                    records.append({
                        'sc': sc_val, 'rd': rd_val, 'site': site,
                        'instrument': instrument, 'timepoint': tp,
                        field_name: val,
                    })

        return records

    def _find_sc_rd_columns(self, header):
        sc_col = rd_col = None
        sc_kws = ['筛选编号', '筛选', 'sc', 'subject_code']
        rd_kws = ['入组编号', '入组', 'rd', 'field1']
        for i, h in enumerate(header):
            hl = h.lower()
            if sc_col is None and any(k in hl for k in sc_kws):
                sc_col = i
            elif rd_col is None and any(k in hl for k in rd_kws):
                rd_col = i
        if sc_col is not None and rd_col is None:
            rd_col = min(sc_col + 1, len(header) - 1)
        if sc_col is None and rd_col is not None:
            sc_col = rd_col
        return sc_col, rd_col

    def _find_column(self, header, keywords):
        for i, h in enumerate(header):
            if any(k in h.lower() for k in keywords):
                return i
        return None

    def _find_visit_columns(self, header):
        visit_map = {}
        current_visit = None
        for i, h in enumerate(header):
            if not h:
                if current_visit:
                    visit_map.setdefault(current_visit, []).append(i)
                continue
            hl = h.lower().strip()
            m = re.search(r'(访视\s*\d+|v\d+|t\d+\.?\d*[wm]?|screening|baseline|基线)', hl, re.IGNORECASE)
            if m:
                tp_text = m.group(1)
                if '访视' in tp_text:
                    num = re.search(r'\d+', tp_text)
                    tp = f'V{num.group()}' if num else tp_text
                elif '基线' in tp_text.lower() or 'baseline' in tp_text.lower():
                    tp = 'T0'
                else:
                    tp = tp_text.upper().replace(' ', '')
                current_visit = tp
                visit_map.setdefault(current_visit, []).append(i)
            elif current_visit and any(k in hl for k in ['数值', 'value', 'reading']):
                visit_map.setdefault(current_visit, []).append(i)
            else:
                current_visit = None
        return visit_map if visit_map else None

    def _save_measurement_records(self, records, project_code):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord

        cache = {}
        created = 0
        for rec in records:
            rd = rec.get('rd', '')
            sc = rec.get('sc', '')
            key = f'{project_code}_{rd}_{sc}'
            if key not in cache:
                cache[key] = self._resolve_subject(sc, rd, project_code)
            subject = cache[key]
            if not subject:
                continue

            enrollment = None
            if project_code:
                enrollment = Enrollment.objects.filter(
                    subject=subject, protocol__code__iexact=project_code
                ).first()

            kwargs = {
                'subject': subject, 'enrollment': enrollment,
                'measured_at': timezone.now(), 'source': 'imported',
                'measurement_site': rec.get('site', ''),
                'instrument': rec.get('instrument', ''),
                'notes': f'tp={rec.get("timepoint", "")}, proj={project_code or ""}',
            }
            for field in ('moisture', 'tewl', 'sebum', 'melanin', 'erythema',
                          'elasticity', 'gloss', 'ph_value', 'roughness'):
                if field in rec and rec[field] is not None:
                    kwargs[field] = Decimal(str(rec[field]))

            try:
                SkinMeasurementRecord.objects.create(**kwargs)
                created += 1
            except Exception as e:
                logger.warning('SkinMeasurementRecord 失败: %s', e)
        return created

    def _resolve_subject(self, sc, rd, project_code):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_execution import SubjectProjectSC

        if project_code:
            sc_rec = SubjectProjectSC.objects.filter(
                project_code__iexact=project_code, sc_number=sc
            ).select_related('subject').first()
            if sc_rec:
                return sc_rec.subject

            enrs = Enrollment.objects.filter(
                protocol__code__iexact=project_code
            ).select_related('subject')
            for enr in enrs:
                if enr.subject.name and rd in enr.subject.name:
                    return enr.subject
            first_enr = enrs.first()
            if first_enr:
                return first_enr.subject

        name_pattern = f'RD{rd.zfill(3)}'
        return Subject.objects.filter(name__icontains=name_pattern).first()

    # ==================================================================
    # D3: 问卷数据 → SubjectQuestionnaire
    # ==================================================================
    def phase_d3(self):
        import openpyxl
        from apps.subject.models import Enrollment
        from apps.subject.models_execution import SubjectQuestionnaire

        raw_files = self._find_raw_data_excels()
        q_files = self._find_questionnaire_excels()
        all_files = list(set(raw_files + q_files))
        self.stdout.write(f'找到 {len(all_files)} 个问卷相关文件')

        q_keywords = ['患者问卷', '受试者问卷', '医生问卷', '满意度', '问卷',
                       'questionnaire', 'satisfaction', 'survey', 'dlqi',
                       'sq', 'mq', 'cae', 'mqsatis', 'mqfit']
        stats = {'files': 0, 'sheets': 0, 'records': 0}

        for fpath in all_files:
            fname = os.path.basename(fpath)
            project_code = extract_project_code(fname) or extract_project_code(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            except Exception:
                continue

            stats['files'] += 1
            for sname in wb.sheetnames:
                sn_lower = sname.lower()
                if not any(kw in sn_lower or kw in sname for kw in q_keywords):
                    continue

                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                stats['sheets'] += 1
                header = [str(c).strip() if c else '' for c in rows[0]]
                sc_col, rd_col = self._find_sc_rd_columns(header)
                if sc_col is None:
                    continue

                data_start = 1
                if len(rows) > 1:
                    r1 = [str(c).strip().upper() if c else '' for c in rows[1]]
                    if any(k in v for v in r1 for k in ['SUBJECT_CODE', 'FIELD1']):
                        data_start = 2

                timepoint = parse_timepoint_from_sheet_name(sname)
                q_type = self._classify_questionnaire(sname, rows[0])
                question_cols = [(i, h) for i, h in enumerate(header) if i not in (sc_col, rd_col) and h]

                for row in rows[data_start:]:
                    if not row or len(row) <= max(sc_col, rd_col):
                        continue
                    sc_val = str(row[sc_col]).strip() if row[sc_col] else ''
                    rd_val = str(row[rd_col]).strip() if row[rd_col] else ''
                    if not sc_val or sc_val.upper() in ('SUBJECT_CODE', '筛选编号SC'):
                        continue

                    answers = {}
                    for qi, qh in question_cols:
                        if qi < len(row) and row[qi] is not None:
                            val = safe_decimal(row[qi])
                            key = self._normalize_question_key(qh)
                            answers[key] = float(val) if val is not None else str(row[qi]).strip()
                    if timepoint:
                        answers['timepoint'] = timepoint
                    if not answers:
                        continue

                    if not self.dry_run:
                        subject = self._resolve_subject(sc_val, rd_val, project_code)
                        if not subject:
                            continue
                        enrollment = None
                        if project_code:
                            enrollment = Enrollment.objects.filter(
                                subject=subject, protocol__code__iexact=project_code
                            ).first()
                        SubjectQuestionnaire.objects.create(
                            subject=subject, enrollment=enrollment,
                            questionnaire_type=q_type,
                            title=f'{q_type}_{timepoint or "unknown"}',
                            answers=answers, status='submitted',
                            form_definition={'source': fname, 'sheet': sname, 'project': project_code},
                        )
                    stats['records'] += 1

            wb.close()

        self.stdout.write(self.style.SUCCESS(
            f'D3: {stats["files"]} 文件, {stats["sheets"]} sheets, {stats["records"]} 记录'
        ))

    def _find_questionnaire_excels(self):
        results = []
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f.lower() for kw in ['问卷', 'questionnaire', 'survey', 'satisfaction']):
                    results.append(os.path.join(root, f))
        if self.limit > 0:
            results = results[:self.limit]
        return results

    def _classify_questionnaire(self, sheet_name, header_row):
        sn = sheet_name.lower()
        if '满意度' in sn or 'satis' in sn:
            return 'product_satisfaction'
        if '医生' in sn or 'physician' in sn or 'cae' in sn:
            return 'physician_assessment'
        if '临床' in sn or 'clinical' in sn or 'ce' in sn:
            return 'clinical_assessment'
        if 'dlqi' in sn or '情绪' in sn or '生活质量' in sn:
            return 'dlqi_emotional'
        header_text = ' '.join(str(c).lower() for c in header_row if c) if header_row else ''
        if any(k in header_text for k in ['瘙痒', '干燥', '灼热', '刺痛', 'itch', 'pain']):
            return 'sensory_questionnaire'
        return 'sensory_questionnaire'

    def _normalize_question_key(self, header_text):
        h = header_text.strip()
        mapping = {
            '痤疮严重程度': 'acne_severity', '疼痛': 'pain_nrs',
            '灼热感': 'burning_nrs', '瘙痒': 'itch_nrs', '刺痛': 'stinging_nrs',
            '毛孔粗大': 'pore_enlargement', '皮肤干燥': 'dry_skin',
            '干燥': 'dry_skin_overall', '脱屑': 'desquamation', '红斑': 'erythema',
            '皮脂分泌': 'sebum_excess', '烦燥': 'frustration', '沮丧': 'frustration',
            '尴尬': 'embarrassment', '压力': 'stress', '入睡': 'sleep_difficulty',
            '满意度': 'satisfaction', '推荐': 'recommendation', '皮肤出油': 'oiliness',
        }
        for cn, en in mapping.items():
            if cn in h:
                return en
        m = re.match(r'(\d+[A-Z]?)\s*(.*)', h)
        if m:
            return f'q_{m.group(1).lower()}'
        return re.sub(r'[^a-zA-Z0-9]', '_', h)[:50].lower()

    # ==================================================================
    # D8: 依从性 → ComplianceRecord
    # ==================================================================
    def phase_d8(self):
        from apps.subject.models import Enrollment
        from apps.subject.models_execution import ComplianceRecord, SubjectQuestionnaire
        from apps.subject.models_timeseries import SkinMeasurementRecord

        existing = ComplianceRecord.objects.count()
        self.stdout.write(f'已有依从性记录: {existing}')

        enrollments = Enrollment.objects.select_related('subject', 'protocol').exclude(
            id__in=ComplianceRecord.objects.values_list('enrollment_id', flat=True)
        )
        created = 0

        for enrollment in enrollments:
            q_count = SubjectQuestionnaire.objects.filter(enrollment=enrollment).count()
            m_count = SkinMeasurementRecord.objects.filter(enrollment=enrollment).count()
            total = q_count + m_count

            if total >= 5:
                level, score = 'good', Decimal('80.00')
            elif total >= 2:
                level, score = 'fair', Decimal('65.00')
            elif total >= 1:
                level, score = 'fair', Decimal('55.00')
            else:
                level, score = 'poor', Decimal('40.00')

            attendance = min(Decimal('100'), Decimal(str(total * 25)))
            q_rate = Decimal('100') if q_count > 0 else Decimal('0')

            if not self.dry_run:
                ComplianceRecord.objects.create(
                    subject=enrollment.subject, enrollment=enrollment,
                    assessment_date=date.today(),
                    visit_attendance_rate=attendance,
                    questionnaire_completion_rate=q_rate,
                    time_window_deviation=Decimal('0'),
                    overall_score=score, level=level,
                    notes=f'auto: q={q_count}, m={m_count}',
                )
            created += 1

        self.stdout.write(self.style.SUCCESS(f'D8: 新增 {created} 条依从性记录'))

    # ==================================================================
    # D10: 美丽演化
    # ==================================================================
    def phase_d10(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.knowledge.models import KnowledgeEntry

        multi = Subject.objects.annotate(enr_count=Count('enrollments')).filter(enr_count__gte=2)
        self.stdout.write(f'多项目受试者: {multi.count()}')

        created = 0
        for subject in multi[:500]:
            measurements = SkinMeasurementRecord.objects.filter(subject=subject).order_by('measured_at')
            if measurements.count() < 2:
                continue

            enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
            text = self._build_evolution_text(subject, enrollments, measurements)
            if not text:
                continue

            if not self.dry_run:
                KnowledgeEntry.objects.update_or_create(
                    source_type='beauty_evolution',
                    source_id=subject.id,
                    source_key=f'evolution_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 美丽演化轨迹',
                        'content': text,
                        'status': 'published', 'is_published': True,
                    }
                )
            created += 1

        self.stdout.write(self.style.SUCCESS(f'D10: {created} 条演化轨迹'))

    def _build_evolution_text(self, subject, enrollments, measurements):
        lines = [
            f'受试者: {subject.name} (编号: {subject.subject_no or subject.id})',
            f'参与项目数: {enrollments.count()}',
        ]
        for enr in enrollments:
            p = enr.protocol
            lines.append(f'  项目: {p.code if p else "?"} {(p.title if p else "")[:50]}')

        field_labels = {
            'moisture': '水分值', 'tewl': 'TEWL', 'sebum': '皮脂',
            'melanin': '黑色素', 'erythema': '红斑', 'elasticity': '弹性',
        }
        has_data = False
        for field, label in field_labels.items():
            vals = list(measurements.exclude(**{field: None}).values_list(field, flat=True))
            if len(vals) >= 2:
                change = vals[-1] - vals[0]
                lines.append(f'{label}: {vals[0]} → {vals[-1]} (变化: {change:+})')
                has_data = True

        return '\n'.join(lines) if has_data else None

    # ==================================================================
    # Profile: 美丽画像 → KnowledgeEntry
    # ==================================================================
    def phase_profile(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'受试者总数: {total}')

        created = updated = 0
        batch_size = 200
        for i in range(0, total, batch_size):
            batch = subjects[i:i + batch_size]
            for subject in batch:
                text = self._build_beauty_profile(subject)
                if not text:
                    continue

                if self.dry_run:
                    created += 1
                    continue

                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_beauty_profile',
                    source_id=subject.id,
                    source_key=f'beauty_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 美丽档案',
                        'content': text,
                        'status': 'published', 'is_published': True,
                    }
                )
                created += 1 if is_new else 0
                updated += 0 if is_new else 1

            self.stdout.write(f'  进度: {min(i + batch_size, total)}/{total}')

        self.stdout.write(self.style.SUCCESS(f'Profile: 新建 {created}, 更新 {updated}'))

    def _build_beauty_profile(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord

        lines = [
            f'受试者编号: {subject.subject_no or subject.id}',
            f'基本信息: {subject.get_gender_display() or "未知"}, '
            f'{subject.age or "未知"}岁, Fitzpatrick {subject.skin_type or "未知"}型',
        ]

        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        if enrollments.exists():
            projects = [f'{e.protocol.code}({e.protocol.title[:30]})'
                        for e in enrollments[:5] if e.protocol]
            lines.append(f'参与项目: {", ".join(projects)}')

        measurements = SkinMeasurementRecord.objects.filter(subject=subject).order_by('-measured_at')
        if measurements.exists():
            latest = measurements.first()
            parts = []
            for field, label, unit in [
                ('moisture', '水分', 'a.u.'), ('tewl', 'TEWL', 'g/h/m²'),
                ('sebum', '皮脂', 'μg/cm²'), ('melanin', '黑色素', ''),
                ('erythema', '红斑', ''), ('elasticity', '弹性', ''),
            ]:
                val = getattr(latest, field, None)
                if val is not None:
                    parts.append(f'{label} {val} {unit}'.strip())
            if parts:
                lines.append(f'皮肤基线: {", ".join(parts)}')

        questionnaires = SubjectQuestionnaire.objects.filter(subject=subject)
        sensory = questionnaires.filter(questionnaire_type='sensory_questionnaire').first()
        if sensory and sensory.answers:
            parts = []
            for key, label in [('itch_nrs', '瘙痒'), ('dry_skin_overall', '干燥'),
                                ('burning_nrs', '灼烧'), ('stinging_nrs', '刺痛')]:
                val = sensory.answers.get(key)
                if val is not None:
                    parts.append(f'{label} {val}/10')
            if parts:
                tp = sensory.answers.get('timepoint', '')
                lines.append(f'感官评价({tp}): {", ".join(parts)}')

        compliance = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if compliance:
            lines.append(
                f'依从性: 到访率 {compliance.visit_attendance_rate}%, '
                f'问卷完成率 {compliance.questionnaire_completion_rate}%, '
                f'等级 {compliance.get_level_display()}'
            )

        return '\n'.join(lines) if len(lines) > 2 else None

    # ==================================================================
    # KG: 知识图谱 — 受试者→项目/肤质/年龄段关联
    # ==================================================================
    def phase_kg(self):
        from apps.subject.models import Subject, Enrollment
        from apps.knowledge.models import KnowledgeEntity, KnowledgeRelation

        enrollments = Enrollment.objects.select_related('subject', 'protocol').all()
        entities_created = relations_created = 0

        for enrollment in enrollments:
            subject = enrollment.subject
            protocol = enrollment.protocol
            if not protocol:
                continue

            subj_entity, s_new = KnowledgeEntity.objects.get_or_create(
                entity_type='instance',
                uri=f'cnkis:subject/{subject.subject_no or subject.id}',
                namespace='cnkis',
                defaults={
                    'label': subject.subject_no or f'受试者{subject.id}',
                    'label_en': subject.subject_no or '',
                    'definition': f'{subject.name}, {subject.get_gender_display()}, {subject.age}岁',
                    'properties': {'subject_id': subject.id, 'skin_type': subject.skin_type},
                }
            )
            if s_new:
                entities_created += 1

            proj_entity = KnowledgeEntity.objects.filter(
                entity_type='project',
                uri__icontains=protocol.code,
            ).first()
            if not proj_entity:
                proj_entity, p_new = KnowledgeEntity.objects.get_or_create(
                    entity_type='project',
                    uri=f'cnkis:project/{protocol.code}',
                    namespace='cnkis',
                    defaults={
                        'label': protocol.code,
                        'label_en': protocol.code,
                        'definition': protocol.title[:200],
                    }
                )
                if p_new:
                    entities_created += 1

            _, r_new = KnowledgeRelation.objects.get_or_create(
                subject=subj_entity, object=proj_entity,
                relation_type='custom',
                predicate_uri='cnkis:enrolled_in',
                defaults={
                    'source': 'auto_fill_beauty',
                    'metadata': {'enrollment_id': enrollment.id},
                }
            )
            if r_new:
                relations_created += 1

            if subject.skin_type:
                skin_entity, sk_new = KnowledgeEntity.objects.get_or_create(
                    entity_type='concept',
                    uri=f'cnkis:skin_type/fitzpatrick_{subject.skin_type}',
                    namespace='cnkis',
                    defaults={
                        'label': f'Fitzpatrick {subject.skin_type}型',
                        'label_en': f'Fitzpatrick Type {subject.skin_type}',
                        'definition': f'Fitzpatrick 皮肤分型 {subject.skin_type}',
                    }
                )
                if sk_new:
                    entities_created += 1
                KnowledgeRelation.objects.get_or_create(
                    subject=subj_entity, object=skin_entity,
                    relation_type='has_property',
                    predicate_uri='cnkis:has_skin_type',
                    defaults={'source': 'auto_fill_beauty'},
                )

            if subject.age:
                age_group = (
                    '18-25' if subject.age < 25 else
                    '25-35' if subject.age < 35 else
                    '35-45' if subject.age < 45 else
                    '45-55' if subject.age < 55 else '55+'
                )
                age_entity, _ = KnowledgeEntity.objects.get_or_create(
                    entity_type='concept',
                    uri=f'cnkis:age_group/{age_group}',
                    namespace='cnkis',
                    defaults={
                        'label': f'{age_group}岁',
                        'label_en': f'Age {age_group}',
                        'definition': f'{age_group}岁年龄段',
                    }
                )
                KnowledgeRelation.objects.get_or_create(
                    subject=subj_entity, object=age_entity,
                    relation_type='part_of',
                    predicate_uri='cnkis:belongs_to_age_group',
                    defaults={'source': 'auto_fill_beauty'},
                )

        self.stdout.write(self.style.SUCCESS(
            f'KG: 新增 {entities_created} 实体, {relations_created} 关系'
        ))
