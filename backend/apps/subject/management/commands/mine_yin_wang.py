"""
从殷淑雯&王梦丹的受试者数据中心全量挖掘

核心数据源（按价值排序）：
  1. 受试者测试项目信息周报 (2020-2021)
     → 真实姓名 + 脱敏手机 + 性别 + 项目编号 + 测试时间 + 测试部位
  2. 合格受试者基础信息 (per project)
     → SC/RD编号 + 年龄 + 性别 + 临床指标
  3. FGD/名单/报名总表
     → 受试者编号 + 姓名 + 手机号
  4. 结算单/礼金发放表
     → 姓名 + 手机 + 结算金额
  5. IM 招募群聊
     → 招募过程数字（到访/初筛/合格/入组）+ 受试者编号

Usage:
  python manage.py mine_yin_wang --phase all
  python manage.py mine_yin_wang --phase weekly_report   # 受试者测试信息周报
  python manage.py mine_yin_wang --phase qualified       # 合格受试者基础信息
  python manage.py mine_yin_wang --phase roster          # 各类名单
  python manage.py mine_yin_wang --phase settlement      # 结算单
  python manage.py mine_yin_wang --phase im_stats        # IM招募统计
  python manage.py mine_yin_wang --phase profile         # 生成知识档案
  python manage.py mine_yin_wang --stats
"""
import os
import re
import logging
from datetime import date
from decimal import Decimal
from collections import defaultdict

from django.core.management.base import BaseCommand
from django.db import connection

logger = logging.getLogger(__name__)
MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')
PROJECT_P = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)


def norm(v):
    return str(v).strip() if v is not None else ''


def safe_int(v):
    try:
        return int(re.sub(r'[^\d]', '', str(v)))
    except Exception:
        return None


def phone_last4(phone_str):
    s = re.sub(r'[^\d]', '', str(phone_str))
    return s[-4:] if len(s) >= 4 else s


def extract_project(text):
    m = PROJECT_P.search(str(text))
    return m.group(0).upper() if m else ''


class Command(BaseCommand):
    help = '从殷淑雯&王梦丹数据中心挖掘受试者全量信息'

    def add_arguments(self, parser):
        parser.add_argument('--phase', default='all',
                            choices=['all', 'weekly_report', 'qualified', 'roster',
                                     'settlement', 'im_stats', 'profile'])
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--stats', action='store_true')

    def handle(self, *args, **options):
        if options['stats']:
            self._show_stats()
            return

        self.dry_run = options['dry_run']
        phase = options['phase']
        self.counters = defaultdict(int)

        phases = ['weekly_report', 'qualified', 'roster', 'settlement', 'im_stats', 'profile'] \
            if phase == 'all' else [phase]

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}', lambda: None)()

        self.stdout.write(self.style.SUCCESS(
            f'\n=== 汇总 ===\n'
            f'  新建受试者: {self.counters["new"]}\n'
            f'  更新受试者: {self.counters["updated"]}\n'
            f'  知识档案: {self.counters["profiles"]}\n'
            f'  IM招募记录: {self.counters["im_records"]}'
        ))

    # ==================================================================
    # 1. 受试者测试项目信息周报（最高价值）
    # ==================================================================
    def phase_weekly_report(self):
        import openpyxl

        files = self._find_files(['受试者测试项目信息周报'])
        self.stdout.write(f'找到 {len(files)} 个周报文件')

        all_records = []
        for fpath in files:
            fname = os.path.basename(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception as e:
                logger.warning('无法打开 %s: %s', fname, e)
                continue

            if not rows:
                continue

            # 表头: 受试者名称, 手机号（脱敏）, 性别, 项目编号, 项目开始时间, 项目结束时间, 测试部位
            header = [norm(c) for c in rows[0]]
            col = {}
            for i, h in enumerate(header):
                if '名称' in h or '姓名' in h:
                    col['name'] = i
                elif '手机' in h:
                    col['phone_masked'] = i
                elif '性别' in h:
                    col['gender'] = i
                elif '项目编号' in h:
                    col['project_code'] = i
                elif '开始' in h:
                    col['start_date'] = i
                elif '结束' in h:
                    col['end_date'] = i
                elif '部位' in h:
                    col['test_site'] = i

            if 'name' not in col or 'project_code' not in col:
                continue

            for row in rows[1:]:
                if not row:
                    continue
                rec = {f: norm(row[i]) for f, i in col.items() if i < len(row) and row[i] is not None}
                name = rec.get('name', '')
                project_code = rec.get('project_code', '')
                if not name or not project_code:
                    continue
                # 脱敏手机转成后4位
                pm = rec.get('phone_masked', '')
                last4 = re.search(r'(\d{4})\s*$', pm)
                rec['phone_last4'] = last4.group(1) if last4 else ''
                rec['source'] = 'weekly_report'
                rec['source_file'] = fname
                all_records.append(rec)

        # 去重（同姓名+项目）
        seen = set()
        unique = []
        for rec in all_records:
            key = f"{rec.get('name','')}_{rec.get('project_code','')}_{rec.get('phone_last4','')}"
            if key not in seen:
                seen.add(key)
                unique.append(rec)

        self.stdout.write(f'解析到 {len(all_records)} 条, 去重后 {len(unique)} 条')

        if not self.dry_run:
            saved = self._upsert_from_weekly(unique)
            self.counters['new'] += saved

        self.stdout.write(self.style.SUCCESS(f'周报: 处理 {len(unique)} 条'))

    def _upsert_from_weekly(self, records):
        from apps.subject.models import Subject, Enrollment
        from apps.protocol.models import Protocol

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name', '').strip()
            project_code = rec.get('project_code', '').strip()
            if not name:
                continue

            gender_raw = rec.get('gender', '')
            gender = 'female' if '女' in gender_raw else ('male' if '男' in gender_raw else '')

            try:
                # 查找现有受试者（按姓名+手机尾号）
                phone4 = rec.get('phone_last4', '')
                subject = None

                if phone4:
                    subject = Subject.objects.filter(
                        name=name, phone__endswith=phone4
                    ).first()

                if subject is None:
                    subject = Subject.objects.filter(name=name).first()

                if subject is None:
                    subject = Subject(
                        name=name[:100], gender=gender,
                        source_channel='database', status='completed',
                    )
                    subject.save()
                    saved += 1
                    self.counters['new'] += 1
                else:
                    changed = False
                    if not subject.gender and gender:
                        subject.gender = gender
                        changed = True
                    if changed:
                        subject.save(update_fields=['gender', 'update_time'])
                    self.counters['updated'] += 1

                # 关联项目入组
                if project_code:
                    if project_code not in proto_cache:
                        proto_cache[project_code] = Protocol.objects.filter(
                            code__iexact=project_code
                        ).first()
                    protocol = proto_cache.get(project_code)
                    if protocol:
                        from apps.subject.models import EnrollmentStatus
                        from django.utils import timezone
                        Enrollment.objects.get_or_create(
                            subject=subject, protocol=protocol,
                            defaults={
                                'status': EnrollmentStatus.COMPLETED,
                                'enrolled_at': timezone.now(),
                            }
                        )

                # 写入知识条目（含项目测试记录）
                self._write_subject_entry(subject, rec)

            except Exception as e:
                logger.warning('weekly_report upsert 失败: %s, name=%s', e, name)

        return saved

    def _write_subject_entry(self, subject, rec):
        from apps.knowledge.models import KnowledgeEntry

        project_code = rec.get('project_code', '')
        start = rec.get('start_date', '')
        end = rec.get('end_date', '')
        site = rec.get('test_site', '')
        phone4 = rec.get('phone_last4', '')
        gender = subject.get_gender_display() if subject.gender else '?'

        content = (
            f"受试者: {subject.name}, {gender}\n"
            f"手机尾号: {phone4}\n"
            f"参与项目: {project_code}\n"
            f"测试时间: {start} ~ {end}\n"
            f"测试部位: {site}\n"
            f"数据来源: 受试者测试项目信息周报"
        )

        key = f"weekly_{subject.id}_{project_code}"
        KnowledgeEntry.objects.update_or_create(
            source_type='subject_project_record',
            source_id=subject.id,
            source_key=key[:120],
            defaults={
                'entry_type': 'lesson_learned',
                'title': f'{subject.name} - {project_code} 测试记录',
                'content': content,
                'status': 'published',
                'is_published': True,
            }
        )
        self.counters['profiles'] += 1

    # ==================================================================
    # 2. 合格受试者基础信息 (SC/RD + 年龄 + 性别)
    # ==================================================================
    def phase_qualified(self):
        import openpyxl

        files = self._find_files(['合格受试者基础信息', '合格受试者清单', '合格受试者名单'])
        self.stdout.write(f'找到 {len(files)} 个合格受试者文件')

        for fpath in files:
            fname = os.path.basename(fpath)
            project_code = extract_project(fname) or extract_project(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    records = self._parse_qualified_sheet(rows, project_code, fname)
                    if records:
                        self.stdout.write(f'  {fname}/{sname}: {len(records)} 条')
                        if not self.dry_run:
                            self._upsert_qualified(records)
                wb.close()
            except Exception as e:
                logger.warning('qualified 解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS('Qualified 完成'))

    def _parse_qualified_sheet(self, rows, project_code, fname):
        if not rows:
            return []
        header = [norm(c) for c in rows[0]]
        col = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if 'sc' == hl or '筛选编号' in h:
                col['sc'] = i
            elif 'rd' == hl or '入组编号' in h:
                col['rd'] = i
            elif 'age' == hl or '年龄' in h:
                col['age'] = i
            elif 'gender' == hl or '性别' in h:
                col['gender'] = i
        if not col:
            return []
        records = []
        for row in rows[1:]:
            if not row:
                continue
            rec = {f: norm(row[i]) for f, i in col.items() if i < len(row)}
            if not rec.get('sc') and not rec.get('rd'):
                continue
            rec['project_code'] = project_code
            rec['source_file'] = fname
            records.append(rec)
        return records

    def _upsert_qualified(self, records):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_execution import SubjectProjectSC
        from apps.protocol.models import Protocol

        proto_cache = {}
        for rec in records:
            sc = rec.get('sc', '')
            rd = rec.get('rd', '')
            project_code = rec.get('project_code', '')
            age = safe_int(rec.get('age', ''))
            gender_raw = rec.get('gender', '')
            gender = 'female' if '女' in gender_raw or gender_raw.lower() == 'f' else \
                     ('male' if '男' in gender_raw or gender_raw.lower() == 'm' else '')

            try:
                # 按 project+RD 查受试者
                cand_no = f'{project_code}-{rd}' if rd and project_code else ''
                subject = Subject.objects.filter(subject_no=cand_no).first() if cand_no else None

                if subject is None:
                    subject = Subject(
                        name=f'{project_code}-{rd}' if rd else f'{project_code}-SC{sc}',
                        gender=gender, age=age,
                        source_channel='database', status='completed',
                    )
                    if cand_no:
                        subject.subject_no = cand_no[:20]
                    subject.save()
                    self.counters['new'] += 1
                else:
                    changed = False
                    if not subject.age and age:
                        subject.age = age
                        changed = True
                    if not subject.gender and gender:
                        subject.gender = gender
                        changed = True
                    if changed:
                        subject.save(update_fields=['age', 'gender', 'update_time'])
                    self.counters['updated'] += 1

                # SC号记录
                if sc and project_code:
                    SubjectProjectSC.objects.get_or_create(
                        subject=subject, project_code=project_code,
                        defaults={'sc_number': sc, 'rd_number': rd}
                    )

            except Exception as e:
                logger.debug('qualified upsert 跳过: %s', e)

    # ==================================================================
    # 3. 各类名单 (FGD / 报名 / 受试者名单)
    # ==================================================================
    def phase_roster(self):
        import openpyxl

        kws = ['受试者名单', 'FGD受试者', '初步甄别', '到访名单', '报名总',
               '知情用', '受试者基础信息']
        files = self._find_files(kws)
        self.stdout.write(f'找到 {len(files)} 个名单文件')
        total = 0

        for fpath in files:
            fname = os.path.basename(fpath)
            project_code = extract_project(fname) or extract_project(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                records = self._parse_roster_sheet(rows, project_code, fname)
                if records:
                    self.stdout.write(f'  {fname}: {len(records)} 条')
                    if not self.dry_run:
                        self._upsert_roster(records)
                    total += len(records)
                wb.close()
            except Exception as e:
                logger.warning('roster 解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS(f'Roster: {total} 条'))

    def _parse_roster_sheet(self, rows, project_code, fname):
        if len(rows) < 2:
            return []
        # 找表头行
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            text = ' '.join(norm(c) for c in row if c)
            if any(k in text for k in ['姓名', '手机', '编号', 'ID', 'SC', 'RD']):
                header_idx = i
                break
        header = [norm(c) for c in rows[header_idx]]
        col = {}
        for i, h in enumerate(header):
            if '姓名' in h and 'name' not in col:
                col['name'] = i
            elif '手机' in h and 'phone' not in col:
                col['phone'] = i
            elif '编号' in h or h == 'ID':
                col['id_col'] = i
            elif '性别' in h:
                col['gender'] = i
            elif '年龄' in h:
                col['age'] = i

        records = []
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            rec = {}
            for f, ci in col.items():
                if ci < len(row) and row[ci] is not None:
                    rec[f] = norm(row[ci])
            name = rec.get('name', '')
            phone = rec.get('phone', '')
            if not name and not phone:
                continue
            if name in ('姓名', '合计', '总计', ''):
                continue
            rec['project_code'] = project_code
            rec['source_file'] = fname
            records.append(rec)
        return records

    def _upsert_roster(self, records):
        from apps.subject.models import Subject
        for rec in records:
            name = rec.get('name', '')
            phone = rec.get('phone', '')[:20] if rec.get('phone') else ''
            if not name:
                continue
            try:
                subject, created = Subject.objects.get_or_create(
                    name=name, phone=phone,
                    defaults={
                        'source_channel': 'database', 'status': 'pre_screened',
                        'gender': ('female' if '女' in rec.get('gender', '') else
                                   'male' if '男' in rec.get('gender', '') else ''),
                        'age': safe_int(rec.get('age', '')),
                    }
                )
                if created:
                    self.counters['new'] += 1
                else:
                    self.counters['updated'] += 1
            except Exception:
                pass

    # ==================================================================
    # 4. 结算单解析
    # ==================================================================
    def phase_settlement(self):
        import openpyxl

        kws = ['结算', '礼金', '兼职联络员劳务', '员工受试者结算', '受试者银行卡', '劳务报酬']
        files = self._find_files(kws)
        self.stdout.write(f'找到 {len(files)} 个结算文件')
        total = 0

        for fpath in files:
            fname = os.path.basename(fpath)
            project_code = extract_project(fname) or extract_project(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(max_row=min(ws.max_row, 300), values_only=True))
                records = self._parse_settlement_sheet(rows, project_code, fname)
                if records:
                    if not self.dry_run:
                        self._upsert_settlement(records)
                    total += len(records)
                wb.close()
            except Exception as e:
                logger.warning('settlement 解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS(f'Settlement: {total} 条'))

    def _parse_settlement_sheet(self, rows, project_code, fname):
        if not rows:
            return []
        hi = 0
        for i, row in enumerate(rows[:6]):
            text = ' '.join(norm(c) for c in row if c)
            if any(k in text for k in ['姓名', '手机', '金额', '银行']):
                hi = i
                break
        header = [norm(c) for c in rows[hi]]
        col = {}
        for i, h in enumerate(header):
            if '姓名' in h and 'name' not in col:
                col['name'] = i
            elif '手机' in h or '电话' in h:
                col['phone'] = i
            elif '金额' in h or '费用' in h or '报酬' in h:
                col['amount'] = i
        if 'name' not in col:
            return []
        records = []
        for row in rows[hi + 1:]:
            if not row:
                continue
            rec = {f: norm(row[ci]) for f, ci in col.items() if ci < len(row)}
            name = rec.get('name', '')
            if not name or name in ('姓名', '合计', '小计', '总计'):
                continue
            rec['project_code'] = project_code
            rec['source_file'] = fname
            records.append(rec)
        return records

    def _upsert_settlement(self, records):
        from apps.subject.models import Subject
        for rec in records:
            name = rec.get('name', '')
            phone = rec.get('phone', '')[:20] if rec.get('phone') else ''
            if not name:
                continue
            try:
                _, created = Subject.objects.get_or_create(
                    name=name, phone=phone,
                    defaults={'source_channel': 'database', 'status': 'completed'}
                )
                if created:
                    self.counters['new'] += 1
                else:
                    self.counters['updated'] += 1
            except Exception:
                pass

    # ==================================================================
    # 5. IM 招募统计
    # ==================================================================
    def phase_im_stats(self):
        from apps.secretary.models import PersonalContext
        from apps.knowledge.models import KnowledgeEntry
        import json

        visit_p = re.compile(r'(?:到访|来访|约访)\D{0,5}(\d+)')
        screen_p = re.compile(r'(?:初筛|粗筛|筛选了?)\D{0,5}(\d+)')
        qual_p = re.compile(r'(?:初筛合格|合格受试者?)\D{0,5}(\d+)')
        enroll_p = re.compile(r'(?:入组了?|已入组|今天入组)\D{0,5}(\d+)')
        drop_p = re.compile(r'(?:脱落|退出了?|不合格)\D{0,5}(\d+)')

        recruit_chats = [
            '特化招募', '组3招募', '组2招募群', '组1 招募沟通', '组7&组9 招募',
            '华山植发-招募重点跟踪群组', '组4-招募沟通群', 'C06招募沟通群',
            '忍者-组3', '头发测试研究团队', '组3运营', '招募天团🤙🏼🤙🏼🤙🏼',
            '压力防脱测试招募渠道扩展沟通', '组3-国家大事专用群',
        ]

        ctxs = PersonalContext.objects.filter(
            source_type='im',
            metadata__chat_name__in=recruit_chats,
        ).filter(
            raw_content__contains='入组'
        ) | PersonalContext.objects.filter(
            source_type='im',
            metadata__chat_name__in=recruit_chats,
        ).filter(
            raw_content__contains='到访'
        )

        self.stdout.write(f'招募IM消息: {ctxs.count():,}')

        # 按项目聚合
        ps = defaultdict(lambda: defaultdict(int))
        daily = []
        processed = 0

        for ctx in ctxs.iterator(chunk_size=500):
            content = ctx.raw_content or ''
            if isinstance(content, str) and (content.startswith('{') or content.startswith('[')):
                try:
                    obj = json.loads(content)
                    content = obj.get('text', '') if isinstance(obj, dict) else ''
                except Exception:
                    content = ''

            projs = [p.upper() for p in PROJECT_P.findall(content)]
            rec = {
                'date': ctx.created_at.date().isoformat() if ctx.created_at else '',
                'chat': (ctx.metadata or {}).get('chat_name', ''),
                'projects': projs,
            }
            for field, pat in [('visited', visit_p), ('screened', screen_p),
                                ('qualified', qual_p), ('enrolled', enroll_p),
                                ('dropout', drop_p)]:
                m = pat.search(content)
                if m:
                    try:
                        rec[field] = int(m.group(1))
                    except Exception:
                        pass

            if any(k in rec for k in ['visited', 'enrolled', 'qualified']):
                daily.append(rec)
                for proj in (projs or ['未知']):
                    for f in ['visited', 'screened', 'qualified', 'enrolled', 'dropout']:
                        ps[proj][f] += rec.get(f, 0)

            processed += 1

        self.counters['im_records'] = len(daily)

        # 写入知识条目
        if not self.dry_run and daily:
            tv = sum(r.get('visited', 0) for r in daily)
            te = sum(r.get('enrolled', 0) for r in daily)
            tq = sum(r.get('qualified', 0) for r in daily)
            ts = sum(r.get('screened', 0) for r in daily)
            td = sum(r.get('dropout', 0) for r in daily)

            proj_lines = []
            for proj, s in sorted(ps.items(), key=lambda x: x[1].get('enrolled', 0), reverse=True)[:50]:
                if sum(s.values()) > 0:
                    proj_lines.append(
                        f"{proj}: 到访{s.get('visited', 0)} 初筛{s.get('screened', 0)} "
                        f"合格{s.get('qualified', 0)} 入组{s.get('enrolled', 0)} 脱落{s.get('dropout', 0)}"
                    )

            summary = (
                f"招募IM数据汇总（来源：殷淑雯&王梦丹招募群聊）\n"
                f"分析记录数: {len(daily)}\n"
                f"累计到访: {tv}\n"
                f"累计初筛: {ts}\n"
                f"累计合格: {tq}\n"
                f"累计入组: {te}\n"
                f"累计脱落: {td}\n\n"
                f"=== 按项目分布 ===\n"
                + '\n'.join(proj_lines)
            )

            KnowledgeEntry.objects.update_or_create(
                source_type='recruit_im_stats',
                source_id=0,
                source_key='recruit_im_stats_yin_wang',
                defaults={
                    'entry_type': 'lesson_learned',
                    'title': '招募部门IM群聊 — 受试者招募统计（殷淑雯&王梦丹）',
                    'content': summary,
                    'status': 'published',
                    'is_published': True,
                }
            )

        self.stdout.write(self.style.SUCCESS(f'IM统计: {len(daily)} 条有效记录'))

    # ==================================================================
    # 6. 生成全景知识档案
    # ==================================================================
    def phase_profile(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'受试者总数: {total:,}')

        created = updated = 0
        for i in range(0, total, 500):
            for subject in subjects[i:i + 500]:
                text = self._build_lifecycle_profile(subject)
                if not text:
                    continue
                if self.dry_run:
                    created += 1
                    continue
                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_full_lifecycle',
                    source_id=subject.id,
                    source_key=f'lifecycle_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 受试者全档案',
                        'content': text,
                        'status': 'published',
                        'is_published': True,
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            self.stdout.write(f'  进度: {min(i + 500, total)}/{total}')

        self.counters['profiles'] = created + updated
        self.stdout.write(self.style.SUCCESS(f'Profile: 新建 {created}, 更新 {updated}'))

    def _build_lifecycle_profile(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord

        lines = [
            f'受试者: {subject.name}',
            f'编号: {subject.subject_no or subject.id}',
            f'性别: {subject.get_gender_display() or "?"}',
            f'年龄: {subject.age or "?"}',
        ]

        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        if enrollments.exists():
            projs = [e.protocol.code for e in enrollments if e.protocol]
            lines.append(f'参与项目({enrollments.count()}): {", ".join(projs[:5])}')

        m_count = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m_count:
            lines.append(f'仪器测量: {m_count} 次')

        q_count = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q_count:
            lines.append(f'问卷记录: {q_count} 份')

        c = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if c:
            lines.append(f'依从性: {c.get_level_display()}, 到访率 {c.visit_attendance_rate}%')

        return '\n'.join(lines) if len(lines) > 3 else None

    # ==================================================================
    # 工具方法
    # ==================================================================
    def _find_files(self, kws):
        results = []
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f for kw in kws):
                    results.append(os.path.join(root, f))
        return list(set(results))

    def _show_stats(self):
        c = connection.cursor()
        print('\n=== 受试者全量统计 ===')
        for table, label in [
            ('t_subject', '受试者总数'),
            ('t_enrollment', '入组记录'),
            ('t_subject_project_sc', 'SC/RD映射'),
            ('t_skin_measurement_record', '仪器测量'),
            ('t_subject_questionnaire', '问卷'),
            ('t_subject_compliance', '依从性'),
        ]:
            c.execute(f'SELECT COUNT(*) FROM {table}')
            print(f'  {label:20s}: {c.fetchone()[0]:>10,}')

        for src in ['subject_full_lifecycle', 'subject_beauty_profile',
                    'subject_project_record', 'recruit_im_stats']:
            c.execute(f"SELECT COUNT(*) FROM t_knowledge_entry WHERE source_type='{src}'")
            print(f'  KE/{src:25s}: {c.fetchone()[0]:>10,}')

        c.execute("SELECT COUNT(*) FROM t_knowledge_entity WHERE entity_type='instance'")
        print(f'  {"受试者知识实体":20s}: {c.fetchone()[0]:>10,}')
