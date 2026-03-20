"""
从飞书全量数据挖掘受试者信息

三大核心数据源：
  A. China-Norm 照片采集表（database sheet）
     → 姓名缩写、RD编号、年龄、性别、肤质、Fitzpatrick分型、项目编号
  B. 报名总表（Worksheet sheet）
     → 受试者编号、姓名、手机号、邮箱、报名时间
  C. 结算/礼金发放表
     → 受试者姓名/手机/结算金额/结算项目

执行流程：
  1. 扫描服务器 /data/media 中所有相关 Excel
  2. 解析后写入 t_subject + t_subject_profile + t_subject_skin_profile
  3. 写入 t_knowledge_entry（source_type=subject_full_lifecycle）
  4. 写入知识图谱（受试者 → 项目 → 阶段关系）

Usage:
  python manage.py mine_feishu_subjects --phase all
  python manage.py mine_feishu_subjects --phase chinanorm   # China-Norm 数据库
  python manage.py mine_feishu_subjects --phase signup      # 报名总表
  python manage.py mine_feishu_subjects --phase settlement  # 结算表
  python manage.py mine_feishu_subjects --phase im          # IM 招募群聊
  python manage.py mine_feishu_subjects --phase profile     # 生成全景档案
  python manage.py mine_feishu_subjects --phase kg          # 知识图谱
  python manage.py mine_feishu_subjects --stats             # 查看统计
"""
import os
import re
import hashlib
import logging
from datetime import datetime, date
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.db import transaction, connection
from django.utils import timezone

logger = logging.getLogger(__name__)

MEDIA_ROOT = os.environ.get('MEDIA_ROOT', '/data/media')

PROJECT_CODE_PATTERN = re.compile(r'[CMOW]\d{5,9}', re.IGNORECASE)

SKIN_TYPE_MAP = {
    'dry': 'dry', '干性': 'dry', '干': 'dry', 'dry skin': 'dry',
    'oily': 'oily', '油性': 'oily', '油': 'oily',
    'normal': 'normal', '中性': 'normal',
    'combo': 'combo', 'combination': 'combo', '混合': 'combo',
    'combo/dry': 'combo_dry', 'combi-dry': 'combo_dry',
    'combo/oily': 'combo_oily', 'combi-oily': 'combo_oily',
    'sensitive': 'sensitive', '敏感': 'sensitive',
}

GENDER_MAP = {
    'female': 'female', '女': 'female', 'f': 'female',
    'male': 'male', '男': 'male', 'm': 'male',
}

FITZPATRICK_MAP = {
    'i': 'I', 'ii': 'II', 'iii': 'III', 'iv': 'IV', 'v': 'V', 'vi': 'VI',
    '1': 'I', '2': 'II', '3': 'III', '4': 'IV', '5': 'V', '6': 'VI',
}


def make_fingerprint(data: dict) -> str:
    key = f"{data.get('name_initials', '')}{data.get('gender', '')}{data.get('age', '')}{data.get('project_code', '')}{data.get('rd_number', '')}"
    return hashlib.md5(key.encode()).hexdigest()[:16]


def norm_str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def parse_age(v):
    try:
        s = re.sub(r'[^\d]', '', str(v))
        a = int(s)
        return a if 10 <= a <= 90 else None
    except Exception:
        return None


def norm_gender(v):
    s = norm_str(v).lower()
    return GENDER_MAP.get(s, GENDER_MAP.get(s[:1], ''))


def norm_skin_type(v):
    s = norm_str(v).lower().strip()
    for k, val in SKIN_TYPE_MAP.items():
        if k in s:
            return val
    return ''


def norm_fitzpatrick(v):
    s = norm_str(v).lower().replace(' ', '')
    for k, val in FITZPATRICK_MAP.items():
        if s == k:
            return val
    return ''


def extract_project_code(text: str) -> str:
    m = PROJECT_CODE_PATTERN.search(str(text))
    return m.group(0).upper() if m else ''


class Command(BaseCommand):
    help = '从飞书全量数据挖掘受试者全生命周期信息'

    def add_arguments(self, parser):
        parser.add_argument('--phase', default='all',
                            choices=['all', 'chinanorm', 'signup', 'settlement', 'im', 'profile', 'kg'])
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--limit', type=int, default=0)
        parser.add_argument('--stats', action='store_true', help='仅显示统计')

    def handle(self, *args, **options):
        if options['stats']:
            self._show_stats()
            return

        self.dry_run = options['dry_run']
        self.limit = options['limit']
        phase = options['phase']
        phases = ['chinanorm', 'signup', 'settlement', 'im', 'profile', 'kg'] if phase == 'all' else [phase]

        self.counters = {'new_subjects': 0, 'updated': 0, 'profiles': 0, 'kg_entities': 0, 'kg_relations': 0}

        for p in phases:
            self.stdout.write(self.style.NOTICE(f'\n{"=" * 60}\n  Phase: {p}\n{"=" * 60}'))
            getattr(self, f'phase_{p}', lambda: None)()

        self.stdout.write(self.style.SUCCESS(
            f'\n=== 汇总 ===\n'
            f'  新建受试者: {self.counters["new_subjects"]}\n'
            f'  更新受试者: {self.counters["updated"]}\n'
            f'  档案知识条目: {self.counters["profiles"]}\n'
            f'  知识图谱实体: {self.counters["kg_entities"]}\n'
            f'  知识图谱关系: {self.counters["kg_relations"]}'
        ))

    # ==================================================================
    # A: China-Norm 照片采集数据库（主力来源）
    # ==================================================================
    def phase_chinanorm(self):
        import openpyxl

        files = self._find_chinanorm_files()
        self.stdout.write(f'找到 {len(files)} 个 China-Norm 文件')
        total_parsed = 0

        for fpath in files:
            fname = os.path.basename(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                if 'database' not in wb.sheetnames:
                    wb.close()
                    continue
                ws = wb['database']
                records = self._parse_chinanorm_sheet(ws, fpath)
                if records:
                    self.stdout.write(f'  {fname}: {len(records)} 条')
                    if not self.dry_run:
                        saved = self._upsert_subjects(records, source='chinanorm')
                        total_parsed += saved
                    else:
                        total_parsed += len(records)
                wb.close()
            except Exception as e:
                logger.warning('China-Norm 解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS(f'ChianNorm: {total_parsed} 条记录'))

    def _find_chinanorm_files(self):
        results = []
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith('.xlsx'):
                    continue
                if 'china-norm' in f.lower() or 'chinanorm' in f.lower():
                    if '照片及受试者' in f or 'database' in f.lower() or '受试者信息' in f:
                        results.append(os.path.join(root, f))
        if self.limit:
            results = results[:self.limit]
        return results

    def _parse_chinanorm_sheet(self, ws, fpath):
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        # 找表头行（包含 Subject # 或 受试者编号）
        header_row_idx = 0
        for i, row in enumerate(rows[:5]):
            row_text = ' '.join(norm_str(c) for c in row if c)
            if any(k in row_text for k in ['Subject #', '受试者编号', 'Subject Initials', '受试者姓名']):
                header_row_idx = i
                break

        header = [norm_str(c) for c in rows[header_row_idx]]

        col_map = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if any(k in hl for k in ['cro study', '测试机构项目', '项目编号']):
                col_map['project_code'] = i
            elif any(k in hl for k in ['subject initials', '姓名缩写', 'subject ini']):
                col_map['name_initials'] = i
            elif any(k in hl for k in ['subject #', '受试者编号', 'subject no']):
                col_map['rd_number'] = i
            elif 'age' in hl or '年龄' in hl:
                col_map['age'] = i
            elif 'gender' in hl or '性别' in hl:
                col_map['gender'] = i
            elif 'skin type' in hl or '肤质类型' in hl:
                col_map['skin_type'] = i
            elif 'fitzpatrick' in hl or '日光分型' in hl:
                col_map['fitzpatrick'] = i
            elif 'elc study' in hl or 'study #' in hl:
                col_map['elc_study'] = i
            elif 'questionnaire' in hl or '问卷' in hl:
                col_map['has_questionnaire'] = i
            elif 'photo release' in hl or '肖像使用' in hl or 'authorization' in hl:
                col_map['photo_release'] = i

        if 'rd_number' not in col_map and 'name_initials' not in col_map:
            return []

        records = []
        for row in rows[header_row_idx + 2:]:  # skip header + example row
            if not row:
                continue
            rec = {}
            for field, ci in col_map.items():
                if ci < len(row) and row[ci] is not None:
                    rec[field] = norm_str(row[ci])

            rd = rec.get('rd_number', '')
            initials = rec.get('name_initials', '')
            if not rd and not initials:
                continue
            if rd.lower() in ('e.g. rd001', '', 'subject #'):
                continue

            project_code = rec.get('project_code', '')
            if not project_code:
                project_code = extract_project_code(fpath)

            record = {
                'name_initials': initials,
                'rd_number': rd,
                'project_code': project_code,
                'elc_study': rec.get('elc_study', ''),
                'age': parse_age(rec.get('age', '')),
                'gender': norm_gender(rec.get('gender', '')),
                'skin_type': norm_skin_type(rec.get('skin_type', '')),
                'fitzpatrick': norm_fitzpatrick(rec.get('fitzpatrick', '')),
                'has_questionnaire': rec.get('has_questionnaire', '').upper() == 'Y',
                'photo_release': rec.get('photo_release', '').upper() == 'Y',
                'source': 'chinanorm',
                'source_file': os.path.basename(fpath),
            }
            records.append(record)

        return records

    # ==================================================================
    # B: 报名总表
    # ==================================================================
    def phase_signup(self):
        import openpyxl

        files = self._find_signup_files()
        self.stdout.write(f'找到 {len(files)} 个报名总表')
        total = 0

        for fpath in files:
            fname = os.path.basename(fpath)
            project_code = extract_project_code(fname) or extract_project_code(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                # 找含 'Worksheet' 或第一个有数据的sheet
                sheet = None
                for sname in ['Worksheet', 'Sheet1', 'Sheet2']:
                    if sname in wb.sheetnames:
                        sheet = wb[sname]
                        break
                if sheet is None:
                    sheet = wb.active

                rows = list(sheet.iter_rows(values_only=True))
                records = self._parse_signup_sheet(rows, project_code, fpath)
                if records:
                    self.stdout.write(f'  {fname}: {len(records)} 条')
                    if not self.dry_run:
                        saved = self._upsert_subjects(records, source='signup')
                        total += saved
                    else:
                        total += len(records)
                wb.close()
            except Exception as e:
                logger.warning('报名表解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS(f'Signup: {total} 条记录'))

    def _find_signup_files(self):
        results = []
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if ('报名' in f and '总' in f) or ('名单' in f and '总' in f):
                    results.append(os.path.join(root, f))
        # 同时找「知情用」类型的文件（含报名者名单）
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if '知情用' in f or '受试者名单' in f or '志愿者名单' in f:
                    results.append(os.path.join(root, f))
        results = list(set(results))
        if self.limit:
            results = results[:self.limit]
        return results

    def _parse_signup_sheet(self, rows, project_code, fpath):
        if len(rows) < 2:
            return []

        # 找数据行（含 ID 列）
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            row_text = ' '.join(norm_str(c) for c in row if c)
            if 'ID' in row_text and ('姓名' in row_text or '手机' in row_text or '受试者编号' in row_text):
                header_idx = i
                break

        header = [norm_str(c) for c in rows[header_idx]]
        col_map = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if h == 'ID' or h == '序号':
                col_map['signup_id'] = i
            elif '受试者编号' in h:
                col_map['subject_no'] = i
            elif '姓名编码' in h or 'name code' in hl:
                col_map['name_code'] = i
            elif '手机号后4位' in h:
                col_map['phone_last4'] = i
            elif h == '姓名' or 'name' == hl:
                col_map['name'] = i
            elif '邮箱' in h or 'email' in hl:
                col_map['email'] = i
            elif '手机号' in h and 'phone' not in col_map:
                col_map['phone'] = i
            elif '打开时间' in h or '开始时间' in h:
                col_map['start_time'] = i
            elif '完成时间' in h:
                col_map['finish_time'] = i
            elif '获奖' in h or '奖励' in h:
                col_map['reward'] = i

        if 'signup_id' not in col_map and 'subject_no' not in col_map and 'name_code' not in col_map:
            return []

        records = []
        fname = os.path.basename(fpath)
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            rec = {}
            for field, ci in col_map.items():
                if ci < len(row) and row[ci] is not None:
                    rec[field] = norm_str(row[ci])

            signup_id = rec.get('signup_id', '')
            name_code = rec.get('name_code', '')
            name = rec.get('name', '')
            phone = rec.get('phone', '')

            if not signup_id and not name_code and not name:
                continue
            # 跳过表头重复行
            if signup_id in ('ID', '') and not name:
                continue

            record = {
                'signup_id': signup_id,
                'name_code': name_code,
                'name': name or name_code,
                'phone': phone,
                'phone_last4': rec.get('phone_last4', ''),
                'email': rec.get('email', ''),
                'project_code': project_code,
                'signup_time': rec.get('finish_time', rec.get('start_time', '')),
                'reward': rec.get('reward', ''),
                'source': 'signup',
                'source_file': fname,
            }
            records.append(record)

        return records

    # ==================================================================
    # C: 结算/礼金发放表
    # ==================================================================
    def phase_settlement(self):
        import openpyxl

        files = self._find_settlement_files()
        self.stdout.write(f'找到 {len(files)} 个结算/礼金文件')
        total = 0

        for fpath in files:
            fname = os.path.basename(fpath)
            if not fname.endswith(('.xlsx', '.xls')) or fname.startswith('~'):
                continue
            project_code = extract_project_code(fname) or extract_project_code(fpath)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(max_row=min(ws.max_row, 500), values_only=True))
                records = self._parse_settlement_sheet(rows, project_code, fpath)
                if records:
                    self.stdout.write(f'  {fname}: {len(records)} 条')
                    if not self.dry_run:
                        saved = self._upsert_subjects_from_settlement(records)
                        total += saved
                    else:
                        total += len(records)
                wb.close()
            except Exception as e:
                logger.warning('结算表解析失败 %s: %s', fname, e)

        self.stdout.write(self.style.SUCCESS(f'Settlement: {total} 条记录'))

    def _find_settlement_files(self):
        results = []
        kws = ['员工受试者结算', '受试者银行卡', '礼金', '兼职联络员劳务', '劳务报酬']
        for root, _dirs, files in os.walk(MEDIA_ROOT):
            for f in files:
                if f.startswith('~') or not f.endswith(('.xlsx', '.xls')):
                    continue
                if any(kw in f for kw in kws):
                    results.append(os.path.join(root, f))
        if self.limit:
            results = results[:self.limit]
        return results

    def _parse_settlement_sheet(self, rows, project_code, fpath):
        if len(rows) < 2:
            return []

        header_idx = 0
        for i, row in enumerate(rows[:6]):
            row_text = ' '.join(norm_str(c) for c in row if c)
            if any(k in row_text for k in ['姓名', '手机', '金额', '银行卡']):
                header_idx = i
                break

        header = [norm_str(c) for c in rows[header_idx]]
        col_map = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if '姓名' in h and 'name' not in col_map:
                col_map['name'] = i
            elif '手机' in h or '电话' in h:
                col_map['phone'] = i
            elif '金额' in h or '费用' in h or '报酬' in h:
                col_map['amount'] = i
            elif '银行卡' in h or '卡号' in h:
                col_map['bank_card'] = i
            elif '项目' in h or 'project' in hl:
                col_map['project'] = i
            elif '备注' in h:
                col_map['notes'] = i

        if 'name' not in col_map and 'phone' not in col_map:
            return []

        fname = os.path.basename(fpath)
        records = []
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            rec = {}
            for field, ci in col_map.items():
                if ci < len(row) and row[ci] is not None:
                    rec[field] = norm_str(row[ci])

            name = rec.get('name', '')
            phone = rec.get('phone', '')
            if not name and not phone:
                continue
            if name in ('姓名', '合计', '小计', '总计', ''):
                continue

            amount = None
            try:
                amount_str = re.sub(r'[^\d.]', '', rec.get('amount', ''))
                if amount_str:
                    amount = Decimal(amount_str)
            except Exception:
                pass

            records.append({
                'name': name,
                'phone': phone,
                'bank_card': rec.get('bank_card', ''),
                'amount': amount,
                'project': rec.get('project', project_code),
                'notes': rec.get('notes', ''),
                'source': 'settlement',
                'source_file': fname,
            })

        return records

    def _upsert_subjects_from_settlement(self, records):
        from apps.subject.models import Subject, SubjectSourceChannel
        saved = 0
        for rec in records:
            name = rec.get('name', '')
            phone = rec.get('phone', '')
            if not name:
                continue
            try:
                subject, created = Subject.objects.get_or_create(
                    name=name, phone=phone[:20] if phone else '',
                    defaults={
                        'source_channel': SubjectSourceChannel.DATABASE,
                        'status': 'completed',
                    }
                )
                if created:
                    self.counters['new_subjects'] += 1
                    saved += 1
                else:
                    self.counters['updated'] += 1
                    saved += 1
            except Exception as e:
                logger.debug('settlement upsert 跳过: %s', e)
        return saved

    # ==================================================================
    # D: IM 招募群聊提取
    # ==================================================================
    def phase_im(self):
        from apps.secretary.models import PersonalContext

        recruit_groups = ['招募', '组1', '组2', '组3', '组4', '组5', '组6', '组7', '组8', '组9',
                          '特化', '普化', '初筛']

        contexts = PersonalContext.objects.filter(
            source_type='im',
            metadata__chat_name__isnull=False,
        ).filter(
            raw_content__iregex=r'(SC\d{3}|RD\d{3}|受试者|报名|初筛|入组|结算|礼金)'
        )

        # 过滤出招募相关群聊
        total = contexts.count()
        self.stdout.write(f'含受试者关键词的IM消息: {total:,}')

        name_rd_pattern = re.compile(
            r'((?:SC|RD)\d{3,6})[^\n]{0,50}?([\u4e00-\u9fff]{2,4})',
            re.IGNORECASE
        )
        phone_pattern = re.compile(r'1[3-9]\d{9}')

        extracted = []
        processed = 0
        for ctx in contexts.iterator(chunk_size=500):
            content = ctx.raw_content or ''
            chat_name = (ctx.metadata or {}).get('chat_name', '')

            # 提取 SC/RD编号与姓名
            for m in name_rd_pattern.finditer(content):
                rd_or_sc = m.group(1).upper()
                possible_name = m.group(2)
                if not possible_name:
                    continue
                extracted.append({
                    'rd_or_sc': rd_or_sc,
                    'name': possible_name,
                    'chat_name': chat_name,
                    'source': 'im_recruit',
                })

            # 提取手机号（用于后续关联）
            phones = phone_pattern.findall(content)
            for phone in phones[:3]:
                extracted.append({
                    'phone': phone,
                    'chat_name': chat_name,
                    'source': 'im_recruit',
                })

            processed += 1
            if self.limit and processed >= self.limit:
                break

        self.stdout.write(f'从IM提取: {len(extracted)} 条受试者信号')
        if not self.dry_run:
            saved = self._upsert_im_subjects(extracted)
            self.stdout.write(self.style.SUCCESS(f'IM: 写入 {saved} 条'))

    def _upsert_im_subjects(self, records):
        from apps.subject.models import Subject, SubjectSourceChannel
        saved = 0
        for rec in records:
            name = rec.get('name', '')
            phone = rec.get('phone', '')
            if not name and not phone:
                continue
            try:
                subject, created = Subject.objects.get_or_create(
                    name=name or '未知',
                    phone=phone[:20] if phone else '',
                    defaults={
                        'source_channel': SubjectSourceChannel.OTHER,
                        'status': 'pre_screened',
                    }
                )
                if created:
                    self.counters['new_subjects'] += 1
                    saved += 1
            except Exception:
                pass
        return saved

    # ==================================================================
    # 核心：upsert Subject + SkinProfile
    # ==================================================================
    def _upsert_subjects(self, records, source='unknown'):
        from apps.subject.models import Subject, SubjectSourceChannel, SubjectStatus
        from apps.subject.models_domain import SkinProfile
        from apps.protocol.models import Protocol

        proto_cache = {}
        saved = 0

        for rec in records:
            name = rec.get('name') or rec.get('name_initials') or rec.get('name_code') or '未知'
            rd_number = rec.get('rd_number', '')
            project_code = rec.get('project_code', '')

            # 构造唯一标识
            if rd_number and project_code:
                unique_key = f'{project_code}_{rd_number}'
            elif rec.get('phone'):
                unique_key = rec['phone']
            elif rec.get('email'):
                unique_key = rec['email']
            else:
                unique_key = make_fingerprint(rec)

            try:
                # 跳过没有任何有效标识的记录
                phone = rec.get('phone', '')[:20]
                email = rec.get('email', '')[:100]
                signup_id = rec.get('signup_id', '')
                if not rd_number and not phone and not email and not signup_id:
                    continue

                subject = None
                cand_no = ''
                if rd_number and project_code:
                    cand_no = f'{project_code}-{rd_number}'[:20]
                    subject = Subject.objects.filter(subject_no=cand_no).first()

                if subject is None and phone:
                    subject = Subject.objects.filter(phone=phone).first()

                if subject is None and email:
                    subject = Subject.objects.filter(
                        subjectprofile__email=email
                    ).first()

                gender = rec.get('gender', '')
                age = rec.get('age')

                if subject is None:
                    # 没有姓名时用报名ID作为临时名称
                    display_name = name[:100] if name and name not in ('未知',) else (
                        f'{project_code}-报名{signup_id}' if signup_id and project_code else
                        f'报名{signup_id}' if signup_id else '未知'
                    )
                    subject = Subject(
                        name=display_name,
                        gender=gender or '',
                        age=age,
                        phone=phone,
                        source_channel=SubjectSourceChannel.DATABASE,
                        status=SubjectStatus.COMPLETED if source in ('chinanorm', 'settlement') else SubjectStatus.PRE_SCREENED,
                    )
                    if cand_no:
                        subject.subject_no = cand_no
                    if not self.dry_run:
                        subject.save()
                    self.counters['new_subjects'] += 1
                    saved += 1
                else:
                    changed = False
                    if not subject.gender and gender:
                        subject.gender = gender
                        changed = True
                    if not subject.age and age:
                        subject.age = age
                        changed = True
                    if not self.dry_run and changed:
                        subject.save(update_fields=['gender', 'age', 'update_time'])
                    self.counters['updated'] += 1
                    saved += 1

                # 更新 SkinProfile
                if not self.dry_run and subject.pk:
                    skin_type = rec.get('skin_type', '')
                    fitzpatrick = rec.get('fitzpatrick', '')
                    if skin_type or fitzpatrick:
                        skin_profile, _ = SkinProfile.objects.get_or_create(subject=subject)
                        update_fields = []
                        if fitzpatrick and not skin_profile.fitzpatrick_type:
                            skin_profile.fitzpatrick_type = fitzpatrick
                            update_fields.append('fitzpatrick_type')
                        if skin_type and not skin_profile.skin_type_u_zone:
                            skin_profile.skin_type_u_zone = skin_type
                            update_fields.append('skin_type_u_zone')
                        if update_fields:
                            skin_profile.save(update_fields=update_fields + ['update_time'])

                    # 关联 Protocol
                    if project_code:
                        if project_code not in proto_cache:
                            proto_cache[project_code] = Protocol.objects.filter(code__iexact=project_code).first()
                        protocol = proto_cache.get(project_code)
                        if protocol:
                            from apps.subject.models import Enrollment, EnrollmentStatus
                            Enrollment.objects.get_or_create(
                                subject=subject, protocol=protocol,
                                defaults={
                                    'status': EnrollmentStatus.COMPLETED,
                                    'enrolled_at': timezone.now(),
                                }
                            )

            except Exception as e:
                logger.warning('upsert 失败: %s, rec=%s', e, rec)

        return saved

    # ==================================================================
    # Profile: 生成全景档案知识条目
    # ==================================================================
    def phase_profile(self):
        from apps.subject.models import Subject, Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord
        from apps.subject.models_domain import SkinProfile
        from apps.knowledge.models import KnowledgeEntry

        subjects = Subject.objects.filter(is_deleted=False)
        total = subjects.count()
        self.stdout.write(f'处理 {total} 名受试者')

        created = updated = 0
        batch = 500
        for i in range(0, total, batch):
            chunk = subjects[i:i + batch]
            for subject in chunk:
                text = self._build_full_lifecycle_profile(subject)
                if not text:
                    continue

                if self.dry_run:
                    created += 1
                    continue

                _, is_new = KnowledgeEntry.objects.update_or_create(
                    source_type='subject_full_lifecycle',
                    source_id=subject.id,
                    source_key=f'lifecycle_{subject.id}',
                    defaults={
                        'entry_type': 'lesson_learned',
                        'title': f'{subject.name} 受试者全档案',
                        'content': text,
                        'status': 'published',
                        'is_published': True,
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            self.stdout.write(f'  进度: {min(i+batch, total)}/{total}')

        self.counters['profiles'] = created + updated
        self.stdout.write(self.style.SUCCESS(f'Profile: 新建 {created}, 更新 {updated}'))

    def _build_full_lifecycle_profile(self, subject):
        from apps.subject.models import Enrollment
        from apps.subject.models_timeseries import SkinMeasurementRecord
        from apps.subject.models_execution import SubjectQuestionnaire, ComplianceRecord
        from apps.subject.models_domain import SkinProfile

        lines = [
            f'受试者编号: {subject.subject_no or subject.id}',
            f'基本信息: {subject.get_gender_display() or "未知性别"}, {subject.age or "?"}岁',
        ]
        if subject.phone:
            lines.append(f'联系方式: 手机尾号 {subject.phone[-4:] if len(subject.phone) >= 4 else ""}')

        # 皮肤档案
        skin = SkinProfile.objects.filter(subject=subject).first()
        if skin:
            skin_parts = []
            if skin.fitzpatrick_type:
                skin_parts.append(f'Fitzpatrick {skin.fitzpatrick_type}型')
            if skin.skin_type_u_zone:
                skin_parts.append(f'{skin.skin_type_u_zone}性肌')
            if skin.moisture_baseline:
                skin_parts.append(f'水分基线 {skin.moisture_baseline}')
            if skin_parts:
                lines.append(f'皮肤特征: {", ".join(skin_parts)}')

        # 项目参与历史
        enrollments = Enrollment.objects.filter(subject=subject).select_related('protocol')
        if enrollments.exists():
            projects = []
            for e in enrollments[:8]:
                if e.protocol:
                    status = e.get_status_display()
                    projects.append(f'{e.protocol.code}[{status}]')
            lines.append(f'参与项目({enrollments.count()}个): {", ".join(projects)}')

        # 仪器测量
        m_count = SkinMeasurementRecord.objects.filter(subject=subject).count()
        if m_count > 0:
            latest = SkinMeasurementRecord.objects.filter(subject=subject).order_by('-measured_at').first()
            m_parts = []
            for field, label in [('moisture', '水分'), ('tewl', 'TEWL'), ('sebum', '皮脂'), ('elasticity', '弹性')]:
                val = getattr(latest, field, None)
                if val:
                    m_parts.append(f'{label} {val}')
            lines.append(f'仪器测量({m_count}次): {", ".join(m_parts) if m_parts else "有记录"}')

        # 问卷
        q_count = SubjectQuestionnaire.objects.filter(subject=subject).count()
        if q_count > 0:
            lines.append(f'问卷记录: {q_count} 份')

        # 依从性
        compliance = ComplianceRecord.objects.filter(subject=subject).order_by('-assessment_date').first()
        if compliance:
            lines.append(f'依从性: {compliance.get_level_display()}, 到访率 {compliance.visit_attendance_rate}%')

        return '\n'.join(lines) if len(lines) > 2 else None

    # ==================================================================
    # KG: 知识图谱 — 受试者全生命周期
    # ==================================================================
    def phase_kg(self):
        from apps.subject.models import Subject, Enrollment
        from apps.knowledge.models import KnowledgeEntity, KnowledgeRelation

        self.stdout.write('构建受试者全生命周期知识图谱...')
        enrollments = Enrollment.objects.select_related('subject', 'protocol').all()
        total = enrollments.count()
        self.stdout.write(f'处理 {total} 条入组关系')

        e_created = r_created = 0

        for enrollment in enrollments:
            subject = enrollment.subject
            protocol = enrollment.protocol
            if not protocol:
                continue

            # 受试者实体
            uri = f'cnkis:subject/{subject.subject_no or subject.id}'
            subj_entity, s_new = KnowledgeEntity.objects.get_or_create(
                uri=uri, namespace='cnkis',
                defaults={
                    'entity_type': 'instance',
                    'label': subject.subject_no or f'受试者{subject.id}',
                    'definition': f'{subject.name}, {subject.get_gender_display() or ""}, {subject.age or "?"}岁',
                    'properties': {
                        'subject_id': subject.id, 'gender': subject.gender,
                        'age': subject.age, 'skin_type': subject.skin_type,
                    },
                }
            )
            if s_new:
                e_created += 1

            # 项目实体
            proj_entity = KnowledgeEntity.objects.filter(
                entity_type='project', uri__contains=protocol.code,
            ).first()
            if not proj_entity:
                proj_entity, p_new = KnowledgeEntity.objects.get_or_create(
                    uri=f'cnkis:project/{protocol.code}', namespace='cnkis',
                    defaults={
                        'entity_type': 'project',
                        'label': protocol.code,
                        'definition': protocol.title[:200],
                    }
                )
                if p_new:
                    e_created += 1

            # 参与项目关系（enrolled_in）
            _, r_new = KnowledgeRelation.objects.get_or_create(
                subject=subj_entity, object=proj_entity,
                predicate_uri='cnkis:enrolled_in',
                relation_type='custom',
                defaults={'source': 'mine_feishu', 'metadata': {
                    'enrollment_id': enrollment.id,
                    'status': enrollment.status,
                }},
            )
            if r_new:
                r_created += 1

            # 肤质分型关系
            if subject.skin_type:
                skin_entity, _ = KnowledgeEntity.objects.get_or_create(
                    uri=f'cnkis:fitzpatrick/{subject.skin_type}', namespace='cnkis',
                    defaults={
                        'entity_type': 'concept',
                        'label': f'Fitzpatrick {subject.skin_type}',
                        'definition': f'Fitzpatrick 皮肤分型 {subject.skin_type}',
                    }
                )
                KnowledgeRelation.objects.get_or_create(
                    subject=subj_entity, object=skin_entity,
                    predicate_uri='cnkis:has_skin_type', relation_type='has_property',
                    defaults={'source': 'mine_feishu'},
                )

            # 年龄段关系
            if subject.age:
                group = (
                    '18-25' if subject.age < 25 else
                    '25-35' if subject.age < 35 else
                    '35-45' if subject.age < 45 else
                    '45-55' if subject.age < 55 else '55+'
                )
                age_entity, _ = KnowledgeEntity.objects.get_or_create(
                    uri=f'cnkis:age_group/{group}', namespace='cnkis',
                    defaults={
                        'entity_type': 'concept',
                        'label': f'{group}岁', 'definition': f'{group}岁年龄段',
                    }
                )
                KnowledgeRelation.objects.get_or_create(
                    subject=subj_entity, object=age_entity,
                    predicate_uri='cnkis:belongs_to_age_group', relation_type='part_of',
                    defaults={'source': 'mine_feishu'},
                )

        self.counters['kg_entities'] = e_created
        self.counters['kg_relations'] = r_created
        self.stdout.write(self.style.SUCCESS(
            f'KG: 新增 {e_created} 实体, {r_created} 关系'
        ))

    # ==================================================================
    # 统计
    # ==================================================================
    def _show_stats(self):
        c = connection.cursor()
        print('\n=== 受试者数据统计 ===')
        for table, label in [
            ('t_subject', '受试者'),
            ('t_enrollment', '入组记录'),
            ('t_subject_skin_profile', '皮肤档案'),
            ('t_skin_measurement_record', '仪器测量'),
            ('t_subject_questionnaire', '问卷'),
            ('t_subject_compliance', '依从性'),
        ]:
            try:
                c.execute(f'SELECT COUNT(*) FROM {table}')
                print(f'  {label:20s}: {c.fetchone()[0]:>10,}')
            except Exception:
                pass

        c.execute("SELECT COUNT(*) FROM t_knowledge_entry WHERE source_type='subject_full_lifecycle'")
        print(f'  {"全景档案条目":20s}: {c.fetchone()[0]:>10,}')
        c.execute("SELECT COUNT(*) FROM t_knowledge_entry WHERE source_type='subject_beauty_profile'")
        print(f'  {"美丽画像条目":20s}: {c.fetchone()[0]:>10,}')
        c.execute("SELECT COUNT(*) FROM t_knowledge_entity WHERE entity_type='instance'")
        print(f'  {"图谱受试者实体":20s}: {c.fetchone()[0]:>10,}')
        c.execute("SELECT COUNT(*) FROM t_knowledge_relation WHERE predicate_uri='cnkis:enrolled_in'")
        print(f'  {"图谱参与关系":20s}: {c.fetchone()[0]:>10,}')
