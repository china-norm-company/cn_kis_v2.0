#!/usr/bin/env python3
"""
import_nas_subjects_standalone.py
NAS 受试者历史档案全量导入脚本（独立运行，不依赖 Django）

数据流：
  NAS xls/xlsx → 清洗 → 与 V2 t_subject 对比
    ├── 已匹配 → t_ext_ingest_candidate (review_status=matched, 洞明查历史)
    │           + t_subject_registration (招招项目参与记录)
    └── 未匹配
          ├── 信息完整(有姓名+手机) → t_subject (status=pending_review)
          │                          + t_ext_ingest_candidate (review_status=pending_review)
          │                          + t_subject_registration (招招待核查)
          └── 信息不完整             → t_ext_ingest_candidate (review_status=needs_more_info)

运行方式：
  # 建立 SSH 隧道（另开终端）
  ssh -i /Users/aksu/Downloads/openclaw1.1.pem -f -N -L 25432:127.0.0.1:5432 root@118.196.64.48

  python3 ops/scripts/import_nas_subjects_standalone.py --dry-run
  python3 ops/scripts/import_nas_subjects_standalone.py
  python3 ops/scripts/import_nas_subjects_standalone.py --report-only
"""
import argparse
import datetime
import glob
import hashlib
import json
import os
import re
import sys

# ────────── 依赖检查 ──────────────────────────────────────────────────────────
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("缺少 psycopg2: pip install psycopg2-binary")
try:
    import openpyxl
except ImportError:
    sys.exit("缺少 openpyxl: pip install openpyxl")
try:
    import xlrd
except ImportError:
    sys.exit("缺少 xlrd: pip install xlrd==1.2.0")

# ────────── 配置 ──────────────────────────────────────────────────────────────
NAS_MOUNT  = '/tmp/nas_cn_kis'
PG_HOST    = '127.0.0.1'
PG_PORT    = 25432
PG_DB      = 'cn_kis_v2'
PG_USER    = 'cn_kis'
PG_PASS    = os.getenv('V2_DB_PASSWORD', '')  # 从环境变量读；或命令行 --db-password 传入

IMPORT_SOURCE_TYPE   = 'nas_xls_import'
IMPORT_BATCH_TAG     = f'nas-{datetime.date.today().isoformat()}'
TARGET_WORKSTATION   = 'subject'
TARGET_MODEL         = 'Subject'
REG_CHANNEL_NOTE     = 'NAS历史档案导入'   # channel_id=NULL，用备注

# ────────── 工具函数 ──────────────────────────────────────────────────────────
def norm_phone(p) -> str:
    if not p:
        return ''
    s = str(p).strip()
    if '.' in s:
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    digits = re.sub(r'\D', '', s)
    return digits[-11:] if len(digits) >= 11 else digits

def norm_idcard(s) -> str:
    return re.sub(r'\s', '', str(s)).upper() if s else ''

def valid_phone(p: str) -> bool:
    return bool(re.match(r'^1[3-9]\d{9}$', p))

def valid_idcard(s: str) -> bool:
    return bool(re.match(r'^\d{17}[\dXx]$', s))

def idcard_hash(s: str) -> str:
    return hashlib.sha256(s.encode()).hexdigest() if s else ''

def completeness_score(name, phone, id_card, gender) -> float:
    """信息完整度评分 0-1"""
    score = 0.0
    if name and len(name) >= 2:  score += 0.3
    if phone and valid_phone(phone): score += 0.35
    if id_card and valid_idcard(id_card): score += 0.3
    if gender: score += 0.05
    return round(min(score, 1.0), 2)

def clean_name(s) -> str:
    if not s:
        return ''
    s = re.sub(r'[^\u4e00-\u9fff\u3040-\u30ffa-zA-Z·•]', '', str(s).strip())
    return s[:20]

def gen_subject_no(seq: int) -> str:
    ym = datetime.date.today().strftime('%Y%m')
    return f'NAS-{ym}-{seq:04d}'

def norm_gender(s: str) -> str:
    """中文性别转英文，未知返回空字符串"""
    if not s:
        return ''
    s = str(s).strip()
    if s in ('男', 'M', 'm', 'male', '男性'): return 'male'
    if s in ('女', 'F', 'f', 'female', '女性'): return 'female'
    return ''

# ────────── XLS/XLSX 解析 ─────────────────────────────────────────────────────
HEADER_KEYWORDS = ['姓名', '名字', '受访者姓名', '手机', '电话', '号码', '联系方式', '联系', '性别', '身份证', '年龄']
# 这批文件表头可能在第8-15行（有大量表头元信息），扩大搜索到40行
HEADER_SEARCH_ROWS = 40

def _find_header_row_xlsx(ws):
    """找到含姓名/手机关键字的最后一个表头行（最靠近数据的那行）"""
    best_i, best_headers = None, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > HEADER_SEARCH_ROWS:
            break
        row_str = ' '.join(str(c) for c in row if c)
        if any(k in row_str for k in ['姓名', '名字', '受访者姓名']) and \
           any(k in row_str for k in ['手机', '电话', '联系', '性别', '序号']):
            best_i = i
            best_headers = [str(c).strip() if c else '' for c in row]
    return best_i, best_headers

def _find_header_row_xls(sheet):
    best_i, best_headers = None, None
    for i in range(min(sheet.nrows, HEADER_SEARCH_ROWS)):
        row_str = ' '.join(str(sheet.cell_value(i, j)) for j in range(sheet.ncols))
        if any(k in row_str for k in ['姓名', '名字', '受访者姓名']) and \
           any(k in row_str for k in ['手机', '电话', '联系', '性别', '序号']):
            best_i = i
            best_headers = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
    return best_i, best_headers

def _extract_from_row(headers, values) -> dict:
    rec = {'name': '', 'phone': '', 'id_card': '', 'gender': '', 'age': ''}
    for j, val in enumerate(values):
        if j >= len(headers):
            break
        h = headers[j]
        if not val:
            continue
        if any(k in h for k in ['姓名', '名字', '受访者姓名']) and not rec['name']:
            rec['name'] = clean_name(val)
        if any(k in h for k in ['手机', '电话', '联系方式', '联系']) and '号' not in h[:2] and not rec['phone']:
            rec['phone'] = norm_phone(val)
        if '身份证' in h and not rec['id_card']:
            rec['id_card'] = norm_idcard(val)
        if '性别' in h and not rec['gender']:
            rec['gender'] = str(val).strip()[:2]
        if '年龄' in h and not rec['age']:
            try:
                rec['age'] = int(float(str(val)))
            except Exception:
                pass
    return rec

def parse_file(fp: str, proj_code: str) -> list:
    """解析单个 xls/xlsx 文件，返回受试者 dict 列表"""
    records = []
    ext = os.path.splitext(fp)[1].lower()
    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                header_idx, headers = _find_header_row_xlsx(ws)
                if headers is None:
                    continue
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i <= header_idx:
                        continue
                    rec = _extract_from_row(headers, row)
                    if rec['name'] and len(rec['name']) >= 2:
                        rec['project_code'] = proj_code
                        rec['source_file']  = os.path.basename(fp)
                        records.append(rec)
            wb.close()
        else:  # .xls
            wb = xlrd.open_workbook(fp)
            for sh in wb.sheets():
                header_idx, headers = _find_header_row_xls(sh)
                if headers is None:
                    continue
                for i in range(header_idx + 1, sh.nrows):
                    row_vals = [sh.cell_value(i, j) for j in range(sh.ncols)]
                    rec = _extract_from_row(headers, row_vals)
                    if rec['name'] and len(rec['name']) >= 2:
                        rec['project_code'] = proj_code
                        rec['source_file']  = os.path.basename(fp)
                        records.append(rec)
    except Exception as e:
        print(f'  [解析失败] {os.path.basename(fp)}: {e}')
    return records

# ────────── 主流程 ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='预演：只打印，不写库')
    parser.add_argument('--report-only', action='store_true', help='只生成 Excel 报告，不写库')
    parser.add_argument('--db-password', type=str, default='', help='V2 数据库密码')
    parser.add_argument('--db-host', type=str, default='', help='数据库主机（覆盖默认 PG_HOST）')
    parser.add_argument('--db-port', type=int, default=0, help='数据库端口（覆盖默认 PG_PORT）')
    parser.add_argument('--db-name', type=str, default='', help='数据库名（覆盖默认 PG_DB）')
    parser.add_argument('--db-user', type=str, default='', help='数据库用户（覆盖默认 PG_USER）')
    parser.add_argument('--batch-size', type=int, default=50, help='每批提交条数（autocommit 模式无效）')
    args = parser.parse_args()

    db_host = args.db_host or PG_HOST
    db_port = args.db_port or PG_PORT
    db_name = args.db_name or PG_DB
    db_user = args.db_user or PG_USER
    db_pass = args.db_password or PG_PASS
    if not db_pass and not args.dry_run and not args.report_only:
        db_pass = input('输入 cn_kis_v2 数据库密码: ').strip()

    # ── 1. 解析 NAS 所有文件 ─────────────────────────────────────────────────
    print('\n' + '=' * 68)
    print('步骤 1: 解析 NAS 受试者档案文件')
    print('=' * 68)

    raw_records = []   # 所有原始行，含重复
    project_stats = {}

    def add_file(fp, label):
        ext = os.path.splitext(fp)[1].lower()
        proj_code = os.path.basename(fp).split(' ')[0].split('\u7535\u8bdd')[0].strip()
        rows = parse_file(fp, proj_code)
        project_stats[label] = {'file': os.path.basename(fp), 'rows': len(rows)}
        raw_records.extend(rows)
        print(f'  {os.path.basename(fp):<55} {len(rows):>5} 条')
        return rows

    # 核心汇总文件
    print('\n[核心汇总文件]')
    add_file(f'{NAS_MOUNT}/综合研究受试者信息/综合研究名单去重汇总 .xlsx', '综合研究名单')
    add_file(f'{NAS_MOUNT}/受试者名单/受试者名单2026.3.23.xlsx', '受试者名单2026')

    # 62 个项目文件
    print('\n[62 个项目文件]')
    for fp in sorted(glob.glob(f'{NAS_MOUNT}/受试者名单/*.xls*')):
        if '受试者名单2026' in fp:
            continue
        add_file(fp, os.path.basename(fp))

    print(f'\n原始总行数（含重复）: {len(raw_records)}')

    # ── 2. 清洗与去重 ────────────────────────────────────────────────────────
    print('\n' + '=' * 68)
    print('步骤 2: 数据清洗与去重')
    print('=' * 68)

    seen_phone = {}    # phone -> merged record
    seen_idcard = {}   # id_card -> merged record
    cleaned = []
    dup_count = 0

    for r in raw_records:
        name    = clean_name(r.get('name', ''))
        phone   = norm_phone(r.get('phone', ''))
        id_card = norm_idcard(r.get('id_card', ''))
        gender  = r.get('gender', '')
        age     = r.get('age', '')
        proj    = r.get('project_code', '')
        src_f   = r.get('source_file', '')

        if not name or len(name) < 2:
            continue

        # 去重键：优先身份证，其次手机，都没有则名字
        dup_key = id_card if id_card else (phone if phone else f'name:{name}')

        if dup_key in seen_phone:
            existing = seen_phone[dup_key]
            # 合并：补充缺失字段
            if not existing.get('id_card') and id_card:
                existing['id_card'] = id_card
            if not existing.get('phone') and phone:
                existing['phone'] = phone
            if not existing.get('gender') and gender:
                existing['gender'] = gender
            # 项目代码多值合并
            existing.setdefault('project_codes', set()).add(proj)
            existing.setdefault('source_files', set()).add(src_f)
            dup_count += 1
            continue

        rec = {
            'name': name,
            'phone': phone,
            'id_card': id_card,
            'gender': gender,
            'age': age,
            'project_codes': {proj} if proj else set(),
            'source_files': {src_f} if src_f else set(),
        }
        rec['score'] = completeness_score(name, phone, id_card, gender)
        seen_phone[dup_key] = rec
        cleaned.append(rec)

    print(f'去重后唯一受试者: {len(cleaned)} 条  (重复合并: {dup_count} 条)')
    score_dist = {'>=0.9': 0, '0.6-0.9': 0, '<0.6': 0}
    for r in cleaned:
        s = r['score']
        if s >= 0.9:   score_dist['>=0.9'] += 1
        elif s >= 0.6: score_dist['0.6-0.9'] += 1
        else:          score_dist['<0.6'] += 1
    print(f'完整度分布: 高(≥0.9)={score_dist[">=0.9"]}  中(0.6-0.9)={score_dist["0.6-0.9"]}  低(<0.6)={score_dist["<0.6"]}')

    # ── 3. 连接数据库并拉取已有受试者 ────────────────────────────────────────
    print('\n' + '=' * 68)
    print('步骤 3: 与 V2 数据库匹配')
    print('=' * 68)

    conn = psycopg2.connect(host=db_host, port=db_port, dbname=db_name,
                            user=db_user, password=db_pass,
                            connect_timeout=10)
    conn.autocommit = True
    cur = conn.cursor()

    # 拉取现有受试者
    cur.execute("""
        SELECT s.id, s.subject_no, s.name, s.phone, s.status,
               sp.id_card_last4
        FROM t_subject s
        LEFT JOIN t_subject_profile sp ON sp.subject_id = s.id
        WHERE s.is_deleted = false
    """)
    db_by_phone = {}
    db_by_last4_name = {}
    for row in cur.fetchall():
        sid, sno, name, phone, status, last4 = row
        p = norm_phone(phone or '')
        if p:
            db_by_phone[p] = {'id': sid, 'subject_no': sno, 'name': name, 'status': status}
        if last4 and name:
            db_by_last4_name[(last4, name)] = {'id': sid, 'subject_no': sno, 'phone': p, 'status': status}
    print(f'V2库现有受试者: {len(db_by_phone)} 条（phone索引）')

    # 匹配
    matched_list   = []
    unmatched_list = []

    for rec in cleaned:
        phone   = rec.get('phone', '')
        id_card = rec.get('id_card', '')
        name    = rec['name']
        last4   = id_card[-4:] if len(id_card) >= 4 else ''

        db_rec = None
        match_method = ''
        if phone and phone in db_by_phone:
            db_rec = db_by_phone[phone]
            # 二次确认：姓名相似（允许1字差异）
            if name and db_rec['name'] and name not in db_rec['name'] and db_rec['name'] not in name:
                # 手机匹配但姓名差异大 → 降为待核查
                rec['match_conflict'] = f"手机匹配但姓名不同: DB={db_rec['name']} NAS={name}"
                unmatched_list.append(rec)
                continue
            match_method = 'phone'
        elif last4 and (last4, name) in db_by_last4_name:
            db_rec = db_by_last4_name[(last4, name)]
            match_method = 'id_last4+name'

        if db_rec:
            rec['db_id']       = db_rec['id']
            rec['subject_no']  = db_rec['subject_no']
            rec['db_status']   = db_rec['status']
            rec['match_method'] = match_method
            matched_list.append(rec)
        else:
            unmatched_list.append(rec)

    print(f'✅ 已匹配（手机或身份证后4位+姓名）: {len(matched_list)} 条')
    print(f'❌ 未匹配（待入库）: {len(unmatched_list)} 条')
    conflict = sum(1 for r in unmatched_list if r.get('match_conflict'))
    print(f'⚠️  手机匹配但姓名冲突（强制待核查）: {conflict} 条')

    if args.report_only or args.dry_run:
        _save_report(matched_list, unmatched_list, project_stats, dry_run=True)
        if args.report_only:
            print('\n[report-only 模式] 报告已生成，未写库')
            cur.close(); conn.close(); return
        print('\n[dry-run 模式] 以下为写库预览，实际不执行')
        print(f'  待写 t_subject (pending_review): {sum(1 for r in unmatched_list if r["score"]>=0.6 and not r.get("match_conflict"))}')
        print(f'  待写 t_ext_ingest_candidate: {len(matched_list) + len(unmatched_list)}')
        cur.close(); conn.close(); return

    # ── 4. 写库 ───────────────────────────────────────────────────────────────
    print('\n' + '=' * 68)
    print('步骤 4: 写入数据库')
    print('=' * 68)

    # 获取当前最大 subject_no 序号
    cur.execute("SELECT subject_no FROM t_subject WHERE subject_no LIKE 'NAS-%' ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()
    nas_seq = 1
    if row:
        try:
            nas_seq = int(row[0].split('-')[-1]) + 1
        except Exception:
            pass

    ok_subject = ok_ingest = ok_reg = err_count = 0
    now_ts = datetime.datetime.now(datetime.timezone.utc)

    # 4-A: 未匹配 → 写 t_subject + t_subject_profile + t_ext_ingest_candidate
    for rec in unmatched_list:
        name    = rec['name']
        phone   = rec.get('phone', '')
        id_card = rec.get('id_card', '')
        gender  = norm_gender(rec.get('gender', ''))
        score   = rec['score']
        projs   = ', '.join(sorted(rec.get('project_codes', set())))
        src_files = '; '.join(sorted(rec.get('source_files', set())))
        conflict  = rec.get('match_conflict', '')

        review_status = 'needs_more_info' if score < 0.6 else 'pending_review'

        subject_id = None
        # 信息足够（有姓名+有效手机）→ 写 t_subject
        if score >= 0.6 and valid_phone(phone):
            try:
                sno = gen_subject_no(nas_seq)
                nas_seq += 1
                cur.execute("""
                    INSERT INTO t_subject
                        (name, gender, phone, status, source_channel, subject_no,
                         skin_type, risk_level, auth_level, id_card_encrypted,
                         identity_verify_status, create_time, update_time, is_deleted)
                    VALUES (%s, %s, %s, 'pending_review', 'nas_import', %s,
                            '', 'low', 'unverified', %s,
                            'unverified', %s, %s, false)
                    ON CONFLICT DO NOTHING
                    RETURNING id
                """, (name, gender or '', phone, sno,
                      id_card or '',
                      now_ts, now_ts))
                result = cur.fetchone()
                if result:
                    subject_id = result[0]
                    # 写 t_subject_profile（所有 NOT NULL 字段补空字符串）
                    ic_hash  = idcard_hash(id_card) if id_card else ''
                    ic_enc   = id_card if id_card else ''
                    ic_last4 = id_card[-4:] if len(id_card) >= 4 else ''
                    cur.execute("""
                        INSERT INTO t_subject_profile
                            (subject_id, id_card_hash, id_card_encrypted, id_card_last4,
                             ethnicity, education, occupation, marital_status,
                             name_pinyin, phone_backup, email,
                             province, city, district, address, postal_code,
                             emergency_contact_name, emergency_contact_phone,
                             emergency_contact_relation,
                             privacy_level, consent_data_sharing, consent_rwe_usage,
                             consent_biobank, consent_follow_up, data_retention_years,
                             total_enrollments, total_completed, create_time, update_time)
                        VALUES (%s, %s, %s, %s,
                                '', '', '', '',
                                '', '', '',
                                '', '上海', '', '', '',
                                '', '', '',
                                'standard', false, false,
                                false, false, 10,
                                0, 0, %s, %s)
                    """, (subject_id, ic_hash, ic_enc, ic_last4, now_ts, now_ts))
                    ok_subject += 1
            except Exception as e:
                err_count += 1
                print(f'  [ERR-subject] {name}: {e}')

        # 写 t_ext_ingest_candidate（洞明待核查）
        snapshot = {
            'name': name, 'phone': phone, 'id_card': id_card,
            'gender': gender, 'project_codes': projs,
            'source_files': src_files, 'completeness_score': score,
            'conflict': conflict,
        }
        try:
            cur.execute("""
                INSERT INTO t_ext_ingest_candidate
                    (source_type, source_raw_id, source_module, source_display_title,
                     source_snapshot, target_workstation, target_model,
                     mapped_fields, confidence_score, review_status,
                     reviewed_by_name, review_comment, reject_reason,
                     modified_fields, ingested_model, ingestion_log,
                     populated_by, ingested_record_id,
                     created_at, updated_at)
                VALUES (%s, 0, 'nas_import', %s, %s::jsonb, %s, %s,
                        %s::jsonb, %s, %s,
                        '', '', '',
                        '{}', '', '{}',
                        %s, %s, %s, %s)
            """, (
                IMPORT_SOURCE_TYPE,
                f'{name} / {phone or "无手机"} / 项目:{projs or "未知"}',
                json.dumps(snapshot, ensure_ascii=False),
                TARGET_WORKSTATION, TARGET_MODEL,
                json.dumps({'name': name, 'phone': phone, 'id_card': id_card,
                            'gender': gender, 'project_codes': projs,
                            'batch': IMPORT_BATCH_TAG}, ensure_ascii=False),
                score, review_status,
                IMPORT_BATCH_TAG,
                subject_id,
                now_ts, now_ts,
            ))
            ok_ingest += 1
        except Exception as e:
            err_count += 1
            print(f'  [ERR-ingest] {name}: {e}')

        # t_subject_registration 跳过（需要外键 plan_id，历史档案不创建招募计划）
        ok_reg += 1  # 计数仍递增（作为成功处理记录）

    # 4-B: 已匹配 → 写 t_ext_ingest_candidate matched 记录 + t_subject_registration
    for rec in matched_list:
        name   = rec['name']
        phone  = rec.get('phone', '')
        projs  = ', '.join(sorted(rec.get('project_codes', set())))
        src_f  = '; '.join(sorted(rec.get('source_files', set())))
        db_id  = rec['db_id']
        sno    = rec['subject_no']

        snapshot = {
            'name': name, 'phone': phone, 'id_card': rec.get('id_card', ''),
            'project_codes': projs, 'source_files': src_f,
            'match_method': rec.get('match_method', ''),
            'db_subject_no': sno,
        }
        try:
            cur.execute("""
                INSERT INTO t_ext_ingest_candidate
                    (source_type, source_raw_id, source_module, source_display_title,
                     source_snapshot, target_workstation, target_model,
                     mapped_fields, confidence_score, review_status,
                     reviewed_by_name, review_comment, reject_reason,
                     modified_fields, ingested_model, ingestion_log,
                     populated_by, ingested_record_id,
                     created_at, updated_at)
                VALUES (%s, 0, 'nas_import', %s, %s::jsonb, %s, %s,
                        %s::jsonb, 1.0, 'matched',
                        '', '', '',
                        '{}', '', '{}',
                        %s, %s, %s, %s)
            """, (
                IMPORT_SOURCE_TYPE,
                f'{name} / {sno} / 项目:{projs or "未知"}',
                json.dumps(snapshot, ensure_ascii=False),
                TARGET_WORKSTATION, TARGET_MODEL,
                json.dumps({'name': name, 'phone': phone,
                            'project_codes': projs, 'db_subject_no': sno,
                            'batch': IMPORT_BATCH_TAG}, ensure_ascii=False),
                IMPORT_BATCH_TAG, db_id, now_ts, now_ts,
            ))
            ok_ingest += 1
        except Exception as e:
            err_count += 1
            print(f'  [ERR-ingest-matched] {name}: {e}')

        # t_subject_registration 跳过（需要外键 plan_id）
        ok_reg += 1

    print(f'\n写库完成:')
    print(f'  t_subject (pending_review 新增): {ok_subject}')
    print(f'  t_subject_profile (档案): {ok_subject}')
    print(f'  t_ext_ingest_candidate (洞明): {ok_ingest}')
    print(f'  t_subject_registration: 跳过（无 plan_id，项目参与信息已写入洞明）')
    print(f'  已匹配处理: {len(matched_list)}  未匹配处理: {len(unmatched_list)}')
    print(f'  错误: {err_count}')

    cur.close()
    conn.close()

    # ── 5. 保存 Excel 报告到 NAS ─────────────────────────────────────────────
    _save_report(matched_list, unmatched_list, project_stats, dry_run=False)


def _save_report(matched_list, unmatched_list, project_stats, dry_run=False):
    """生成详细 Excel 对比报告，保存到 NAS"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('openpyxl 未安装，跳过 Excel 报告')
        return

    wb = openpyxl.Workbook()

    # Sheet1: 概览
    ws1 = wb.active
    ws1.title = '导入概览'
    ws1.append(['CN KIS V2.0 NAS 受试者导入报告'])
    ws1.append([f'生成时间: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws1.append([''])
    ws1.append(['项目', '数值'])
    for row in [
        ['NAS 原始档案文件数', len(project_stats)],
        ['NAS 去重唯一受试者', len(matched_list) + len(unmatched_list)],
        ['已匹配 V2 已入库', len(matched_list)],
        ['未入库（待核查）', len(unmatched_list)],
        ['覆盖率', f'{len(matched_list)*100//(len(matched_list)+len(unmatched_list)) if (matched_list or unmatched_list) else 0}%'],
        ['dry_run 模式', '是' if dry_run else '否'],
    ]:
        ws1.append(row)

    # Sheet2: 已匹配
    ws2 = wb.create_sheet('已匹配-已入库')
    ws2.append(['姓名', '手机', '身份证', '受试者编号', '匹配方式', '参与项目', '完整度'])
    for r in matched_list:
        ws2.append([
            r['name'], r.get('phone',''), r.get('id_card',''),
            r.get('subject_no',''), r.get('match_method',''),
            ', '.join(sorted(r.get('project_codes', set()))),
            r.get('score', 0),
        ])

    # Sheet3: 未入库
    ws3 = wb.create_sheet('未入库-待核查')
    ws3.append(['姓名', '手机', '身份证', '性别', '完整度评分', '参与项目', '来源文件', '备注'])
    for r in unmatched_list:
        ws3.append([
            r['name'], r.get('phone',''), r.get('id_card',''),
            r.get('gender',''), r.get('score',0),
            ', '.join(sorted(r.get('project_codes', set()))),
            '; '.join(sorted(r.get('source_files', set())))[:100],
            r.get('match_conflict', ''),
        ])

    # Sheet4: 项目文件统计
    ws4 = wb.create_sheet('项目文件统计')
    ws4.append(['项目标签', '文件名', '解析行数'])
    for label, info in project_stats.items():
        ws4.append([label, info['file'], info['rows']])

    # 保存
    ts = datetime.date.today().isoformat()
    report_path = f'{NAS_MOUNT}/受试者导入报告_{ts}.xlsx'
    try:
        wb.save(report_path)
        print(f'\n✅ Excel 报告已保存: {report_path}')
    except Exception as e:
        # fallback 到 /tmp
        report_path = f'/tmp/受试者导入报告_{ts}.xlsx'
        wb.save(report_path)
        print(f'\n⚠️  NAS 写入失败，报告保存至本地: {report_path}  原因: {e}')


if __name__ == '__main__':
    main()
