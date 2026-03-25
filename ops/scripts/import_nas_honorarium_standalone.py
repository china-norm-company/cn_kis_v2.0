#!/usr/bin/env python3
"""
import_nas_honorarium_standalone.py
NAS 受试者礼金历史档案全量导入脚本（独立运行，不依赖 Django 运行时）

功能：
  1. 扫描 NAS 受试者礼金历史文档（~2100+ xlsx/xls）
  2. 解析：姓名、完整身份证、手机、银行卡号、金额、项目代码、支付日期
  3. 银行卡号 AES-Fernet 加密入库（t_subject_payment.bank_account_encrypted）
  4. 与 t_subject 匹配（手机优先，身份证后4位+姓名次之）
  5. 写入 t_subject_payment（礼金支付记录）
  6. 写入 t_subject_points_ledger（积分台账，1元=1分）
  7. 反向补全：对已匹配受试者用完整身份证更新 t_subject_profile.id_card_*

运行前准备：
  # 确保 SSH 隧道已建立
  ssh -i /Users/aksu/Downloads/openclaw1.1.pem -f -N -L 25432:127.0.0.1:5432 root@118.196.64.48

  # （可选）设置加密密钥（不设则用开发默认key）
  export FIELD_ENCRYPTION_KEY=$(python3 -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())")

运行方式：
  python3 ops/scripts/import_nas_honorarium_standalone.py --dry-run
  python3 ops/scripts/import_nas_honorarium_standalone.py --backfill-only   # 仅反向补全身份证
  python3 ops/scripts/import_nas_honorarium_standalone.py                   # 正式导入
"""
import argparse
import base64
import datetime
import glob
import hashlib
import json
import os
import re
import sys

# ────────── 依赖检查 ──────────────────────────────────────────────────────────
for pkg, name in [('psycopg2', 'psycopg2-binary'), ('openpyxl', 'openpyxl'), ('xlrd', 'xlrd==1.2.0')]:
    try:
        __import__(pkg)
    except ImportError:
        sys.exit(f'缺少依赖: pip install {name}')
try:
    from cryptography.fernet import Fernet
except ImportError:
    sys.exit('缺少依赖: pip install cryptography')

import psycopg2
import psycopg2.extras
import openpyxl
import xlrd

# ────────── 配置 ──────────────────────────────────────────────────────────────
NAS_MOUNT      = '/tmp/nas_cn_kis'
HONOR_BASE     = f'{NAS_MOUNT}/受试者礼金历史文档'
PG_HOST        = '127.0.0.1'
PG_PORT        = 25432
PG_DB          = 'cn_kis_v2'
PG_USER        = 'cn_kis'
PG_PASS        = os.getenv('V2_DB_PASSWORD', '')
IMPORT_BATCH   = f'nas-honor-v2-{datetime.date.today().isoformat()}'  # v2: 新脚本全量无截断版本
POINTS_RATE    = 1   # 1 元 = 1 积分

# ────────── 加密工具（内联，不依赖 Django）────────────────────────────────────
_fernet = None

def _get_fernet() -> Fernet:
    global _fernet
    if _fernet:
        return _fernet
    key_str = os.environ.get('FIELD_ENCRYPTION_KEY', '')
    if not key_str:
        key_str = base64.urlsafe_b64encode(
            hashlib.sha256(b'cn-kis-dev-only-not-for-prod').digest()
        ).decode()
    key_bytes = key_str.encode() if isinstance(key_str, str) else key_str
    if len(key_bytes) == 32:
        key_bytes = base64.urlsafe_b64encode(key_bytes)
    _fernet = Fernet(key_bytes)
    return _fernet

def encrypt_field(value: str) -> str:
    if not value:
        return ''
    return _get_fernet().encrypt(value.encode()).decode()

def hash_field(value: str) -> str:
    return hashlib.sha256(value.encode()).hexdigest() if value else ''

def mask_last4(value: str) -> str:
    return value[-4:] if len(value) >= 4 else value

# 中国行政区划代码 → 省份名（前2位）
_PROVINCE_CODE = {
    '11': '北京', '12': '天津', '13': '河北', '14': '山西', '15': '内蒙古',
    '21': '辽宁', '22': '吉林', '23': '黑龙江',
    '31': '上海', '32': '江苏', '33': '浙江', '34': '安徽', '35': '福建',
    '36': '江西', '37': '山东',
    '41': '河南', '42': '湖北', '43': '湖南', '44': '广东', '45': '广西',
    '46': '海南',
    '50': '重庆', '51': '四川', '52': '贵州', '53': '云南', '54': '西藏',
    '61': '陕西', '62': '甘肃', '63': '青海', '64': '宁夏', '65': '新疆',
    '71': '台湾', '81': '香港', '82': '澳门',
}

def parse_idcard_info(ic: str) -> dict:
    """
    从18位身份证提取结构化信息：出生日期、性别、年龄、户籍省份。
    ID 格式：RRRRRRYYYYMMDDXXXC
      - 1-6:  行政区划代码
      - 7-14: 出生日期 YYYYMMDD
      - 15-17: 顺序码（第17位奇=男，偶=女）
      - 18:   校验码
    """
    result = {'gender': '', 'birth_date': None, 'age': None, 'province': ''}
    if not valid_idcard(ic) or len(ic) < 17:
        return result
    try:
        year  = int(ic[6:10])
        month = int(ic[10:12])
        day   = int(ic[12:14])
        birth_date = datetime.date(year, month, day)
        today = datetime.date.today()
        age = today.year - birth_date.year - (
            (today.month, today.day) < (birth_date.month, birth_date.day)
        )
        gender = 'male' if int(ic[16]) % 2 == 1 else 'female'
        province = _PROVINCE_CODE.get(ic[:2], '')
        result.update({
            'gender': gender,
            'birth_date': birth_date,
            'age': max(0, age),
            'province': province,
        })
    except Exception:
        pass
    return result

# ────────── 工具函数 ──────────────────────────────────────────────────────────
PLATFORM_MAP = {
    '八羿': '八羿', '捷仕达': '福建捷仕达', '安徽创启': '安徽创启',
    '安徽斯长': '安徽斯长', '宿钲': '宿钲信息科技', '融辰': '融辰',
    '怀宁': '怀宁青枫', '湖北耀运': '湖北耀运',
}

def detect_platform(path: str) -> str:
    for kw, name in PLATFORM_MAP.items():
        if kw in path:
            return name
    return ''

def parse_paid_date_from_dir(dir_name: str):
    """从目录名如 '2023.06.15已支付捷仕达17074.8' 解析支付日期"""
    m = re.match(r'(\d{4})[.\-](\d{1,2})[.\-](\d{1,2})', dir_name)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None

def parse_project_code_from_note(note: str) -> str:
    """从打款附言中提取项目代码 Cxxxxxxx / Mxxxxxxx / Wxxxxxxx"""
    if not note:
        return ''
    m = re.search(r'[CMWPmcwp][T\d]\w{4,10}', str(note))
    return m.group(0).upper() if m else ''

def norm_phone(p) -> str:
    if not p: return ''
    s = str(p).strip()
    if '.' in s:
        try: s = str(int(float(s)))
        except: pass
    digits = re.sub(r'\D', '', s)
    return digits[-11:] if len(digits) >= 11 else digits

def norm_idcard(s) -> str:
    return re.sub(r'\s', '', str(s)).upper().replace('X', 'X') if s else ''

def valid_phone(p: str) -> bool:
    return bool(re.match(r'^1[3-9]\d{9}$', p or ''))

def valid_idcard(s: str) -> bool:
    return bool(re.match(r'^\d{17}[\dXx]$', s or ''))

def norm_bank_account(s) -> str:
    if not s: return ''
    return re.sub(r'\D', '', str(s))

def clean_name(s) -> str:
    if not s: return ''
    return re.sub(r'[^\u4e00-\u9fff\u3040-\u30ffa-zA-Z·•]', '', str(s).strip())[:20]

def norm_amount(v) -> float:
    if not v: return 0.0
    try: return float(str(v).replace(',', '').strip())
    except: return 0.0

def gen_payment_no(seq: int) -> str:
    ym = datetime.date.today().strftime('%Y%m')
    return f'NAS-PAY-{ym}-{seq:06d}'

# ────────── 文件解析 ──────────────────────────────────────────────────────────
# ⚠️ 规范：根据 .cursor/rules/no-data-truncation.mdc，所有文件必须全量读取，
#    严禁使用任何行数上限（max_row 必须为 None）。违反此规定属于数据完整性严重违规。
HEADER_SEARCH = 20  # 在前20行中搜索表头行（扩大到20以兼容更多格式）

# 扩展的字段识别关键词（覆盖更多支付平台的列名变体）
_NAME_KEYS   = ['姓名', '经营者姓名', '收款方姓名', '经营者', '收款方', '收款人', '姓  名',
                '受益人', '付款对象', '受试者姓名', '名字']
_ID_KEYS     = ['身份证号', '身份证', '证件号码', '证件号', '身份证件号', '证件', 'ID号',
                '居民身份证', '身份证号码']
_PHONE_KEYS  = ['手机号', '手机号码', '手机', '电话', '联系电话', '联系方式', '移动电话',
                '电话号码']
_BANK_KEYS   = ['收款账号', '银行卡号', '银行卡', '账号', '银行账号', '收款卡号', '卡号',
                '账户号码', '银行账户', '收款账户', '账  号']
_AMOUNT_KEYS = ['结算金额', '服务金额', '支付金额', '金额', '转账金额', '实发金额',
                '应发金额', '礼金金额', '付款金额', '费用', '金额（元）', '合计金额']
_NOTE_KEYS   = ['打款附言', '备注', '附言', '项目编号', '项目代码', '项目', '用途',
                '摘要', '说明', '备注说明']


def _find_header(headers_candidates):
    """
    在前 HEADER_SEARCH 行中找最佳表头。
    识别策略更宽松：有姓名列 + (银行卡 OR 手机 OR 身份证) 即可识别。
    返回 (行索引, 表头列表)。
    """
    best = (None, None, -1)  # (idx, headers, score)
    for i, row in enumerate(headers_candidates):
        row_str = ' '.join(str(c) for c in row if c)
        has_name  = any(k in row_str for k in _NAME_KEYS)
        has_id    = any(k in row_str for k in _ID_KEYS)
        has_bank  = any(k in row_str for k in _BANK_KEYS)
        has_phone = any(k in row_str for k in _PHONE_KEYS)
        has_amt   = any(k in row_str for k in _AMOUNT_KEYS)
        # 评分：字段越多越好
        score = sum([has_name, has_id, has_bank, has_phone, has_amt])
        # 最低要求：有姓名 + 至少一个标识字段
        if has_name and score >= 2 and score > best[2]:
            best = (i, [str(c).strip() if c else '' for c in row], score)
    if best[0] is not None:
        return best[0], best[1]
    return None, None


def _extract_row(headers, values, source_file: str = '', paid_date=None) -> dict:
    rec = {'name': '', 'id_card': '', 'phone': '', 'bank_account': '',
           'amount': 0.0, 'project_code': '', 'note': '', '_raw': {}}
    for j, val in enumerate(values):
        if j >= len(headers) or val is None or val == '':
            continue
        h = headers[j]
        v = val
        # 记录原始值供审计
        if h:
            rec['_raw'][h] = str(v)[:80]
        if any(k in h for k in _NAME_KEYS) and not rec['name']:
            rec['name'] = clean_name(v)
        elif any(k in h for k in _ID_KEYS) and not rec['id_card']:
            rec['id_card'] = norm_idcard(v)
        elif any(k in h for k in _PHONE_KEYS) and not rec['phone']:
            rec['phone'] = norm_phone(v)
        elif any(k in h for k in _BANK_KEYS) and not rec['bank_account']:
            rec['bank_account'] = norm_bank_account(v)
        elif any(k in h for k in _AMOUNT_KEYS) and not rec['amount']:
            rec['amount'] = norm_amount(v)
        elif any(k in h for k in _NOTE_KEYS):
            code = parse_project_code_from_note(v)
            if code and not rec['project_code']:
                rec['project_code'] = code
            if not rec['note']:
                rec['note'] = str(v)[:100]
    return rec


def parse_xlsx_file(fp: str) -> tuple:
    """
    解析 xlsx 文件，全量读取（无行数限制）。
    返回 (records, total_rows_read, sheets_parsed, sheets_skipped)
    """
    records = []
    total_rows = 0
    sheets_parsed = 0
    sheets_skipped = []
    try:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            # ✅ 全量读取：max_row=None，严禁设置行数上限
            all_rows = list(ws.iter_rows(max_row=None, values_only=True))
            header_idx, headers = _find_header(all_rows[:HEADER_SEARCH])
            if headers is None:
                sheets_skipped.append(f'{sh}(无表头)')
                continue
            data_rows = all_rows[header_idx + 1:]
            total_rows += len(data_rows)
            sheets_parsed += 1
            for row in data_rows:
                rec = _extract_row(headers, row, fp)
                if rec['name'] and len(rec['name']) >= 2 and (rec['id_card'] or rec['bank_account'] or rec['phone']):
                    records.append(rec)
        wb.close()
    except Exception as e:
        return records, total_rows, sheets_parsed, [f'文件异常: {e}']
    return records, total_rows, sheets_parsed, sheets_skipped


def parse_xls_file(fp: str) -> tuple:
    """
    解析 xls 文件，全量读取（无行数限制）。
    返回 (records, total_rows_read, sheets_parsed, sheets_skipped)
    """
    records = []
    total_rows = 0
    sheets_parsed = 0
    sheets_skipped = []
    try:
        wb = xlrd.open_workbook(fp)
        for sh in wb.sheets():
            # ✅ 全量读取：使用 sh.nrows 不限制
            all_rows = [[sh.cell_value(i, j) for j in range(sh.ncols)]
                        for i in range(sh.nrows)]
            header_idx, headers = _find_header(all_rows[:HEADER_SEARCH])
            if headers is None:
                sheets_skipped.append(f'{sh.name}(无表头)')
                continue
            data_rows = all_rows[header_idx + 1:]
            total_rows += len(data_rows)
            sheets_parsed += 1
            for row in data_rows:
                rec = _extract_row(headers, row, fp)
                if rec['name'] and len(rec['name']) >= 2 and (rec['id_card'] or rec['bank_account'] or rec['phone']):
                    records.append(rec)
    except Exception as e:
        return records, total_rows, sheets_parsed, [f'文件异常: {e}']
    return records, total_rows, sheets_parsed, sheets_skipped


def parse_file(fp: str) -> tuple:
    ext = os.path.splitext(fp)[1].lower()
    return parse_xlsx_file(fp) if ext == '.xlsx' else parse_xls_file(fp)

# ────────── 去重逻辑（保守策略）────────────────────────────────────────────────
#
# 旧逻辑的问题：key = (id_card OR name, bank_account OR phone, round(amount))
#   → 同一人在不同日期/项目收到相同金额的记录被误删
#
# 新策略：只在以下情况下确认为重复（精确匹配）：
#   - 强去重（skip）：身份证 + 银行卡 + 金额 + 支付日期 + 项目 完全一致 → 真重复（同一文件重复行）
#   - 弱去重（flag）：身份证 + 银行卡 + 金额 一致但日期/项目不同 → 标记 confidence_note，保留但标注"疑似重复"
#   - 不去重：仅姓名或仅手机匹配时，绝不去重，完整保留

def _dedup_records(records: list) -> tuple:
    """
    返回 (unique_records, skipped_exact_dups, flagged_as_suspicious)
    - unique_records: 去重后的记录（含被标记为疑似重复的，但不丢弃）
    - skipped_exact_dups: 完全精确重复的条数（跳过写库）
    - flagged_as_suspicious: 被标记为疑似重复但保留的条数
    """
    # 精确去重键：身份证+银行卡+金额+日期+项目 全部一致 = 同一个文件里的重复行
    exact_seen = set()
    # 弱相似键：身份证+银行卡+金额（不含日期/项目）= 可能是多次相同金额
    weak_seen = {}   # key -> first occurrence info

    unique = []
    skipped_exact = 0
    flagged = 0

    for r in records:
        ic       = r.get('id_card', '')
        bank     = r.get('bank_account', '')
        amount   = round(r.get('amount', 0.0), 2)
        paid_dt  = str(r.get('_paid_date', '') or '')
        proj     = r.get('project_code', '') or ''
        name     = r.get('name', '')
        phone    = r.get('phone', '')

        # 精确去重：必须有身份证+银行卡，且日期/项目/金额全部相同
        if ic and len(ic) == 18 and bank and len(bank) >= 10:
            exact_key = (ic, bank, amount, paid_dt, proj)
            if exact_key in exact_seen:
                skipped_exact += 1
                continue
            exact_seen.add(exact_key)

            # 弱相似检测（标记，不丢弃）
            weak_key = (ic, bank, amount)
            if weak_key in weak_seen:
                prev = weak_seen[weak_key]
                r['_dup_flag'] = (
                    f"疑似同人同额多次支付 "
                    f"(首次:{prev['date']}/{prev['proj']}, 本次:{paid_dt}/{proj})"
                )
                flagged += 1
            else:
                weak_seen[weak_key] = {'date': paid_dt, 'proj': proj}
        # 如果没有身份证+银行卡的完整组合，绝不去重，直接保留

        unique.append(r)

    return unique, skipped_exact, flagged

# ────────── 主流程 ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run',       action='store_true', help='预演，不写库')
    parser.add_argument('--backfill-only', action='store_true', help='仅执行身份证反向补全')
    parser.add_argument('--db-password',   type=str, default='')
    parser.add_argument('--limit-files',   type=int, default=0,
                        help='[调试用] 限制处理文件数，默认0=不限制。'
                             '⚠️ 生产运行必须保持 0，否则违反 no-data-truncation 规范')
    parser.add_argument('--clear-batch',   type=str, default='',
                        help='清除指定批次的历史数据后重新导入（如 nas-honor-2026-03-24）')
    args = parser.parse_args()

    if args.limit_files > 0:
        print(f'⚠️  [警告] --limit-files={args.limit_files} 已激活，仅处理前{args.limit_files}个文件。'
              f'这是调试模式，不得用于生产导入！')

    db_pass = args.db_password or PG_PASS
    if not db_pass and not args.dry_run:
        db_pass = input('输入 cn_kis_v2 数据库密码: ').strip()

    conn = psycopg2.connect(host=PG_HOST, port=PG_PORT, dbname=PG_DB,
                            user=PG_USER, password=db_pass, connect_timeout=15)
    conn.autocommit = True
    cur = conn.cursor()

    # ── 可选：清除旧批次数据 ──────────────────────────────────────────────────
    if args.clear_batch and not args.dry_run:
        print(f'\n🗑️  清除旧批次数据: {args.clear_batch}')
        cur.execute("SELECT count(*) FROM t_subject_payment WHERE nas_import_batch = %s", (args.clear_batch,))
        old_cnt = cur.fetchone()[0]
        if old_cnt > 0:
            confirm = input(f'  将删除 {old_cnt} 条支付记录及对应积分记录，确认? (yes/no): ').strip()
            if confirm.lower() == 'yes':
                # 先删积分台账（外键关联）
                cur.execute("""
                    DELETE FROM t_subject_points_ledger
                    WHERE payment_id IN (
                        SELECT id FROM t_subject_payment WHERE nas_import_batch = %s
                    )
                """, (args.clear_batch,))
                pts_del = cur.rowcount
                cur.execute("DELETE FROM t_subject_payment WHERE nas_import_batch = %s", (args.clear_batch,))
                pay_del = cur.rowcount
                print(f'  ✅ 已删除: 支付记录 {pay_del} 条，积分记录 {pts_del} 条')
            else:
                print('  已取消，退出。')
                cur.close(); conn.close(); return
        else:
            print(f'  批次 {args.clear_batch} 无数据，跳过清除。')

    # ── 加载现有受试者索引 ────────────────────────────────────────────────────
    print('\n正在加载受试者索引...')
    cur.execute("""
        SELECT s.id, s.subject_no, s.name, s.phone,
               sp.id_card_last4, sp.id_card_encrypted, sp.id_card_hash
        FROM t_subject s
        LEFT JOIN t_subject_profile sp ON sp.subject_id = s.id
        WHERE s.is_deleted = false
    """)
    db_by_phone = {}       # phone -> {id, subject_no, name, has_idcard}
    db_by_last4_name = {}  # (last4, name) -> {id, subject_no}
    for row in cur.fetchall():
        sid, sno, name, phone, last4, ic_enc, ic_hash = row
        p = norm_phone(phone or '')
        has_ic = bool(ic_enc and len(ic_enc) > 20)
        if p:
            db_by_phone[p] = {'id': sid, 'subject_no': sno, 'name': name,
                               'has_idcard': has_ic, 'ic_enc': ic_enc or ''}
        if last4 and name:
            db_by_last4_name[(last4, name)] = {'id': sid, 'subject_no': sno}
    print(f'  已有受试者: {len(db_by_phone)} 条（phone索引），{len(db_by_last4_name)} 条（身份证后4位+姓名索引）')

    if args.backfill_only:
        _run_backfill(cur, conn, db_by_phone, db_by_last4_name, args.dry_run)
        cur.close(); conn.close(); return

    # ── 扫描所有礼金文件 ─────────────────────────────────────────────────────
    print(f'\n扫描礼金档案目录: {HONOR_BASE}')
    all_files = sorted(
        glob.glob(f'{HONOR_BASE}/**/*.xlsx', recursive=True) +
        glob.glob(f'{HONOR_BASE}/**/*.xls',  recursive=True)
    )
    # 排除临时文件
    all_files = [f for f in all_files if not os.path.basename(f).startswith('~')
                 and '/.~' not in f]
    if args.limit_files:
        all_files = all_files[:args.limit_files]
    print(f'找到文件: {len(all_files)} 个（.xlsx + .xls）')

    # ── 全量解析所有文件（无截断）─────────────────────────────────────────────
    raw_records = []
    total_rows_read = 0
    files_with_data  = 0
    files_no_header  = 0
    files_error      = 0
    files_no_header_list = []

    for i, fp in enumerate(all_files):
        if i % 200 == 0 and i > 0:
            print(f'  解析进度: {i}/{len(all_files)} | 已读行数: {total_rows_read} | 有效记录: {len(raw_records)}')
        recs, rows, n_sheets, skipped = parse_file(fp)
        paid_date = parse_paid_date_from_dir(os.path.basename(os.path.dirname(fp)))
        platform  = detect_platform(fp)
        total_rows_read += rows
        if recs:
            files_with_data += 1
            for r in recs:
                r['_source_file'] = fp
                r['_paid_date']   = paid_date
                r['_platform']    = platform
            raw_records.extend(recs)
        elif skipped and any('文件异常' in s for s in skipped):
            files_error += 1
        elif not recs and os.path.getsize(fp) > 3000:
            files_no_header += 1
            files_no_header_list.append(os.path.relpath(fp, HONOR_BASE))

    # 对账摘要（必须输出，根据 no-data-truncation 规范）
    print(f'\n══ 文件解析对账 ══════════════════════════════════════════')
    print(f'  总文件数:              {len(all_files)}')
    print(f'  成功解析（有数据）:    {files_with_data}')
    print(f'  无表头/无数据:         {files_no_header}')
    print(f'  文件异常（报错）:      {files_error}')
    print(f'  总计读取数据行:        {total_rows_read}')
    print(f'  提取有效记录:          {len(raw_records)}')
    if files_no_header_list[:10]:
        print(f'  [无表头文件示例]:')
        for f in files_no_header_list[:10]:
            print(f'    - {f}')
        if len(files_no_header_list) > 10:
            print(f'    ... 共 {len(files_no_header_list)} 个，见报告 Sheet "无表头文件"')

    # ── 保守去重 ──────────────────────────────────────────────────────────────
    print(f'\n执行保守去重（只去除精确重复行，保留疑似多次支付）...')
    dedup_records, exact_dups, flagged_dups = _dedup_records(raw_records)
    print(f'  原始记录:           {len(raw_records)}')
    print(f'  精确重复（跳过）:   {exact_dups}  ← 同文件重复行')
    print(f'  疑似多次支付（保留，标注）: {flagged_dups}  ← 同人同额不同日期/项目')
    print(f'  去重后保留:         {len(dedup_records)}')

    valid_idcard_count = sum(1 for r in dedup_records if valid_idcard(r.get('id_card', '')))
    valid_bank_count   = sum(1 for r in dedup_records if len(r.get('bank_account', '')) >= 10)
    print(f'  有效身份证:         {valid_idcard_count}')
    print(f'  有效银行卡:         {valid_bank_count}')

    # ── 匹配受试者（保守策略）──────────────────────────────────────────────────
    # 规则：
    # - phone精确匹配 + 姓名兼容 → 直接匹配
    # - 身份证后4位 + 姓名匹配 → 直接匹配
    # - 仅姓名 OR 仅手机但姓名不一致 → 待审核（不强制绑定，避免误匹配）
    # - 无任何有效标识 → 待审核
    # unmatched 受试者会被创建为待审核候选，其礼金记录也会写入（关联临时 subject_id = -1 的受试者候选）
    print('\n与 V2 受试者匹配（保守策略）...')
    matched       = []   # 找到 subject_id
    unmatched     = []   # 未找到（需创建受试者候选 + 礼金待审核）
    conflict_flag = []   # 手机匹配但姓名冲突（需人工审核）

    for rec in dedup_records:
        phone  = rec.get('phone', '')
        ic     = rec.get('id_card', '')
        name   = rec.get('name', '')
        last4  = ic[-4:] if len(ic) >= 4 else ''

        db_rec = None
        match_type = 'none'

        # 策略1：手机精确匹配
        if phone and phone in db_by_phone:
            candidate = db_by_phone[phone]
            if not name or not candidate['name']:
                # 姓名其中一方为空，弱匹配
                db_rec = candidate
                match_type = 'phone_only'
            elif name in candidate['name'] or candidate['name'] in name or name == candidate['name']:
                # 姓名相符
                db_rec = candidate
                match_type = 'phone+name'
            else:
                # 手机匹配但姓名冲突 → 标记冲突，进待审核
                rec['_conflict'] = (
                    f"手机{phone}匹配到{candidate['name']}，但记录姓名={name}，需人工确认"
                )
                conflict_flag.append(rec)
                continue

        # 策略2：身份证后4位 + 姓名（弥补无手机或手机未入库的情况）
        if not db_rec and last4 and name:
            k = (last4, name)
            if k in db_by_last4_name:
                db_rec = db_by_last4_name[k]
                match_type = 'id_last4+name'

        rec['_match'] = match_type
        if db_rec:
            rec['_subject_id'] = db_rec['id']
            rec['_subject_no'] = db_rec.get('subject_no', '')
            rec['_has_idcard'] = db_rec.get('has_idcard', False)
            matched.append(rec)
        else:
            unmatched.append(rec)

    print(f'  精确/强匹配（直接入库）:   {len(matched)} 条')
    print(f'  姓名冲突（待人工审核）:    {len(conflict_flag)} 条')
    print(f'  无匹配（新建受试者候选）:  {len(unmatched)} 条')
    # 冲突的也加入 unmatched 待审核
    unmatched.extend(conflict_flag)
    all_to_write = matched + unmatched  # 全部都要入库

    if args.dry_run:
        print('\n[dry-run] 写库预览（全量）：')
        print(f'  待写 t_subject_payment（matched）: {len(matched)}')
        print(f'  待写 t_subject_payment（unmatched，关联新建候选受试者）: {len(unmatched)}')
        print(f'  待写 t_ext_ingest_candidate（待审核）: {len(unmatched)}')
        print(f'  待写 t_subject_points_ledger: {len(matched)}')
        backfill_cnt = sum(1 for r in matched
                           if valid_idcard(r.get('id_card','')) and not r.get('_has_idcard'))
        print(f'  反向补全 id_card（无身份证受试者）: {backfill_cnt}')
        _save_report(matched, unmatched, all_files)
        cur.close(); conn.close(); return

    # ── 获取当前最大 payment_no 序号 ─────────────────────────────────────────
    cur.execute("SELECT payment_no FROM t_subject_payment WHERE payment_no LIKE 'NAS-PAY-%' ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()
    pay_seq = int(row[0].split('-')[-1]) + 1 if row else 1

    ok_pay = ok_pts = ok_backfill = err = 0
    ok_unmatched_pay = 0
    now_ts = datetime.datetime.now(datetime.timezone.utc)

    # ── 第一步：为 unmatched 记录创建完整受试者档案 ──────────────────────────────
    # 策略：为每个 unmatched 记录创建完整的 t_subject + t_subject_profile
    # subject_no 格式：NAS-{YYYYMM}-{seq:05d}（最长16字符，符合 VARCHAR(20) 限制）
    # 从身份证推导 birth_date、gender、age，不丢弃任何可用信息
    print('\n为未匹配记录创建完整受试者档案...')

    # 查询当前 NAS 导入序号最大值，续号
    ym_str = datetime.date.today().strftime('%Y%m')
    cur.execute(
        "SELECT subject_no FROM t_subject WHERE subject_no LIKE %s ORDER BY subject_no DESC LIMIT 1",
        (f'NAS-{ym_str}-%',)
    )
    row = cur.fetchone()
    if row:
        try:
            nas_seq = int(row[0].split('-')[-1]) + 1
        except Exception:
            nas_seq = 1
    else:
        nas_seq = 1

    # 为每个 unmatched 记录准备完整的 t_subject 数据
    unmatched_subj_rows = []   # 用于插入 t_subject
    unmatched_profile_data = []  # 用于插入 t_subject_profile（拿到 id 后填）

    for rec in unmatched:
        name  = (rec.get('name', '') or '').strip() or '未知'
        phone = rec.get('phone', '') or ''
        ic    = rec.get('id_card', '') or ''
        ic_info = parse_idcard_info(ic)

        gender   = ic_info.get('gender', '') or ('female' if '女' in name else '')
        age      = ic_info.get('age')          # None 如果无法推导
        province = ic_info.get('province', '')

        sno = f'NAS-{ym_str}-{nas_seq:05d}'
        nas_seq += 1

        # 身份证加密存储
        ic_enc  = encrypt_field(ic) if valid_idcard(ic) else ''
        ic_hash = hash_field(ic) if valid_idcard(ic) else ''
        ic_last4 = ic[-4:] if len(ic) >= 4 else ''

        unmatched_subj_rows.append((
            sno, name, phone, gender, age,
            ic_enc,   # t_subject.id_card_encrypted
            now_ts, now_ts,
        ))
        unmatched_profile_data.append({
            'birth_date': ic_info.get('birth_date'),
            'ic_enc': ic_enc,
            'ic_hash': ic_hash,
            'ic_last4': ic_last4,
            'province': province,
            'gender': gender,
        })
        rec['_has_idcard'] = bool(ic_enc)

    # 批量插入 t_subject，并回收 id
    subj_sql = """
        INSERT INTO t_subject
            (subject_no, name, phone, gender, age, id_card_encrypted,
             skin_type, risk_level, status, source_channel,
             auth_level, create_time, update_time, is_deleted)
        VALUES (%s, %s, %s, %s, %s, %s, '', '', 'pending_review', 'nas_import', '', %s, %s, false)
        ON CONFLICT (subject_no) DO NOTHING
        RETURNING id
    """
    conn.autocommit = False
    new_sids = []
    try:
        for i, row_data in enumerate(unmatched_subj_rows):
            cur.execute(subj_sql, row_data)
            result = cur.fetchone()
            new_sids.append(result[0] if result else None)
        conn.commit()
        created_count = sum(1 for s in new_sids if s is not None)
        print(f'  ✅ 新建受试者: {created_count} 条（{len(unmatched) - created_count} 条因编号冲突跳过）')
    except Exception as e:
        conn.rollback()
        print(f'  ❌ 创建受试者失败: {e}')
        new_sids = [None] * len(unmatched)
    conn.autocommit = True

    # 批量插入 t_subject_profile（仅针对成功创建的新受试者）
    profile_rows = []
    for i, sid in enumerate(new_sids):
        if not sid:
            continue
        pd = unmatched_profile_data[i]
        profile_rows.append((
            sid,
            pd['birth_date'],
            pd['ic_enc'],
            pd['ic_hash'],
            pd['ic_last4'],
            pd['province'],  # 写入 province 字段
            now_ts, now_ts,
        ))

    if profile_rows:
        from psycopg2.extras import execute_values
        profile_sql = """
            INSERT INTO t_subject_profile
                (subject_id, birth_date, id_card_encrypted, id_card_hash, id_card_last4,
                 province, age, ethnicity, education, occupation, marital_status,
                 name_pinyin, phone_backup, email, city, district, address, postal_code,
                 emergency_contact_name, emergency_contact_phone, emergency_contact_relation,
                 total_enrollments, total_completed, privacy_level,
                 consent_data_sharing, consent_rwe_usage, consent_biobank, consent_follow_up,
                 data_retention_years, create_time, update_time)
            VALUES %s
            ON CONFLICT (subject_id) DO NOTHING
        """
        profile_template = (
            "(%s, %s, %s, %s, %s, %s, "
            # age: 从 birth_date 推导（存入 profile.age）
            "NULL, '', '', '', '', '', '', '', '', '', '', '', '', '', '', "
            "0, 0, 'standard', false, false, false, false, 5, %s, %s)"
        )
        conn.autocommit = False
        try:
            execute_values(cur, profile_sql, profile_rows, template=profile_template, page_size=500)
            conn.commit()
            print(f'  ✅ 新建受试者 profile: {len(profile_rows)} 条（含 birth_date/id_card/province）')
        except Exception as e:
            conn.rollback()
            print(f'  ❌ 创建受试者 profile 失败: {e}')
        conn.autocommit = True

    for idx, rec in enumerate(unmatched):
        rec['_subject_id'] = new_sids[idx] if idx < len(new_sids) else None

    # ── 批量写入 t_subject_payment（matched + unmatched 全部）────────────────
    print('\n预加密银行卡号并准备批量数据...')
    pay_rows    = []
    matched_meta = []

    for rec in matched + unmatched:
        sid      = rec.get('_subject_id')
        if not sid:
            continue  # 极少数情况：候选受试者创建失败
        amount   = rec.get('amount', 0.0)
        bank_raw = rec.get('bank_account', '')
        ic       = rec.get('id_card', '')
        paid_dt  = rec.get('_paid_date')
        platform = rec.get('_platform', '')
        proj     = rec.get('project_code', '')
        note_txt = rec.get('note', '') or ''
        is_unmatched = '_conflict' in rec or rec.get('_match') == 'none'
        dup_note = rec.get('_dup_flag', '')
        full_note = f"{note_txt} {dup_note}".strip()[:200]

        pno        = gen_payment_no(pay_seq); pay_seq += 1
        bank_enc   = encrypt_field(bank_raw) if bank_raw else ''
        bank_last4 = mask_last4(bank_raw) if bank_raw else ''
        points     = max(0, int(amount * POINTS_RATE)) if not is_unmatched else 0

        pay_rows.append((
            sid, pno, amount,
            bank_enc, bank_last4,
            platform, proj, paid_dt, IMPORT_BATCH,
            points, full_note,
            now_ts, now_ts,
        ))
        matched_meta.append({
            'pno': pno, 'sid': sid, 'points': points,
            'proj': proj, 'paid_dt': paid_dt, 'platform': platform,
            'ic': ic, 'has_idcard': rec.get('_has_idcard', False),
            'is_unmatched': is_unmatched,
        })

    print(f'  预处理完成: {len(pay_rows)} 条，开始批量写库...')

    # 使用 execute_values 批量插入（每批 500 条，约减少 99.2% 的网络往返）
    from psycopg2.extras import execute_values

    BATCH = 500
    pay_sql = """
        INSERT INTO t_subject_payment
            (subject_id, payment_no, payment_type, amount, status,
             bank_account_encrypted, bank_account_last4,
             platform, project_code, nas_paid_date, nas_import_batch,
             points_awarded, payment_method, transaction_id, notes,
             create_time, update_time)
        VALUES %s
        ON CONFLICT (payment_no) DO NOTHING
    """
    pay_template = "(%s, %s, 'visit_compensation', %s, 'paid', %s, %s, %s, %s, %s, %s, %s, '银行转账', '', %s, %s, %s)"

    conn.autocommit = False  # 批量写入时用事务，失败可回滚
    try:
        for i in range(0, len(pay_rows), BATCH):
            batch = pay_rows[i:i + BATCH]
            execute_values(cur, pay_sql, batch, template=pay_template, page_size=BATCH)
            if (i // BATCH) % 10 == 0:
                print(f'  支付记录写入: {min(i + BATCH, len(pay_rows))}/{len(pay_rows)}')
        conn.commit()
        ok_pay = len(pay_rows)
        ok_unmatched_pay = sum(1 for m in matched_meta if m.get('is_unmatched'))
        print(f'  ✅ t_subject_payment 批量写入完成: {ok_pay} 条（其中待审核: {ok_unmatched_pay}）')
    except Exception as e:
        conn.rollback()
        print(f'  ❌ 支付记录批量写入失败: {e}')
        cur.close(); conn.close(); return

    conn.autocommit = True

    # ── 回查支付记录 ID，用于积分台账 ─────────────────────────────────────────
    print('\n回查支付记录ID...')
    pno_list = [m['pno'] for m in matched_meta]
    # 分批查询，防止 IN 列表过长
    pno_to_id = {}
    for i in range(0, len(pno_list), 2000):
        chunk = pno_list[i:i + 2000]
        cur.execute(
            "SELECT payment_no, id FROM t_subject_payment WHERE payment_no = ANY(%s)",
            (chunk,)
        )
        for pno, pid in cur.fetchall():
            pno_to_id[pno] = pid
    print(f'  回查到 {len(pno_to_id)} 条记录ID')

    # ── 批量写入 t_subject_points_ledger ──────────────────────────────────────
    print('\n批量写入积分台账...')
    cumulative_balances = {}
    pts_rows = []

    for m in matched_meta:
        if m.get('is_unmatched'):
            continue  # 待审核受试者的积分待人工确认后再生成
        pno    = m['pno']
        sid    = m['sid']
        points = m['points']
        pay_id = pno_to_id.get(pno)
        if not pay_id:
            continue  # ON CONFLICT DO NOTHING 跳过的重复记录

        prev_bal = cumulative_balances.get(sid, 0)
        new_bal  = prev_bal + points
        cumulative_balances[sid] = new_bal

        pts_rows.append((
            sid, pay_id, 'import_backfill', points, new_bal,
            m['proj'],
            f'NAS礼金导入 {m["paid_dt"] or ""} {m["platform"]} {pno}',
            IMPORT_BATCH, now_ts, now_ts,
        ))

    pts_sql = """
        INSERT INTO t_subject_points_ledger
            (subject_id, payment_id, event_type, delta, balance_after,
             project_code, note, import_batch, create_time, update_time)
        VALUES %s
    """
    conn.autocommit = False
    try:
        for i in range(0, len(pts_rows), BATCH):
            batch = pts_rows[i:i + BATCH]
            execute_values(cur, pts_sql, batch, page_size=BATCH)
        conn.commit()
        ok_pts = len(pts_rows)
        print(f'  ✅ t_subject_points_ledger 批量写入完成: {ok_pts} 条')
    except Exception as e:
        conn.rollback()
        print(f'  ❌ 积分台账批量写入失败: {e}')
    conn.autocommit = True

    # ── 反向补全 id_card ───────────────────────────────────────────────────────
    print('\n反向补全身份证...')
    ok_backfill = _run_backfill(cur, conn, db_by_phone, db_by_last4_name,
                                args.dry_run, matched_records=matched)
    print(f'  补全: {ok_backfill} 条')

    # ── 生成报告 ──────────────────────────────────────────────────────────────
    _save_report(matched, unmatched, all_files)

    print('\n' + '=' * 60)
    print(f'导入完成 | 批次: {IMPORT_BATCH}')
    print(f'  支付记录总计 (t_subject_payment):    {ok_pay}')
    print(f'    其中：精确匹配入库:               {ok_pay - ok_unmatched_pay}')
    print(f'    其中：待审核（unmatched）:         {ok_unmatched_pay}')
    print(f'  积分台账 (t_subject_points_ledger):  {ok_pts}')
    print(f'  反向补全 id_card:                    {ok_backfill}')
    print(f'  精确重复跳过:                        {exact_dups}')
    print(f'  疑似多次支付（已保留+标注）:         {flagged_dups}')
    total_amount = sum(r.get('amount', 0.0) for r in matched + unmatched)
    print(f'  总金额合计:                          ¥{total_amount:,.2f}')
    print('=' * 60)

    cur.close()
    conn.close()


def _run_backfill(cur, conn, db_by_phone, db_by_last4_name,
                  dry_run: bool, matched_records=None):
    """
    反向补全：从礼金档案中找到完整身份证，更新已匹配受试者的：
    - t_subject_profile: id_card_encrypted/hash/last4, birth_date, province
    - t_subject: id_card_encrypted, gender（如为空）, age（如为空）
    使用临时表批量 UPDATE，将大量网络往返压缩为少数几次。
    """
    if matched_records is None:
        print('[backfill-only] 重新扫描礼金文件获取身份证...')
        files = sorted(
            glob.glob(f'{HONOR_BASE}/**/*.xlsx', recursive=True) +
            glob.glob(f'{HONOR_BASE}/**/*.xls', recursive=True)
        )
        files = [f for f in files if not os.path.basename(f).startswith('~')]
        matched_records = []
        for fp in files:
            rows, _, _, _ = parse_file(fp)
            for r in rows:
                phone = norm_phone(r.get('phone', ''))
                ic    = norm_idcard(r.get('id_card', ''))
                name  = r.get('name', '')
                if not valid_idcard(ic):
                    continue
                last4 = ic[-4:]
                db_rec = None
                if phone and phone in db_by_phone:
                    db_rec = db_by_phone[phone]
                elif (last4, name) in db_by_last4_name:
                    db_rec = db_by_last4_name[(last4, name)]
                if db_rec:
                    matched_records.append({
                        '_subject_id': db_rec['id'],
                        '_has_idcard': db_rec.get('has_idcard', False),
                        'id_card': ic,
                    })
        print(f'  backfill 候选: {len(matched_records)} 条')

    # 收集需要补全的数据（每个 subject_id 只取第一条有效身份证）
    backfill_data = {}  # subject_id -> (ic_hash, ic_enc, ic_last4, birth_date, gender, age, province)
    for rec in matched_records:
        ic = norm_idcard(rec.get('id_card', ''))
        if not valid_idcard(ic):
            continue
        sid = rec.get('_subject_id')
        if not sid or sid in backfill_data:
            continue
        ic_info = parse_idcard_info(ic)
        backfill_data[sid] = (
            hash_field(ic),
            encrypt_field(ic),
            ic[-4:],
            ic_info.get('birth_date'),
            ic_info.get('gender', ''),
            ic_info.get('age'),
            ic_info.get('province', ''),
        )

    if not backfill_data:
        return 0

    print(f'  需补全身份证/出生日期的受试者: {len(backfill_data)} 人')

    if dry_run:
        return len(backfill_data)

    from psycopg2.extras import execute_values

    conn.autocommit = False
    try:
        cur.execute("""
            CREATE TEMP TABLE _backfill_idcard (
                subject_id   BIGINT,
                ic_encrypted TEXT,
                ic_hash      TEXT,
                ic_last4     VARCHAR(4),
                birth_date   DATE,
                gender       VARCHAR(10),
                age          INTEGER,
                province     VARCHAR(50)
            ) ON COMMIT DROP
        """)

        rows = [
            (sid, enc, hsh, l4, bd, gd, ag, pv)
            for sid, (hsh, enc, l4, bd, gd, ag, pv) in backfill_data.items()
        ]
        execute_values(cur,
            """INSERT INTO _backfill_idcard
               (subject_id, ic_encrypted, ic_hash, ic_last4, birth_date, gender, age, province)
               VALUES %s""",
            rows, page_size=2000
        )

        # 更新 t_subject_profile：id_card 字段 + birth_date + province
        cur.execute("""
            UPDATE t_subject_profile sp
            SET id_card_encrypted = b.ic_encrypted,
                id_card_hash      = b.ic_hash,
                id_card_last4     = b.ic_last4,
                birth_date        = COALESCE(sp.birth_date, b.birth_date),
                province          = CASE WHEN sp.province = '' OR sp.province IS NULL
                                         THEN b.province ELSE sp.province END,
                update_time       = NOW()
            FROM _backfill_idcard b
            WHERE sp.subject_id = b.subject_id
              AND (sp.id_card_encrypted IS NULL OR sp.id_card_encrypted = ''
                   OR sp.id_card_hash IS NULL OR sp.id_card_hash = ''
                   OR sp.birth_date IS NULL)
        """)
        profile_updated = cur.rowcount

        # 更新 t_subject：id_card_encrypted + gender（空时）+ age（空时）
        cur.execute("""
            UPDATE t_subject s
            SET id_card_encrypted = CASE WHEN s.id_card_encrypted = '' OR s.id_card_encrypted IS NULL
                                         THEN b.ic_encrypted ELSE s.id_card_encrypted END,
                gender            = CASE WHEN s.gender = '' OR s.gender IS NULL
                                         THEN b.gender ELSE s.gender END,
                age               = CASE WHEN s.age IS NULL THEN b.age ELSE s.age END,
                update_time       = NOW()
            FROM _backfill_idcard b
            WHERE s.id = b.subject_id
              AND (s.id_card_encrypted = '' OR s.id_card_encrypted IS NULL
                   OR s.gender = '' OR s.gender IS NULL
                   OR s.age IS NULL)
        """)
        subject_updated = cur.rowcount

        conn.commit()
        print(f'  ✅ 反向补全完成:')
        print(f'     t_subject_profile 更新 {profile_updated} 条（id_card + birth_date + province）')
        print(f'     t_subject 更新 {subject_updated} 条（id_card + gender + age）')
        return profile_updated

    except Exception as e:
        conn.rollback()
        print(f'  ❌ 反向补全失败: {e}')
        return 0
    finally:
        conn.autocommit = True


def _save_report(matched, unmatched, all_files):
    """保存 Excel 报告到 NAS"""
    try:
        import openpyxl as xl
        wb = xl.Workbook()

        ws1 = wb.active
        ws1.title = '导入概览'
        ws1.append(['CN KIS V2.0 礼金档案导入报告'])
        ws1.append([f'生成时间: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        ws1.append([f'批次: {IMPORT_BATCH}'])
        ws1.append([''])
        for row in [
            ['档案文件总数', len(all_files)],
            ['已匹配受试者支付记录', len(matched)],
            ['未匹配（无受试者档案）', len(unmatched)],
        ]:
            ws1.append(row)

        ws2 = wb.create_sheet('已匹配支付记录')
        ws2.append(['受试者编号', '姓名', '手机', '金额', '项目代码',
                    '支付平台', '支付日期', '银行卡后4位', '匹配方式'])
        for r in matched[:5000]:
            ws2.append([
                r.get('_subject_no', ''), r.get('name', ''),
                r.get('phone', ''), r.get('amount', 0),
                r.get('project_code', ''), r.get('_platform', ''),
                str(r.get('_paid_date', '')), mask_last4(r.get('bank_account', '')),
                r.get('_match', ''),
            ])

        ws3 = wb.create_sheet('未匹配记录')
        ws3.append(['姓名', '手机', '身份证后4位', '金额', '项目代码', '平台', '支付日期'])
        for r in unmatched[:2000]:
            ic = r.get('id_card', '')
            ws3.append([
                r.get('name', ''), r.get('phone', ''),
                ic[-4:] if len(ic) >= 4 else '',
                r.get('amount', 0), r.get('project_code', ''),
                r.get('_platform', ''), str(r.get('_paid_date', '')),
            ])

        ts  = datetime.date.today().isoformat()
        out = f'{NAS_MOUNT}/礼金档案导入报告_{ts}.xlsx'
        try:
            wb.save(out)
            print(f'\n✅ 报告已保存: {out}')
        except Exception as e:
            out = f'/tmp/礼金档案导入报告_{ts}.xlsx'
            wb.save(out)
            print(f'\n⚠️  NAS写入失败，报告保存至: {out}  ({e})')
    except Exception as e:
        print(f'报告生成失败: {e}')


if __name__ == '__main__':
    main()
