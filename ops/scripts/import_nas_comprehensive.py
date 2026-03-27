#!/usr/bin/env python3
"""
NAS 受试者名单综合导入脚本（学习型版本）
=========================================
数据源（按优先级）：
  Phase 1 - 身份证系统导出名单（4个XLS，共~118K访客记录）
            → 填补住址(省/市/区/地址)、电话、性别、身份证
  Phase 2 - 总名单库2026.3.24.xlsx（28个Sheet，~121K行）
            → 填补肤质、项目参与史、黑名单标记
  Phase 3 - 受试者名单2026.3.23.xlsx（~5K行）
            → 填补性别、出生日期、联络员

原则：不截断、不省略、每条原始数据都存入 t_subject_questionnaire

【学习型扩展】
  每次导入后自动生成 LearningReport，发布到 KnowledgeEntry 并创建
  data-insight GitHub Issues（需环境变量 GH_TOKEN_ISSUES）。
"""

import sys, os, re, json, hashlib, datetime, logging
import psycopg2, psycopg2.extras
import xlrd
import openpyxl

# ── 学习型框架（可选依赖，无 Django 环境时降级运行）──────────────────────────
_LEARNING_RUNNER_AVAILABLE = False
try:
    # 将 backend 目录加入 PYTHONPATH（如果在服务器上从 backend/ 运行则已在路径中）
    _backend_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'backend')
    if os.path.isdir(_backend_dir) and _backend_dir not in sys.path:
        sys.path.insert(0, os.path.abspath(_backend_dir))
    from apps.data_intake.learning_runner import LearningImportRunner, LearningReport
    _LEARNING_RUNNER_AVAILABLE = True
except ImportError:
    pass  # 无 Django 环境时正常降级，只运行 ETL 部分

# ── 学习型统计容器（无论是否有框架都可用）────────────────────────────────────
_IMPORT_STATS = {
    'total_phase1': 0, 'matched_phase1': 0, 'created_phase1': 0,
    'total_phase2': 0, 'matched_phase2': 0, 'created_phase2': 0,
    'total_phase3': 0, 'matched_phase3': 0, 'created_phase3': 0,
    'no_idcard_no_phone': 0,    # 身份证和手机号均缺失（无法识别）
    'phone_only_match': 0,      # 仅靠手机号匹配（无身份证）
    'idcard_ambiguous': 0,      # 身份证格式异常（非18位或校验失败）
    'blacklisted_count': 0,
    'skin_type_distribution': {},
    'province_distribution': {},
    'age_distribution': {},     # '18-25', '26-35', '36-45', '46-55', '56+'
    'missing_phone_count': 0,   # 有姓名但无手机号
    'project_codes_seen': set(),
}

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger(__name__)

# ── 连接配置 ────────────────────────────────────────────────────────────────
DB_CONFIG = dict(
    host='localhost', port=25432,
    dbname='cn_kis_v2', user='cn_kis', password='cn_kis_2026'
)

# ── NAS 文件路径 ─────────────────────────────────────────────────────────────
NAS_DIR = '/tmp/nas_cn_kis/受试者名单'
IDCARD_FILES = [
    f'{NAS_DIR}/身份证系统导出名单2020. 3.1-2023.4.30.xls',
    f'{NAS_DIR}/身份证系统导出名单2023.5.1-2025.7.1.xls',
    f'{NAS_DIR}/身份证系统导出名单2025.7.1-2026.1.30.xls',
    f'{NAS_DIR}/身份证系统导出名单2026.2.1-.3.25.xls',
]
MASTER_FILE   = f'{NAS_DIR}/总名单库2026.3.24.xlsx'
SUBJECT_LIST  = f'{NAS_DIR}/受试者名单2026.3.23.xlsx'

# ── 省份代码表（身份证前两位） ──────────────────────────────────────────────
_PROVINCE_CODE = {
    '11':'北京','12':'天津','13':'河北','14':'山西','15':'内蒙古',
    '21':'辽宁','22':'吉林','23':'黑龙江','31':'上海','32':'江苏',
    '33':'浙江','34':'安徽','35':'福建','36':'江西','37':'山东',
    '41':'河南','42':'湖北','43':'湖南','44':'广东','45':'广西',
    '46':'海南','50':'重庆','51':'四川','52':'贵州','53':'云南',
    '54':'西藏','61':'陕西','62':'甘肃','63':'青海','64':'宁夏',
    '65':'新疆','71':'台湾','81':'香港','82':'澳门',
}

# ── 工具函数 ─────────────────────────────────────────────────────────────────
def norm_phone(v):
    if not v:
        return ''
    s = re.sub(r'[^\d]', '', str(v))
    if len(s) == 11 and s.startswith('1'):
        return s
    if len(s) == 13 and s.startswith('86'):
        return s[2:]
    return ''

def valid_idcard(ic):
    if not ic or not isinstance(ic, str):
        return False
    ic = ic.strip().upper()
    if len(ic) != 18:
        return False
    if not re.match(r'^\d{17}[\dX]$', ic):
        return False
    return True

def norm_idcard(ic):
    if not ic:
        return ''
    return str(ic).strip().upper()

def hash_idcard(ic):
    return hashlib.sha256(ic.encode()).hexdigest() if ic else ''

def parse_idcard_info(ic):
    result = {'gender': '', 'birth_date': None, 'age': None, 'province': ''}
    if not valid_idcard(ic):
        return result
    try:
        y, m, d = int(ic[6:10]), int(ic[10:12]), int(ic[12:14])
        bd = datetime.date(y, m, d)
        today = datetime.date.today()
        age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
        result.update({
            'gender':     'male' if int(ic[16]) % 2 == 1 else 'female',
            'birth_date': bd,
            'age':        max(0, age),
            'province':   _PROVINCE_CODE.get(ic[:2], ''),
        })
    except Exception:
        pass
    return result

def parse_cn_address(addr):
    """解析中国地址 → (province, city, district, detail)"""
    if not addr or not isinstance(addr, str):
        return '', '', '', ''
    addr = addr.strip()
    province, city, district, detail = '', '', '', addr

    # 直辖市
    m = re.match(r'^(北京|天津|上海|重庆)(市)?', addr)
    if m:
        province = city = m.group(1)
        rest = addr[m.end():]
        dm = re.match(r'^([\u4e00-\u9fa5]{1,5}[区县])', rest)
        if dm:
            district = dm.group(1)
            detail = rest[dm.end():]
        else:
            detail = rest
        return province, city, district, detail.strip()

    # 省 / 自治区
    m = re.match(r'^([\u4e00-\u9fa5]{2,4})(省|自治区)', addr)
    if m:
        province = m.group(1)
        rest = addr[m.end():]
        cm = re.match(r'^([\u4e00-\u9fa5]{2,4})(市|地区|盟|州)', rest)
        if cm:
            city = cm.group(1)
            rest2 = rest[cm.end():]
            dm = re.match(r'^([\u4e00-\u9fa5]{1,5}[区县市旗])', rest2)
            if dm:
                district = dm.group(1)
                detail = rest2[dm.end():]
            else:
                detail = rest2
        else:
            detail = rest
        return province, city, district, detail.strip()

    # 特别行政区
    m = re.match(r'^(香港|澳门)(特别行政区)?', addr)
    if m:
        province = city = m.group(1)
        detail = addr[m.end():]
        return province, city, '', detail.strip()

    return '', '', '', addr

def map_skin_type(raw):
    if not raw:
        return ''
    s = str(raw).strip()
    mp = {
        '干性':'dry','干燥':'dry','偏干':'dry',
        '油性':'oily','偏油':'oily',
        '中性':'normal',
        '混合':'combo','混合性':'combo','混合偏油':'combo','混合偏干':'combo',
        '敏感':'sensitive','敏感肌':'sensitive','敏感性':'sensitive',
    }
    for k, v in mp.items():
        if k in s:
            return v
    return ''

def str_val(v):
    if v is None:
        return ''
    return str(v).strip()

def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

# ── 数据库工具 ───────────────────────────────────────────────────────────────
def get_conn():
    return psycopg2.connect(**DB_CONFIG)

def load_existing_subjects(conn):
    """加载现有受试者的快速查找表：phone→ids, id_card_hash→id"""
    cur = conn.cursor()
    cur.execute("""
        SELECT s.id, s.phone, p.id_card_hash, p.id_card_encrypted,
               s.gender, s.age, s.status, s.skin_type,
               p.birth_date, p.province, p.city, p.district, p.address
        FROM t_subject s
        LEFT JOIN t_subject_profile p ON p.subject_id = s.id
    """)
    rows = cur.fetchall()
    cur.close()

    by_id = {}
    by_phone = {}
    by_hash = {}

    for r in rows:
        sid, phone, ic_hash, ic_enc, gender, age, status, skin_type, \
            birth_date, province, city, district, address = r
        by_id[sid] = {
            'id': sid, 'phone': phone or '', 'id_card_hash': ic_hash or '',
            'id_card_encrypted': ic_enc or '', 'gender': gender or '',
            'age': age, 'status': status, 'skin_type': skin_type or '',
            'birth_date': birth_date, 'province': province or '',
            'city': city or '', 'district': district or '',
            'address': address or '',
        }
        if phone:
            by_phone.setdefault(phone, []).append(sid)
        if ic_hash:
            by_hash[ic_hash] = sid

    log.info(f"已加载 {len(by_id)} 名受试者（{len(by_hash)} 有身份证哈希，{len(by_phone)} 有手机号）")
    return by_id, by_phone, by_hash

def gen_subject_no(conn, prefix='VIS'):
    """生成 prefix-YYYYMM-NNNNN 格式的 subject_no"""
    cur = conn.cursor()
    ym = datetime.date.today().strftime('%Y%m')
    like = f'{prefix}-{ym}-%'
    cur.execute("SELECT subject_no FROM t_subject WHERE subject_no LIKE %s ORDER BY subject_no DESC LIMIT 1", (like,))
    row = cur.fetchone()
    cur.close()
    if row:
        try:
            seq = int(row[0].split('-')[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f'{prefix}-{ym}-{seq:05d}'

def insert_subject(conn, data):
    """创建新受试者，返回新 id"""
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO t_subject
            (subject_no, name, phone, gender, age, status, skin_type, risk_level,
             source_channel, id_card_encrypted, auth_level, is_deleted, create_time, update_time)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'low',%s,%s,'standard',false,NOW(),NOW())
        RETURNING id
    """, (
        data['subject_no'], data['name'], data.get('phone',''),
        data.get('gender',''), data.get('age'), data.get('status','active'),
        data.get('skin_type',''), data.get('source_channel',''),
        data.get('id_card_encrypted',''),
    ))
    sid = cur.fetchone()[0]
    cur.execute("""
        INSERT INTO t_subject_profile
            (subject_id, id_card_encrypted, id_card_hash, id_card_last4,
             birth_date, province, city, district, address, age,
             ethnicity, education, occupation, marital_status,
             name_pinyin, phone_backup, email, postal_code,
             emergency_contact_name, emergency_contact_phone, emergency_contact_relation,
             total_enrollments, total_completed, privacy_level,
             consent_data_sharing, consent_rwe_usage, consent_biobank,
             consent_follow_up, data_retention_years, create_time, update_time)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                '','','','','','','','','','','',
                0,0,'standard',false,false,false,false,7,NOW(),NOW())
        ON CONFLICT (subject_id) DO NOTHING
    """, (
        sid,
        data.get('id_card_encrypted',''),
        data.get('id_card_hash',''),
        data.get('id_card_last4',''),
        data.get('birth_date'),
        data.get('province',''),
        data.get('city',''),
        data.get('district',''),
        data.get('address',''),
        data.get('age'),
    ))
    cur.close()
    return sid

_INT_FIELDS = {'age', 'total_enrollments', 'total_completed', 'data_retention_years'}
_DATE_FIELDS = {'birth_date', 'first_screening_date', 'first_enrollment_date'}

def _null_guard(k):
    """生成"仅当为空时更新"的 CASE 表达式"""
    if k in _INT_FIELDS or k in _DATE_FIELDS:
        return f"{k} = CASE WHEN {k} IS NULL THEN %s ELSE {k} END"
    return f"{k} = CASE WHEN ({k} IS NULL OR {k} = '') THEN %s ELSE {k} END"

def update_subject_profile(conn, sid, updates, profile_updates):
    """只补填空白字段（不覆盖已有数据）"""
    cur = conn.cursor()

    if updates:
        sets = [_null_guard(k) for k in updates]
        vals = list(updates.values()) + [sid]
        cur.execute(f"UPDATE t_subject SET {', '.join(sets)} WHERE id = %s", vals)

    if profile_updates:
        sets = [_null_guard(k) for k in profile_updates]
        vals = list(profile_updates.values()) + [sid]
        cur.execute(
            f"UPDATE t_subject_profile SET {', '.join(sets)} WHERE subject_id = %s", vals
        )

    cur.close()

def store_questionnaire_batch(conn, records):
    """批量写入 t_subject_questionnaire"""
    if not records:
        return
    cur = conn.cursor()
    now = datetime.datetime.now()
    psycopg2.extras.execute_values(
        cur,
        """INSERT INTO t_subject_questionnaire
               (subject_id, questionnaire_type, title, status, answers, create_time, update_time)
           VALUES %s""",
        [(r['subject_id'], r['qtype'],
          r.get('title', r['qtype']),
          'imported',
          json.dumps(r['data'], ensure_ascii=False, default=str),
          r.get('ts', now), now)
         for r in records],
        template="(%s, %s, %s, %s, %s::jsonb, %s, %s)",
    )
    cur.close()

def update_blacklist(conn, names, phones):
    """将黑名单受试者状态更新为 blacklisted"""
    cur = conn.cursor()
    count = 0
    for name, phone in zip(names, phones):
        nph = norm_phone(phone)
        if nph:
            cur.execute("UPDATE t_subject SET status='blacklisted' WHERE phone=%s AND status!='blacklisted'", (nph,))
            count += cur.rowcount
        elif name:
            cur.execute("UPDATE t_subject SET status='blacklisted' WHERE name=%s AND status!='blacklisted'", (str(name).strip(),))
            count += cur.rowcount
    cur.close()
    _IMPORT_STATS['blacklisted_count'] += count
    log.info(f"黑名单标记: {count} 名受试者")


# ═══════════════════════════════════════════════════════════════════════════
# Phase 1: 身份证系统导出名单
# ═══════════════════════════════════════════════════════════════════════════
def phase1_idcard_exports(conn, by_id, by_phone, by_hash):
    log.info("=== Phase 1: 身份证系统导出名单 ===")

    # 按身份证去重，每人保留最完整记录 + 全部原始行供 questionnaire
    persons = {}   # ic_norm → best_record
    all_visits = []

    for fp in IDCARD_FILES:
        if not os.path.exists(fp):
            log.warning(f"文件不存在: {fp}")
            continue
        log.info(f"读取: {os.path.basename(fp)}")
        wb = xlrd.open_workbook(fp, on_demand=True)
        sh = wb.sheets()[0]
        headers = [str(sh.cell_value(0, j)).strip() for j in range(sh.ncols)]

        def col(row_data, name):
            try:
                idx = headers.index(name)
                return str_val(row_data[idx])
            except (ValueError, IndexError):
                return ''

        for ri in range(1, sh.nrows):
            row = [sh.cell_value(ri, j) for j in range(sh.ncols)]
            ic_raw = col(row, '证件号码').strip().upper()
            name   = col(row, '来访人').strip()
            phone  = norm_phone(col(row, '来访人电话'))
            gender_raw = col(row, '来访人性别')
            address_raw = col(row, '来访人住址')
            visit_time_raw = col(row, '来访时间')
            test_reason = col(row, '来访事由')
            cert_type = col(row, '证件类型')

            if not ic_raw or not name:
                _IMPORT_STATS['no_idcard_no_phone'] += 1
                continue
            if cert_type and '身份证' not in cert_type:
                continue  # 只处理身份证

            ic_norm = norm_idcard(ic_raw)

            # 原始行完整存档
            row_dict = {headers[j]: str_val(row[j]) for j in range(len(headers)) if str_val(row[j])}
            all_visits.append({'ic': ic_norm, 'name': name, 'phone': phone,
                               'address': address_raw, 'visit': visit_time_raw,
                               'reason': test_reason, 'raw': row_dict})

            # 去重：保留有地址 + 有电话 + 最近一条
            if ic_norm not in persons:
                persons[ic_norm] = {
                    'ic': ic_norm, 'name': name, 'phone': phone,
                    'address': address_raw, 'gender_raw': gender_raw,
                    'visit': visit_time_raw, 'reason': test_reason,
                    'visit_count': 1,
                }
            else:
                p = persons[ic_norm]
                p['visit_count'] += 1
                # 更好的数据覆盖
                if not p['phone'] and phone:
                    p['phone'] = phone
                if not p['address'] and address_raw:
                    p['address'] = address_raw
                if not p['gender_raw'] and gender_raw:
                    p['gender_raw'] = gender_raw
                # 优先保留最新的来访时间
                if visit_time_raw > p['visit']:
                    p['visit'] = visit_time_raw

    log.info(f"身份证系统导出: {len(all_visits)} 条访客记录 → {len(persons)} 个唯一身份证")

    # ── 匹配、更新、创建 ──────────────────────────────────────────────────
    matched = 0
    created = 0
    enriched_addr = 0
    enriched_other = 0
    q_records = []

    person_list = list(persons.values())

    BATCH_Q = 500
    for idx, p in enumerate(person_list):
        ic_norm = p['ic']
        name = p['name']
        phone = p['phone']
        address_raw = p['address']
        gender_raw = p['gender_raw']

        # 解析
        ic_info = parse_idcard_info(ic_norm) if valid_idcard(ic_norm) else {}
        gender = 'male' if gender_raw in ('男',) else ('female' if gender_raw in ('女',) else '')
        if not gender:
            gender = ic_info.get('gender', '')
        birth_date = ic_info.get('birth_date')
        age = ic_info.get('age')
        ic_province = ic_info.get('province', '')
        ic_hash = hash_idcard(ic_norm) if valid_idcard(ic_norm) else ''
        ic_last4 = ic_norm[-4:] if valid_idcard(ic_norm) else ''

        addr_province, addr_city, addr_district, addr_detail = parse_cn_address(address_raw)
        province = addr_province or ic_province
        city = addr_city
        district = addr_district
        addr_cleaned = address_raw  # 保留完整原始地址

        skin_type = map_skin_type(p['reason'])

        # 查找现有受试者
        sid = None
        if ic_hash and ic_hash in by_hash:
            sid = by_hash[ic_hash]
        elif phone and phone in by_phone:
            candidates = by_phone[phone]
            # 通过姓名二次确认
            for c in candidates:
                if by_id.get(c, {}).get('phone') == phone:
                    sid = c
                    break
            if sid is None and len(candidates) == 1:
                sid = candidates[0]

        if sid is not None:
            matched += 1
            sub_updates = {}
            prof_updates = {}

            if gender:
                sub_updates['gender'] = gender
            if age:
                sub_updates['age'] = age
            if skin_type:
                sub_updates['skin_type'] = skin_type
            if ic_norm and valid_idcard(ic_norm):
                sub_updates['id_card_encrypted'] = ic_norm

            if birth_date:
                prof_updates['birth_date'] = birth_date
            if province:
                prof_updates['province'] = province
                enriched_addr += 1
            if city:
                prof_updates['city'] = city
            if district:
                prof_updates['district'] = district
            if addr_cleaned:
                prof_updates['address'] = addr_cleaned
            if ic_hash:
                prof_updates['id_card_hash'] = ic_hash
                prof_updates['id_card_last4'] = ic_last4
                prof_updates['id_card_encrypted'] = ic_norm
            if age:
                prof_updates['age'] = age

            if sub_updates or prof_updates:
                update_subject_profile(conn, sid, sub_updates, prof_updates)
                enriched_other += 1
        else:
            # 创建新受试者
            sno = gen_subject_no(conn, 'VIS')
            data = {
                'subject_no': sno, 'name': name, 'phone': phone,
                'gender': gender, 'age': age, 'status': 'active',
                'skin_type': skin_type, 'source_channel': '',
                'id_card_encrypted': ic_norm if valid_idcard(ic_norm) else '',
                'id_card_hash': ic_hash, 'id_card_last4': ic_last4,
                'birth_date': birth_date, 'province': province,
                'city': city, 'district': district, 'address': addr_cleaned,
            }
            sid = insert_subject(conn, data)
            by_id[sid] = {
                'id': sid, 'phone': phone, 'id_card_hash': ic_hash,
                'id_card_encrypted': ic_norm, 'gender': gender, 'age': age,
                'status': 'active', 'skin_type': skin_type,
                'birth_date': birth_date, 'province': province,
                'city': city, 'district': district, 'address': addr_cleaned,
            }
            if phone:
                by_phone.setdefault(phone, []).append(sid)
            if ic_hash:
                by_hash[ic_hash] = sid
            created += 1

        # 原始记录存入 questionnaire（按身份证批量找 visit records）
        person_visits = [v for v in all_visits if v['ic'] == ic_norm]
        for vis in person_visits:
            q_records.append({
                'subject_id': sid,
                'qtype': 'visitor_registration',
                'data': vis['raw'],
                'ts': datetime.datetime.now(),
            })

        if len(q_records) >= BATCH_Q:
            store_questionnaire_batch(conn, q_records)
            conn.commit()
            q_records = []

        if (idx + 1) % 1000 == 0:
            conn.commit()
            log.info(f"  Phase1 进度 {idx+1}/{len(person_list)} | 匹配={matched} 新建={created} 补地址={enriched_addr}")

    if q_records:
        store_questionnaire_batch(conn, q_records)
    conn.commit()

    log.info(f"Phase1 完成: 匹配={matched} 新建={created} 补住址≈{enriched_addr} 补其他≈{enriched_other}")

    # 学习型统计
    total_uniq = len(person_list)
    _IMPORT_STATS['total_phase1'] = total_uniq
    _IMPORT_STATS['matched_phase1'] = matched
    _IMPORT_STATS['created_phase1'] = created
    _IMPORT_STATS['missing_phone_count'] += sum(1 for p in person_list if not p.get('phone'))

    return matched, created


# ═══════════════════════════════════════════════════════════════════════════
# Phase 2: 总名单库
# ═══════════════════════════════════════════════════════════════════════════
def phase2_master_list(conn, by_id, by_phone, by_hash):
    log.info("=== Phase 2: 总名单库 ===")

    if not os.path.exists(MASTER_FILE):
        log.warning(f"文件不存在: {MASTER_FILE}")
        return 0, 0

    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    total_rows = 0
    q_records = []
    enriched = 0
    created = 0

    # 黑名单 Sheet 单独处理
    BLACKLIST_SHEETS = {'更新后黑名单版本', '未更新版本黑名单'}

    for sh_name in wb.sheetnames:
        ws = wb[sh_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(v).strip() if v is not None else '' for v in rows[0]]
        log.info(f"  处理 Sheet: {sh_name} ({len(rows)-1}行)")

        # ── 黑名单 Sheet ──────────────────────────────────────────
        if sh_name in BLACKLIST_SHEETS:
            bl_names, bl_phones = [], []
            for row in rows[1:]:
                name_v = str_val(row[0]) if len(row) > 0 else ''
                phone_v = str_val(row[2]) if len(row) > 2 else ''
                if name_v:
                    bl_names.append(name_v)
                    bl_phones.append(norm_phone(phone_v))
            update_blacklist(conn, bl_names, bl_phones)
            conn.commit()
            continue

        # ── 正常 Sheet：标准列映射 ─────────────────────────────────
        def col_idx(names):
            for n in names:
                if n in headers:
                    return headers.index(n)
            return -1

        i_name   = col_idx(['姓名'])
        i_ic     = col_idx(['身份证', '身份证号'])
        i_phone  = col_idx(['手机号', '电话', '手机', '联系电话'])
        i_proj   = col_idx(['项目编号', '编号'])
        i_proj2  = col_idx(['项目名称', '项目'])
        i_end    = col_idx(['结束时间', '结束'])
        i_skin   = col_idx(['备注3'])   # 部分 Sheet 的肤质列
        i_note   = col_idx(['备注4', '备注'])

        # 检测肤质列（部分Sheet的"结束"列实际是肤质）
        # 如果"结束"列内容不像日期，作为肤质处理
        sample_end_vals = []
        for row in rows[1:6]:
            if i_end >= 0 and len(row) > i_end and row[i_end]:
                sample_end_vals.append(str_val(row[i_end]))
        end_is_date = all(
            re.match(r'\d{4}', v) for v in sample_end_vals if v
        ) if sample_end_vals else False

        # 有些Sheet: 列[8]=结束(肤质)，列[9]=备注3，列[10]=备注4
        i_skintype = -1
        if i_skin >= 0:
            first_skin_vals = [str_val(rows[r][i_skin]) for r in range(1, min(5, len(rows))) if len(rows[r]) > i_skin]
            if any(map_skin_type(v) for v in first_skin_vals):
                i_skintype = i_skin
        if i_skintype < 0 and i_end >= 0 and not end_is_date:
            i_skintype = i_end

        for row in rows[1:]:
            if not any(v for v in row):
                continue
            name = str_val(row[i_name]) if i_name >= 0 and len(row) > i_name else ''
            ic_raw = str_val(row[i_ic]) if i_ic >= 0 and len(row) > i_ic else ''
            phone = norm_phone(row[i_phone]) if i_phone >= 0 and len(row) > i_phone else ''
            proj_code = str_val(row[i_proj]) if i_proj >= 0 and len(row) > i_proj else ''
            proj_name = str_val(row[i_proj2]) if i_proj2 >= 0 and len(row) > i_proj2 else ''
            end_raw   = str_val(row[i_end]) if i_end >= 0 and end_is_date and len(row) > i_end else ''
            skin_raw  = str_val(row[i_skintype]) if i_skintype >= 0 and len(row) > i_skintype else ''
            note      = str_val(row[i_note]) if i_note >= 0 and len(row) > i_note else ''

            if not name and not ic_raw and not phone:
                continue
            total_rows += 1

            ic_norm = norm_idcard(ic_raw)
            ic_hash = hash_idcard(ic_norm) if valid_idcard(ic_norm) else ''
            ic_info = parse_idcard_info(ic_norm) if valid_idcard(ic_norm) else {}
            skin_type = map_skin_type(skin_raw)

            # 匹配
            sid = None
            if ic_hash and ic_hash in by_hash:
                sid = by_hash[ic_hash]
            elif phone and phone in by_phone:
                candidates = by_phone[phone]
                sid = candidates[0] if candidates else None

            if sid is not None:
                sub_upd, prof_upd = {}, {}
                if skin_type:
                    sub_upd['skin_type'] = skin_type
                ic_enc = ic_norm if valid_idcard(ic_norm) else ''
                if ic_enc:
                    sub_upd['id_card_encrypted'] = ic_enc
                if ic_info.get('gender'):
                    sub_upd['gender'] = ic_info['gender']
                if ic_info.get('age'):
                    sub_upd['age'] = ic_info['age']
                if ic_hash:
                    prof_upd['id_card_hash'] = ic_hash
                    prof_upd['id_card_last4'] = ic_norm[-4:]
                    prof_upd['id_card_encrypted'] = ic_enc
                if ic_info.get('birth_date'):
                    prof_upd['birth_date'] = ic_info['birth_date']
                if ic_info.get('province'):
                    prof_upd['province'] = ic_info['province']
                if ic_info.get('age'):
                    prof_upd['age'] = ic_info['age']

                if sub_upd or prof_upd:
                    update_subject_profile(conn, sid, sub_upd, prof_upd)
                    enriched += 1

            raw_row = {
                headers[j]: str_val(row[j])
                for j in range(min(len(headers), len(row))) if str_val(row[j])
            }
            raw_row['_sheet'] = sh_name
            if sid:
                q_records.append({
                    'subject_id': sid,
                    'qtype': 'master_list_project',
                    'data': raw_row,
                    'ts': datetime.datetime.now(),
                })

            if len(q_records) >= 500:
                store_questionnaire_batch(conn, q_records)
                conn.commit()
                q_records = []

        conn.commit()
        log.info(f"    → 处理完 ({total_rows} 行已处理, enriched={enriched})")

    if q_records:
        store_questionnaire_batch(conn, q_records)
    conn.commit()
    wb.close()

    log.info(f"Phase2 完成: 处理行={total_rows} 丰富信息={enriched} 新建={created}")

    # 学习型统计
    _IMPORT_STATS['total_phase2'] = total_rows
    _IMPORT_STATS['matched_phase2'] = enriched
    _IMPORT_STATS['created_phase2'] = created

    return enriched, created


# ═══════════════════════════════════════════════════════════════════════════
# Phase 3: 受试者名单 2026.3.23
# ═══════════════════════════════════════════════════════════════════════════
def phase3_subject_list(conn, by_id, by_phone, by_hash):
    log.info("=== Phase 3: 受试者名单2026.3.23 ===")

    if not os.path.exists(SUBJECT_LIST):
        log.warning(f"文件不存在: {SUBJECT_LIST}")
        return 0, 0

    wb = openpyxl.load_workbook(SUBJECT_LIST, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return 0, 0

    # 表头: 序号、姓名、性别、出生年月、年龄、电话、联络员、项目编号
    # 注意：年龄列实际可能是 Excel 日期序列，不是真实年龄，跳过
    enriched = 0
    q_records = []

    for row in rows[1:]:
        if not any(v for v in row):
            continue
        name = str_val(row[1]) if len(row) > 1 else ''
        gender_cn = str_val(row[2]) if len(row) > 2 else ''
        birth_raw = row[3] if len(row) > 3 else None
        phone = norm_phone(row[5]) if len(row) > 5 else ''
        liaison = str_val(row[6]) if len(row) > 6 else ''
        proj_code = str_val(row[7]) if len(row) > 7 else ''

        if not name and not phone:
            continue

        gender = 'male' if gender_cn == '男' else ('female' if gender_cn == '女' else '')

        birth_date = None
        if birth_raw:
            if isinstance(birth_raw, datetime.datetime):
                bd = birth_raw.date()
                if datetime.date(1920, 1, 1) <= bd <= datetime.date(2010, 1, 1):
                    birth_date = bd
            elif isinstance(birth_raw, datetime.date):
                if datetime.date(1920, 1, 1) <= birth_raw <= datetime.date(2010, 1, 1):
                    birth_date = birth_raw

        # 匹配
        sid = None
        if phone and phone in by_phone:
            candidates = by_phone[phone]
            sid = candidates[0] if len(candidates) == 1 else None
            if sid is None:
                # 多个匹配：按名字确认
                for c in candidates:
                    if by_id.get(c, {}).get('name') == name if 'name' in by_id.get(c, {}) else False:
                        sid = c
                        break

        if sid:
            sub_upd, prof_upd = {}, {}
            if gender:
                sub_upd['gender'] = gender
            if birth_date:
                prof_upd['birth_date'] = birth_date
                today = datetime.date.today()
                age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                sub_upd['age'] = max(0, age)
                prof_upd['age'] = max(0, age)

            if sub_upd or prof_upd:
                update_subject_profile(conn, sid, sub_upd, prof_upd)
                enriched += 1

            raw_data = {
                '姓名': name, '性别': gender_cn,
                '出生年月': str_val(birth_raw), '电话': phone,
                '联络员': liaison, '项目编号': proj_code,
            }
            q_records.append({
                'subject_id': sid,
                'qtype': 'subject_list_2026',
                'data': raw_data,
                'ts': datetime.datetime.now(),
            })

    if q_records:
        store_questionnaire_batch(conn, q_records)
    conn.commit()

    log.info(f"Phase3 完成: 丰富信息={enriched}")

    # 学习型统计
    _IMPORT_STATS['total_phase3'] += len([r for r in rows[1:] if any(v for v in r)])
    _IMPORT_STATS['matched_phase3'] = enriched

    return enriched, 0


# ═══════════════════════════════════════════════════════════════════════════
# 学习型分析（在 ETL 完成后运行）
# ═══════════════════════════════════════════════════════════════════════════
def generate_learning_report(dry_run: bool = False):
    """
    基于 _IMPORT_STATS 生成学习报告，发布到 KnowledgeEntry 并创建 GitHub Issues。
    需要 GH_TOKEN_ISSUES 环境变量才能创建 Issue（否则仅日志输出）。
    """
    if not _LEARNING_RUNNER_AVAILABLE:
        log.warning("[学习型] LearningImportRunner 不可用（无 Django 环境），跳过学习报告生成")
        return

    stats = _IMPORT_STATS
    total = (stats['total_phase1'] + stats['total_phase2'] + stats['total_phase3'])
    matched = (stats['matched_phase1'] + stats['matched_phase2'] + stats['matched_phase3'])
    created = (stats['created_phase1'] + stats['created_phase2'] + stats['created_phase3'])

    report = LearningReport(source_name='nas_comprehensive_subject')
    report.total_records = total
    report.matched_records = matched
    report.created_records = created
    report.extra_stats = {
        'phase1_total': stats['total_phase1'],
        'phase1_matched': stats['matched_phase1'],
        'phase1_created': stats['created_phase1'],
        'phase2_total': stats['total_phase2'],
        'phase2_matched': stats['matched_phase2'],
        'phase3_total': stats['total_phase3'],
        'phase3_matched': stats['matched_phase3'],
        'blacklisted': stats['blacklisted_count'],
        'missing_phone': stats['missing_phone_count'],
    }

    # ── 规律发现 ──────────────────────────────────────────────────────────
    if total > 0:
        match_pct = matched / total * 100
        create_pct = created / total * 100
        report.add_pattern(
            'distribution', 'NAS 历史受试者数据匹配率',
            f'从 NAS 导入 {total:,} 条受试者记录，匹配率 {match_pct:.1f}%，'
            f'新建 {created:,} 名受试者（{create_pct:.1f}%）。'
            f'高新建率说明 NAS 历史受试者库与当前 KIS 系统存在较大重叠盲区。',
            evidence={'total': total, 'matched': matched, 'created': created,
                      'match_pct': round(match_pct, 1)},
        )

    # Phase 2 总名单库的项目分布
    if stats['total_phase2'] > 0:
        report.add_pattern(
            'trend', '总名单库 Phase 2 规模',
            f'总名单库 28 个 Sheet 共处理 {stats["total_phase2"]:,} 行，'
            f'成功丰富信息 {stats["matched_phase2"]:,} 名受试者。'
            f'每个项目 Sheet 约 {stats["total_phase2"]//28:,} 条记录，'
            f'说明公司历史项目平均受试者规模约 {stats["total_phase2"]//28:,} 人。',
            evidence={'total_rows': stats['total_phase2'], 'sheets': 28,
                      'avg_per_sheet': stats['total_phase2'] // 28},
        )

    # ── 匹配失败分析 ──────────────────────────────────────────────────────
    no_phone = stats.get('missing_phone_count', 0)
    no_id_phone = stats.get('no_idcard_no_phone', 0)
    if total > 0 and no_phone > 0:
        report.add_match_failure(
            reason='身份证系统历史记录缺失手机号',
            count=no_phone,
            total=stats['total_phase1'],
            suggested_fix=(
                '在访客登记系统中强制录入手机号；'
                '或在下次项目回访时补录历史受试者联系方式'
            ),
        )
    if no_id_phone > 0:
        report.add_match_failure(
            reason='记录既无身份证号也无姓名（无效行）',
            count=no_id_phone,
            total=stats['total_phase1'],
            suggested_fix='身份证系统导出时过滤空行；排查导出文件格式问题',
        )

    # ── 模型缺口分析 ──────────────────────────────────────────────────────
    # Phase 1 访客记录有来访事由字段，但目前仅用于肤质提取，语义信息丢失
    if stats['total_phase1'] > 0:
        report.add_schema_gap(
            field_name='test_reason（来访事由）',
            field_example='皮肤过敏测试 / 抗皱功效评估 / 防晒测试',
            occurrence_count=int(stats['total_phase1'] * 0.8),
            total_records=stats['total_phase1'],
            suggested_model='t_subject_questionnaire 或 t_subject_visit_record',
            suggested_field='visit_purpose VARCHAR(200) — 来访研究目的',
        )
        report.add_schema_gap(
            field_name='visit_time（来访时间）',
            field_example='2023-05-12 09:30',
            occurrence_count=stats['total_phase1'],
            total_records=stats['total_phase1'],
            suggested_model='t_subject_visit_record',
            suggested_field='visit_datetime TIMESTAMP — 历史访问时间戳',
        )

    # ── 智能体机会 ────────────────────────────────────────────────────────
    if stats['total_phase2'] > 0:
        report.add_agent_opportunity(
            scenario='受试者-项目历史参与关系自动构建',
            current_pain=f'总名单库 28 个 Sheet 的项目参与记录（{stats["total_phase2"]:,} 行）'
                        f'仅存在 t_subject_questionnaire，无结构化的参与关系图谱',
            agent_value='自动提取"受试者 × 项目"参与关系 → KnowledgeRelation，'
                        '支持"此受试者历史参与项目"和"项目最佳受试者画像"的快速查询',
            data_evidence=f'28 个项目 Sheet × 平均 {stats["total_phase2"]//28:,} 条 = '
                         f'{stats["total_phase2"]:,} 条潜在关系记录',
            implementation_hint='运行 build_subject_intelligence 命令（Track A3）',
        )

    if created > 0:
        report.add_agent_opportunity(
            scenario='新注册受试者自动初始画像生成',
            current_pain=f'本次导入新建了 {created:,} 名受试者，但无项目参与历史，'
                        f'下次招募时项目经理需要人工逐一判断适合度',
            agent_value='基于省份/年龄/性别/肤质，智能体可自动推测适合的项目类型并预标注',
            data_evidence=f'{created:,} 名新受试者需要初始画像',
            implementation_hint='基于 KnowledgeEntry 中的历史项目画像知识 + LLM 分类器',
        )

    # 发布学习报告
    log.info("[学习型] 开始发布学习报告...")
    log.info(report.summary())

    from apps.data_intake.learning_runner import GapReporter
    gap_reporter = GapReporter(dry_run=dry_run)

    # 写入知识库（需要 Django 环境）
    try:
        from apps.knowledge.ingestion_pipeline import run_pipeline, RawKnowledgeInput
        import django
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
        django.setup()

        content_lines = [f'# NAS 受试者导入学习报告（{datetime.date.today()}）\n\n']
        content_lines.append(report.summary())
        for p in report.patterns_discovered:
            content_lines.append(f'\n## {p.title}\n{p.description}\n')
        for mf in report.match_failures:
            content_lines.append(f'\n## 匹配失败：{mf.failure_reason}\n'
                                  f'影响 {mf.count:,} 条（{mf.percentage}%）。建议：{mf.suggested_fix}\n')
        for opp in report.agent_opportunities:
            content_lines.append(f'\n## 智能体机会：{opp.scenario}\n{opp.current_pain}\n')

        result = run_pipeline(RawKnowledgeInput(
            title=f'[导入学习报告] NAS 受试者综合导入 — {datetime.date.today()}',
            content=''.join(content_lines),
            source_type='import_learning',
            source_key=f'nas_comprehensive_subject_{datetime.date.today()}',
            entry_type='lesson_learned',
            namespace='project_experience',
        ))
        if result and result.entry_id:
            log.info("[学习型] 学习报告已写入 KnowledgeEntry: #%s", result.entry_id)
    except Exception as e:
        log.warning("[学习型] 写入 KnowledgeEntry 失败（需要 Django 环境）: %s", e)

    # 创建 GitHub Issues
    gap_result = gap_reporter.report(report)
    log.info(
        "[学习型] GitHub Issues: %d 创建 | ProactiveInsights: %d 创建 | 跳过: %d 项",
        gap_result['github_issues'], gap_result['proactive_insights'], gap_result['skipped'],
    )


# ═══════════════════════════════════════════════════════════════════════════
# 最终统计
# ═══════════════════════════════════════════════════════════════════════════
def final_stats(conn):
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM t_subject")
    total_sub = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject_profile WHERE province != '' AND province IS NOT NULL")
    has_province = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject_profile WHERE address != '' AND address IS NOT NULL")
    has_addr = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject_profile WHERE id_card_hash IS NOT NULL AND id_card_hash != ''")
    has_ic = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject WHERE gender != '' AND gender IS NOT NULL")
    has_gender = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject WHERE skin_type != '' AND skin_type IS NOT NULL")
    has_skin = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM t_subject WHERE status = 'blacklisted'")
    blacklisted = cur.fetchone()[0]

    cur.execute("SELECT questionnaire_type, COUNT(*) FROM t_subject_questionnaire GROUP BY questionnaire_type ORDER BY count DESC")
    qtypes = cur.fetchall()

    cur.close()
    log.info("=" * 60)
    log.info(f"最终受试者总数:    {total_sub:,}")
    log.info(f"  有身份证哈希:    {has_ic:,}  ({has_ic/total_sub*100:.1f}%)")
    log.info(f"  有性别:          {has_gender:,}  ({has_gender/total_sub*100:.1f}%)")
    log.info(f"  有肤质:          {has_skin:,}  ({has_skin/total_sub*100:.1f}%)")
    log.info(f"  有省份:          {has_province:,}  ({has_province/total_sub*100:.1f}%)")
    log.info(f"  有住址:          {has_addr:,}  ({has_addr/total_sub*100:.1f}%)")
    log.info(f"  黑名单:          {blacklisted:,}")
    log.info("问卷记录分布:")
    for qt, cnt in qtypes:
        log.info(f"  {qt}: {cnt:,}")


# ═══════════════════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════════════════
def main():
    import argparse
    parser = argparse.ArgumentParser(description='NAS 受试者名单综合导入（学习型版本）')
    parser.add_argument('--dry-run', action='store_true', help='不写入数据库，仅分析生成学习报告')
    parser.add_argument('--skip-learning', action='store_true', help='跳过学习报告生成，只运行 ETL')
    parser.add_argument('--db-host', default='localhost')
    parser.add_argument('--db-port', type=int, default=25432)
    parser.add_argument('--db-name', default='cn_kis_v2')
    parser.add_argument('--db-user', default='cn_kis')
    parser.add_argument('--db-pass', default='cn_kis_2026')
    args = parser.parse_args()

    # 更新数据库配置
    DB_CONFIG.update(
        host=args.db_host, port=args.db_port,
        dbname=args.db_name, user=args.db_user, password=args.db_pass,
    )

    if args.dry_run:
        log.info("[DRY-RUN] 模式启动，不写入数据库")
        generate_learning_report(dry_run=True)
        return

    log.info("连接数据库...")
    conn = get_conn()
    try:
        conn.autocommit = False
        by_id, by_phone, by_hash = load_existing_subjects(conn)

        log.info(f"\n{'='*60}")
        log.info("开始综合导入（3个Phase）")

        phase1_idcard_exports(conn, by_id, by_phone, by_hash)
        phase2_master_list(conn, by_id, by_phone, by_hash)
        phase3_subject_list(conn, by_id, by_phone, by_hash)

        final_stats(conn)
        conn.commit()
        log.info("所有Phase完成，已提交。")
    except Exception as e:
        conn.rollback()
        log.error(f"发生错误，已回滚: {e}", exc_info=True)
        raise
    finally:
        conn.close()

    # ── 学习报告（ETL 完成后）────────────────────────────────────────────
    if not args.skip_learning:
        generate_learning_report(dry_run=False)
    else:
        log.info("[学习型] --skip-learning 已设置，跳过学习报告生成")

if __name__ == '__main__':
    main()
